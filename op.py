# -*- coding: utf-8 -*-
# VERSION V4.1 - Correction finale appel fonction save_kpis_to_excel
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(layout="wide", page_title="Dashboard KPI V4.1")

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT COR_EGAL"]
ALL_KPI = QK + PK
CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT COR_EGAL":100}
ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois."}
LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]
MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    filepath="KPI_Historique.xlsx"
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F"); tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    return df

def compute_variance(hist_df):
    if hist_df.empty: return pd.DataFrame(), "", ""
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame(), "", ""
    prev_date=dates[-2]; curr_date=dates[-1]
    perf_df=hist_df[hist_df["_section"]=="perf"].copy(); qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    pp=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
    cp=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
    pq=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
    cq=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
    common=list(set(pp.index)&set(cp.index)&set(pq.index)&set(cq.index))
    if not common: return pd.DataFrame(), prev_date, curr_date
    rows=[]
    for p in common:
        sp_prev=float(pp.loc[p,"Score Performance"]) if "Score Performance" in pp.columns and p in pp.index else np.nan
        sp_curr=float(cp.loc[p,"Score Performance"]) if "Score Performance" in cp.columns and p in cp.index else np.nan
        sq_prev=float(pq.loc[p,"Score Qualite"]) if "Score Qualite" in pq.columns and p in pq.index else np.nan
        sq_curr=float(cq.loc[p,"Score Qualite"]) if "Score Qualite" in cq.columns and p in cq.index else np.nan
        sp_diff=sp_curr-sp_prev if not pd.isna(sp_prev) and not pd.isna(sp_curr) else np.nan
        sp_pct=(sp_diff/sp_prev*100) if not pd.isna(sp_prev) and sp_prev!=0 else np.nan
        sq_diff=sq_curr-sq_prev if not pd.isna(sq_prev) and not pd.isna(sq_curr) else np.nan
        sq_pct=(sq_diff/sq_prev*100) if not pd.isna(sq_prev) and sq_prev!=0 else np.nan
        def gt(d): return "🟢 Amelioration" if not pd.isna(d) and d>0.5 else ("🔴 Degradation" if not pd.isna(d) and d<-0.5 else "🟡 Stable")
        rows.append({"Poste de travail":p,"Perf Prec.":sp_prev,"Perf Act.":sp_curr,"Ecart Perf":sp_diff,"Ecart % Perf":sp_pct,"Tendance Perf":gt(sp_diff),"Qual Prec.":sq_prev,"Qual Act.":sq_curr,"Ecart Qual":sq_diff,"Ecart % Qual":sq_pct,"Tendance Qual":gt(sq_diff)})
    return pd.DataFrame(rows), prev_date, curr_date

def html_variance_table(var_df):
    if var_df.empty: return ""
    h='<table class="tw vt"><thead><tr><th>Poste de travail</th><th>Perf Prec.</th><th>Perf Act.</th><th>Ecart</th><th>%</th><th>Tendance</th><th>Qual Prec.</th><th>Qual Act.</th><th>Ecart</th><th>%</th><th>Tendance</th></tr></thead><tbody>'
    for _,r in var_df.iterrows():
        h+='<tr><td class="poste-cell">%s</td>'%r["Poste de travail"]
        for b in ["Perf","Qual"]:
            h+='<td>%.1f%%</td>'%r[f"{b} Prec."] if not pd.isna(r[f"{b} Prec."]) else '<td>—</td>'
            h+='<td>%.1f%%</td>'%r[f"{b} Act."] if not pd.isna(r[f"{b} Act."]) else '<td>—</td>'
            d=r[f"Ecart {b}"]
            if not pd.isna(d):
                c="#276749" if d>0.5 else ("#c53030" if d<-0.5 else "#975a16")
                h+='<td style="color:%s;font-weight:700">%+.1f</td>'%(c,d)
            else: h+='<td>—</td>'
            p=r[f"Ecart % {b}"]
            if not pd.isna(p):
                c="#276749" if p>0.5 else ("#c53030" if p<-0.5 else "#975a16")
                h+='<td style="color:%s;font-weight:700">%+.1f%%</td>'%(c,p)
            else: h+='<td>—</td>'
            h+='<td style="text-align:center;font-size:14px">%s</td>'%r[f"Tendance {b}"]
        h+='</tr>'
    return h+'</tbody></table>'

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy(); qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date=dates[i-1]; curr_date=dates[i]
        pp=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        cp_d=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        pq=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        cq=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,pd_d,cd_d,kl in [("Performance",pp,cp_d,QK+["Score Performance"]),("Qualite",pq,cq,PK+["Score Qualite"])]:
            for poste in set(pd_d.index)&set(cd_d.index):
                for kpi in kl:
                    if kpi not in pd_d.columns or kpi not in cd_d.columns: continue
                    try: pv=float(pd_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(cd_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=diff/pv*100 if pv!=0 else (100 if cv!=0 else 0)
                    trend="stabilite" if abs(diff)<=0.5 else ("hausse" if diff>0.5 else "baisse")
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,"Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),"Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5; j=j[j["Significatif"]].copy()
    def sf(r):
        if r["Tendance"]=="hausse" and r["KPI"] not in LOWER_BETTER: return "Amelioration"
        elif r["Tendance"]=="baisse" and r["KPI"] in LOWER_BETTER: return "Amelioration"
        else: return "Degradation"
    j["Sens"]=j.apply(sf,axis=1)
    return j.sort_values(["Date actuelle","Sens","Ecart %"],ascending=[True,False,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy(); s=0
        for _,r in pv.iterrows(): s=s+(-r["Ecart %"]) if r["KPI"] in LOWER_BETTER else s+r["Ecart %"]
        scores[poste]=s
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    top=pd.DataFrame(ranked[:5],columns=["Poste","Score variation"])
    bot=pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"]) if len(ranked)>5 else pd.DataFrame(columns=["Poste","Score variation"])
    return top,bot

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:10px 0 4px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}
    .stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}.stl.om{border-left-color:#2f855a}
    .stl.th{border-left-color:#2b6cb0}.stl.bp{border-left-color:#6b46c1}.stl.be{border-left-color:#c05621}.stl.v{border-left-color:#553c9a}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.atp thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.atq thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.vt thead th{background:linear-gradient(135deg,#553c9a,#805ad5)}
    .tw.bt-oms thead th{background:linear-gradient(135deg,#276749,#2f855a)}
    .tw.bt-th thead th{background:linear-gradient(135deg,#2b6cb0,#4299e1)}
    .tw.bt-bp thead th{background:linear-gradient(135deg,#553c9a,#805ad5)}
    .tw.bt-be thead th{background:linear-gradient(135deg,#9c4221,#dd6b20)}
    .tw .poste-cell{background:#f7fafc;font-weight:700;color:#1a202c;white-space:nowrap;min-width:200px;max-width:220px;position:sticky;left:0;z-index:5;border-right:2px solid #e2e8f0}
    .tw tbody tr:nth-child(even) .poste-cell{background:#edf2f7}
    .tw tbody tr:hover .poste-cell{background:#dbeafe!important}
    .tw tbody td{padding:5px 8px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important}
    .tr .poste-cell{background:#e2e8f0!important;color:#1a202c!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}.gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}.gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}.gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:5px}.gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}.cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}.cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .sec-sep{height:1px;background:linear-gradient(90deg,transparent,#cbd5e0,transparent);margin:12px 0}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg{grid-template-columns:1fr}}
    </style>""",unsafe_allow_html=True)

def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass

    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False

    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def contient_mot(t,lm):
        t=str(t)
        for l in lm:
            for m in l.split():
                if m in t: return True
        return False

    def cat_age(a):
        if a<=1: return "<1 mois"
        elif a>=3: return ">3 mois"
        return "1 mois < <3 mois"

    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)

    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)

    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
        return df

    def is_lb(k): return k in LOWER_BETTER

    def std_cols(df):
        df.columns = [str(c).strip() for c in df.columns]
        mp = {"Statut_ot":"Statut OT","StatutOT":"Statut OT","Poste de travail princ.":"Poste travail princ.",
              "N° appel pl.entret.":"Nº appel pl.entret.","Date de debut planifiee":"Date de début planifiée",
              "Total couts budgétés":"Total coûts budgétés","Total couts reels":"Total coûts réels",
              "Statut systeme":"Statut système","Debut reel":"Début réel","Fin reelle":"Fin réelle","N° avis":"Avis","Nº avis":"Avis"}
        df.columns = [mp.get(c, c) for c in df.columns]
        return df

    def map_statut_ot(s):
        if pd.isna(s): return "AUTRE"
        u = str(s).upper().strip()
        if "CLRE" in u: return "CLOT"
        if "TECO" in u: return "TCLO"
        if "REL" in u or "PRRT" in u: return "LANC"
        if "CRTD" in u: return "CRÉÉ"
        return "AUTRE"

    def calc_kpis(df_i,av_i,now,posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
        df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
        for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc]=pd.to_datetime(df[dc],errors='coerce')
                df[am]=((now.year-df[dc].dt.year)*12+(now.month-df[dc].dt.month)).round(2)
                df[ac]=df[am].apply(cat_age)
            else: df[am]=np.nan; df[ac]="Inconnu"
        df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
        df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
        df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
        df["OT COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
        res['dfp']=df

        an=cpiv(df,df["Nº appel pl.entret."].fillna(0)==0,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["Total"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=ckpi(an["TCLO"],an["Total"])

        pr=cpiv(df,df["Statut OT"]=="CRÉÉ","ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"])
        pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0)
        pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)

        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"])
        pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0)
        pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)

        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"])
        ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0)
        ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)

        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])

        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])

        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])

        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT COR_EGAL","OT COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"])
            res[kn.lower().replace(" ","_")]=pv

        avf=av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip()=="")].copy()
        res['avf']=avf

        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1)
        tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])

        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],
            "OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],
            "OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],
            "OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],
            "Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT COR_EGAL":res['ot_cor_egal']["OT COR_EGAL"]
        })
        return res

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            if val>=80: return "background:#c6efce;color:#006100;font-weight:600"
            elif val>=75: return "background:#ffeb9c;color:#9c6500;font-weight:600"
            else: return "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            if val>=85: return "background:#c6efce;color:#006100;font-weight:600"
            elif val>=80: return "background:#ffeb9c;color:#9c6500;font-weight:600"
            else: return "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="appel avis approuvé":
            if val>=95: return "background:#c6efce;color:#006100;font-weight:600"
            elif val>=90: return "background:#ffeb9c;color:#9c6500;font-weight:600"
            else: return "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT COR_EGAL"]:
            if val>=100: return "background:#c6efce;color:#006100;font-weight:600"
            elif val>=95: return "background:#ffeb9c;color:#9c6500;font-weight:600"
            else: return "background:#ffc7ce;color:#9c0006;font-weight:600"
        return ""

    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        if val>=90: return "background:#c6efce;color:#006100;font-weight:700"
        elif val>=80: return "background:#ffeb9c;color:#9c6500;font-weight:700"
        else: return "background:#ffc7ce;color:#9c0006;font-weight:700"

    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT COR_EGAL"]: return 1 if a>=95 else 0
        return 0

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                else:
                    s=cs(v) if sc_col and c in sc_col else ks(v,c)
                    h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'

    def html_ano_transposed(ano_rows,kpi_list,css_class,cell_class):
        matrix={}
        for r in ano_rows:
            if r.get("_t")=="total": continue
            kpi=r.get("KPI",""); poste=r.get("Poste de travail","")
            val=r.get("Valeur",0); ecart=r.get("Ecart",0); cible=r.get("Cible",0)
            if kpi not in matrix: matrix[kpi]={}
            matrix[kpi][poste]={"val":val,"ecart":ecart,"cible":cible}
        if not matrix: return ""
        ordered_kpis=[k for k in kpi_list if k in matrix]
        extra_kpis=sorted([k for k in matrix.keys() if k not in ordered_kpis])
        all_kpis=ordered_kpis+extra_kpis
        all_postes=sorted(set(p for d in matrix.values() for p in d.keys()))
        h='<table class="tw %s"><thead><tr><th style="min-width:200px;text-align:left">KPI</th>'%css_class
        for p in all_postes: h+='<th style="min-width:90px">%s</th>'%p
        h+='<th style="min-width:60px">Nb</th></tr></thead><tbody>'
        grand_total=0
        for kpi in all_kpis:
            h+='<tr><td class="%s">%s</td>'%(cell_class,kpi); row_count=0
            for p in all_postes:
                if p in matrix[kpi]:
                    d=matrix[kpi][p]; ev=d["ecart"]; vl=d["val"]; cb=d["cible"]; row_count+=1
                    try: ae=abs(float(ev))
                    except: ae=0
                    if ae>10: s="background:#fed7d7;color:#9b2c2c;font-weight:700"
                    elif ae>5: s="background:#ffeb9c;color:#975a16;font-weight:600"
                    else: s="background:#feebc8;color:#9c4221;font-weight:600"
                    h+='<td style="%s" title="Val: %.1f%% | Cible: %.0f%%">%.1f<br><span style="font-size:10px;opacity:.8">(%+.1f)</span></td>'%(s,vl,cb,vl,ev)
                else: h+='<td style="color:#cbd5e0">—</td>'
            grand_total+=row_count
            h+='<td style="font-weight:800;text-align:center;%s">%d</td>'%("color:#c53030" if row_count>3 else "color:#4a5568",row_count)+'</tr>'
        h+='<tr class="tr"><td class="%s">TOTAL</td>'%cell_class
        for p in all_postes:
            cnt=sum(1 for kpi in all_kpis if p in matrix.get(kpi,{}))
            h+='<td style="text-align:center;font-weight:800">%d</td>'%cnt
        h+='<td style="text-align:center;font-weight:900;font-size:14px">%d</td></tr>'%grand_total
        return h+'</tbody></table>'

    def html_backlog_pivot(pivot_df, css_class, pct_col_name=None, pct_values=None):
        if pivot_df.empty: return '<div class="es">Aucune donnee</div>'
        cats = list(pivot_df.columns); all_cols = ["Poste de travail"] + cats
        if pct_col_name: all_cols.append(pct_col_name)
        all_cols.append("Total")
        h = '<table class="tw %s"><thead><tr>' % css_class
        for c in all_cols: h += '<th>%s</th>' % c
        h += '</tr></thead><tbody>'
        grand_totals = {c: int(pivot_df[c].sum()) if c in pivot_df.columns else 0 for c in cats}
        grand_total_all = sum(grand_totals.values())
        for idx, poste in enumerate(pivot_df.index):
            h += '<tr><td class="poste-cell">%s</td>' % poste; row_total = 0
            for c in cats:
                v = int(pivot_df.loc[poste, c]) if c in pivot_df.columns else 0
                h += '<td style="text-align:center">%d</td>' % v; row_total += v
            if pct_col_name and pct_values and idx < len(pct_values):
                pv = pct_values[idx]
                try: pf = float(str(pv).replace('%','').strip()); h += '<td style="text-align:center;%s">%s</td>' % (cs("%.1f" % pf), pv)
                except: h += '<td style="text-align:center">%s</td>' % pv
            h += '<td style="text-align:center;font-weight:800;background:#edf2f7">%d</td>' % row_total + '</tr>'
        h += '<tr class="tr"><td class="poste-cell">Total</td>'
        for c in cats: h += '<td style="text-align:center">%d</td>' % grand_totals[c]
        if pct_col_name:
            if pct_values:
                try: avg_pct = round(np.mean([float(str(v).replace('%','').strip()) for v in pct_values if str(v).replace('%','').strip().lstrip('-').isdigit()]), 1)
                except: avg_pct = 0
                h += '<td style="text-align:center;%s">%s%%</td>' % (cs("%.1f" % avg_pct), avg_pct)
            else: h += '<td style="text-align:center">—</td>'
        h += '<td style="text-align:center;font-weight:900;font-size:14px">%d</td>' % grand_total_all + '</tr>'
        return h + '</tbody></table>'

    def html_actions_table(kpi_list,actuals,targets,act_map):
        h='<table class="tw at"><thead><tr><th>KPI</th><th>Valeur</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action</th></tr></thead><tbody>'
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            st_s="background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
            ec_clr="#276749" if met else "#c53030"; action="Objectif atteint" if met else act_map.get(k,"")
            h+='<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td></tr>'%(k,av,tv,ec_clr,diff,st_s,"ATTEINT" if met else "NON ATTEINT",action)
        return h+'</tbody></table>'

    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'
        return h

    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        for p in sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True):
            pv=pscores.get(p,0); qv=qscores.get(p,0)
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div></div>'%(p,min(max(pv,0),100),pv,min(max(qv,0),100),qv)
        return h+'</div>'

    def filter_ot_by_keyword(df, keyword):
        mask = pd.Series(False, index=df.index); kw = str(keyword).upper().strip()
        for col in df.columns:
            if df[col].dtype == 'object': mask = mask | df[col].astype(str).str.upper().str.contains(kw, na=False)
        return df[mask]

    def anl_pie_chart(data,names_col,values_col,title,colors=None,min_pct=3.0):
        if data.empty: return None
        df=data[[names_col,values_col]].dropna().copy()
        df[values_col]=pd.to_numeric(df[values_col],errors='coerce').fillna(0)
        total=df[values_col].sum()
        if total==0: return None
        df["_pct"]=df[values_col]/total*100
        big=df[df["_pct"]>=min_pct].copy(); small=df[df["_pct"]<min_pct].copy()
        has_small=len(small)>=1 and small[values_col].sum()>0
        if not has_small:
            fig=px.pie(df,names=names_col,values=values_col,title="<b>%s</b>"%title,color_discrete_sequence=colors or px.colors.qualitative.Set2)
            fig.update_traces(textposition='inside',textinfo='percent+label+value',textfont_size=12,pull=[0.02]*len(df))
            fig.update_layout(margin=dict(t=60,b=50,l=20,r=20),height=440,autosize=True,title_font_size=14,legend=dict(font_size=11,orientation="h",yanchor="bottom",y=-0.12))
            return fig
        else:
            others_label="Autres (%d)"%len(small); others_row=pd.DataFrame([{names_col:others_label,values_col:small[values_col].sum(),"_pct":small["_pct"].sum()}])
            main_df=pd.concat([big,others_row],ignore_index=True); sub_df=small.sort_values(values_col,ascending=False).copy()
            base_colors=colors or px.colors.qualitative.Set2; main_colors=[]
            for i in range(len(main_df)): main_colors.append("#CBD5E0" if i==len(main_df)-1 else base_colors[i%len(base_colors)])
            sub_colors=[base_colors[(len(big)+i)%len(base_colors)] for i in range(len(sub_df))]
            fig=make_subplots(rows=1,cols=2,specs=[[{"type":"pie"},{"type":"pie"}]],subplot_titles=["<b>%s</b>"%title,"<b>Detail 'Autres' (%d)</b>"%len(small)],horizontal_spacing=0.08)
            fig.add_trace(go.Pie(labels=main_df[names_col].tolist(),values=main_df[values_col].tolist(),textinfo='percent+label+value',textposition='inside',textfont_size=12,marker_colors=main_colors,pull=[0.03 if i==len(main_df)-1 else 0.01 for i in range(len(main_df))]),row=1,col=1)
            fig.add_trace(go.Pie(labels=sub_df[names_col].tolist(),values=sub_df[values_col].tolist(),textinfo='percent+label+value',textposition='inside',textfont_size=11,marker_colors=sub_colors,hole=0.3),row=1,col=2)
            fig.update_layout(margin=dict(t=60,b=50,l=10,r=10),height=440,autosize=True,title_font_size=14,showlegend=True,legend=dict(font_size=10,orientation="h",yanchor="bottom",y=-0.06))
            return fig

    def make_rank_bar_chart(top_df,bottom_df,kpi_name,target_val,color_top,color_bottom):
        all_items=[]
        if not top_df.empty:
            for _,r in top_df.iterrows(): all_items.append({"Poste":r["Poste"],"Score":r["Score"],"Groupe":"Top 5"})
        if not bottom_df.empty:
            for _,r in bottom_df.iterrows(): all_items.append({"Poste":r["Poste"],"Score":r["Score"],"Groupe":"Bottom 5"})
        if not all_items: return None
        rdf=pd.DataFrame(all_items).sort_values("Score",ascending=True)
        bar_colors=[color_top if g=="Top 5" else color_bottom for g in rdf["Groupe"]]
        fig=go.Figure()
        fig.add_trace(go.Bar(y=rdf["Poste"],x=rdf["Score"],orientation='h',marker_color=bar_colors,text=["%.1f"%s for s in rdf["Score"]],textposition='outside',textfont=dict(size=12,color="#1a202c",family="Inter"),hovertemplate="<b>%{y}</b><br>Score: %{x:.1f}<extra></extra>"))
        fig.add_vline(x=target_val,line_dash="dash",line_width=2.5,line_color="#e53e3e",annotation_text="Objectif: %d"%target_val if target_val!=0 else "Seuil: 0",annotation_position="top right",annotation_font=dict(size=12,color="#e53e3e",family="Inter",weight="bold"),annotation_bgcolor="rgba(255,255,255,0.85)",annotation_bordercolor="#e53e3e",annotation_borderwidth=1)
        fig.update_layout(title="<b>%s</b>"%kpi_name,height=max(220,len(rdf)*48+100),margin=dict(l=180,r=80,t=60,b=30),xaxis=dict(range=[min(-10,rdf["Score"].min()*1.15),max(10,rdf["Score"].max()*1.15)],title="Variance (points)",gridcolor="#edf2f7",zeroline=True),yaxis=dict(tickfont=dict(size=12,family="Inter")),plot_bgcolor="white",font=dict(family="Inter"),showlegend=False,bargap=0.35)
        return fig

    def make_kpi_bar_chart_with_target(kpi_list,actuals,targets,title,color_ok,color_fail):
        names=[]; vals=[]; colors=[]; cibles=[]
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); met=av<=tv if is_lb(k) else av>=tv
            names.append(k); vals.append(round(av,1)); colors.append(color_ok if met else color_fail); cibles.append(tv)
        fig=go.Figure()
        fig.add_trace(go.Bar(x=names,y=vals,marker_color=colors,text=["%.1f%%"%v for v in vals],textposition='outside',textfont=dict(size=11,color="#1a202c",family="Inter"),hovertemplate="<b>%{x}</b><br>Valeur: %{y:.1f}%%<extra></extra>",name="Valeur"))
        for i,(n,tv) in enumerate(zip(names,cibles)):
            fig.add_shape(type="line", x0=i-0.4, x1=i+0.4, y0=tv, y1=tv, line=dict(color="#e53e3e", width=2.5, dash="dash")))
            fig.add_annotation(x=i,y=tv,text="Cible %d%%"%tv,showarrow=False,yshift=8,font=dict(size=9,color="#e53e3e",family="Inter",weight="bold"),bgcolor="rgba(255,255,255,0.8)",bordercolor="#e53e3e",borderwidth=0.5)
        fig.update_layout(title="<b>%s</b>"%title,height=420,margin=dict(l=20,r=20,t=60,b=140),xaxis=dict(tickangle=-45,tickfont=dict(size=10,family="Inter")),yaxis=dict(range=[0,max(max(vals),max(cibles))*1.2],title="%",gridcolor="#edf2f7"),plot_bgcolor="white",font=dict(family="Inter"),showlegend=False,bargap=0.3)
        return fig

    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div></div>""",unsafe_allow_html=True)
        st.markdown("---")
        show_filters=st.checkbox("Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf"); ot_f=None; av_f=None; apm=[]
            if unf: ot_f=st.file_uploader("Fichier OT",type=["xlsx"],key="uot"); av_f=st.file_uploader("Fichier AVIS",type=["xlsx"],key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try: _t=excr(pd.read_excel("ot.xlsx")); apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**"); sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            st.markdown("**🏭 Atelier**"); sa=st.multiselect("Atelier",["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"],["All"],key="sa")
            st.markdown("**🏢 Division**"); sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=None; av_f=None; apm=[]; sp=["All"]; sa=["All"]; sd=["All"]; dr=(datetime(2025,1,1).date(),datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try: _t=excr(pd.read_excel("ot.xlsx")); apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception: pass

    # ===================== DATA LOADING =====================
    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf: raw_ot=pd.read_excel(ot_f); raw_av=pd.read_excel(av_f)
            else: raw_ot=pd.read_excel("ot.xlsx"); raw_av=pd.read_excel("avis.xlsx")
            raw_ot = std_cols(raw_ot); raw_av = std_cols(raw_av)
            if "Statut OT" not in raw_ot.columns:
                if "Statut système" in raw_ot.columns: raw_ot["Statut OT"] = raw_ot["Statut système"].apply(map_statut_ot)
                else: raw_ot["Statut OT"] = "CRÉÉ"
            if "Avis" not in raw_av.columns:
                for c in raw_av.columns:
                    if "avis" in c.lower(): raw_av["Avis"] = raw_av[c]; break
                else: raw_av["Avis"] = raw_av.index
            raw_ot=excr(raw_ot); raw_av=excr(raw_av)
            for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
            for c in ["Créé le","Début souhaité","Date de la clôture"]:
                if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
            if not apm: apm=sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp_sel=apm
            else: sp_sel=[p for p in sp if p in apm]
            df_ot=raw_ot[raw_ot["Poste travail princ."].isin(sp_sel)].copy(); df_av=raw_av[raw_av["Poste travail princ."].isin(sp_sel)].copy()
            if sa and "All" not in sa:
                afilter="|".join(sa)
                if "Poste travail princ." in df_ot.columns: df_ot=df_ot[df_ot["Poste travail princ."].str.contains(afilter,case=False,na=False)]
                if "Poste travail princ." in df_av.columns: df_av=df_av[df_av["Poste travail princ."].str.contains(afilter,case=False,na=False)]
            if sd and "All" not in sd:
                dfilter="|".join(sd)
                if "Poste travail princ." in df_ot.columns: df_ot=df_ot[df_ot["Poste travail princ."].str.contains(dfilter,case=False,na=False)]
                if "Poste travail princ." in df_av.columns: df_av=df_av[df_av["Poste travail princ."].str.contains(dfilter,case=False,na=False)]
            if dr and len(dr)==2:
                d1=pd.Timestamp(dr[0]); d2=pd.Timestamp(dr[1])
                if "Date de début planifiée" in df_ot.columns: df_ot=df_ot[(df_ot["Date de début planifiée"].isna())|(df_ot["Date de début planifiée"]>=d1)&(df_ot["Date de début planifiée"]<=d2)]
            posts=sorted(sp_sel); now=datetime.now(); kp=calc_kpis(df_ot,df_av,now,posts); ckdf=kp['ckdf']; dfp=kp['dfp']; avf=kp['avf']
            perf_scores={}; qual_scores={}
            for p in posts:
                if p in ckdf.index:
                    ps=0; qs=0
                    for k in QK: a=ckdf.loc[p,k] if k in ckdf.columns else 0; t=CIBLE.get(k,100); ps+=gscore(k,a,t)
                    for k in PK: a=ckdf.loc[p,k] if k in ckdf.columns else 0; t=CIBLE.get(k,100); qs+=gscore(k,a,t)
                    perf_scores[p]=ps/len(QK)*100 if QK else 0; qual_scores[p]=qs/len(PK)*100 if PK else 0
                else: perf_scores[p]=0; qual_scores[p]=0
            tot_ot=len(dfp); tot_av=len(avf); avg_p=np.mean(list(perf_scores.values())) if perf_scores else 0; avg_q=np.mean(list(qual_scores.values())) if qual_scores else 0
            global_actuals={k:round(ckdf[k].mean(),1) if k in ckdf.columns else 0 for k in ALL_KPI}
            perf_ranked=sorted(perf_scores.items(),key=lambda x:x[1],reverse=True)
            top5_perf=pd.DataFrame(perf_ranked[:5],columns=["Poste","Score"]); bot5_perf=pd.DataFrame(perf_ranked[-5:][::-1],columns=["Poste","Score"]) if len(perf_ranked)>5 else pd.DataFrame(columns=["Poste","Score"])
            qual_ranked=sorted(qual_scores.items(),key=lambda x:x[1],reverse=True)
            top5_qual=pd.DataFrame(qual_ranked[:5],columns=["Poste","Score"]); bot5_qual=pd.DataFrame(qual_ranked[-5:][::-1],columns=["Poste","Score"]) if len(qual_ranked)>5 else pd.DataFrame(columns=["Poste","Score"])
            ot_statut_df=dfp.groupby("Statut OT",dropna=False).size().reset_index(name="Nombre"); ot_statut_df["Statut OT"]=ot_statut_df["Statut OT"].fillna("INCONNU")
            avis_statut_df=avf.groupby("Statut utilisateur",dropna=False).size().reset_index(name="Nombre"); avis_statut_df["Statut utilisateur"]=avis_statut_df["Statut utilisateur"].fillna("INCONNU")
            ano_p_rows=[]; ano_q_rows=[]
            for p in posts:
                if p not in ckdf.index: continue
                for k in QK:
                    v=ckdf.loc[p,k]; t=CIBLE.get(k,100)
                    if not (v<=t if is_lb(k) else v>=t): ano_p_rows.append({"Poste de travail":p,"KPI":k,"Valeur":round(v,1),"Cible":t,"Ecart":round(v-t,1)})
                for k in PK:
                    v=ckdf.loc[p,k]; t=CIBLE.get(k,100)
                    if not (v<=t if is_lb(k) else v>=t): ano_q_rows.append({"Poste de travail":p,"KPI":k,"Valeur":round(v,1),"Cible":t,"Ecart":round(v-t,1)})
            tot_ano_p=len(ano_p_rows); tot_ano_q=len(ano_q_rows)
            ano_p_excel=ano_p_rows.copy(); ano_p_excel.append({"Poste de travail":"Total","KPI":"","Valeur":"","Cible":"","Ecart":tot_ano_p,"_t":"total"})
            ano_q_excel=ano_q_rows.copy(); ano_q_excel.append({"Poste de travail":"Total","KPI":"","Valeur":"","Cible":"","Ecart":tot_ano_q,"_t":"total"})
            pcols=["Poste de travail"]+QK+["Score Performance"]; prows=[]
            for p in posts:
                if p in ckdf.index:
                    r={"Poste de travail":p}
                    for k in QK: r[k]=round(ckdf.loc[p,k],1) if k in ckdf.columns else 0
                    r["Score Performance"]=round(perf_scores[p],1); prows.append(r)
            prows.append({"Poste de travail":"CIBLE","_t":"cible"})
            for k in QK: prows[-1][k]=CIBLE.get(k,"")
            prows[-1]["Score Performance"]="≥80%"
            prows.append({"Poste de travail":"Moyenne","_t":"total"})
            for k in QK: prows[-1][k]=round(ckdf[k].mean(),1) if k in ckdf.columns else 0
            prows[-1]["Score Performance"]=round(avg_p,1)
            qcols=["Poste de travail"]+PK+["Score Qualite"]; qrows=[]
            for p in posts:
                if p in ckdf.index:
                    r={"Poste de travail":p}
                    for k in PK: r[k]=round(ckdf.loc[p,k],1) if k in ckdf.columns else 0
                    r["Score Qualite"]=round(qual_scores[p],1); qrows.append(r)
            qrows.append({"Poste de travail":"CIBLE","_t":"cible"})
            for k in PK: qrows[-1][k]=CIBLE.get(k,"")
            qrows[-1]["Score Qualite"]="≥80%"
            qrows.append({"Poste de travail":"Moyenne","_t":"total"})
            for k in PK: qrows[-1][k]=round(ckdf[k].mean(),1) if k in ckdf.columns else 0
            qrows[-1]["Score Qualite"]=round(avg_q,1)
            # CORRECTION FINALE ICI : ano_p_cols devient ano_p_c
            save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_excel,ano_p_c=["Poste de travail","KPI","Valeur","Cible","Ecart"],ano_q_r=ano_q_excel,ano_q_c=["Poste de travail","KPI","Valeur","Cible","Ecart"],sheet_name=fichier_date)
            df_oms = filter_ot_by_keyword(dfp, "OMS"); oms_pivot = pd.DataFrame()
            if not df_oms.empty and "Statut OT" in df_oms.columns:
                oms_pivot = pd.pivot_table(df_oms, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0).loc[oms_pivot.sum(axis=1) > 0]
            df_thermo = filter_ot_by_keyword(dfp, "THERMOGRAPHIE"); thermo_pivot = pd.DataFrame()
            if not df_thermo.empty and "Statut OT" in df_thermo.columns:
                thermo_pivot = pd.pivot_table(df_thermo, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0).loc[thermo_pivot.sum(axis=1) > 0]
            df_cre = dfp[dfp["Statut OT"] == "CRÉÉ"] if "Statut OT" in dfp.columns else pd.DataFrame(); bp_pivot = pd.DataFrame(); bp_pct = []
            if not df_cre.empty:
                bp_pivot = pd.pivot_table(df_cre, index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
                for c in ["CARACTERISE", "NON CARACTERISE"]: bp_pivot[c] = bp_pivot.get(c, 0)
                bp_pivot["Total"] = bp_pivot["CARACTERISE"] + bp_pivot["NON CARACTERISE"]; bp_pivot = bp_pivot.loc[bp_pivot["Total"] > 0]
                for p in bp_pivot.index:
                    t = bp_pivot.loc[p, "Total"]; bp_pct.append(round(bp_pivot.loc[p, "CARACTERISE"] / t * 100, 1) if t > 0 else "0.0%")
            df_exec = dfp[(dfp["Statut OT"] == "LANC") & (dfp["Contient SOPL"] == 1)] if "Statut OT" in dfp.columns else pd.DataFrame(); be_pivot = pd.DataFrame()
            if not df_exec.empty and "aex" in df_exec.columns:
                be_pivot = pd.pivot_table(df_exec, index="Poste travail princ.", columns="aex", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
                for c in ["<1 mois", ">3 mois", "1 mois < <3 mois"]: be_pivot[c] = be_pivot.get(c, 0)
                be_pivot["Total"] = be_pivot[["<1 mois", "1 mois < <3 mois", ">3 mois"]].sum(axis=1); be_pivot = be_pivot.loc[be_pivot["Total"] > 0]

            # ===================== AFFICHAGE =====================
            st.markdown('<div class="mh"><h1>📊 Tableau de Bord KPI — Maintenance</h1><span class="db">📅 %s</span></div>'%fichier_date,unsafe_allow_html=True)
            st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%d</div><div class="cl">OT Total</div></div><div class="cc c2"><div class="cv">%d</div><div class="cl">Avis Sans OT</div></div><div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div><div class="cc c4"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div></div>'%(tot_ot,tot_av,avg_p,avg_q),unsafe_allow_html=True)

            tab_synthese, tab_performance, tab_qualite, tab_anomalies, tab_variance, tab_backlog, tab_journal, tab_classement = st.tabs([
                "📋 Synthese & Actions","⚡ Performance","🎯 Qualite","⚠️ Anomalies","📈 Variance","📦 Appel Analyse Backlog","📖 Journal","🏆 Classement"
            ])

            with tab_synthese:
                st.markdown('<div class="stl s">Synthese & Actions</div>',unsafe_allow_html=True)
                c1,c2=st.columns(2)
                with c1:
                    fig_ot=anl_pie_chart(ot_statut_df,"Statut OT","Nombre","OT par Statut",colors=["#38a169","#3182ce","#d69e2e","#e53e3e","#805ad5","#CBD5E0"])
                    if fig_ot: st.plotly_chart(fig_ot,use_container_width=True)
                with c2:
                    fig_av=anl_pie_chart(avis_statut_df,"Statut utilisateur","Nombre","Avis par Statut",colors=["#38a169","#3182ce","#d69e2e","#e53e3e","#805ad5","#CBD5E0"])
                    if fig_av: st.plotly_chart(fig_av,use_container_width=True)
                st.markdown('<div class="stl p">Performance</div>',unsafe_allow_html=True)
                fig_perf_bar=make_kpi_bar_chart_with_target(QK,global_actuals,CIBLE,"Performance","#38a169","#e53e3e")
                if fig_perf_bar: st.plotly_chart(fig_perf_bar,use_container_width=True)
                st.markdown('<div class="stl q">Qualite</div>',unsafe_allow_html=True)
                fig_qual_bar=make_kpi_bar_chart_with_target(PK,global_actuals,CIBLE,"Qualite","#3182ce","#e53e3e")
                if fig_qual_bar: st.plotly_chart(fig_qual_bar,use_container_width=True)
                st.markdown('<div class="stl p">Top/Bottom Performance</div>',unsafe_allow_html=True)
                fig_tp=make_rank_bar_chart(top5_perf,bot5_perf,"Top / Bottom 5 Performance",80,"#38a169","#e53e3e")
                if fig_tp: st.plotly_chart(fig_tp,use_container_width=True)
                st.markdown('<div class="stl q">Top/Bottom Qualite</div>',unsafe_allow_html=True)
                fig_tq=make_rank_bar_chart(top5_qual,bot5_qual,"Top / Bottom 5 Qualite",80,"#3182ce","#e53e3e")
                if fig_tq: st.plotly_chart(fig_tq,use_container_width=True)
                na_p=[k for k in QK if not (global_actuals.get(k,0)<=CIBLE.get(k,100) if is_lb(k) else global_actuals.get(k,0)>=CIBLE.get(k,100))]
                na_q=[k for k in PK if not (global_actuals.get(k,0)<=CIBLE.get(k,100) if is_lb(k) else global_actuals.get(k,0)>=CIBLE.get(k,100))]
                ca1,ca2=st.columns(2)
                with ca1:
                    st.markdown('<div class="ca"><div class="ct" style="color:#38a169">Actions Performance</div>',unsafe_allow_html=True)
                    if na_p: st.markdown(html_actions_table(na_p,global_actuals,CIBLE,ACT_MAP),unsafe_allow_html=True)
                    else: st.markdown('<div class="es">✅ Tous atteints</div>',unsafe_allow_html=True)
                    st.markdown('</div>',unsafe_allow_html=True)
                with ca2:
                    st.markdown('<div class="ca"><div class="ct" style="color:#3182ce">Actions Qualite</div>',unsafe_allow_html=True)
                    if na_q: st.markdown(html_actions_table(na_q,global_actuals,CIBLE,ACT_MAP),unsafe_allow_html=True)
                    else: st.markdown('<div class="es">✅ Tous atteints</div>',unsafe_allow_html=True)
                    st.markdown('</div>',unsafe_allow_html=True)
                st.markdown(html_grouped_bars(posts,perf_scores,qual_scores,"Performance vs Qualite par Poste"),unsafe_allow_html=True)

            with tab_performance:
                st.markdown('<div class="stl p">Indicateurs de Performance</div>',unsafe_allow_html=True)
                st.markdown(html_table(prows,pcols,"pt",sc_col=set(QK+["Score Performance"])),unsafe_allow_html=True)
            with tab_qualite:
                st.markdown('<div class="stl q">Indicateurs de Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_table(qrows,qcols,"qt",sc_col=set(PK+["Score Qualite"])),unsafe_allow_html=True)
            with tab_anomalies:
                st.markdown('<div class="stl a">Anomalies</div>',unsafe_allow_html=True)
                if tot_ano_p==0 and tot_ano_q==0: st.markdown('<div class="es" style="padding:30px">✅ Aucune anomalie</div>',unsafe_allow_html=True)
                else:
                    if tot_ano_p > 0:
                        st.markdown('<div style="font-size:14px;font-weight:800;color:#276749;margin:8px 0 4px 0;padding-left:10px;border-left:3px solid #38a169">⚠️ Anomalies Performance (%d)</div>'%tot_ano_p,unsafe_allow_html=True)
                        st.markdown(html_ano_transposed(ano_p_rows,QK,"atp","kpi-cell"),unsafe_allow_html=True)
                    else: st.markdown('<div class="es">✅ Aucune anomalie performance</div>',unsafe_allow_html=True)
                    if tot_ano_q > 0:
                        st.markdown('<div style="font-size:14px;font-weight:800;color:#2b6cb0;margin:8px 0 4px 0;padding-left:10px;border-left:3px solid #3182ce">⚠️ Anomalies Qualite (%d)</div>'%tot_ano_q,unsafe_allow_html=True)
                        st.markdown(html_ano_transposed(ano_q_rows,PK,"atq","kpi-cell"),unsafe_allow_html=True)
                    else: st.markdown('<div class="es">✅ Aucune anomalie qualite</div>',unsafe_allow_html=True)

            # ============ PAGE VARIANCE ============
            with tab_variance:
                st.markdown('<div class="stl v">📈 Analyse de Variance — Comparaison Historique</div>',unsafe_allow_html=True)
                hist_path="KPI_Historique.xlsx"; hist_df=load_historical_kpis(hist_path)
                if hist_df.empty: st.markdown('<div class="es" style="padding:30px">Pas assez d\'historique. Le fichier <b>KPI_Historique.xlsx</b> doit contenir au moins 2 feuilles (dates).</div>',unsafe_allow_html=True)
                else:
                    var_df, prev_d, curr_d = compute_variance(hist_df)
                    if var_df.empty: st.markdown('<div class="es">Impossible de calculer la variance (postes manquants entre les 2 dernieres dates).</div>',unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="font-size:13px;color:#4a5568;margin-bottom:8px">Comparaison entre <b>%s</b> et <b>%s</b></div>'%(prev_d, curr_d),unsafe_allow_html=True)
                        st.markdown(html_variance_table(var_df),unsafe_allow_html=True)
                        fig_comp = go.Figure()
                        postes_v = var_df["Poste de travail"].tolist()
                        fig_comp.add_trace(go.Bar(name=f'Perf {prev_d}', x=postes_v, y=var_df["Perf Prec."], marker_color='#90cdf4'))
                        fig_comp.add_trace(go.Bar(name=f'Perf {curr_d}', x=postes_v, y=var_df["Perf Act."], marker_color='#2b6cb0'))
                        fig_comp.add_trace(go.Bar(name=f'Qual {prev_d}', x=postes_v, y=var_df["Qual Prec."], marker_color='#9ae6b4'))
                        fig_comp.add_trace(go.Bar(name=f'Qual {curr_d}', x=postes_v, y=var_df["Qual Act."], marker_color='#276749'))
                        fig_comp.update_layout(barmode='group', title=f"<b>Comparaison Postes : {prev_d} vs {curr_d}</b>", xaxis_title="Poste", yaxis_title="Score (%)", plot_bgcolor="white", height=500, margin=dict(l=50,r=50,t=80,b=150), xaxis=dict(tickangle=-45))
                        st.plotly_chart(fig_comp, use_container_width=True)
                        top5_vp = var_df.nlargest(5, "Ecart Perf")[["Poste de travail", "Ecart Perf"]].rename(columns={"Poste de travail": "Poste", "Ecart Perf": "Score"})
                        bot5_vp = var_df.nsmallest(5, "Ecart Perf")[["Poste de travail", "Ecart Perf"]].rename(columns={"Poste de travail": "Poste", "Ecart Perf": "Score"})
                        top5_vq = var_df.nlargest(5, "Ecart Qual")[["Poste de travail", "Ecart Qual"]].rename(columns={"Poste de travail": "Poste", "Ecart Qual": "Score"})
                        bot5_vq = var_df.nsmallest(5, "Ecart Qual")[["Poste de travail", "Ecart Qual"]].rename(columns={"Poste de travail": "Poste", "Ecart Qual": "Score"})
                        cv1,cv2=st.columns(2)
                        with cv1:
                            fig_vtp=make_rank_bar_chart(top5_vp,bot5_vp,"Variance Performance (Top/Bottom 5)",0,"#38a169","#e53e3e")
                            if fig_vtp: st.plotly_chart(fig_vtp,use_container_width=True)
                        with cv2:
                            fig_vtq=make_rank_bar_chart(top5_vq,bot5_vq,"Variance Qualite (Top/Bottom 5)",0,"#3182ce","#e53e3e")
                            if fig_vtq: st.plotly_chart(fig_vtq,use_container_width=True)
                        st.markdown('<div class="stl a">Recommandations</div>',unsafe_allow_html=True)
                        rec_h='<div class="ca"><div class="ct">Actions et Recommandations</div>'
                        imp_p=var_df[var_df["Ecart Perf"]>2]; deg_p=var_df[var_df["Ecart Perf"]<-2]
                        imp_q=var_df[var_df["Ecart Qual"]>2]; deg_q=var_df[var_df["Ecart Qual"]<-2]
                        if not imp_p.empty: rec_h+='<div style="margin-bottom:8px"><b style="color:#276749">🟢 Amelioration Significative Performance :</b><ul style="margin-top:4px;padding-left:20px;color:#4a5568">'
                        for _,r in imp_p.iterrows(): rec_h+='<li><b>%s</b> (+%.1f pts) → Maintenir les bonnes pratiques.</li>'%(r["Poste de travail"],r["Ecart Perf"])
                        rec_h+='</ul></div>'
                        if not deg_p.empty: rec_h+='<div style="margin-bottom:8px"><b style="color:#c53030">🔴 Degradation Significative Performance :</b><ul style="margin-top:4px;padding-left:20px;color:#4a5568">'
                        for _,r in deg_p.iterrows(): rec_h+='<li><b>%s</b> (%.1f pts) → Analyser les causes et planifier des actions correctives.</li>'%(r["Poste de travail"],r["Ecart Perf"])
                        rec_h+='</ul></div>'
                        if not imp_q.empty: rec_h+='<div style="margin-bottom:8px"><b style="color:#2b6cb0">🟢 Amelioration Significative Qualite :</b><ul style="margin-top:4px;padding-left:20px;color:#4a5568">'
                        for _,r in imp_q.iterrows(): rec_h+='<li><b>%s</b> (+%.1f pts) → Capitaliser sur les demarches qualite.</li>'%(r["Poste de travail"],r["Ecart Qual"])
                        rec_h+='</ul></div>'
                        if not deg_q.empty: rec_h+='<div style="margin-bottom:8px"><b style="color:#c53030">🔴 Degradation Significative Qualite :</b><ul style="margin-top:4px;padding-left:20px;color:#4a5568">'
                        for _,r in deg_q.iterrows(): rec_h+='<li><b>%s</b> (%.1f pts) → Renforcer le suivi et verifier les caracterisations.</li>'%(r["Poste de travail"],r["Ecart Qual"])
                        rec_h+='</ul></div>'
                        if imp_p.empty and deg_p.empty and imp_q.empty and deg_q.empty: rec_h+='<div class="es">Aucune variation significative (ecart > 2 pts) detectee.</div>'
                        rec_h+='</div>'
                        st.markdown(rec_h,unsafe_allow_html=True)

            with tab_backlog:
                st.markdown('<div class="stl om">🔍 OMS</div>',unsafe_allow_html=True)
                if oms_pivot.empty: st.markdown('<div class="es">Aucun OT "OMS"</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div style="font-size:12px;color:#718096;margin-bottom:4px">%d OT "OMS"</div>'%int(oms_pivot.sum().sum()),unsafe_allow_html=True)
                    c1,c2=st.columns([3,2])
                    with c1: st.markdown(html_backlog_pivot(oms_pivot,"bt-oms"),unsafe_allow_html=True)
                    with c2:
                        d=oms_pivot.sum().reset_index(); d.columns=["Statut OT","Nombre"]
                        f=anl_pie_chart(d,"Statut OT","Nombre","OMS par Statut",colors=["#38a169","#3182ce","#d69e2e","#e53e3e","#805ad5","#CBD5E0"])
                        if f: st.plotly_chart(f,use_container_width=True)
                st.markdown('<div class="sec-sep"></div>',unsafe_allow_html=True)
                st.markdown('<div class="stl th">🌡️ THERMOGRAPHIE</div>',unsafe_allow_html=True)
                if thermo_pivot.empty: st.markdown('<div class="es">Aucun OT "THERMOGRAPHIE"</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div style="font-size:12px;color:#718096;margin-bottom:4px">%d OT "THERMOGRAPHIE"</div>'%int(thermo_pivot.sum().sum()),unsafe_allow_html=True)
                    c1,c2=st.columns([3,2])
                    with c1: st.markdown(html_backlog_pivot(thermo_pivot,"bt-th"),unsafe_allow_html=True)
                    with c2:
                        d=thermo_pivot.sum().reset_index(); d.columns=["Statut OT","Nombre"]
                        f=anl_pie_chart(d,"Statut OT","Nombre","THERMOGRAPHIE par Statut",colors=["#3182ce","#38a169","#d69e2e","#e53e3e","#805ad5","#CBD5E0"])
                        if f: st.plotly_chart(f,use_container_width=True)
                st.markdown('<div class="sec-sep"></div>',unsafe_allow_html=True)
                st.markdown('<div class="stl bp">📋 Backlog Preparation</div>',unsafe_allow_html=True)
                if bp_pivot.empty: st.markdown('<div class="es">Aucun OT en attente de preparation.</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div style="font-size:12px;color:#718096;margin-bottom:4px">%d OT en attente</div>'%int(bp_pivot["Total"].sum()),unsafe_allow_html=True)
                    c1,c2=st.columns([3,2])
                    with c1: st.markdown(html_backlog_pivot(bp_pivot[["CARACTERISE","NON CARACTERISE"]],"bt-bp","% Caractérisé",bp_pct),unsafe_allow_html=True)
                    with c2:
                        d=pd.DataFrame({"Statut":["CARACTERISE","NON CARACTERISE"],"Nombre":[int(bp_pivot["CARACTERISE"].sum()),int(bp_pivot["NON CARACTERISE"].sum())]})
                        f=anl_pie_chart(d,"Statut","Nombre","Backlog Prep",colors=["#38a169","#e53e3e"])
                        if f: st.plotly_chart(f,use_container_width=True)
                st.markdown('<div class="sec-sep"></div>',unsafe_allow_html=True)
                st.markdown('<div class="stl be">⚙️ Backlog Execution</div>',unsafe_allow_html=True)
                if be_pivot.empty: st.markdown('<div class="es">Aucun OT en execution.</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div style="font-size:12px;color:#718096;margin-bottom:4px">%d OT en execution</div>'%int(be_pivot["Total"].sum()),unsafe_allow_html=True)
                    c1,c2=st.columns([3,2])
                    with c1: st.markdown(html_backlog_pivot(be_pivot[["<1 mois","1 mois < <3 mois",">3 mois"]],"bt-be"),unsafe_allow_html=True)
                    with c2:
                        d=pd.DataFrame({"Age":["<1 mois","1 mois < <3 mois",">3 mois"],"Nombre":[int(be_pivot["<1 mois"].sum()),int(be_pivot["1 mois < <3 mois"].sum()),int(be_pivot[">3 mois"].sum())]})
                        f=anl_pie_chart(d,"Age","Nombre","Backlog Exec",colors=["#38a169","#d69e2e","#e53e3e"])
                        if f: st.plotly_chart(f,use_container_width=True)

            with tab_journal:
                st.markdown('<div class="stl s">Journal des Variations Significatives</div>',unsafe_allow_html=True)
                hist_path="KPI_Historique.xlsx"; hist_df=load_historical_kpis(hist_path)
                if hist_df.empty: st.markdown('<div class="es">Pas assez d\'historique (min 2 periodes)</div>',unsafe_allow_html=True)
                else:
                    var_df=calculate_variations(hist_df)
                    if var_df.empty: st.markdown('<div class="es">Aucune variation calculee</div>',unsafe_allow_html=True)
                    else:
                        jrn=generate_journal(var_df)
                        if jrn.empty: st.markdown('<div class="es">Aucune variation significative (>=5%%)</div>',unsafe_allow_html=True)
                        else:
                            st.dataframe(jrn.drop(columns=["Significatif"],errors="ignore"),use_container_width=True,height=500)
                            with st.expander("📥 Export"): export_btn(jrn.drop(columns=["Significatif"],errors="ignore"),"journal_variations.xlsx")

            with tab_classement:
                st.markdown('<div class="stl c">Classement Global</div>',unsafe_allow_html=True)
                combined_scores={p:(perf_scores.get(p,0)+qual_scores.get(p,0))/2 for p in posts}
                st.markdown(html_classement(combined_scores,"#805ad5"),unsafe_allow_html=True)
                st.markdown('<div class="stl p">Classement Performance</div>',unsafe_allow_html=True)
                st.markdown(html_classement(perf_scores,"#38a169"),unsafe_allow_html=True)
                st.markdown('<div class="stl q">Classement Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_classement(qual_scores,"#3182ce"),unsafe_allow_html=True)
                hist_path="KPI_Historique.xlsx"; hist_df=load_historical_kpis(hist_path)
                if not hist_df.empty:
                    var_df=calculate_variations(hist_df)
                    if not var_df.empty:
                        top_h,bot_h=calculate_rankings(var_df)
                        st.markdown('<div class="stl s">Evolution Historique</div>',unsafe_allow_html=True)
                        c1,c2=st.columns(2)
                        with c1:
                            if not top_h.empty:
                                st.markdown("**🟢 Top 5 Progression**"); st.dataframe(top_h,use_container_width=True)
                        with c2:
                            if not bot_h.empty:
                                st.markdown("**🔴 Bottom 5 Regression**"); st.dataframe(bot_h,use_container_width=True)

        except Exception as e:
            st.error("Erreur de chargement: %s"%str(e))
            st.info("Verifiez que les fichiers ot.xlsx et avis.xlsx sont presents.")
    else:
        if unf:
            st.markdown("""<div style="text-align:center;padding:60px;color:#718096"><div style="font-size:64px;margin-bottom:16px">📁</div><h2 style="color:#1e3a5f">Chargement de fichiers</h2><p>Veuillez charger les fichiers OT et AVIS.</p></div>""",unsafe_allow_html=True)
        else:
            st.markdown("""<div style="text-align:center;padding:60px;color:#718096"><div style="font-size:64px;margin-bottom:16px">📂</div><h2 style="color:#1e3a5f">Fichiers non trouves</h2><p>Placez les fichiers <b>ot.xlsx</b> et <b>avis.xlsx</b> a la racine.</p></div>""",unsafe_allow_html=True)

if __name__=="__main__":
    main()
