# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import io, locale, random, time, os, hashlib, json, base64
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI", initial_sidebar_state="expanded")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois",
      "Performance Graissage","Performance Inspection","Performance Appels Systématiques"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL",
      "OT Fiabilité","Total Avis de Panne"]
ALL_KPI = QK + PK

CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100,
         "Performance Graissage":95,"Performance Inspection":95,"Performance Appels Systématiques":95,
         "OT Fiabilité":100,"Total Avis de Panne":100}

ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "Performance Graissage":"Ameliorer le taux de realisation des OT de graissage (Type 350).",
           "Performance Inspection":"Ameliorer le taux de realisation des OT d'inspection (Types 290,300,310).",
           "Performance Appels Systématiques":"Ameliorer le taux de realisation des appels systematiques (Type 360).",
           "OT Fiabilité":"Maintenir la fiabilite des OT a 100%.",
           "Total Avis de Panne":"Maintenir le suivi des avis de panne a 100%."}

# Mapping pour l'attribution automatique des responsables
KPI_RESP_MAP = {
    "TAUX_REALISATION_CORRECTIF/PT": "Chef d'atelier",
    "OT préparation <1 mois": "Préparateur BM",
    "OT préparation 1mois< <3mois": "Préparateur BM",
    "OT préparation >3 mois": "Préparateur BM",
    "OT planification <1 mois": "Planificateur BM",
    "OT planification 1mois< <3mois": "Planificateur BM",
    "OT planification >3 mois": "Planificateur BM",
    "OT exécution <1 mois": "Chef d'atelier",
    "OT exécution 1mois< <3mois": "Chef d'atelier",
    "OT exécution >3 mois": "Chef d'atelier",
    "appel avis approuvé": "Chef d'atelier",
    "OT LANC ESTIME": "Fiabilité",
    "Backlog préparation caractérisé": "Préparateur BM",
    "Backlog planification caractérisé": "Planificateur BM",
    "OT CONFIME": "Agent de saisie",
    "OT_COR_EGAL": "Agent de saisie",
    "Performance Graissage": "Chef d'atelier",
    "Performance Inspection": "Chef d'atelier",
    "Performance Appels Systématiques": "Chef d'atelier",
    "OT Fiabilité": "Fiabilité",
    "Total Avis de Panne": "Fiabilité"
}

LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]

CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def get_logo_base64():
    for path in ["logo.png", "./logo.png", "../logo.png"]:
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    return base64.b64encode(f.read()).decode()
            except Exception:
                pass
    return None

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return "18/06/2026"

# --- HELPERS FOR PREPARE_DATA ---
def contient_mot(t,lm):
    t=str(t); return any(m in t for l in lm for m in l.split())
def cat_age(a):
    if pd.isna(a): return "Inconnu"
    if a<=1: return "<1 mois"
    elif a>=3: return ">3 mois"
    return "1 mois < <3 mois"
def excr(df):
    if "Poste travail princ." in df.columns:
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
    return df

# ============================================================
# CACHED HEAVY DATA PREPARATION
# ============================================================
@st.cache_data(show_spinner=False)
def prepare_data(ot_bytes, av_bytes, date_str):
    raw_ot = pd.read_excel(io.BytesIO(ot_bytes))
    raw_av = pd.read_excel(io.BytesIO(av_bytes))
    raw_ot = excr(raw_ot)
    raw_av = excr(raw_av)
    
    for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
        if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
    for c in ["Créé le","Début souhaité","Date de la clôture"]:
        if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
        
    now_ts = pd.to_datetime(date_str, format="%d/%m/%Y", errors='coerce')
    if pd.isna(now_ts): now_ts = pd.Timestamp.now()
    
    df = raw_ot.copy()
    
    # Heavy feature engineering
    df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
    df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
    df["Type Carac Prep"]=df["Statut utilisateur"].apply(lambda x: next((kw for kw in MP_KW if kw in str(x)), "NON CARACTERISE"))
    df["Type Carac Plan"]=df["Statut utilisateur"].apply(lambda x: next((kw for kw in MPLAN_KW if kw in str(x)), "NON CARACTERISE"))
    
    for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
        if dc in df.columns:
            df[am]=((now_ts.year-df[dc].dt.year)*12+(now_ts.month-df[dc].dt.month)).round(2)
            df[ac]=df[am].apply(cat_age)
        else: df[am]=np.nan; df[ac]="Inconnu"
        
    df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
    df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
    df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
    df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
    df["_tw_num"]=pd.to_numeric(df.get("Type de travail",pd.Series(dtype=float)),errors="coerce")
    
    if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
    
    avf = raw_av[(raw_av["Ordre"].isna())|(raw_av["Ordre"].astype(str).str.strip()=="")].copy()
    
    apm = sorted(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
    
    return df, avf, apm, now_ts

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("Cible","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    sens = "Stable"
                    if trend != "stabilite":
                        if (trend == "hausse" and kpi not in LOWER_BETTER) or (trend == "baisse" and kpi in LOWER_BETTER):
                            sens = "Amelioration"
                        else:
                            sens = "Degradation"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend, "Sens":sens})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    return j.sort_values(["Date actuelle","Ecart %"],ascending=[True,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:10px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden;display:flex;align-items:center;gap:12px}
    .mh h1{color:#fff;font-size:40px;font-weight:800;margin:0;display:inline;flex:1}
    .mh .logo{height:47px;width:auto;max-width:140px;object-fit:contain;flex-shrink:0;border-radius:4px}
    .mh .db{background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:21px;font-weight:500;border:1px solid rgba(255,255,255,.2);white-space:nowrap;flex-shrink:0}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .cc.c5{border-top:3px solid #2b6cb0}.cc.c5 .cv{color:#2b6cb0}
    .cc.c6{border-top:3px solid #3182ce}.cc.c6 .cv{color:#3182ce}
    .cc.c7{border-top:3px solid #d69e2e}.cc.c7 .cv{color:#975a16}
    .cc.c8{border-top:3px solid #ecc94b}.cc.c8 .cv{color:#b7791f}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw.omt thead th{background:linear-gradient(135deg,#6b46c1,#805ad5)}
    .tw.tht thead th{background:linear-gradient(135deg,#9b2c2c,#e53e3e)}
    .tw thead th:first-child{z-index:11;left:0}
    .tw tbody td:first-child{position:sticky;left:0;background:#fff;z-index:5;border-right:1px solid #edf2f7;color:#1a202c !important}
    .tw tbody tr:nth-child(even) td:first-child{background:#f7fafc}
    .tw tbody tr:hover td:first-child{background:#ebf8ff}
    .tw.cb td:first-child{background:#2b6cb0!important;color:#fff!important}
    .tw.tr td:first-child{background:#e2e8f0!important;color:#1a202c!important}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap;color:#1a202c !important}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;color:#1a202c !important;font-weight:800!important;font-size:12px!important}
    
    .stTabs [data-baseweb="tab-list"]{gap:6px;background:#e2e8f0;padding:6px;border-radius:8px;margin-bottom:8px}
    .stTabs [data-baseweb="tab"]{
        border-radius:6px;
        padding:12px 22px;
        font-weight:700;
        font-size:20px;
        line-height:1.5;
        min-height:48px;
    }
    .stTabs [data-baseweb="tab"] span,
    .stTabs [data-baseweb="tab"] > div{
        font-size:22px !important;
    }
    .stTabs [aria-selected="true"]{
        background:#fff!important;
        color:var(--p)!important;
        box-shadow:0 3px 8px rgba(0,0,0,.1);
        font-size:21px;
    }
    .stTabs [data-baseweb="tab"] svg{width:22px;height:22px}
    
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:6px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:26px;background:#edf2f7;border-radius:4px;overflow:visible;position:relative}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    
    .car .target-mark{
        position:absolute;
        top:-4px;
        bottom:-4px;
        width:3px;
        background:#e53e3e;
        z-index:20;
        transform:translateX(-50%);
        box-shadow:0 0 6px rgba(229,62,62,.9),0 0 2px rgba(0,0,0,.4);
        border-radius:2px;
    }
    .car .target-mark::before{
        content:"";
        position:absolute;
        top:-5px;
        left:50%;
        transform:translateX(-50%);
        width:0;height:0;
        border-left:5px solid transparent;
        border-right:5px solid transparent;
        border-top:6px solid #e53e3e;
    }
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .car .cav-tgt{font-size:10px;font-weight:700;color:#1a202c;min-width:42px;text-align:right;padding-left:4px;opacity:.7}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1;position:relative}
    .gbr-target{position:absolute;left:90%;top:-4px;bottom:-4px;width:3px;background:#e53e3e;z-index:10;box-shadow:0 0 6px rgba(229,62,62,.8);border-radius:2px}
    .gbr-target-label{position:absolute;left:90%;top:-20px;transform:translateX(-50%);font-size:9px;font-weight:800;color:#fff;background:#e53e3e;padding:1px 5px;border-radius:3px;white-space:nowrap;z-index:11;box-shadow:0 1px 3px rgba(0,0,0,.2)}
    .gbr-w{flex:1;height:22px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:10px;font-size:12px;font-weight:700;align-items:center}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .gbr-legend .target-icon{display:inline-block;width:3px;height:14px;background:#e53e3e;border-radius:1px;box-shadow:0 0 3px rgba(229,62,62,.6)}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    
    /* ===== SIDEBAR BACKGROUND BLEU FONCE ===== */
    div[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#1e40af 0%,#1e3a8a 50%,#1e3a5f 100%)!important;
    }
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label,div[data-testid="stSidebar"] .stTextInput label{color:rgba(255,255,255,.9)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.1);border-radius:6px;padding:5px 10px;margin-bottom:5px;border:1px solid rgba(255,255,255,.15)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div,div[data-testid="stSidebar"] .stTextInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .synth-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px}
    .synth-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;padding:5px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .synth-tbl tbody td{padding:4px 8px;border-bottom:1px solid #edf2f7;text-align:center;color:#1a202c !important}
    .synth-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .synth-tbl tbody tr:hover td{background:#ebf8ff!important}
    .synth-tbl .poste-cell{text-align:left;font-weight:700;white-space:nowrap;min-width:140px;color:#1a202c !important}
    div[data-testid="stHorizontalBlock"]{align-items:center!important}
    
    /* ===== SUPPRIMER UNIQUEMENT LES BOUTONS DE PARTAGE/MENU STREAMLIT (Garder le bouton Sidebar) ===== */
    [data-testid="stHeaderActionElements"]{display:none !important;}
    [data-testid="stActionButtonContainer"]{display:none !important;}
    
    .footer {
        text-align: center;
        margin-top: 30px;
        padding: 15px;
        color: #718096;
        font-size: 13px;
        border-top: 1px solid #e2e8f0;
        font-weight: 600;
    }

    /* ===== AGRANDIR LA TAILLE DU TEXTE DANS LE PLAN D' ACTIONS (sans scroll horizontal) ===== */
    div[data-testid="stDataEditor"] table, 
    div[data-testid="stDataEditor"] th, 
    div[data-testid="stDataEditor"] td {
        font-size: 18px !important;
        line-height: 1.4 !important;
        white-space: normal !important; /* Permet le retour à la ligne */
        word-wrap: break-word !important;
    }
    div[data-testid="stDataEditor"] [data-testid="stMarkdownContainer"] {
        font-size: 18px !important;
    }
    /* Empêcher la barre de défilement horizontale et s'adapter à la page */
    div[data-testid="stDataEditor"] [data-testid="stTable"] {
        overflow-x: hidden !important;
        width: 100% !important;
    }

    @media(max-width:768px){
        .cr{grid-template-columns:repeat(2,1fr)}
        .mh{padding:8px 10px;gap:8px}
        .mh h1{font-size:18px}
        .mh .logo{height:35px;max-width:70px}
        .mh .db{font-size:11px;padding:2px 8px}
        .cg,.dgrid{grid-template-columns:1fr}
        .car{flex-wrap:wrap;gap:2px}
        .car .cal{width:100%;text-align:left;padding-right:0;margin-bottom:2px}
        .car .cab{flex:1 1 70%}
        .car .cav-out,.car .cav-tgt{flex:1 1 15%;min-width:40px}
        .gbr{flex-direction:column;align-items:flex-start;gap:4px}
        .gbr-l{width:100%;margin-bottom:2px}
        .gbr-g{width:100%;flex-wrap:wrap}
        .gbr-w{flex:1 1 45%}
        .gbr-v{flex:1 1 10%;min-width:40px}
        .tw{font-size:10px}
        .tw thead th,.tw tbody td{padding:3px 4px}
        .stl{font-size:13px}
        .stTabs [data-baseweb="tab"]{padding:8px 12px;font-size:15px}
        .stTabs [data-baseweb="tab"] span{font-size:16px !important}
        div[data-testid="stDataEditor"] table, div[data-testid="stDataEditor"] th, div[data-testid="stDataEditor"] td { font-size: 14px !important; }
    }
    
    /* ===== CSS POUR L'IMPRESSION PDF (MASQUE L'INTERFACE STREAMLIT) ===== */
    @media print {
        /* Cacher toute l'interface Streamlit */
        section[data-testid="stSidebar"], 
        header[data-testid="stHeader"], 
        div[data-testid="stToolbar"], 
        div[data-testid="stHeaderActionElements"],
        footer, 
        .stDeployButton, 
        #MainMenu { display: none !important; }
        
        /* S'assurer que le contenu prend toute la largeur sans marge */
        .main .block-container { 
            padding-top: 0 !important; 
            padding-left: 0 !important; 
            padding-right: 0 !important; 
            max-width: 100% !important; 
        }
        
        /* Cacher tous les boutons (y compris le bouton d'impression et d'export) */
        .stButton, .stDownloadButton { display: none !important; }
        
        /* Forcer l'impression des couleurs de fond (pour vos tableaux et cartes) */
        * {
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }
        
        /* Gérer les onglets : cacher la barre d'onglets mais garder le contenu visible */
        .stTabs [data-baseweb="tab-list"] { display: none !important; }
        .stTabs [data-baseweb="tab-panel"] { 
            display: block !important; 
            page-break-inside: avoid; 
        }
        
        /* Éviter de couper les tableaux */
        .tw, .synth-tbl { page-break-inside: avoid; overflow: visible !important; }
    }
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def get_text_col(df):
        for c in ["Désignation","Designation","Désignation OT","Texte ordre","Texte","Description","Libellé","Libelle"]:
            if c in df.columns: return c
        for c in df.columns:
            if df[c].dtype=='object' and any(kw in str(c).lower() for kw in ['sign','text','desc','libell']):
                return c
        return None
    def build_statut_pivot(df_sub, posts):
        if df_sub.empty:
            return pd.DataFrame(index=posts, columns=["CRÉÉ","LANC","CLOT","TCLO","Total"]).fillna(0).astype(int)
        piv=pd.pivot_table(df_sub, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0)
        for s in ["CRÉÉ","LANC","CLOT","TCLO"]:
            if s not in piv.columns: piv[s]=0
        piv["Total"]=piv[["CRÉÉ","LANC","CLOT","TCLO"]].sum(axis=1)
        return piv.reindex(posts, fill_value=0).fillna(0).astype(int)
    def html_statut_pivot(piv_df, table_class):
        cols=["Poste de travail","CRÉÉ","LANC","CLOT","TCLO","Total"]
        h='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:600">%s</td>'%poste
            for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
                h+='<td style="text-align:center">%d</td>'%int(row.get(c,0))
            h+='<td style="text-align:center;font-weight:800">%d</td>'%int(row.get("Total",0))
            h+='</tr>'
        h+='<tr class="tr"><td>Total</td>'
        for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
            h+='<td style="text-align:center">%d</td>'%int(piv_df[c].sum())
        h+='<td style="text-align:center">%d</td>'%int(piv_df["Total"].sum())
        h+='</tr></tbody></table>'
        return h
        
    def show_pie_pair(piv_df, title_prefix):
        global_counts=piv_df[["CRÉÉ","LANC","CLOT","TCLO"]].sum()
        global_counts=global_counts[global_counts>0]
        realised=global_counts.get("CLOT",0)+global_counts.get("TCLO",0)
        not_realised=global_counts.sum()-realised
        if not global_counts.empty:
            fig1=px.pie(global_counts, names=global_counts.index, values=global_counts.values,
                title="%s — Par Statut OT"%title_prefix,
                color_discrete_sequence=["#e53e3e","#d69e2e","#38a169","#3182ce"])
            fig1.update_traces(textposition='inside',textinfo='percent+value',textfont_size=11)
            fig1.update_layout(margin=dict(t=40,b=10,l=10,r=10),height=300,legend=dict(font_size=10,orientation="h",yanchor="bottom",y=-0.1))
            st.plotly_chart(fig1,use_container_width=True)
        else:
            st.markdown('<div class="es">Aucune donnee</div>',unsafe_allow_html=True)
            
        if global_counts.sum()>0:
            pie2_data=pd.DataFrame({"Statut":["Réalisés (CLOT+TCLO)","Non Réalisés"],"Nombre":[realised,not_realised]})
            fig2=px.pie(pie2_data, names="Statut", values="Nombre",
                title="%s — Réalisés vs Non Réalisés"%title_prefix,
                color="Statut", color_discrete_map={"Réalisés (CLOT+TCLO)":"#38a169","Non Réalisés":"#e53e3e"})
            fig2.update_traces(textposition='inside',textinfo='percent+value',textfont_size=11)
            fig2.update_layout(margin=dict(t=40,b=10,l=10,r=10),height=300,legend=dict(font_size=10,orientation="h",yanchor="bottom",y=-0.1))
            st.plotly_chart(fig2,use_container_width=True)
        else:
            st.markdown('<div class="es">Aucune donnee</div>',unsafe_allow_html=True)

    def show_simple_pie(piv_df, title, keep_non_carac=False):
        if not keep_non_carac and "NON CARACTERISE" in piv_df.columns:
            piv_df = piv_df.drop(columns=["NON CARACTERISE"])
        counts = piv_df.sum()
        counts = counts[counts > 0]
        if not counts.empty:
            color_map = {
                "CARACTERISE": "#38a169",
                "NON CARACTERISE": "#e53e3e"
            }
            colors = [color_map.get(str(c), None) for c in counts.index]
            fig = px.pie(
                counts, names=counts.index, values=counts.values, title=title,
                color=counts.index,
                color_discrete_map={k: v for k, v in zip(counts.index, colors)} if any(colors) else None
            )
            fig.update_traces(textposition='inside', textinfo='percent+value', textfont_size=11)
            fig.update_layout(margin=dict(t=40, b=10, l=10, r=10), height=280,
                              legend=dict(font_size=10, orientation="h", yanchor="bottom", y=-0.1))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.markdown('<div class="es">Aucune donnee</div>', unsafe_allow_html=True)

    def calc_kpis(df_i, av_i, now_ts, posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        res['dfp']=df
        filt_corr=(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)
        an=cpiv(df,filt_corr,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["OT_CLOTURES"]=an["CLOT"]+an["TCLO"]
        an["TOTAL_OT"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=np.where(an["TOTAL_OT"]==0,100.0,ckpi(an["OT_CLOTURES"],an["TOTAL_OT"]))
        
        pr=cpiv(df,(df["Statut OT"]=="CRÉÉ")&(df["Statut utilisateur"].str.contains("CRPR",na=False)),"ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False)),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
            
        avf=av.copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        
        g_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_den=df[(df["Contient SOPL"]==1)&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_df=pd.DataFrame({"_n":g_num,"_d":g_den}).reindex(posts,fill_value=0)
        g_df["Performance Graissage"]=np.where(g_df["_d"]==0,100.0,(g_df["_n"]/g_df["_d"])*100)
        
        ins_types=[290,300,310]
        ins_base=(df["_tw_num"].isin(ins_types))&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        ins_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_den=df[(df["Contient SOPL"]==1)&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_df=pd.DataFrame({"_n":ins_num,"_d":ins_den}).reindex(posts,fill_value=0)
        ins_df["Performance Inspection"]=np.where(ins_df["_d"]==0,100.0,(ins_df["_n"]/ins_df["_d"])*100)
        
        sys_base=(df["_tw_num"]==360)&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now_ts)
        sys_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_den=df[(df["Contient SOPL"]==1)&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_df=pd.DataFrame({"_n":sys_num,"_d":sys_den}).reindex(posts,fill_value=0)
        sys_df["Performance Appels Systématiques"]=np.where(sys_df["_d"]==0,100.0,(sys_df["_n"]/sys_df["_d"])*100)
        
        fiab_s=pd.Series(100.0,index=posts); avpan_s=pd.Series(100.0,index=posts)
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "Performance Graissage":g_df["Performance Graissage"],"Performance Inspection":ins_df["Performance Inspection"],"Performance Appels Systématiques":sys_df["Performance Appels Systématiques"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"],
            "OT Fiabilité":fiab_s,"Total Avis de Panne":avpan_s
        })
        return res

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT Fiabilité","Total Avis de Panne"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else "background:#ffeb9c;color:#9c6500;font-weight:600"
        return ""
        
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "background:#c6efce;color:#006100;font-weight:600"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        if k in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]: return 1 if a>=95 else 0
        if k in ["OT Fiabilité","Total Avis de Panne"]: return 1 if a>=100 else 0
        return 0
    def is_lb(k): return k in LOWER_BETTER

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                else: s=cs(v) if sc_col and c in sc_col else ks(v,c); h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
        
    def html_anomaly_table(rows,cols,tc):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="tr" if r.get("Poste de travail")=="Total" else ""
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if c=="Poste de travail": h+='<td style="font-weight:700">%s</td>'%v
                elif c=="Total Anomalies": h+='<td style="text-align:center;font-weight:800">%s</td>'%v
                else:
                    s=kas(v)
                    h+='<td style="%s;text-align:center">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'

    def html_actions_table(kpi_list,actuals,targets,act_map):
        h='<table class="tw at"><thead><tr><th>KPI</th><th>Valeur Actuelle</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action Recommandee</th></tr></thead><tbody>'
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            status="ATTEINT" if met else "NON ATTEINT"
            st_s="background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
            ec_clr="#276749" if met else "#c53030"
            action="Objectif atteint" if met else act_map.get(k,"")
            h+='<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td></tr>'%(k,av,tv,ec_clr,diff,st_s,status,action)
        return h+'</tbody></table>'
    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'; return h
    def html_kpi_bars(kpi_list,actuals,targets,title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,title)
        h+='<div class="gbr-legend"><span><span class="target-icon"></span> Cible</span></div>'
        for k in kpi_list:
            av=actuals.get(k,0)
            tv=targets.get(k,100)
            met=av<=tv if is_lb(k) else av>=tv
            bw=min(max(av,0),100)
            bg=color_ok if met else color_fail
            tv_pos=min(max(tv,0),100)
            h+=('<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div><div class="target-mark" style="left:%s%%"></div></div><div class="cav-out">%.1f%%</div><div class="cav-tgt">/%.0f%%</div></div>')%(k,bw,bg,tv_pos,av,tv)
        return h+'</div>'
        
    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span><span><span class="target-icon"></span> Cible 90%%</span></div>'
        sorted_posts = sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True)
        for idx,p in enumerate(sorted_posts):
            pv,qv=pscores.get(p,0),qscores.get(p,0)
            label_html = '<div class="gbr-target-label">90%%</div>' if idx==0 else ''
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-target"></div>%s<div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>'%(p,label_html,min(max(pv,0),100),pv,min(max(qv,0),100),qv)
        return h+'</div>'
        
    def html_synthese_table(synth_data,kpi_list,posts):
        h='<table class="synth-tbl"><thead><tr><th style="min-width:160px;text-align:left">Poste de travail</th>'
        for kpi in kpi_list: h+='<th>%s</th>'%kpi
        h+='</tr></thead><tbody>'
        for poste in posts:
            h+='<tr><td class="poste-cell">%s</td>'%poste
            for kpi in kpi_list:
                info=synth_data.get(poste,{}).get(kpi,{})
                diff=info.get("diff","—")
                if diff != "—":
                    try:
                        d = float(diff)
                        if d > 0: clr="#c6efce"
                        elif d < 0: clr="#ffc7ce"
                        else: clr=""
                        h+='<td style="background:%s;text-align:center;font-weight:700">%s</td>'%(clr,diff)
                    except:
                        h+='<td style="text-align:center">—</td>'
                else:
                    h+='<td style="text-align:center">—</td>'
            h+='</tr>'
        h+='</tbody></table>'
        return h
        
    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
    def html_generic_pivot(piv_df, table_class, title):
        piv_df = piv_df.copy()
        piv_df["Total"] = piv_df.sum(axis=1)
        cols = ["Poste de travail"] + [str(c) for c in piv_df.columns]
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:600">%s</td>'%poste
            for c in piv_df.columns:
                h+='<td style="text-align:center">%d</td>'%int(row.get(c,0))
            h+='</tr>'
        h+='<tr class="tr"><td>Total</td>'
        for c in piv_df.columns:
            h+='<td style="text-align:center">%d</td>'%int(piv_df[c].sum())
        h+='</tr></tbody></table></div>'
        return h

    # ===================== LOAD CACHED DATA =====================
    ot_bytes, av_bytes = None, None
    if os.path.exists("ot.xlsx") and os.path.exists("avis.xlsx"):
        with open("ot.xlsx", "rb") as f: ot_bytes = f.read()
        with open("avis.xlsx", "rb") as f: av_bytes = f.read()

    if ot_bytes and av_bytes:
        df_full, av_full, apm, now_ts = prepare_data(ot_bytes, av_bytes, fichier_date)
    else:
        df_full, av_full, apm, now_ts = pd.DataFrame(), pd.DataFrame(), [], pd.Timestamp.now()

    # ===================== SIDEBAR =====================
    with st.sidebar:
        logo_b64 = get_logo_base64()
        if logo_b64:
            st.markdown('<div style="display:flex;justify-content:center;padding:10px 0 15px 0;border-bottom:1px solid rgba(255,255,255,0.1);margin-bottom:10px;"><img src="data:image/png;base64,%s" style="max-width:100%%;height:auto;max-height:200px;object-fit:contain;border-radius:4px;"></div>'%logo_b64,unsafe_allow_html=True)
        else:
            st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",unsafe_allow_html=True)
        
        if st.button("🔄 Rafraîchir le cache", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

        if st.button("🖥️ Mode Présentation (Slide/PDF)", use_container_width=True):
            components.html("<script>window.parent.print();</script>", height=0, width=0)
            
        show_filters=st.checkbox("✅ Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf")
            ot_f=av_f=None
            if unf:
                pwd = st.text_input("Mot de passe administrateur", type="password")
                if pwd == "779900":
                    ot_f = st.file_uploader("Fichier OT", type=["xlsx"], key="uot")
                    av_f = st.file_uploader("Fichier AVIS", type=["xlsx"], key="uav")
                    new_date = st.text_input("Entrez la date (JJ/MM/AAAA)", value=fichier_date)
                    if st.button("💾 Sauvegarder et Appliquer"):
                        try:
                            datetime.strptime(new_date, "%d/%m/%Y")
                            if ot_f is not None:
                                with open("ot.xlsx", "wb") as f: f.write(ot_f.getbuffer())
                            if av_f is not None:
                                with open("avis.xlsx", "wb") as f: f.write(av_f.getbuffer())
                            with open("date.txt", "w", encoding="utf-8") as f: f.write(new_date)
                            st.success("Fichiers et date mis à jour avec succès !")
                            time.sleep(2)
                            st.cache_data.clear()
                            st.rerun()
                        except ValueError:
                            st.error("Format de date invalide. Veuillez utiliser JJ/MM/AAAA.")
                elif pwd != "":
                    st.error("Mot de passe incorrect.")
            else:
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            st.markdown("**🏭 Atelier**")
            sa=st.multiselect("Atelier",["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"],["All"],key="sa")
            st.markdown("**🏢 Division**")
            sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=av_f=None; sp=["All"]; sa=["All"]; sd=["All"]
            dr=(datetime(2025,1,1).date(),datetime.today().date())

    # ===================== APPLY FAST FILTERS & CALCULATIONS =====================
    if not df_full.empty:
        try:
            if unf and ot_f is not None and av_f is not None:
                df_full, av_full, apm, now_ts = prepare_data(ot_f.getvalue(), av_f.getvalue(), fichier_date)
                
            if "All" in sp or not sp: sp=apm
            if "All" in sa or not sa: sa=["All"]
            if "All" in sd or not sd: sd=["All"]
            sdt=pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt=pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p=str(poste).upper()
                if "All" not in sa:
                    m=False
                    if "Sulfurique (PS)" in sa and "PS" in p: m=True
                    if "Phosphorique (PP)" in sa and "PP" in p: m=True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m=True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m=True
                    if not m: return False
                if "All" not in sd:
                    m=False
                    if "SF1" in sd and "SF1" in p: m=True
                    if "SF2" in sd and "SF2" in p: m=True
                    if not m: return False
                return True

            vp=[p for p in apm if mf(p) and p in sp]
            
            df = df_full[(df_full["Poste travail princ."].isin(vp)) & (df_full["Date de début planifiée"].between(sdt,edt))].copy()
            avdf = av_full[av_full["Poste travail princ."].isin(vp)].copy()
            if "Créé le" in avdf.columns:
                avdf = avdf[avdf["Créé le"].between(sdt,edt)]
                
            df_dash = df_full[df_full["Poste travail princ."].isin(vp)].copy()
            avdf_dash = av_full[av_full["Poste travail princ."].isin(vp)].copy()
            
            res = calc_kpis(df, avdf, now_ts, vp)
            res_d = calc_kpis(df_dash, avdf_dash, now_ts, vp)

            ckdf=res['ckdf']; dfp=res['dfp']; avf=res['avf']; ckdf_d=res_d['ckdf']
            pa={k:round(ckdf[k].mean(),2) for k in QK}; qa={k:round(ckdf[k].mean(),2) for k in PK}
            pscores={}; qscores={}
            for poste in ckdf.index:
                r=ckdf.loc[poste]
                pscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0

            sf1_posts = [p for p in vp if str(p).startswith("SF1")]
            sf2_posts = [p for p in vp if str(p).startswith("SF2")]
            sf1_p_score = np.mean([pscores[p] for p in sf1_posts]) if sf1_posts else 0
            sf1_q_score = np.mean([qscores[p] for p in sf1_posts]) if sf1_posts else 0
            sf2_p_score = np.mean([pscores[p] for p in sf2_posts]) if sf2_posts else 0
            sf2_q_score = np.mean([qscores[p] for p in sf2_posts]) if sf2_posts else 0

            # ANOMALIES
            ano_map = {}
            ano_map["TAUX_REALISATION_CORRECTIF/PT"] = dfp[(dfp["Nº appel pl.entret."].fillna(0)==0)&(dfp["Contient SOPL"]==1)&(~dfp["Statut OT"].isin(["CLOT","TCLO"]))].groupby("Poste travail princ.")["Ordre"].count()
            prep_filt = (dfp["Statut OT"]=="CRÉÉ")&(dfp["Statut utilisateur"].str.contains("CRPR",na=False))
            ano_map["OT préparation <1 mois"] = dfp[prep_filt & (dfp["ap"]!="<1 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT préparation >3 mois"] = dfp[prep_filt & (dfp["ap"]==">3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT préparation 1mois< <3mois"] = dfp[prep_filt & (dfp["ap"]=="1 mois < <3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            plan_filt = (dfp["Statut OT"]=="LANC")&(dfp["Statut utilisateur"].str.contains("ATPL",case=False,na=False))
            ano_map["OT planification <1 mois"] = dfp[plan_filt & (dfp["alp"]!="<1 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT planification >3 mois"] = dfp[plan_filt & (dfp["alp"]==">3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT planification 1mois< <3mois"] = dfp[plan_filt & (dfp["alp"]=="1 mois < <3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            exec_filt = (dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==1)
            ano_map["OT exécution <1 mois"] = dfp[exec_filt & (dfp["aex"]!="<1 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT exécution >3 mois"] = dfp[exec_filt & (dfp["aex"]==">3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT exécution 1mois< <3mois"] = dfp[exec_filt & (dfp["aex"]=="1 mois < <3 mois")].groupby("Poste travail princ.")["Ordre"].count()
            perf_filt = (dfp["Contient SOPL"]==1)&(~dfp["Statut OT"].isin(["CLOT","TCLO"]))
            ano_map["Performance Graissage"] = dfp[perf_filt & (dfp["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Performance Inspection"] = dfp[perf_filt & (dfp["_tw_num"].isin([290,300,310]))&(dfp["Date de début planifiée"]<=now_ts)].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Performance Appels Systématiques"] = dfp[perf_filt & (dfp["_tw_num"]==360)&(dfp["Date de début planifiée"]<=now_ts)].groupby("Poste travail princ.")["Ordre"].count()
            avf_tot = avf.groupby("Poste travail princ.")["Avis"].count()
            avf_aprv = avf[avf["Statut utilisateur"].isin(["APRV","APRV AVAU"])].groupby("Poste travail princ.")["Avis"].count()
            ano_map["appel avis approuvé"] = avf_tot.sub(avf_aprv, fill_value=0)
            ano_map["OT LANC ESTIME"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["OT LANC ESTIME"]=="NON")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Backlog préparation caractérisé"] = dfp[(dfp["Statut OT"]=="CRÉÉ")&(dfp["Backlog preparation"]=="NON CARACTERISE")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["Backlog planification caractérisé"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["Backlog planification"]=="NON CARACTERISE")].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT CONFIME"] = dfp[dfp["OT CONFIME"]=="NON"].groupby("Poste travail princ.")["Ordre"].count()
            ano_map["OT_COR_EGAL"] = dfp[dfp["OT_COR_EGAL"]=="NON"].groupby("Poste travail princ.")["Ordre"].count()
            
            ano_p_rows = []
            for poste in vp:
                r = {"Poste de travail": poste}
                total = 0
                for kpi in QK:
                    cnt = int(ano_map.get(kpi, pd.Series()).get(poste, 0))
                    r[kpi] = cnt
                    total += cnt
                r["Total Anomalies"] = total
                ano_p_rows.append(r)
            tot_row_p = {"Poste de travail": "Total"}
            tot_tot = 0
            for kpi in QK:
                s = sum([r[kpi] for r in ano_p_rows])
                tot_row_p[kpi] = s
                tot_tot += s
            tot_row_p["Total Anomalies"] = tot_tot
            ano_p_rows.append(tot_row_p)

            ano_q_rows = []
            for poste in vp:
                r = {"Poste de travail": poste}
                total = 0
                for kpi in PK:
                    if kpi in ["OT Fiabilité", "Total Avis de Panne"]:
                        cnt = 0
                    else:
                        cnt = int(ano_map.get(kpi, pd.Series()).get(poste, 0))
                    r[kpi] = cnt
                    total += cnt
                r["Total Anomalies"] = total
                ano_q_rows.append(r)
            tot_row_q = {"Poste de travail": "Total"}
            tot_tot = 0
            for kpi in PK:
                s = sum([r[kpi] for r in ano_q_rows])
                tot_row_q[kpi] = s
                tot_tot += s
            tot_row_q["Total Anomalies"] = tot_tot
            ano_q_rows.append(tot_row_q)

            # ANOMALIES DETAILED EXPORT
            anomaly_dfs = {}
            anomaly_dfs["TAUX_REALISATION_CORRECTIF/PT"] = dfp[(dfp["Nº appel pl.entret."].fillna(0)==0)&(dfp["Contient SOPL"]==1)&(~dfp["Statut OT"].isin(["CLOT","TCLO"]))].copy()
            anomaly_dfs["OT préparation <1 mois"] = dfp[prep_filt & (dfp["ap"]!="<1 mois")].copy()
            anomaly_dfs["OT préparation >3 mois"] = dfp[prep_filt & (dfp["ap"]==">3 mois")].copy()
            anomaly_dfs["OT préparation 1mois< <3mois"] = dfp[prep_filt & (dfp["ap"]=="1 mois < <3 mois")].copy()
            anomaly_dfs["OT planification <1 mois"] = dfp[plan_filt & (dfp["alp"]!="<1 mois")].copy()
            anomaly_dfs["OT planification >3 mois"] = dfp[plan_filt & (dfp["alp"]==">3 mois")].copy()
            anomaly_dfs["OT planification 1mois< <3mois"] = dfp[plan_filt & (dfp["alp"]=="1 mois < <3 mois")].copy()
            anomaly_dfs["OT exécution <1 mois"] = dfp[exec_filt & (dfp["aex"]!="<1 mois")].copy()
            anomaly_dfs["OT exécution >3 mois"] = dfp[exec_filt & (dfp["aex"]==">3 mois")].copy()
            anomaly_dfs["OT exécution 1mois< <3mois"] = dfp[exec_filt & (dfp["aex"]=="1 mois < <3 mois")].copy()
            anomaly_dfs["Performance Graissage"] = dfp[perf_filt & (dfp["_tw_num"]==350)].copy()
            anomaly_dfs["Performance Inspection"] = dfp[perf_filt & (dfp["_tw_num"].isin([290,300,310]))&(dfp["Date de début planifiée"]<=now_ts)].copy()
            anomaly_dfs["Performance Appels Systématiques"] = dfp[perf_filt & (dfp["_tw_num"]==360)&(dfp["Date de début planifiée"]<=now_ts)].copy()
            anomaly_dfs["appel avis approuvé"] = avf[~avf["Statut utilisateur"].isin(["APRV","APRV AVAU"])].copy()
            anomaly_dfs["OT LANC ESTIME"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["OT LANC ESTIME"]=="NON")].copy()
            anomaly_dfs["Backlog préparation caractérisé"] = dfp[(dfp["Statut OT"]=="CRÉÉ")&(dfp["Backlog preparation"]=="NON CARACTERISE")].copy()
            anomaly_dfs["Backlog planification caractérisé"] = dfp[(dfp["Statut OT"]=="LANC")&(dfp["Backlog planification"]=="NON CARACTERISE")].copy()
            anomaly_dfs["OT CONFIME"] = dfp[dfp["OT CONFIME"]=="NON"].copy()
            anomaly_dfs["OT_COR_EGAL"] = dfp[dfp["OT_COR_EGAL"]=="NON"].copy()
            
            ot_cols = ["Ordre","Désignation","Poste technique","Désignation du poste technique","Poste travail princ.","Divis. planification","Statut système","Statut utilisateur","Date de début planifiée"]
            av_cols = ["Avis","Description","Poste technique","Poste travail princ.","Statut système","Statut utilisateur"]
            
            export_anomaly_dfs = {}
            for kpi, df_anom in anomaly_dfs.items():
                if not df_anom.empty:
                    if "Ordre" in df_anom.columns:
                        cols_exist = [c for c in ot_cols if c in df_anom.columns]
                    else:
                        cols_exist = [c for c in av_cols if c in df_anom.columns]
                    df_anom = df_anom[cols_exist].copy()
                    df_anom["kpi action"] = ACT_MAP.get(kpi, "")
                    export_anomaly_dfs[kpi] = df_anom
            
            if export_anomaly_dfs:
                buf_anom = io.BytesIO()
                with pd.ExcelWriter(buf_anom, engine='openpyxl') as writer:
                    for kpi, df_anom in export_anomaly_dfs.items():
                        sheet_name = str(kpi)[:31].replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")
                        df_anom.to_excel(writer, sheet_name=sheet_name, index=False)
                buf_anom.seek(0)
                st.sidebar.download_button("📥 Exporter Anomalies (Excel)", data=buf_anom, file_name="anomalies_kpis.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # TABLE ROWS
            pcols=["Poste de travail"]+QK+["Score Performance"]
            qcols=["Poste de travail"]+PK+["Score Qualite"]
            prows=[]; qrows=[]
            for poste in ckdf.index:
                r=ckdf.loc[poste]; prw={"Poste de travail":poste}
                for k in QK: prw[k]="%.1f"%r[k] if k in r.index else "0.0"
                prw["Score Performance"]="%.2f"%pscores.get(poste,0); prows.append(prw)
                qrw={"Poste de travail":poste}
                for k in PK: qrw[k]="%.1f"%r[k] if k in r.index else "0.0"
                qrw["Score Qualite"]="%.2f"%qscores.get(poste,0); qrows.append(qrw)
            cible_p={"Poste de travail":"CIBLE"}
            for k in QK: cible_p[k]="%.0f"%CIBLE.get(k,100)
            cible_p["Score Performance"]="%.0f"%100; prows.append({"_t":"cible",**cible_p})
            tot_p={"Poste de travail":"Total general"}
            for k in QK:
                vals=[float(rw[k]) for rw in prows if k in rw]
                tot_p[k]="%.1f"%(sum(vals)/len(vals)) if vals else "0.0"
            tot_p["Score Performance"]="%.2f"%(sum(pscores.values())/len(pscores)) if pscores else "0.00"
            prows.append({"_t":"total",**tot_p})
            cible_q={"Poste de travail":"CIBLE"}
            for k in PK: cible_q[k]="%.0f"%CIBLE.get(k,100)
            cible_q["Score Qualite"]="%.0f"%100; qrows.append({"_t":"cible",**cible_q})
            tot_q={"Poste de travail":"Total general"}
            for k in PK:
                vals=[float(rw[k]) for rw in qrows if k in rw]
                tot_q[k]="%.1f"%(sum(vals)/len(vals)) if vals else "0.0"
            tot_q["Score Qualite"]="%.2f"%(sum(qscores.values())/len(qscores)) if qscores else "0.00"
            qrows.append({"_t":"total",**tot_q})

            # BACKLOG PIVOTS
            prep_backlog_df = dfp[dfp["Statut OT"]=="CRÉÉ"].copy()
            plan_backlog_df = dfp[dfp["Statut OT"]=="LANC"].copy()
            
            piv_carac_prep_stat = pd.pivot_table(prep_backlog_df, index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)
            prep_carac_df = prep_backlog_df[prep_backlog_df["Backlog preparation"]=="CARACTERISE"]
            piv_carac_prep_type = pd.pivot_table(prep_carac_df, index="Poste travail princ.", columns="Type Carac Prep", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)
            
            piv_carac_plan_stat = pd.pivot_table(plan_backlog_df, index="Poste travail princ.", columns="Backlog planification", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)
            plan_carac_df = plan_backlog_df[plan_backlog_df["Backlog planification"]=="CARACTERISE"]
            piv_carac_plan_type = pd.pivot_table(plan_carac_df, index="Poste travail princ.", columns="Type Carac Plan", values="Ordre", aggfunc="count", fill_value=0).reindex(vp, fill_value=0)

            text_col=get_text_col(dfp)
            oms_df_sub=dfp[dfp[text_col].astype(str).str.contains("OMS",case=False,na=False)] if text_col else pd.DataFrame()
            thm_df_sub=dfp[dfp[text_col].astype(str).str.contains("THERMO|THERMOGRAPH",case=False,na=False)] if text_col else pd.DataFrame()
            piv_oms=build_statut_pivot(oms_df_sub,vp)
            piv_thm=build_statut_pivot(thm_df_sub,vp)
            piv_all=build_statut_pivot(dfp,vp)

            ano_p_cols = ["Poste de travail"] + QK + ["Total Anomalies"]
            ano_q_cols = ["Poste de travail"] + PK + ["Total Anomalies"]
            save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_rows,ano_p_cols,ano_q_rows,ano_q_cols,fichier_date)

            hist_filepath=os.path.join("kpis","indicateurs_kpis.xlsx")
            hist_df=load_historical_kpis(hist_filepath)
            var_df=calculate_variations(hist_df)
            journal_df=generate_journal(var_df)
            top5_df,bot5_df=calculate_rankings(var_df)

            # SYNTHESE DATA
            synth_perf={}; synth_qual={}
            if not var_df.empty and "Date precedente" in var_df.columns:
                for poste in vp:
                    synth_perf[poste]={}; synth_qual[poste]={}
                    pv=var_df[var_df["Poste"]==poste]
                    for kpi in QK:
                        kpi_v=pv[pv["KPI"]==kpi]
                        if not kpi_v.empty:
                            last=kpi_v.iloc[-1]
                            synth_perf[poste][kpi]={"diff":"%+.1f"%last["Ecart"]}
                        else:
                            synth_perf[poste][kpi]={"diff":"—"}
                    for kpi in PK:
                        kpi_v=pv[pv["KPI"]==kpi]
                        if not kpi_v.empty:
                            last=kpi_v.iloc[-1]
                            synth_qual[poste][kpi]={"diff":"%+.1f"%last["Ecart"]}
                        else:
                            synth_qual[poste][kpi]={"diff":"—"}

            # RENDER
            avg_p_score=sum(pa.values())/len(pa) if pa else 0
            avg_q_score=sum(qa.values())/len(qa) if qa else 0
            total_ano_p=sum([r["Total Anomalies"] for r in ano_p_rows if r.get("Poste de travail")!="Total"])
            total_ano_q=sum([r["Total Anomalies"] for r in ano_q_rows if r.get("Poste de travail")!="Total"])
            total_ot=len(df)

            # Construction des données pour le tableau de recommandations
            plan_actions_data = []
            for kpi, df_anom in anomaly_dfs.items():
                if not df_anom.empty and "Poste travail princ." in df_anom.columns:
                    postes = df_anom["Poste travail princ."].dropna().unique().tolist()
                    postes_str = ", ".join(postes)
                    mapping_resp = KPI_RESP_MAP.get(kpi, "Non assigné")
                    action_text = ACT_MAP.get(kpi, "Vérifier le KPI manuellement")
                    plan_actions_data.append({
                        "KPI en anomalie": kpi,
                        "Poste(s) de travail concerné(s)": postes_str,
                        "Responsable": mapping_resp,
                        "Action recommandée": action_text,
                        "Délai": ""
                    })
            
            df_plan_actions = pd.DataFrame(plan_actions_data)
            if not df_plan_actions.empty:
                df_plan_actions = df_plan_actions[["KPI en anomalie", "Poste(s) de travail concerné(s)", "Responsable", "Action recommandée", "Délai"]]

            logo_b64 = get_logo_base64()
            if logo_b64:
                st.markdown('<div class="mh"><img src="data:image/png;base64,%s" class="logo" alt="Logo"><h1>Tableau de Bord KPIs Performance & Qualite</h1><span class="db">📅 18/06/2026</span></div>'%logo_b64,unsafe_allow_html=True)
            else:
                st.markdown('<div class="mh"><h1>Tableau de Bord KPIs Performance & Qualite</h1><span class="db">📅 18/06/2026</span></div>',unsafe_allow_html=True)
            
            st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%d</div><div class="cl">OT Analyses</div></div><div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance Global</div></div><div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualite Global</div></div><div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies Totales</div></div></div>'%(total_ot,avg_p_score,avg_q_score,total_ano_p+total_ano_q),unsafe_allow_html=True)
            st.markdown('<div class="cr"><div class="cc c5"><div class="cv">%.1f%%</div><div class="cl">Performance SF1</div></div><div class="cc c6"><div class="cv">%.1f%%</div><div class="cl">Qualite SF1</div></div><div class="cc c7"><div class="cv">%.1f%%</div><div class="cl">Performance SF2</div></div><div class="cc c8"><div class="cv">%.1f%%</div><div class="cl">Qualite SF2</div></div></div>'%(sf1_p_score,sf1_q_score,sf2_p_score,sf2_q_score),unsafe_allow_html=True)

            tabs=st.tabs(["🏠 Tableau de Bord","📈 Performance","✅ Qualite","📂 Backlog","📋 Suivi & Evolution","🎯 Recommandations et Plan d'Actions"])

            with tabs[0]:
                st.markdown('<div class="stl p">Scores globaux par poste</div>',unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp,pscores,qscores,"Comparaison Performance / Qualite par poste"),unsafe_allow_html=True)
                col1,col2=st.columns(2)
                with col1:
                    st.markdown('<div class="stl p">Indicateurs de Performance</div>',unsafe_allow_html=True)
                    st.markdown(html_kpi_bars(QK,pa,CIBLE,"Taux moyens — Performance","#38a169","#e53e3e"),unsafe_allow_html=True)
                with col2:
                    st.markdown('<div class="stl q">Indicateurs de Qualite</div>',unsafe_allow_html=True)
                    st.markdown(html_kpi_bars(PK,qa,CIBLE,"Taux moyens — Qualite","#3182ce","#e53e3e"),unsafe_allow_html=True)
                st.markdown('<div class="stl c">Classement Performance</div>',unsafe_allow_html=True)
                st.markdown(html_classement(pscores,"#38a169"),unsafe_allow_html=True)
                st.markdown('<div class="stl c">Classement Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_classement(qscores,"#3182ce"),unsafe_allow_html=True)

            with tabs[1]:
                st.markdown('<div class="stl p">Detail des indicateurs de Performance</div>',unsafe_allow_html=True)
                st.markdown(html_table(prows,pcols,"pt",["Score Performance"]),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Nombre d\'anomalies par KPI et Poste (à traiter pour atteindre 100%)</div>',unsafe_allow_html=True)
                st.markdown(html_anomaly_table(ano_p_rows,ano_p_cols,"at"),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Actions recommandees — Performance</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table(QK,pa,CIBLE,ACT_MAP),unsafe_allow_html=True)

            with tabs[2]:
                st.markdown('<div class="stl q">Detail des indicateurs de Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_table(qrows,qcols,"qt",["Score Qualite"]),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Nombre d\'anomalies par KPI et Poste (à traiter pour atteindre 100%)</div>',unsafe_allow_html=True)
                st.markdown(html_anomaly_table(ano_q_rows,ano_q_cols,"at"),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Actions recommandees — Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table(PK,qa,CIBLE,ACT_MAP),unsafe_allow_html=True)

            with tabs[3]:
                st.markdown('<div class="stl c">Caractérisation Backlog Préparation</div>',unsafe_allow_html=True)
                c1, c2 = st.columns([0.6, 0.4])
                with c1:
                    st.markdown(html_generic_pivot(piv_carac_prep_stat, "omt", "Synthèse Caractérisé / Non Caractérisé"),unsafe_allow_html=True)
                with c2:
                    show_simple_pie(piv_carac_prep_stat, "Répartition Globale Caractérisé / Non Caractérisé", keep_non_carac=True)
                    show_simple_pie(piv_carac_prep_type, "Répartition par Type de Caractérisation", keep_non_carac=False)

                st.markdown('<div class="stl c">Caractérisation Backlog Planification</div>',unsafe_allow_html=True)
                c5, c6 = st.columns([0.6, 0.4])
                with c5:
                    st.markdown(html_generic_pivot(piv_carac_plan_stat, "omt", "Synthèse Caractérisé / Non Caractérisé"),unsafe_allow_html=True)
                with c6:
                    show_simple_pie(piv_carac_plan_stat, "Répartition Globale Caractérisé / Non Caractérisé", keep_non_carac=True)
                    show_simple_pie(piv_carac_plan_type, "Répartition par Type de Caractérisation", keep_non_carac=False)

                st.markdown('<div class="stl p">Statuts OT par Poste de Travail</div>',unsafe_allow_html=True)
                
                st.markdown('<div class="stl s">OT OMS par Poste et Statut OT</div>',unsafe_allow_html=True)
                c_oms1, c_oms2 = st.columns([0.55, 0.45])
                with c_oms1: st.markdown(html_statut_pivot(piv_oms,"omt"),unsafe_allow_html=True)
                with c_oms2: show_pie_pair(piv_oms,"OT OMS")
                
                st.markdown('<div class="stl s">OT Thermographie par Poste et Statut OT</div>',unsafe_allow_html=True)
                c_thm1, c_thm2 = st.columns([0.55, 0.45])
                with c_thm1: st.markdown(html_statut_pivot(piv_thm,"tht"),unsafe_allow_html=True)
                with c_thm2: show_pie_pair(piv_thm,"OT Thermographie")
                
                st.markdown('<div class="stl s">Tous les OT par Poste et Statut OT</div>',unsafe_allow_html=True)
                c_all1, c_all2 = st.columns([0.55, 0.45])
                with c_all1: st.markdown(html_statut_pivot(piv_all,"pt"),unsafe_allow_html=True)
                with c_all2: show_pie_pair(piv_all,"Tous les OT")

            with tabs[4]:
                min_date = var_df["Date precedente"].min() if not var_df.empty else "?"
                max_date = var_df["Date actuelle"].max() if not var_df.empty else "?"
                
                st.markdown(f'<div class="stl s">Synthèse d\'évolution Performance entre {min_date} et {max_date}</div>',unsafe_allow_html=True)
                if synth_perf and any(any(v.get("diff","—")!="—" for v in d.values()) for d in synth_perf.values()):
                    st.markdown(html_synthese_table(synth_perf,QK,vp),unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">Pas assez de donnees historiques pour calculer la synthese Performance. Au moins 2 periodes sont necessaires.</div>',unsafe_allow_html=True)
                    
                st.markdown(f'<div class="stl s">Synthèse d\'évolution Qualité entre {min_date} et {max_date}</div>',unsafe_allow_html=True)
                if synth_qual and any(any(v.get("diff","—")!="—" for v in d.values()) for d in synth_qual.values()):
                    st.markdown(html_synthese_table(synth_qual,PK,vp),unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">Pas assez de donnees historiques pour calculer la synthese Qualite. Au moins 2 periodes sont necessaires.</div>',unsafe_allow_html=True)

                st.markdown("---")
                st.markdown('<div class="stl s">Journal des variations significatives</div>',unsafe_allow_html=True)
                if not journal_df.empty:
                    st.dataframe(journal_df[["Date precedente","Date actuelle","Poste","Type","KPI","Valeur precedente","Valeur actuelle","Ecart %","Sens"]].reset_index(drop=True),use_container_width=True,height=400)
                else:
                    st.markdown('<div class="es">Aucune variation significative detectee (ecart >= 5%% entre deux periodes)</div>',unsafe_allow_html=True)

                if not top5_df.empty:
                    c1,c2=st.columns(2)
                    with c1:
                        st.markdown('<div class="stl p">Top 5 Postes — Amelioration</div>',unsafe_allow_html=True)
                        st.dataframe(top5_df,use_container_width=True)
                    with c2:
                        st.markdown('<div class="stl a">Bottom 5 Postes — Degradation</div>',unsafe_allow_html=True)
                        st.dataframe(bot5_df,use_container_width=True)

            with tabs[5]:
                st.markdown('<div class="stl a">Recommandations et Plan d\'Actions</div>',unsafe_allow_html=True)
                if not df_plan_actions.empty:
                    col_m1, col_m2 = st.columns([1, 4])
                    with col_m1:
                        st.metric(label="🔔 Recommandations Ouvertes", value=len(df_plan_actions))
                    
                    st.write("")
                    
                    # Hauteur augmentée à 1000 pour limiter la barre de défilement verticale si beaucoup de lignes
                    edited_plan_df = st.data_editor(
                        df_plan_actions,
                        column_config={
                            "Délai": st.column_config.TextColumn(
                                "Délai",
                                help="Renseignez le délai de traitement (ex: 15 jours, 30/06/2026)",
                                required=False
                            )
                        },
                        use_container_width=True,
                        hide_index=True,
                        num_rows="fixed",
                        height=1000
                    )
                    
                    st.write("")
                    
                    def to_excel_plan(df):
                        output = io.BytesIO()
                        writer = pd.ExcelWriter(output, engine='openpyxl')
                        df.to_excel(writer, index=False, sheet_name='Plan d actions')
                        writer.close()
                        return output.getvalue()
                    
                    excel_data_plan = to_excel_plan(edited_plan_df)
                    
                    st.download_button(
                        label="📥 Exporter le plan d'actions en Excel",
                        data=excel_data_plan,
                        file_name='plan_d_actions_kpis.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True
                    )
                else:
                    st.markdown('<div class="es">Aucune anomalie détectée. Tous les KPIs sont aux normes ! 🎉</div>', unsafe_allow_html=True)

        except Exception as e:
            st.error("Erreur lors du chargement des donnees : %s"%str(e))
            st.markdown('<div class="es">Veuillez verifier que les fichiers ot.xlsx et avis.xlsx sont presents dans le repertoire.</div>',unsafe_allow_html=True)
    else:
        st.markdown('<div class="es">📁 Veuillez charger les fichiers OT et AVIS via le panneau de filtres.</div>',unsafe_allow_html=True)

    st.markdown('<div class="footer">Bureau Méthodes Maroc Chimie – © 2026 Tous droits réservés</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    main()
