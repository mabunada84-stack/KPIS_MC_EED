# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# PAGE CONFIG — Must be the very first Streamlit command
# ============================================================
st.set_page_config(
    layout="wide",
    page_title="Dashboard KPI"
)

# ============================================================
# MODULE-LEVEL CONSTANTS
# ============================================================
QK = [
    "TAUX_REALISATION_CORRECTIF/PT", "OT préparation <1 mois",
    "OT préparation >3 mois", "OT préparation 1mois< <3mois",
    "OT planification <1 mois", "OT planification >3 mois",
    "OT planification 1mois< <3mois", "OT exécution <1 mois",
    "OT exécution >3 mois", "OT exécution 1mois< <3mois"
]
PK = [
    "appel avis approuvé", "OT LANC ESTIME",
    "Backlog préparation caractérisé", "Backlog planification caractérisé",
    "OT CONFIME", "OT_COR_EGAL"
]
ALL_KPI = QK + PK
CIBLE = {
    "TAUX_REALISATION_CORRECTIF/PT": 85, "OT préparation <1 mois": 80,
    "OT préparation >3 mois": 5, "OT préparation 1mois< <3mois": 15,
    "OT planification <1 mois": 80, "OT planification >3 mois": 5,
    "OT planification 1mois< <3mois": 15, "OT exécution <1 mois": 80,
    "OT exécution >3 mois": 5, "OT exécution 1mois< <3mois": 15,
    "appel avis approuvé": 95, "OT LANC ESTIME": 100,
    "Backlog préparation caractérisé": 100, "Backlog planification caractérisé": 100,
    "OT CONFIME": 100, "OT_COR_EGAL": 100
}
ACT_MAP = {
    "TAUX_REALISATION_CORRECTIF/PT": "Ameliorer le taux de realisation des OT.",
    "OT préparation <1 mois": "Reduire l'age de preparation des OT (< 1 mois).",
    "OT préparation >3 mois": "Traiter les OT avec preparation > 3 mois.",
    "OT planification <1 mois": "Reduire l'age de planification des OT (< 1 mois).",
    "OT planification >3 mois": "Traiter les OT avec planification > 3 mois.",
    "OT exécution <1 mois": "Reduire l'age d'execution des OT (< 1 mois).",
    "OT exécution >3 mois": "Traiter les OT avec execution > 3 mois.",
    "OT LANC ESTIME": "Estimer les couts des OT lances.",
    "Backlog préparation caractérisé": "Caracteriser le backlog de preparation.",
    "Backlog planification caractérisé": "Caracteriser le backlog de planification.",
    "OT CONFIME": "Confirmer les OT termines.",
    "OT_COR_EGAL": "Rapprocher les couts reels et budgetes.",
    "appel avis approuvé": "Creer un OT pour les avis sans ordre."
}
LOWER_BETTER = [
    "OT préparation >3 mois", "OT planification >3 mois", "OT exécution >3 mois",
    "OT préparation 1mois< <3mois", "OT planification 1mois< <3mois",
    "OT exécution 1mois< <3mois"
]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.",
    "Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.",
    "Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.",
    "Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.",
    "Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.",
    "Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.",
    "Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.",
    "Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.",
    "Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.",
    "Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.",
    "Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.",
    "Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.",
    "Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.",
    "Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.",
    "Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.",
    "Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.",
    "Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.",
    "Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.",
    "Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.",
    "Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.",
    "Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.",
    "Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.",
    "Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.",
    "Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.",
    "Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.",
    "Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.",
    "S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.",
    "Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.",
    "Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.",
    "Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.",
    "Zero accident commence par un comportement sur."
]


# ============================================================
# FILE HELPERS
# ============================================================
def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt", "r", encoding="utf-8") as f:
                return f.read().strip()
        except Exception:
            pass
    return datetime.now().strftime("%d/%m/%Y")


def save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_r, ano_p_c, ano_q_r, ano_q_c, sheet_name):
    kpis_dir = "kpis"
    os.makedirs(kpis_dir, exist_ok=True)
    filepath = os.path.join(kpis_dir, "indicateurs_kpis.xlsx")
    sn = str(sheet_name).replace("/", "-").replace("\\", "-").replace("*", "").replace(
        "?", "").replace("[", "").replace("]", "")[:31]
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfl = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    tf = Font(bold=True, size=12, color="1E3A5F")
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    try:
        wb = load_workbook(filepath)
    except Exception:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    if sn in wb.sheetnames:
        del wb[sn]
    ws = wb.create_sheet(sn)
    rn = 1

    def ws_section(title, cols, rows, sr):
        ws.cell(row=sr, column=1, value=title).font = tf
        sr += 1
        for j, c in enumerate(cols, 1):
            cl = ws.cell(row=sr, column=j, value=c)
            cl.font = hf; cl.fill = hfl
            cl.alignment = Alignment(horizontal='center'); cl.border = tb
        sr += 1
        for r in rows:
            for j, c in enumerate(cols, 1):
                cl = ws.cell(row=sr, column=j, value=r.get(c, ""))
                cl.border = tb; cl.alignment = Alignment(horizontal='center')
            sr += 1
        return sr + 1

    rn = ws_section("INDICATEURS DE PERFORMANCE", pcols, prows, rn)
    if ano_p_c and ano_p_r:
        rn = ws_section("ANOMALIES PERFORMANCE", ano_p_c, ano_p_r, rn)
    rn = ws_section("INDICATEURS DE QUALITE", qcols, qrows, rn)
    if ano_q_c and ano_q_r:
        rn = ws_section("ANOMALIES QUALITE", ano_q_c, ano_q_r, rn)
    try:
        wb.save(filepath)
    except Exception:
        pass


# ============================================================
# HISTORICAL KPI LOADER
# ============================================================
def load_historical_kpis(filepath):
    """Charge toutes les feuilles historiques et retourne un DataFrame."""
    if not os.path.exists(filepath):
        return pd.DataFrame()
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame()
    records = []
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            section = None
            headers = None
            for row in rows_data:
                cell0 = str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper():
                    section = "perf"; headers = None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper():
                    section = "qual"; headers = None; continue
                elif "ANOMALIES" in cell0.upper():
                    section = None; continue
                if section and headers is None and cell0:
                    headers = [str(c).strip() if c else "" for c in row]
                    continue
                if section and headers and cell0 and cell0 not in ("CIBLE", "Total general", ""):
                    entry = {"Date": sheet_name}
                    for j, h in enumerate(headers):
                        if j < len(row):
                            entry[h] = row[j]
                    entry["_section"] = section
                    records.append(entry)
        except Exception:
            continue
    wb.close()
    if not records:
        return pd.DataFrame()
    df = pd.DataFrame(records)
    df["Date_parsed"] = pd.to_datetime(df["Date"].str.replace("-", "/"), format="%d/%m/%Y", errors="coerce")
    df = df.sort_values("Date_parsed").reset_index(drop=True)
    return df


def calculate_variations(hist_df):
    """Calcule les variations entre dates consécutives."""
    if hist_df.empty or "Date" not in hist_df.columns:
        return pd.DataFrame()
    dates = sorted(hist_df["Date"].unique())
    if len(dates) < 2:
        return pd.DataFrame()
    # Pivot: rows=Date, columns=(Poste, KPI)
    perf_df = hist_df[hist_df["_section"] == "perf"].copy()
    qual_df = hist_df[hist_df["_section"] == "qual"].copy()
    variations = []
    for i in range(1, len(dates)):
        prev_date = dates[i - 1]
        curr_date = dates[i]
        prev_perf = perf_df[perf_df["Date"] == prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf = perf_df[perf_df["Date"] == curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual = qual_df[qual_df["Date"] == prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual = qual_df[qual_df["Date"] == curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for section_name, prev_d, curr_d, kpi_list in [
            ("Performance", prev_perf, curr_perf, QK + ["Score Performance"]),
            ("Qualite", prev_qual, curr_qual, PK + ["Score Qualite"])
        ]:
            common_posts = set(prev_d.index) & set(curr_d.index)
            for poste in common_posts:
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns:
                        continue
                    try:
                        pv = float(prev_d.loc[poste, kpi])
                    except (ValueError, TypeError, KeyError):
                        continue
                    try:
                        cv = float(curr_d.loc[poste, kpi])
                    except (ValueError, TypeError, KeyError):
                        continue
                    diff = cv - pv
                    pct = (diff / pv * 100) if pv != 0 else (100 if cv != 0 else 0)
                    if kpi in LOWER_BETTER:
                        trend = "hausse" if diff > 0.5 else ("baisse" if diff < -0.5 else "stabilite")
                    else:
                        trend = "hausse" if diff > 0.5 else ("baisse" if diff < -0.5 else "stabilite")
                    variations.append({
                        "Date precedente": prev_date, "Date actuelle": curr_date,
                        "Poste": poste, "Type": section_name, "KPI": kpi,
                        "Valeur precedente": round(pv, 2), "Valeur actuelle": round(cv, 2),
                        "Ecart": round(diff, 2), "Ecart %": round(pct, 2), "Tendance": trend
                    })
    return pd.DataFrame(variations)


def generate_journal(var_df):
    """Genere le journal des evolutions significatives."""
    if var_df.empty:
        return pd.DataFrame()
    journal = var_df.copy()
    journal["Significatif"] = journal["Ecart %"].abs() >= 5
    journal = journal[journal["Significatif"]].copy()
    journal["Sens"] = journal.apply(
        lambda r: "Amelioration" if (
            (r["Tendance"] == "hausse" and r["KPI"] not in LOWER_BETTER) or
            (r["Tendance"] == "baisse" and r["KPI"] in LOWER_BETTER)
        ) else "Degradation", axis=1
    )
    return journal.sort_values(["Date actuelle", "Sens", "Ecart %"], ascending=[True, False, False])


def calculate_rankings(var_df):
    """Calcule le Top 5 ameliorations et Top 5 degradations par poste."""
    if var_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    # Score de variation par poste (somme des écarts % pondérés)
    scores = {}
    for poste in var_df["Poste"].unique():
        pv = var_df[var_df["Poste"] == poste].copy()
        total = 0
        for _, r in pv.iterrows():
            kpi = r["KPI"]
            ec = r["Ecart %"]
            # Pour LOWER_BETTER, une baisse est positive
            if kpi in LOWER_BETTER:
                total -= ec
            else:
                total += ec
        scores[poste] = total
    ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    top5 = pd.DataFrame(ranked[:5], columns=["Poste", "Score variation"])
    bot5 = pd.DataFrame(ranked[-5:][::-1], columns=["Poste", "Score variation"])
    return top5, bot5


# ============================================================
# CSS INJECTION (modifiée selon les 5 points)
# ============================================================
def inject_custom_css():
    st.markdown("""<style>
    /* === POINT 2 : Sidebar 250px === */
    section[data-testid="stSidebar"] {
        width: 250px !important;
    }
    section[data-testid="stSidebar"][aria-expanded="false"] {
        width: 0px !important;
    }

    /* === POINT 5 : Pleine largeur === */
    .main .block-container{
        max-width: 100% !important;
        width: 100% !important;
        padding-left: 0.5rem !important;
        padding-right: 0.5rem !important;
    }

    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.6rem;padding-bottom:.6rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:10px 16px;border-radius:var(--r);margin-bottom:4px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:16px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:2px 10px;border-radius:14px;color:#fff;font-size:10px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:4px;margin-bottom:4px}
    .cc{background:#fff;border-radius:var(--r);padding:8px 10px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:22px;font-weight:900;line-height:1}
    .cc .cl{font-size:7px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:1px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:11px;font-weight:700;color:var(--p);margin:4px 0 1px 0;padding-left:8px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:8px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:7px;text-transform:uppercase;letter-spacing:.3px;padding:3px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw tbody td{padding:2px 3px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:8px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:8px!important}
    .stTabs [data-baseweb="tab-list"]{gap:2px;background:#e2e8f0;padding:2px;border-radius:6px;margin-bottom:3px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:5px 10px;font-weight:600;font-size:10px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:4px 8px;background:#fff;border-radius:5px;margin-bottom:1px;border:1px solid var(--b);font-size:9px}
    .sr .sn{font-weight:700;color:var(--p);min-width:200px;font-size:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:2px 7px;border-radius:12px;font-weight:800;font-size:10px;min-width:40px;text-align:center;margin:0 6px;color:#fff}
    .sr .sa{color:#718096;font-size:8px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .stg{font-size:7px;color:#718096;min-width:50px;text-align:center;white-space:nowrap}
    .sr .sb{font-size:7px;font-weight:700;padding:1px 5px;border-radius:3px;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:8px;margin-top:2px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:10px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:3px;font-size:8px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:240px;font-weight:600;color:var(--p);text-align:right;padding-right:6px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:22px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:8px;font-weight:800;color:#1a202c;min-width:50px;text-align:right;padding-left:4px}
    .gbr{display:flex;align-items:center;padding:2px 0;font-size:8px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:140px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:7px}
    .gbr-g{display:flex;align-items:center;gap:3px;flex:1}
    .gbr-w{flex:1;height:18px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}
    .gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:7px;font-weight:800;min-width:42px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:12px;margin-bottom:4px;font-size:8px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:4px}
    .gbr-legend i{display:inline-block;width:12px;height:12px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:4px}
    .cg>div{background:#fff;border-radius:var(--r);padding:6px 8px;border:1px solid var(--b)}
    .cg .ct{font-size:9px;font-weight:700;margin-bottom:2px;padding-bottom:2px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:2px 0;font-size:8px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:14px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:45px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:4px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:6px 12px;font-weight:700;font-size:11px;width:100%}
    ::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:2px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:9px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:2px 6px;margin-bottom:2px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:10px;color:#718096;font-size:10px}
    .rh{display:flex;align-items:center;justify-content:space-between;margin-bottom:0}
    .rh .stl{margin:0}
    .anl-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:9px;margin:0}
    .anl-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:8px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0}
    .anl-tbl tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7}
    .anl-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .anl-tbl tbody tr:hover td{background:#ebf8ff!important}
    .anl-tbl .tot td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .trend-up{color:#276749;font-weight:800;font-size:12px}
    .trend-down{color:#c53030;font-weight:800;font-size:12px}
    .trend-stable{color:#718096;font-weight:800;font-size:12px}
    .spark-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:6px}
    .spark-card{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .spark-card .sp-title{font-size:9px;font-weight:800;color:var(--p);margin-bottom:2px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .spark-card .sp-sub{font-size:7px;color:#718096;margin-bottom:4px}
    .rank-card{background:#fff;border-radius:var(--r);padding:10px 14px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .rank-card .rank-title{font-size:11px;font-weight:800;margin-bottom:6px;padding-bottom:4px;border-bottom:2px solid var(--b)}
    .rank-row{display:flex;align-items:center;padding:4px 0;font-size:9px;border-bottom:1px solid #f7fafc}
    .rank-row:last-child{border:none}
    .rank-row .rank-num{width:22px;height:22px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:10px;color:#fff;margin-right:8px;flex-shrink:0}
    .rank-row .rank-name{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .rank-row .rank-score{font-weight:900;min-width:60px;text-align:right}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:13px}.mh .db{float:none;display:block;margin-top:2px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:100px}.gbr-l{width:90px}.spark-grid{grid-template-columns:1fr}}
    </style>""", unsafe_allow_html=True)


# ============================================================
# MAIN
# ============================================================
def main():
    try:
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
    except Exception:
        try:
            locale.setlocale(locale.LC_ALL, 'fr_FR')
        except Exception:
            pass
    inject_custom_css()
    fichier_date = get_date_from_file()

    # HSE SPLASH
    if "hse_affiche" not in st.session_state:
        st.session_state.hse_affiche = False
    if not st.session_state.hse_affiche:
        c = random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:42px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:18px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:28px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:32px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>""" % c, unsafe_allow_html=True)
        time.sleep(6)
        st.session_state.hse_affiche = True
        st.rerun()
        st.stop()

    # ---------- HELPERS ----------
    def contient_mot(t, lm):
        t = str(t)
        return any(m in t for l in lm for m in l.split())

    def cat_age(a):
        if a <= 1: return "<1 mois"
        elif a >= 3: return ">3 mois"
        return "1 mois < <3 mois"

    def ckpi(n, d, sz=100):
        return np.where(d == 0, sz, (n / d) * 100)

    def cpiv(df, f, c, p):
        return pd.pivot_table(df[f], index="Poste travail princ.", columns=c,
                              values="Ordre", aggfunc="count", fill_value=0).reindex(p, fill_value=0)

    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur", case=False, na=False)].copy()
        return df

    def get_metier(p):
        p = str(p).upper()
        if "E" in p: return "Electrique"
        if "M" in p: return "Mecanique"
        if "R" in p: return "Instrumentation"
        if "G" in p: return "Genie Civil"
        return "Autre"

    def get_atelier(p):
        p = str(p).upper()
        if "PS" in p: return "Sulfurique"
        if "PP" in p: return "Phosphorique"
        if "TSP" in p or "REX" in p: return "Engrais"
        if "MCP" in p or "DCP" in p: return "Feed"
        return "Autre"

    def get_division(p):
        p = str(p).upper()
        if "SF1" in p: return "SF1"
        if "SF2" in p: return "SF2"
        return "Autre"

    # ---------- CALC KPI ----------
    def calc_kpis(df_i, av_i, now, posts):
        res = {}
        df = df_i.copy()
        av = av_i.copy()
        mp = ["CRPR ATPD", "CRPR ATMR", "CRPR ATER", "CRPR ATRS", "CRPR ATMO",
              "ATPD", "ATMR", "ATER", "ATRS", "ATMO"]
        df["Backlog preparation"] = np.where(
            df["Statut utilisateur"].apply(lambda x: contient_mot(x, mp)), "CARACTERISE", "NON CARACTERISE")
        mplan = ["ATPL ATEI", "ATPL ATAL", "ATPL ATER", "ATPL AGAR", "ATPL ATHS",
                 "ATEI", "ATAL", "ATAS", "AGAR", "ATHS"]
        df["Backlog planification"] = np.where(
            df["Statut utilisateur"].apply(lambda x: contient_mot(x, mplan)), "CARACTERISE", "NON CARACTERISE")
        for dc, am, ac in [('Créé le', "amp", "ap"), ('Date de début planifiée', "amlp", "alp"),
                            ('Date de début planifiée', "amex", "aex")]:
            if dc in df.columns:
                df[dc] = pd.to_datetime(df[dc], errors='coerce')
                df[am] = ((now.year - df[dc].dt.year) * 12 + (now.month - df[dc].dt.month)).round(2)
                df[ac] = df[am].apply(cat_age)
            else:
                df[am] = np.nan
                df[ac] = "Inconnu"
        df["OT CONFIME"] = np.where(
            df["Statut système"].str.contains("CLO", na=False) & df["Statut système"].str.contains("CONF", na=False),
            "OUI", "NON")
        df["Contient SOPL"] = df["Statut utilisateur"].str.contains("SOPL", na=False).map({True: 1, False: 0})
        df["OT LANC ESTIME"] = np.where(df["Total coûts budgétés"].fillna(0) == 0, "NON", "OUI")
        df["OT_COR_EGAL"] = np.where(
            (df["Total coûts budgétés"].fillna(0) - df["Total coûts réels"].fillna(0)) == 0, "OUI", "NON")
        res['dfp'] = df
        an = cpiv(df, df["Nº appel pl.entret."].fillna(0) == 0, "Statut OT", posts)
        for c in ["CLOT", "CRÉÉ", "LANC", "TCLO"]:
            an[c] = an.get(c, 0)
        an["Total"] = an[["CLOT", "CRÉÉ", "LANC", "TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"] = ckpi(an["TCLO"], an["Total"])
        pr = cpiv(df, df["Statut OT"] == "CRÉÉ", "ap", posts)
        for c in ["<1 mois", ">3 mois", "1 mois < <3 mois"]:
            pr[c] = pr.get(c, 0)
        pr["Total"] = pr[["<1 mois", "1 mois < <3 mois", ">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"] = ckpi(pr["<1 mois"], pr["Total"])
        pr["OT préparation >3 mois"] = ckpi(pr[">3 mois"], pr["Total"], 0)
        pr["OT préparation 1mois< <3mois"] = ckpi(pr["1 mois < <3 mois"], pr["Total"], 0)
        pl = cpiv(df, (df["Statut OT"] == "LANC") & (df["Contient SOPL"] == 0), "alp", posts)
        for c in ["<1 mois", ">3 mois", "1 mois < <3 mois"]:
            pl[c] = pl.get(c, 0)
        pl["Total"] = pl[["<1 mois", "1 mois < <3 mois", ">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"] = ckpi(pl["<1 mois"], pl["Total"])
        pl["OT planification >3 mois"] = ckpi(pl[">3 mois"], pl["Total"], 0)
        pl["OT planification 1mois< <3mois"] = ckpi(pl["1 mois < <3 mois"], pl["Total"], 0)
        ex = cpiv(df, (df["Statut OT"] == "LANC") & (df["Contient SOPL"] == 1), "aex", posts)
        for c in ["<1 mois", ">3 mois", "1 mois < <3 mois"]:
            ex[c] = ex.get(c, 0)
        ex["Total"] = ex[["<1 mois", "1 mois < <3 mois", ">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"] = ckpi(ex["<1 mois"], ex["Total"])
        ex["OT exécution >3 mois"] = ckpi(ex[">3 mois"], ex["Total"], 0)
        ex["OT exécution 1mois< <3mois"] = ckpi(ex["1 mois < <3 mois"], ex["Total"], 0)
        la = pd.pivot_table(df[df["Statut OT"] == "LANC"], index="Poste travail princ.",
                            columns="OT LANC ESTIME", values="Ordre", aggfunc="count",
                            fill_value=0).reindex(posts, fill_value=0)
        for c in ["OUI", "NON"]:
            la[c] = la.get(c, 0)
        la["Total"] = la["OUI"] + la["NON"]
        la["OT LANC ESTIME"] = ckpi(la["OUI"], la["Total"])
        pc = pd.pivot_table(df[df["Statut OT"] == "CRÉÉ"], index="Poste travail princ.",
                            columns="Backlog preparation", values="Ordre", aggfunc="count",
                            fill_value=0).reindex(posts, fill_value=0)
        for c in ["CARACTERISE", "NON CARACTERISE"]:
            pc[c] = pc.get(c, 0)
        pc["Total"] = pc["CARACTERISE"] + pc["NON CARACTERISE"]
        pc["Backlog préparation caractérisé"] = ckpi(pc["CARACTERISE"], pc["Total"])
        plc = pd.pivot_table(df[df["Statut OT"] == "LANC"], index="Poste travail princ.",
                             columns="Backlog planification", values="Ordre", aggfunc="count",
                             fill_value=0).reindex(posts, fill_value=0)
        for c in ["CARACTERISE", "NON CARACTERISE"]:
            plc[c] = plc.get(c, 0)
        plc["Total"] = plc["CARACTERISE"] + plc["NON CARACTERISE"]
        plc["Backlog planification caractérisé"] = ckpi(plc["CARACTERISE"], plc["Total"])
        for kn, cn in [("OT CONFIME", "OT CONFIME"), ("OT_COR_EGAL", "OT_COR_EGAL")]:
            pv = pd.pivot_table(df, index="Poste travail princ.", columns=cn,
                                values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["OUI", "NON"]:
                pv[c] = pv.get(c, 0)
            pv["Total"] = pv["OUI"] + pv["NON"]
            pv[cn] = ckpi(pv["OUI"], pv["Total"])
            res[kn.lower().replace(" ", "_")] = pv
        avf = av[(av["Ordre"].isna()) | (av["Ordre"].astype(str).str.strip() == "")].copy()
        res['avf'] = avf
        tca = pd.pivot_table(avf, index="Poste travail princ.", columns="Statut utilisateur",
                             values="Avis", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["APRQ", "APRV", "APRV AVAU", "REJT"]:
            tca[c] = tca.get(c, 0)
        tca["Total"] = tca[["APRQ", "APRV", "APRV AVAU", "REJT"]].sum(axis=1)
        tca["appel avis approuvé"] = ckpi(tca["APRV"], tca["Total"])
        res['ckdf'] = pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT": an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois": pr["OT préparation <1 mois"],
            "OT préparation >3 mois": pr["OT préparation >3 mois"],
            "OT préparation 1mois< <3mois": pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois": pl["OT planification <1 mois"],
            "OT planification >3 mois": pl["OT planification >3 mois"],
            "OT planification 1mois< <3mois": pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois": ex["OT exécution <1 mois"],
            "OT exécution >3 mois": ex["OT exécution >3 mois"],
            "OT exécution 1mois< <3mois": ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé": tca["appel avis approuvé"],
            "OT LANC ESTIME": la["OT LANC ESTIME"],
            "Backlog préparation caractérisé": pc["Backlog préparation caractérisé"],
            "Backlog planification caractérisé": plc["Backlog planification caractérisé"],
            "OT CONFIME": res['ot_confime']["OT CONFIME"],
            "OT_COR_EGAL": res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res

    # ---------- STYLING HELPERS ----------
    def ks(v, c):
        try:
            val = float(v)
        except Exception:
            return ""
        if c in ["OT préparation <1 mois", "OT planification <1 mois", "OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val >= 80 else (
                "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois", "OT planification 1mois< <3mois", "OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val <= 15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois", "OT planification >3 mois", "OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val <= 5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c == "TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val >= 85 else (
                "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c == "appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val >= 95 else (
                "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME", "Backlog préparation caractérisé", "Backlog planification caractérisé", "OT CONFIME", "OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val >= 100 else (
                "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""

    def cs(v):
        try:
            val = float(str(v).replace(' %', '').strip())
        except Exception:
            return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val >= 90 else (
            "background:#ffeb9c;color:#9c6500;font-weight:700" if val >= 80 else "background:#ffc7ce;color:#9c0006;font-weight:700")

    def kas(v):
        try:
            val = int(v)
        except Exception:
            return ""
        if val == 0: return "color:#cbd5e0"
        if val <= 3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val <= 10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"

    def gscore(k, a, t):
        if pd.isna(a) or pd.isna(t):
            return 0
        if k in ["OT préparation <1 mois", "OT planification <1 mois", "OT exécution <1 mois"]:
            return 1 if a >= 75 else 0
        if k in ["OT préparation 1mois< <3mois", "OT planification 1mois< <3mois", "OT exécution 1mois< <3mois"]:
            return 1 if a <= 15 else 0
        if k in ["OT préparation >3 mois", "OT planification >3 mois", "OT exécution >3 mois"]:
            return 1 if a <= 5 else 0
        if k == "TAUX_REALISATION_CORRECTIF/PT":
            return 1 if a >= 80 else 0
        if k == "appel avis approuvé":
            return 1 if a >= 90 else 0
        if k in ["OT LANC ESTIME", "Backlog préparation caractérisé", "Backlog planification caractérisé", "OT CONFIME", "OT_COR_EGAL"]:
            return 1 if a >= 95 else 0
        return 0

    def is_lb(k):
        return k in LOWER_BETTER

    # ---------- HTML RENDERERS ----------
    def html_table(rows, cols, tc, sc_col=None):
        h = '<table class="tw %s"><thead><tr>' % tc + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
        for r in rows:
            rc = "cb" if r.get("_t") == "cible" else ("tr" if r.get("_t") == "total" else "")
            h += '<tr class="%s">' % rc
            for c in cols:
                v = r.get(c, "")
                if r.get("_t") == "cible":
                    h += '<td>%s</td>' % v
                else:
                    s = cs(v) if sc_col and c in sc_col else ks(v, c)
                    h += '<td style="%s">%s</td>' % (s or "", v)
            h += '</tr>'
        return h + '</tbody></table>'

    def html_ano(rows, cols):
        h = '<table class="tw at"><thead><tr>' + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
        for r in rows:
            h += '<tr class="%s">' % ("tr" if r.get("_t") == "total" else "")
            for c in cols:
                v = r.get(c, "")
                h += '<td style="%s">%s</td>' % (kas(v) or "", v)
            h += '</tr>'
        return h + '</tbody></table>'

    def html_synth(kpi_list, actuals, targets, act_map, accent):
        h = ''
        for k in kpi_list:
            av = actuals.get(k, 0)
            tv = targets.get(k, 100)
            met = av <= tv if is_lb(k) else av >= tv
            sbg, sclr = ("#c6efce", "#006100") if met else ("#ffc7ce", "#9c0006")
            scbg = accent if met else "#e53e3e"
            act = "Objectif atteint" if met else act_map.get(k, "")
            h += '<div class="sr"><div class="sn">%s</div><div class="sc" style="background:%s">%.1f%%</div><div class="stg">Cible: %s%%</div><div class="sb" style="color:%s;background:%s">%s</div><div class="sa">%s</div></div>' % (
                k, scbg, av, tv, sclr, sbg, "ATTEINT" if met else "NON ATTEINT", act)
        return h

    def html_classement(scores, accent):
        sp = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        met_p = [(p, s) for p, s in sp if s >= 80]
        not_p = [(p, s) for p, s in sp if s < 80]
        t5 = met_p[:5]
        b5 = not_p[-5:] if len(not_p) > 5 else not_p
        h = '<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i, (p, s) in enumerate(t5):
                h += '<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (
                    accent, i + 1, p, cs("%.2f" % s), s)
        else:
            h += '<div style="padding:4px;font-size:8px;color:#718096">Aucun poste</div>'
        h += '</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i, (p, s) in enumerate(reversed(b5)):
                h += '<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (
                    len(b5) - i, p, cs("%.2f" % s), s)
        else:
            h += '<div style="padding:4px;font-size:8px;color:#38a169">Tous atteints</div>'
        h += '</div></div>'
        return h

    def html_kpi_bars(kpi_list, actuals, targets, title, color_ok, color_fail):
        h = '<div class="ca"><div class="ct" style="color:%s">%s</div>' % (color_ok, title)
        for k in kpi_list:
            av = actuals.get(k, 0)
            tv = targets.get(k, 100)
            met = av <= tv if is_lb(k) else av >= tv
            bw = min(max(av, 0), 100)
            bg = color_ok if met else color_fail
            h += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>' % (
                k, bw, bg, av)
        h += '</div>'
        return h

    def html_bars(data, title, color):
        h = '<div class="ca"><div class="ct" style="color:%s">%s</div>' % (color, title)
        for label, val in sorted(data, key=lambda x: x[1], reverse=True):
            bw = min(max(val, 0), 100)
            h += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>' % (
                label, bw, color, val)
        h += '</div>'
        return h

    def html_grouped_bars(posts, pscores, qscores, title):
        h = '<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>' % title
        h += '<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        sp2 = sorted(posts, key=lambda x: (pscores.get(x, 0) + qscores.get(x, 0)) / 2, reverse=True)
        for p in sp2:
            pv = pscores.get(p, 0)
            qv = qscores.get(p, 0)
            pw = min(max(pv, 0), 100)
            qw = min(max(qv, 0), 100)
            h += '<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>' % (
                p, pw, pv, qw, qv)
        h += '</div>'
        return h

    def anl_pie_chart(data, names_col, values_col, title, colors=None):
        if data.empty:
            return None
        fig = px.pie(data, names=names_col, values=values_col, title=title,
                     color_discrete_sequence=colors or px.colors.qualitative.Set2)
        fig.update_traces(textposition='inside', textinfo='percent+label+value', textfont_size=9)
        # === POINT 4 : height=450, autosize=True ===
        fig.update_layout(margin=dict(t=40, b=10, l=10, r=10), height=450, autosize=True,
                          title_font_size=11,
                          legend=dict(font_size=8, orientation="h", yanchor="bottom", y=-0.15))
        return fig

    def anl_html_table(df_out, pct_col=None, pct_thresh=(80, 60)):
        h = '<table class="anl-tbl"><thead><tr>'
        for c in df_out.columns:
            h += '<th>%s</th>' % c
        h += '</tr></thead><tbody>'
        for idx, row in df_out.iterrows():
            is_tot = str(idx) == "TOTAL" or str(row.iloc[0]) == "TOTAL"
            rc = "tot" if is_tot else ""
            h += '<tr class="%s">' % rc
            for c in df_out.columns:
                v = row[c]
                s = ""
                if pct_col and c == pct_col and not is_tot:
                    try:
                        pv = float(str(v).replace('%', '').strip())
                        s = "g-green" if pv >= pct_thresh[0] else ("g-yellow" if pv >= pct_thresh[1] else "g-red")
                    except Exception:
                        pass
                if isinstance(v, float):
                    v = round(v, 1)
                h += '<td class="%s">%s</td>' % (s, v)
            h += '</tr>'
        return h + '</tbody></table>'

    def export_btn(df, filename):
        buf = io.BytesIO()
        df.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)
        st.download_button("📥 Exporter Excel", data=buf, file_name=filename,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:18px;margin-bottom:2px">⚙️</div><div style="font-size:12px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:8px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",
                    unsafe_allow_html=True)
        st.markdown("---")

        # === POINT 3 : Checkbox pour masquer les filtres ===
        show_filters = st.checkbox("Afficher les filtres", value=True, key="show_filters")

        if show_filters:
            unf = st.toggle("📁 Charger nouveaux fichiers", value=False, key="tf")
            ot_f = av_f = None
            apm = []
            if unf:
                ot_f = st.file_uploader("Fichier OT", type=["xlsx"], key="uot")
                av_f = st.file_uploader("Fichier AVIS", type=["xlsx"], key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t = excr(pd.read_excel("ot.xlsx"))
                        apm = sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(
                            ("SF1", "SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception:
                        pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:5px 8px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:7px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:10px;color:white;font-weight:600;margin-top:1px">📅 %s</div></div>""" % fichier_date,
                            unsafe_allow_html=True)
            st.markdown("---")
            st.markdown("**🎯 Postes**")
            sp = st.multiselect("Poste", ["All"] + apm, ["All"], key="sp")
            st.markdown("**🏭 Atelier**")
            sa = st.multiselect("Atelier",
                                ["All", "Sulfurique (PS)", "Phosphorique (PP)", "Engrais (TSP/REX)", "Feed (MCP/DCP)"],
                                ["All"], key="sa")
            st.markdown("**🏢 Division**")
            sd = st.multiselect("Division", ["All", "SF1", "SF2"], ["All"], key="sd")
            st.markdown("---")
            st.markdown("**📅 Periode**")
            dr = st.date_input("Date debut planifiee",
                               value=(datetime(2025, 1, 1).date(), datetime.today().date()),
                               format="DD/MM/YYYY", key="dr")
        else:
            unf = False
            ot_f = av_f = None
            apm = []
            sp = ["All"]
            sa = ["All"]
            sd = ["All"]
            dr = (datetime(2025, 1, 1).date(), datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try:
                    _t = excr(pd.read_excel("ot.xlsx"))
                    apm = sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(
                        ("SF1", "SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception:
                    pass

    # ===================== DATA LOADING =====================
    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf:
                raw_ot = pd.read_excel(ot_f)
                raw_av = pd.read_excel(av_f)
            else:
                raw_ot = pd.read_excel("ot.xlsx")
                raw_av = pd.read_excel("avis.xlsx")
            raw_ot = excr(raw_ot)
            raw_av = excr(raw_av)
            for c in ["Créé le", "Date de début planifiée", "Date de clôture", "Début réel", "Fin réelle"]:
                if c in raw_ot.columns:
                    raw_ot[c] = pd.to_datetime(raw_ot[c], errors="coerce")
            for c in ["Créé le", "Début souhaité", "Date de la clôture"]:
                if c in raw_av.columns:
                    raw_av[c] = pd.to_datetime(raw_av[c], errors="coerce")
            if not apm:
                apm = sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(
                    ("SF1", "SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp:
                sp = apm
            if "All" in sa or not sa:
                sa = ["All"]
            if "All" in sd or not sd:
                sd = ["All"]
            sdt = pd.to_datetime(dr[0]) if len(dr) == 2 else pd.to_datetime(datetime(2025, 1, 1))
            edt = pd.to_datetime(dr[1]) if len(dr) == 2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p = str(poste).upper()
                if "All" not in sa:
                    m = False
                    if "Sulfurique (PS)" in sa and "PS" in p: m = True
                    if "Phosphorique (PP)" in sa and "PP" in p: m = True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m = True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m = True
                    if not m: return False
                if "All" not in sd:
                    m = False
                    if "SF1" in sd and "SF1" in p: m = True
                    if "SF2" in sd and "SF2" in p: m = True
                    if not m: return False
                return True

            vp = [p for p in apm if mf(p) and p in sp]

            df = raw_ot[(raw_ot["Poste travail princ."].isin(vp)) & (
                raw_ot["Date de début planifiée"].between(sdt, edt))].copy()
            avdf = raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df = excr(df[df["Poste travail princ."].astype(str).str.startswith(
                ("SF1", "SF2"), na=False)].drop_duplicates())
            avdf = excr(avdf[(avdf["Ordre"].isna()) | (
                avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns:
                df["Statut OT"] = df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            df_dash = raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash = excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(
                ("SF1", "SF2"), na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns:
                df_dash["Statut OT"] = df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now = pd.Timestamp.now()
            res = calc_kpis(df, avdf, now, vp)
            ckdf = res['ckdf']
            dfp = res['dfp']
            res_d = calc_kpis(df_dash, avdf, now, vp)
            ckdf_d = res_d['ckdf']

            pa = {k: round(ckdf[k].mean(), 2) for k in QK}
            qa = {k: round(ckdf[k].mean(), 2) for k in PK}
            pa_d = {k: round(ckdf_d[k].mean(), 2) for k in QK}
            qa_d = {k: round(ckdf_d[k].mean(), 2) for k in PK}

            pscores = {}
            qscores = {}
            for poste in ckdf.index:
                r = ckdf.loc[poste]
                pscores[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in QK if k in r.index) / len(QK) * 100) if QK else 0
                qscores[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in PK if k in r.index) / len(PK) * 100) if PK else 0

            pscores_d = {}
            qscores_d = {}
            for poste in ckdf_d.index:
                r = ckdf_d.loc[poste]
                pscores_d[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in QK if k in r.index) / len(QK) * 100) if QK else 0
                qscores_d[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in PK if k in r.index) / len(PK) * 100) if PK else 0

            # ANOMALIES
            all_ano = []
            sub_p = {
                "TAUX_REALISATION_CORRECTIF/PT": lambda d: d[(d["Nº appel pl.entret."].fillna(0) == 0) & (~d["Statut OT"].isin(["CLOT", "TCLO"]))],
                "OT préparation <1 mois": lambda d: d[(d["Statut OT"] == "CRÉÉ") & (d["ap"] != "<1 mois")],
                "OT préparation >3 mois": lambda d: d[(d["Statut OT"] == "CRÉÉ") & (d["ap"] == ">3 mois")],
                "OT planification <1 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 0) & (d["alp"] != "<1 mois")],
                "OT planification >3 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 0) & (d["alp"] == ">3 mois")],
                "OT exécution <1 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 1) & (d["aex"] != "<1 mois")],
                "OT exécution >3 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 1) & (d["aex"] == ">3 mois")]
            }
            sub_q = {
                "OT LANC ESTIME": lambda d: d[(d["Statut OT"] == "LANC") & (d["OT LANC ESTIME"] == "NON")],
                "Backlog préparation caractérisé": lambda d: d[(d["Statut OT"] == "CRÉÉ") & (d["Backlog preparation"] == "NON CARACTERISE")],
                "Backlog planification caractérisé": lambda d: d[(d["Statut OT"] == "LANC") & (d["Backlog planification"] == "NON CARACTERISE")],
                "OT CONFIME": lambda d: d[d["OT CONFIME"] == "NON"],
                "OT_COR_EGAL": lambda d: d[d["OT_COR_EGAL"] == "NON"]
            }
            for poste in vp:
                if poste not in dfp["Poste travail princ."].values:
                    continue
                dp = dfp[dfp["Poste travail princ."] == poste]
                for kn, sf in sub_p.items():
                    vk = ckdf.loc[poste, kn] if poste in ckdf.index else 100
                    if pd.notna(vk) and vk < CIBLE[kn]:
                        cnt = len(sf(dp))
                        if cnt > 0:
                            all_ano.append({"Poste": poste, "KPI": kn, "Nb": cnt, "Type": "P"})
                for kn, sf in sub_q.items():
                    vk = ckdf.loc[poste, kn] if poste in ckdf.index else 100
                    if pd.notna(vk) and vk < CIBLE[kn]:
                        cnt = len(sf(dp))
                        if cnt > 0:
                            all_ano.append({"Poste": poste, "KPI": kn, "Nb": cnt, "Type": "Q"})
                vk_av = ckdf.loc[poste, "appel avis approuvé"] if poste in ckdf.index else 100
                if pd.notna(vk_av) and vk_av < CIBLE["appel avis approuvé"]:
                    cnt = len(res['avf'][res['avf']["Poste travail princ."] == poste])
                    if cnt > 0:
                        all_ano.append({"Poste": poste, "KPI": "appel avis approuvé", "Nb": cnt, "Type": "Q"})

            def build_ano(ano_list, kpi_list):
                if not ano_list:
                    return [], []
                adf = pd.DataFrame(ano_list)
                pv = adf.pivot_table(index="Poste", columns="KPI", values="Nb", aggfunc="sum", fill_value=0).astype(int)
                pv["Total"] = pv.sum(axis=1)
                tot = pv.sum()
                cols = [c for c in kpi_list if c in pv.columns] + ["Total"]
                rows = []
                for idx in pv.index:
                    r = {"_t": "n", "Poste de travail": idx}
                    for c in cols:
                        r[c] = pv.loc[idx, c]
                    rows.append(r)
                tr = {"_t": "total", "Poste de travail": "Total general"}
                for c in cols:
                    tr[c] = int(tot[c])
                rows.append(tr)
                return ["Poste de travail"] + cols, rows

            ano_p_c, ano_p_r = build_ano([a for a in all_ano if a["Type"] == "P"], QK)
            ano_q_c, ano_q_r = build_ano([a for a in all_ano if a["Type"] == "Q"], PK)

            def build_kpi(kpi_list, scores, sname):
                sp2 = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)
                cols = ["Poste de travail"] + kpi_list + [sname]
                rows = []
                cr = {"_t": "cible", "Poste de travail": "CIBLE"}
                for k in kpi_list:
                    cr[k] = CIBLE[k]
                cr[sname] = "100.00 %"
                rows.append(cr)
                for p in sp2:
                    r = {"_t": "n", "Poste de travail": p}
                    for k in kpi_list:
                        r[k] = round(ckdf.loc[p, k], 2) if p in ckdf.index else ""
                    r[sname] = "%.2f %%" % scores[p]
                    rows.append(r)
                tr = {"_t": "total", "Poste de travail": "Total general"}
                for k in kpi_list:
                    tr[k] = round(ckdf[k].mean(), 2)
                tr[sname] = "%.2f %%" % (np.mean(list(scores.values())) if scores else 0)
                rows.append(tr)
                return cols, rows

            pcols, prows = build_kpi(QK, pscores, "Score Performance")
            qcols, qrows = build_kpi(PK, qscores, "Score Qualite")

            # Sauvegarde Excel
            save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_r, ano_p_c, ano_q_r, ano_q_c, fichier_date)

            # Donnees dashboard
            df_sc_d = pd.DataFrame([{"Poste": p, "Perf": pscores_d[p], "Qual": qscores_d[p],
                                     "Metier": get_metier(p), "Atelier": get_atelier(p),
                                     "Division": get_division(p)} for p in vp if p in pscores_d])
            by_at = df_sc_d.groupby("Atelier")[["Perf", "Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()
            by_mt = df_sc_d.groupby("Metier")[["Perf", "Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()
            by_div = df_sc_d.groupby("Division")[["Perf", "Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()

            total_ot = len(df)
            avg_p = np.mean(list(pscores.values())) if pscores else 0
            avg_q = np.mean(list(qscores.values())) if qscores else 0
            total_ano = sum(a["Nb"] for a in all_ano)

            # Designation column
            desig_col = None
            for cn in ["Désignation du travail", "Designation du travail", "Désignation", "Designation", "Description"]:
                if cn in dfp.columns:
                    desig_col = cn
                    break

            # ========== HISTORICAL DATA LOADING ==========
            kpis_file = os.path.join("kpis", "indicateurs_kpis.xlsx")
            hist_df = load_historical_kpis(kpis_file)
            var_df = calculate_variations(hist_df)
            journal_df = generate_journal(var_df)
            top5_imp, top5_deg = calculate_rankings(var_df)

            # ===================== RENDER =====================
            st.markdown('<div class="mh"><h1>📊 KPI Dashboard MC & FEED</h1><div class="db">📅 %s</div></div>' % fichier_date,
                        unsafe_allow_html=True)
            st.markdown("""<div class="cr">
            <div class="cc c1"><div class="cv">%s</div><div class="cl">Total OT Analyses</div></div>
            <div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div>
            <div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div>
            <div class="cc c4"><div class="cv">%s</div><div class="cl">Total Anomalies</div></div>
            </div>""" % (total_ot, avg_p, avg_q, total_ano), unsafe_allow_html=True)

            tab0, tab1, tab2, tab3, tab4 = st.tabs([
                "📊 TABLEAU DE BORD", "📈 INDICATEURS PERFORMANCE",
                "✅ INDICATEUR QUALITE", "🔬 ANALYSE", "📉 SUIVI DES AMELIORATIONS"
            ])

            # ==================== DASHBOARD ====================
            with tab0:
                st.markdown('<div class="stl p">📊 Vue d\'ensemble par poste</div>', unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp, pscores_d, qscores_d, "Performance & Qualite par Poste de Travail"),
                            unsafe_allow_html=True)

                st.markdown('<div class="stl p">🏭 Par Atelier</div>', unsafe_allow_html=True)
                if not by_at.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx, row["Perf"]) for idx, row in by_at.iterrows()],
                                              "Performance par Atelier", "#2b6cb0"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx, row["Qual"]) for idx, row in by_at.iterrows()],
                                              "Qualite par Atelier", "#276749"), unsafe_allow_html=True)

                st.markdown('<div class="stl p">🔧 Par Metier</div>', unsafe_allow_html=True)
                if not by_mt.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx, row["Perf"]) for idx, row in by_mt.iterrows()],
                                              "Performance par Metier", "#2b6cb0"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx, row["Qual"]) for idx, row in by_mt.iterrows()],
                                              "Qualite par Metier", "#276749"), unsafe_allow_html=True)

                st.markdown('<div class="stl p">🏢 Par Division</div>', unsafe_allow_html=True)
                if not by_div.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx, row["Perf"]) for idx, row in by_div.iterrows()],
                                              "Performance par Division", "#2b6cb0"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx, row["Qual"]) for idx, row in by_div.iterrows()],
                                              "Qualite par Division", "#276749"), unsafe_allow_html=True)

                st.markdown('<div class="stl c">Classement des postes</div>', unsafe_allow_html=True)
                st.markdown(html_classement(pscores, "#2b6cb0"), unsafe_allow_html=True)

                # Pie charts
                if "Statut OT" in df.columns:
                    sc = df["Statut OT"].value_counts().reset_index()
                    sc.columns = ["Statut", "Nb"]
                    if not sc.empty:
                        st.markdown('<div class="stl c">Repartition par statut OT</div>', unsafe_allow_html=True)
                        fig_pie = anl_pie_chart(sc, "Statut", "Nb", "Repartition des OT par statut")
                        if fig_pie:
                            st.plotly_chart(fig_pie, use_container_width=True)

            # ==================== PERFORMANCE ====================
            with tab1:
                st.markdown('<div class="stl p">📈 Indicateurs de Performance</div>', unsafe_allow_html=True)
                st.markdown(html_table(prows, pcols, "pt", ["Score Performance"]), unsafe_allow_html=True)
                if ano_p_c and ano_p_r:
                    st.markdown('<div class="stl a">⚠️ Anomalies Performance</div>', unsafe_allow_html=True)
                    st.markdown(html_ano(ano_p_r, ano_p_c), unsafe_allow_html=True)
                st.markdown('<div class="stl p">Synthese Performance</div>', unsafe_allow_html=True)
                st.markdown(html_synth(QK, pa, CIBLE, ACT_MAP, "#276749"), unsafe_allow_html=True)
                st.markdown(html_kpi_bars(QK, pa, CIBLE, "Barres de progression Performance", "#276749", "#e53e3e"),
                            unsafe_allow_html=True)

            # ==================== QUALITE ====================
            with tab2:
                st.markdown('<div class="stl q">✅ Indicateurs de Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_table(qrows, qcols, "qt", ["Score Qualite"]), unsafe_allow_html=True)
                if ano_q_c and ano_q_r:
                    st.markdown('<div class="stl a">⚠️ Anomalies Qualite</div>', unsafe_allow_html=True)
                    st.markdown(html_ano(ano_q_r, ano_q_c), unsafe_allow_html=True)
                st.markdown('<div class="stl q">Synthese Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_synth(PK, qa, CIBLE, ACT_MAP, "#2b6cb0"), unsafe_allow_html=True)
                st.markdown(html_kpi_bars(PK, qa, CIBLE, "Barres de progression Qualite", "#2b6cb0", "#e53e3e"),
                            unsafe_allow_html=True)

            # ==================== ANALYSE ====================
            with tab3:
                st.markdown('<div class="stl c">🔬 Analyse detaillee</div>', unsafe_allow_html=True)
                if not df_sc_d.empty:
                    # Analyse par metier
                    st.markdown('<div class="stl p">Performance par Metier</div>', unsafe_allow_html=True)
                    mt_a = df_sc_d.groupby("Metier")[["Perf", "Qual"]].mean().round(1).reset_index()
                    mt_a.columns = ["Metier", "Score Performance", "Score Qualite"]
                    st.markdown(anl_html_table(mt_a, "Score Performance"), unsafe_allow_html=True)

                    # Analyse par atelier
                    st.markdown('<div class="stl p">Performance par Atelier</div>', unsafe_allow_html=True)
                    at_a = df_sc_d.groupby("Atelier")[["Perf", "Qual"]].mean().round(1).reset_index()
                    at_a.columns = ["Atelier", "Score Performance", "Score Qualite"]
                    st.markdown(anl_html_table(at_a, "Score Performance"), unsafe_allow_html=True)

                    # Analyse par division
                    st.markdown('<div class="stl p">Performance par Division</div>', unsafe_allow_html=True)
                    dv_a = df_sc_d.groupby("Division")[["Perf", "Qual"]].mean().round(1).reset_index()
                    dv_a.columns = ["Division", "Score Performance", "Score Qualite"]
                    st.markdown(anl_html_table(dv_a, "Score Performance"), unsafe_allow_html=True)

                    # Postes sous seuil
                    st.markdown('<div class="stl a">Postes sous seuil critique (<60%%)</div>', unsafe_allow_html=True)
                    low_posts = df_sc_d[(df_sc_d["Perf"] < 60) | (df_sc_d["Qual"] < 60)].sort_values("Perf")
                    if not low_posts.empty:
                        lp = low_posts[["Poste", "Perf", "Qual", "Metier", "Atelier", "Division"]].copy()
                        lp.columns = ["Poste", "Perf %%", "Qual %%", "Metier", "Atelier", "Division"]
                        st.markdown(anl_html_table(lp, "Perf %%"), unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="es">Aucun poste sous le seuil critique</div>', unsafe_allow_html=True)

                # Export
                st.markdown("---")
                exp_data = ckdf.copy()
                exp_data["Score Performance"] = exp_data.index.map(lambda p: "%.2f" % pscores.get(p, 0))
                exp_data["Score Qualite"] = exp_data.index.map(lambda p: "%.2f" % qscores.get(p, 0))
                exp_data = exp_data.reset_index().rename(columns={"index": "Poste"})
                export_btn(exp_data, "kpi_export_%s.xlsx" % fichier_date.replace("/", "-"))

            # ==================== SUIVI DES AMELIORATIONS ====================
            with tab4:
                st.markdown('<div class="mh" style="margin-bottom:8px"><h1>📉 Suivi des Améliorations</h1><div class="db">Historique & Variations</div></div>',
                            unsafe_allow_html=True)

                if hist_df.empty or len(hist_df["Date"].unique()) < 2:
                    st.markdown("""<div class="es" style="padding:40px">
                    <div style="font-size:48px;margin-bottom:12px">📅</div>
                    <div style="font-size:16px;font-weight:700;color:#1e3a5f;margin-bottom:6px">Donnees historiques insuffisantes</div>
                    <div style="font-size:12px;color:#718096">Le suivi des ameliorations necessite au moins 2 periodes d'enregistrement.<br>
                    Les donnees sont sauvegardees automatiquement dans <code>kpis/indicateurs_kpis.xlsx</code> a chaque execution.</div>
                    </div>""", unsafe_allow_html=True)
                else:
                    dates_list = sorted(hist_df["Date"].unique())
                    nb_dates = len(dates_list)

                    # KPI cards
                    nb_ameliorations = len(journal_df[journal_df["Sens"] == "Amelioration"]) if not journal_df.empty else 0
                    nb_degradations = len(journal_df[journal_df["Sens"] == "Degradation"]) if not journal_df.empty else 0
                    st.markdown("""<div class="cr">
                    <div class="cc c2"><div class="cv">%s</div><div class="cl">Periodes enregistrees</div></div>
                    <div class="cc c1"><div class="cv">%s</div><div class="cl">Ameliorations significatives</div></div>
                    <div class="cc c4"><div class="cv">%s</div><div class="cl">Degradations significatives</div></div>
                    <div class="cc c3"><div class="cv">%s</div><div class="cl">Total variations calculees</div></div>
                    </div>""" % (nb_dates, nb_ameliorations, nb_degradations, len(var_df)), unsafe_allow_html=True)

                    # --- CLASSEMENTS ---
                    st.markdown('<div class="stl s">🏆 Classements des postes</div>', unsafe_allow_html=True)
                    rc1, rc2 = st.columns(2)
                    with rc1:
                        if not top5_imp.empty:
                            h = '<div class="rank-card"><div class="rank-title" style="color:#276749;border-bottom-color:#276749">🥇 Top 5 Ameliorations</div>'
                            colors = ["#276749", "#38a169", "#48bb78", "#68d391", "#9ae6b4"]
                            for i, (_, row) in enumerate(top5_imp.iterrows()):
                                h += '<div class="rank-row"><div class="rank-num" style="background:%s">%s</div><div class="rank-name">%s</div><div class="rank-score" style="color:#276749">%+.1f pts</div></div>' % (
                                    colors[i], i + 1, row["Poste"], row["Score variation"])
                            h += '</div>'
                            st.markdown(h, unsafe_allow_html=True)
                    with rc2:
                        if not top5_deg.empty:
                            h = '<div class="rank-card"><div class="rank-title" style="color:#c53030;border-bottom-color:#c53030">🔻 Top 5 Degradations</div>'
                            colors = ["#c53030", "#e53e3e", "#fc8181", "#feb2b2", "#fed7d7"]
                            for i, (_, row) in enumerate(top5_deg.iterrows()):
                                h += '<div class="rank-row"><div class="rank-num" style="background:%s">%s</div><div class="rank-name">%s</div><div class="rank-score" style="color:#c53030">%+.1f pts</div></div>' % (
                                    colors[i], i + 1, row["Poste"], row["Score variation"])
                            h += '</div>'
                            st.markdown(h, unsafe_allow_html=True)

                    # --- TABLE VARIATIONS ---
                    st.markdown('<div class="stl s">📊 Analyse detaillee des variations</div>', unsafe_allow_html=True)

                    # Filtres par periode
                    sc1, sc2, sc3 = st.columns(3)
                    with sc1:
                        sel_type = st.selectbox("Type", ["Tous", "Performance", "Qualite"], key="var_type")
                    with sc2:
                        sel_poste_var = st.selectbox("Poste", ["Tous"] + sorted(var_df["Poste"].unique().tolist()), key="var_poste")
                    with sc3:
                        sel_periode = st.selectbox("Periode", ["Toutes"] + [f"{dates_list[i]} → {dates_list[i+1]}" for i in range(len(dates_list)-1)], key="var_periode")

                    filtered_var = var_df.copy()
                    if sel_type != "Tous":
                        filtered_var = filtered_var[filtered_var["Type"] == sel_type]
                    if sel_poste_var != "Tous":
                        filtered_var = filtered_var[filtered_var["Poste"] == sel_poste_var]
                    if sel_periode != "Toutes":
                        parts = sel_periode.split(" → ")
                        if len(parts) == 2:
                            filtered_var = filtered_var[
                                (filtered_var["Date precedente"] == parts[0].strip()) &
                                (filtered_var["Date actuelle"] == parts[1].strip())]

                    if not filtered_var.empty:
                        # Build HTML table with trend arrows
                        vcols = ["Poste", "Type", "KPI", "Valeur precedente", "Valeur actuelle", "Ecart", "Ecart %", "Tendance"]
                        vh = '<table class="tw st"><thead><tr>' + ''.join('<th>%s</th>' % c for c in vcols) + '</tr></thead><tbody>'
                        for _, r in filtered_var.iterrows():
                            # Tendance arrow
                            kpi = r["KPI"]
                            if kpi in LOWER_BETTER:
                                if r["Tendance"] == "baisse":
                                    arrow = '<span class="trend-up">▲</span>'
                                elif r["Tendance"] == "hausse":
                                    arrow = '<span class="trend-down">▼</span>'
                                else:
                                    arrow = '<span class="trend-stable">►</span>'
                            else:
                                if r["Tendance"] == "hausse":
                                    arrow = '<span class="trend-up">▲</span>'
                                elif r["Tendance"] == "baisse":
                                    arrow = '<span class="trend-down">▼</span>'
                                else:
                                    arrow = '<span class="trend-stable">►</span>'
                            # Color for ecart %
                            ep = r["Ecart %"]
                            if ep > 0.5:
                                ec_s = "color:#276749;font-weight:700" if kpi not in LOWER_BETTER else "color:#c53030;font-weight:700"
                            elif ep < -0.5:
                                ec_s = "color:#c53030;font-weight:700" if kpi not in LOWER_BETTER else "color:#276749;font-weight:700"
                            else:
                                ec_s = "color:#718096"
                            vh += '<tr><td>%s</td><td>%s</td><td style="font-weight:600">%s</td><td>%.2f</td><td>%.2f</td><td>%+.2f</td><td style="%s">%+.2f%%</td><td style="text-align:center">%s</td></tr>' % (
                                r["Poste"], r["Type"], r["KPI"], r["Valeur precedente"], r["Valeur actuelle"],
                                r["Ecart"], ec_s, r["Ecart %"], arrow)
                        vh += '</tbody></table>'
                        st.markdown(vh, unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="es">Aucune variation trouvee pour les filtres selectionnes</div>',
                                    unsafe_allow_html=True)

                    # --- JOURNAL DES EVOLUTIONS ---
                    st.markdown('<div class="stl s">📝 Journal des evolutions significatives</div>', unsafe_allow_html=True)
                    if not journal_df.empty:
                        jcols = ["Date actuelle", "Poste", "Type", "KPI", "Valeur precedente", "Valeur actuelle", "Ecart %", "Sens"]
                        jh = '<table class="tw st"><thead><tr>' + ''.join('<th>%s</th>' % c for c in jcols) + '</tr></thead><tbody>'
                        for _, r in journal_df.iterrows():
                            sens_color = "background:#c6efce;color:#006100;font-weight:700" if r["Sens"] == "Amelioration" else "background:#ffc7ce;color:#9c0006;font-weight:700"
                            jh += '<tr><td>%s</td><td>%s</td><td>%s</td><td style="font-weight:600">%s</td><td>%.2f</td><td>%.2f</td><td>%+.2f%%</td><td style="%s">%s</td></tr>' % (
                                r["Date actuelle"], r["Poste"], r["Type"], r["KPI"],
                                r["Valeur precedente"], r["Valeur actuelle"], r["Ecart %"],
                                sens_color, r["Sens"])
                        jh += '</tbody></table>'
                        st.markdown(jh, unsafe_allow_html=True)

                        # Export journal
                        export_btn(journal_df[["Date precedente", "Date actuelle", "Poste", "Type", "KPI",
                                               "Valeur precedente", "Valeur actuelle", "Ecart", "Ecart %", "Tendance", "Sens"]],
                                   "journal_evolutions.xlsx")
                    else:
                        st.markdown('<div class="es">Aucune evolution significative (>5%%) detectee</div>',
                                    unsafe_allow_html=True)

                    # --- SPARKLINES / COURBES D'EVOLUTION ---
                    st.markdown('<div class="stl s">📈 Courbes d\'evolution par poste</div>', unsafe_allow_html=True)

                    # Filtre pour sparklines
                    sk1, sk2 = st.columns(2)
                    with sk1:
                        sel_score_type = st.selectbox("Indicateur", ["Score Performance", "Score Qualite"], key="spark_type")
                    with sk2:
                        sel_poste_spark = st.multiselect("Postes (vide = tous)", sorted(var_df["Poste"].unique().tolist()), key="spark_postes")

                    score_col = "Score Performance" if sel_score_type == "Score Performance" else "Score Qualite"
                    spark_posts = sel_poste_spark if sel_poste_spark else sorted(var_df["Poste"].unique().tolist())

                    # Build sparkline data
                    spark_data = hist_df[hist_df["_section"] == ("perf" if score_col == "Score Performance" else "qual")].copy()
                    if "Poste de travail" in spark_data.columns and score_col in spark_data.columns:
                        spark_pivot = spark_data.pivot_table(
                            index="Date", columns="Poste de travail",
                            values=score_col, aggfunc="first"
                        ).reindex(dates_list)
                        spark_pivot.index = pd.to_datetime(spark_pivot.index.str.replace("-", "/"), format="%d/%m/%Y", errors="coerce")

                        # Grid of sparkline cards
                        spark_html = '<div class="spark-grid">'
                        for poste in spark_posts:
                            if poste not in spark_pivot.columns:
                                continue
                            series = spark_pivot[poste].dropna()
                            if len(series) < 2:
                                continue
                            first_val = series.iloc[0]
                            last_val = series.iloc[-1]
                            diff = last_val - first_val
                            trend_icon = "▲" if diff > 0.5 else ("▼" if diff < -0.5 else "►")
                            trend_clr = "#276749" if diff > 0.5 else ("#c53030" if diff < -0.5 else "#718096")

                            # Create mini Plotly chart
                            fig_sp = go.Figure()
                            fig_sp.add_trace(go.Scatter(
                                x=series.index, y=series.values,
                                mode='lines+markers',
                                line=dict(color='#2b6cb0' if score_col == "Score Performance" else '#276749', width=2),
                                marker=dict(size=4),
                                fill='tozeroy',
                                fillcolor='rgba(43,108,176,0.08)' if score_col == "Score Performance" else 'rgba(39,103,73,0.08)'
                            ))
                            # Add target line
                            target_val = CIBLE.get(score_col.replace("Score ", ""), 100)
                            fig_sp.add_hline(y=target_val, line_dash="dash", line_color="#e53e3e", line_width=1,
                                             annotation_text="Cible", annotation_position="top right",
                                             annotation_font_size=7, annotation_font_color="#e53e3e")
                            fig_sp.update_layout(
                                height=160, autosize=True,
                                margin=dict(l=35, r=10, t=5, b=25),
                                xaxis=dict(tickfont=dict(size=7), showgrid=False),
                                yaxis=dict(tickfont=dict(size=7), showgrid=True, gridcolor='#edf2f7'),
                                showlegend=False,
                                plot_bgcolor='white'
                            )
                            spark_html += '<div class="spark-card">'
                            spark_html += '<div class="sp-title">%s</div>' % poste
                            spark_html += '<div class="sp-sub">%s: %.1f%% → %.1f%% <span style="color:%s;font-weight:800">%s %+.1f</span></div>' % (
                                score_col, first_val, last_val, trend_clr, trend_icon, diff)
                            spark_html += '</div>'
                        spark_html += '</div>'
                        st.markdown(spark_html, unsafe_allow_html=True)

                        # Plotly charts in columns
                        n_cols = min(3, len(spark_posts))
                        if n_cols > 0:
                            cols_sp = st.columns(n_cols)
                            for idx, poste in enumerate(spark_posts):
                                if poste not in spark_pivot.columns:
                                    continue
                                series = spark_pivot[poste].dropna()
                                if len(series) < 2:
                                    continue
                                with cols_sp[idx % n_cols]:
                                    fig_s = go.Figure()
                                    fig_s.add_trace(go.Scatter(
                                        x=series.index, y=series.values,
                                        mode='lines+markers+text',
                                        text=[f"{v:.1f}" for v in series.values],
                                        textposition='top center',
                                        textfont=dict(size=8),
                                        line=dict(
                                            color='#2b6cb0' if score_col == "Score Performance" else '#276749',
                                            width=2.5),
                                        marker=dict(size=6, color='#fff',
                                                    line=dict(width=2,
                                                              color='#2b6cb0' if score_col == "Score Performance" else '#276749')),
                                        fill='tozeroy',
                                        fillcolor='rgba(43,108,176,0.06)' if score_col == "Score Performance" else 'rgba(39,103,73,0.06)'
                                    ))
                                    target_val = CIBLE.get(score_col.replace("Score ", ""), 100)
                                    fig_s.add_hline(y=target_val, line_dash="dash", line_color="#e53e3e",
                                                     line_width=1.5,
                                                     annotation_text="Cible %.0f%%" % target_val,
                                                     annotation_position="top right",
                                                     annotation_font_size=8, annotation_font_color="#e53e3e")
                                    first_v = series.iloc[0]
                                    last_v = series.iloc[-1]
                                    diff_v = last_v - first_v
                                    arrow = "▲" if diff_v > 0.5 else ("▼" if diff_v < -0.5 else "►")
                                    clr = "#276749" if diff_v > 0.5 else ("#c53030" if diff_v < -0.5 else "#718096")
                                    fig_s.update_layout(
                                        height=220, autosize=True,
                                        title=dict(text='<b>%s</b><br><span style="font-size:10px;color:%s">%s %+.1f pts</span>' % (
                                            poste, clr, arrow, diff_v),
                                                   font_size=10, x=0.01, xanchor='left'),
                                        margin=dict(l=40, r=15, t=50, b=30),
                                        xaxis=dict(tickfont=dict(size=8), showgrid=False,
                                                   tickformat='%d/%m'),
                                        yaxis=dict(tickfont=dict(size=8), showgrid=True,
                                                   gridcolor='#edf2f7', range=[0, 110]),
                                        showlegend=False, plot_bgcolor='white'
                                    )
                                    st.plotly_chart(fig_s, use_container_width=True)
                    else:
                        st.markdown('<div class="es">Donnees insuffisantes pour les courbes</div>', unsafe_allow_html=True)

                    # --- EVOLUTION GLOBALE ---
                    st.markdown('<div class="stl s">📉 Evolution globale des scores moyens</div>', unsafe_allow_html=True)
                    # Compute average scores per date
                    for s_type, s_section, s_color in [
                        ("Performance", "perf", "#2b6cb0"),
                        ("Qualite", "qual", "#276749")
                    ]:
                        sdata = hist_df[hist_df["_section"] == s_section].copy()
                        score_col_name = "Score " + s_type
                        if "Poste de travail" in sdata.columns and score_col_name in sdata.columns:
                            avg_per_date = sdata.groupby("Date")[score_col_name].mean()
                            avg_per_date.index = pd.to_datetime(avg_per_date.index.str.replace("-", "/"),
                                                                format="%d/%m/%Y", errors="coerce")
                            avg_per_date = avg_per_date.sort_index()
                            if len(avg_per_date) >= 2:
                                fig_g = go.Figure()
                                fig_g.add_trace(go.Scatter(
                                    x=avg_per_date.index, y=avg_per_date.values,
                                    mode='lines+markers+text',
                                    text=[f"{v:.1f}%%" for v in avg_per_date.values],
                                    textposition='top center',
                                    textfont=dict(size=10, color=s_color),
                                    line=dict(color=s_color, width=3),
                                    marker=dict(size=8, color='#fff',
                                                line=dict(width=2.5, color=s_color)),
                                    fill='tozeroy',
                                    fillcolor=s_color.replace(')', ',0.06)').replace('rgb', 'rgba')
                                ))
                                fig_g.update_layout(
                                    height=300, autosize=True,
                                    title=dict(text=f"Evolution moyenne {s_type}", font_size=12,
                                               font_color=s_color),
                                    margin=dict(l=50, r=20, t=40, b=30),
                                    xaxis=dict(tickfont=dict(size=9), showgrid=False,
                                               tickformat='%d/%m/%Y'),
                                    yaxis=dict(tickfont=dict(size=9), showgrid=True,
                                               gridcolor='#edf2f7', range=[0, 105]),
                                    showlegend=False, plot_bgcolor='white'
                                )
                                st.plotly_chart(fig_g, use_container_width=True)

        except Exception as e:
            st.error(f"Erreur de chargement : {str(e)}")
            import traceback
            st.code(traceback.format_exc())


if __name__ == "__main__":
    main()
