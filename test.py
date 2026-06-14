# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os
from datetime import datetime
import plotly.express as px
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt", "r", encoding="utf-8") as f:
                return f.read().strip()
        except: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_r, ano_p_c, ano_q_r, ano_q_c, var_p_r, var_p_c, var_q_r, var_q_c, sheet_name):
    kpis_dir = "kpis"
    os.makedirs(kpis_dir, exist_ok=True)
    filepath = os.path.join(kpis_dir, "indicateurs_kpis.xlsx")
    sn = str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfl = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hfl2 = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    tf = Font(bold=True, size=12, color="1E3A5F")
    tb = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb = load_workbook(filepath)
    except:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws = wb.create_sheet(sn)
    rn = 1
    def ws_section(title, cols, rows, sr, hdr_fill=None):
        ws.cell(row=sr, column=1, value=title).font = tf; sr += 1
        for j, c in enumerate(cols, 1):
            cl = ws.cell(row=sr, column=j, value=c); cl.font = hf; cl.fill = hdr_fill or hfl; cl.alignment = Alignment(horizontal='center'); cl.border = tb
        sr += 1
        for r in rows:
            for j, c in enumerate(cols, 1):
                cl = ws.cell(row=sr, column=j, value=r.get(c, "")); cl.border = tb; cl.alignment = Alignment(horizontal='center')
            sr += 1
        return sr + 1
    rn = ws_section("INDICATEURS DE PERFORMANCE", pcols, prows, rn)
    if ano_p_c and ano_p_r: rn = ws_section("ANOMALIES PERFORMANCE", ano_p_c, ano_p_r, rn)
    rn = ws_section("INDICATEURS DE QUALITE", qcols, qrows, rn)
    if ano_q_c and ano_q_r: rn = ws_section("ANOMALIES QUALITE", ano_q_c, ano_q_r, rn)
    if var_p_c and var_p_r: rn = ws_section("VARIANCE PERFORMANCE (Periode vs Reference)", var_p_c, var_p_r, rn, hfl2)
    if var_q_c and var_q_r: rn = ws_section("VARIANCE QUALITE (Periode vs Reference)", var_q_c, var_q_r, rn, hfl2)
    try: wb.save(filepath)
    except: pass

def inject_custom_css():
    st.markdown("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important;padding-top:.6rem;padding-bottom:.6rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:10px 16px;border-radius:var(--r);margin-bottom:4px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{font-size:24px;color:#fff;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:2px 10px;border-radius:14px;color:#fff;font-size:10px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:4px;margin-bottom:4px}
    .cc{background:#fff;border-radius:var(--r);padding:8px 10px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:32px;font-weight:900;line-height:1}
    .cc .cl{font-size:12px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:1px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .cc.c5{border-top:3px solid #e65100}.cc.c5 .cv{color:#e65100}
    .cc.c6{border-top:3px solid #00838f}.cc.c6 .cv{color:#00838f}
    .stl{font-size:16px;font-weight:700;color:var(--p);margin:4px 0 1px 0;padding-left:8px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.v{border-left-color:#e65100}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:12px;text-transform:uppercase;letter-spacing:.3px;padding:6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.vt thead th{background:linear-gradient(135deg,#e65100,#ff8f00)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap;font-size:12px}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:2px;background:#e2e8f0;padding:2px;border-radius:6px;margin-bottom:3px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:5px 10px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:4px 8px;background:#fff;border-radius:5px;margin-bottom:1px;border:1px solid var(--b);font-size:12px}
    .sr .sn{font-weight:700;color:var(--p);min-width:200px;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:2px 7px;border-radius:12px;font-weight:800;font-size:13px;min-width:40px;text-align:center;margin:0 6px;color:#fff}
    .sr .sa{color:#718096;font-size:12px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .stg{font-size:12px;color:#718096;min-width:50px;text-align:center;white-space:nowrap}
    .sr .sb{font-size:12px;font-weight:700;padding:1px 5px;border-radius:3px;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:8px;margin-top:2px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:10px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:3px;font-size:8px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:160px;font-weight:600;color:var(--p);text-align:right;padding-right:6px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:22px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:8px;font-weight:800;color:#1a202c;min-width:50px;text-align:right;padding-left:4px}
    .gbr{display:flex;align-items:center;padding:2px 0;font-size:8px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:140px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:7px}
    .gbr-g{display:flex;align-items:center;gap:3px;flex:1}
    .gbr-w{flex:1;height:18px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}
    .gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gb-v{background:linear-gradient(90deg,#e65100,#ff8f00)}
    .gbr-v{font-size:7px;font-weight:800;min-width:42px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:12px;margin-bottom:4px;font-size:8px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:4px}
    .gbr-legend i{display:inline-block;width:12px;height:12px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:4px}
    .cg>div{background:#fff;border-radius:var(--r);padding:6px 8px;border:1px solid var(--b)}
    .cg .ct{font-size:9px;font-weight:700;margin-bottom:2px;padding-bottom:2px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:2px 0;font-size:8px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:14px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:45px;text-align:right}
    .cgr .pv{font-weight:800;min-width:55px;text-align:right;font-size:7px}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:4px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:6px 12px;font-weight:700;font-size:11px;width:100%}
    ::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:2px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:9px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:2px 6px;margin-bottom:2px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:10px;color:#718096;font-size:10px}
    .anl-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;margin:0}
    .anl-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:12px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0}
    .anl-tbl tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7}
    .anl-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .anl-tbl tbody tr:hover td{background:#ebf8ff!important}
    .anl-tbl .tot td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .v-pos{background:#c6efce;color:#006100;font-weight:700}
    .v-neg{background:#ffc7ce;color:#9c0006;font-weight:700}
    .v-zero{background:#f0f0f0;color:#718096;font-weight:600}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:18px}.mh .db{float:none;display:block;margin-top:2px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:100px}.gbr-l{width:90px}}
    </style>""", unsafe_allow_html=True)

def main():
    try: locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
    except:
        try: locale.setlocale(locale.LC_ALL, 'fr_FR')
        except: pass
    inject_custom_css()
    fichier_date = get_date_from_file()

    qk = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois","OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois","OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois","OT exécution 1mois< <3mois"]
    pk = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
    cible = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,"OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,"OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,"OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,"Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,"OT CONFIME":100,"OT_COR_EGAL":100}
    act_map = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.","OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).","OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.","OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).","OT planification >3 mois":"Traiter les OT avec planification > 3 mois.","OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).","OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.","OT LANC ESTIME":"Estimer les couts des OT lances.","Backlog préparation caractérisé":"Caracteriser le backlog de preparation.","Backlog planification caractérisé":"Caracteriser le backlog de planification.","OT CONFIME":"Confirmer les OT termines.","OT_COR_EGAL":"Rapprocher les couts reels et budgetes.","appel avis approuvé":"Creer un OT pour les avis sans ordre."}

    consignes = ["Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.","Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.","Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.","Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.","Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.","Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.","Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.","Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.","Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.","Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.","Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.","Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.","Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.","Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.","Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.","Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.","Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.","Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.","Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.","Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.","Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.","Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.","Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.","Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.","Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.","Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.","Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.","Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.","La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.","Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche = False
    if not st.session_state.hse_affiche:
        c = random.choice(consignes)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:42px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:18px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:28px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:32px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>""" % c, unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche = True; st.rerun(); st.stop()

    def contient_mot(t, lm):
        t = str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if a <= 1: return "<1 mois"
        elif a >= 3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n, d, sz=100): return np.where(d == 0, sz, (n / d) * 100)
    def cpiv(df, f, c, p):
        return pd.pivot_table(df[f], index="Poste travail princ.", columns=c, values="Ordre", aggfunc="count", fill_value=0).reindex(p, fill_value=0)
    def excr(df):
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur", case=False, na=False)].copy() if "Poste travail princ." in df.columns else df
    def get_metier(p):
        p = str(p).upper()
        if "E" in p: return "Electrique"
        if "M" in p: return "Mecanique"
        if "R" in p: return "Instrumentation"
        if "G" in p: return "Genie Civil"
        return "Autre"
    def get_atelier(p):
        p = str(p).upper()
        if "PS" in p: return "Sulfurique"
        if "PP" in p: return "Phosphorique"
        if "TSP" in p or "REX" in p: return "Engrais"
        if "MCP" in p or "DCP" in p: return "Feed"
        return "Autre"
    def get_division(p):
        p = str(p).upper()
        if "SF1" in p: return "SF1"
        if "SF2" in p: return "SF2"
        return "Autre"

    def calc_kpis(df_i, av_i, now, posts):
        res = {}; df = df_i.copy(); av = av_i.copy()
        mp = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
        df["Backlog preparation"] = np.where(df["Statut utilisateur"].apply(lambda x: contient_mot(x, mp)), "CARACTERISE", "NON CARACTERISE")
        mplan = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
        df["Backlog planification"] = np.where(df["Statut utilisateur"].apply(lambda x: contient_mot(x, mplan)), "CARACTERISE", "NON CARACTERISE")
        for dc, am, ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc] = pd.to_datetime(df[dc], errors='coerce')
                df[am] = ((now.year - df[dc].dt.year)*12 + (now.month - df[dc].dt.month)).round(2)
                df[ac] = df[am].apply(cat_age)
            else: df[am] = np.nan; df[ac] = "Inconnu"
        df["OT CONFIME"] = np.where(df["Statut système"].str.contains("CLO", na=False) & df["Statut système"].str.contains("CONF", na=False), "OUI", "NON")
        df["Contient SOPL"] = df["Statut utilisateur"].str.contains("SOPL", na=False).map({True:1, False:0})
        df["OT LANC ESTIME"] = np.where(df["Total coûts budgétés"].fillna(0) == 0, "NON", "OUI")
        df["OT_COR_EGAL"] = np.where((df["Total coûts budgétés"].fillna(0) - df["Total coûts réels"].fillna(0)) == 0, "OUI", "NON")
        res['dfp'] = df
        an = cpiv(df, df["Nº appel pl.entret."].fillna(0)==0, "Statut OT", posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c] = an.get(c, 0)
        an["Total"] = an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1); an["TAUX_REALISATION_CORRECTIF/PT"] = ckpi(an["TCLO"], an["Total"])
        pr = cpiv(df, df["Statut OT"]=="CRÉÉ", "ap", posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c] = pr.get(c, 0)
        pr["Total"] = pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"] = ckpi(pr["<1 mois"], pr["Total"]); pr["OT préparation >3 mois"] = ckpi(pr[">3 mois"], pr["Total"], 0); pr["OT préparation 1mois< <3mois"] = ckpi(pr["1 mois < <3 mois"], pr["Total"], 0)
        pl = cpiv(df, (df["Statut OT"]=="LANC") & (df["Contient SOPL"]==0), "alp", posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c] = pl.get(c, 0)
        pl["Total"] = pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"] = ckpi(pl["<1 mois"], pl["Total"]); pl["OT planification >3 mois"] = ckpi(pl[">3 mois"], pl["Total"], 0); pl["OT planification 1mois< <3mois"] = ckpi(pl["1 mois < <3 mois"], pl["Total"], 0)
        ex = cpiv(df, (df["Statut OT"]=="LANC") & (df["Contient SOPL"]==1), "aex", posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c] = ex.get(c, 0)
        ex["Total"] = ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"] = ckpi(ex["<1 mois"], ex["Total"]); ex["OT exécution >3 mois"] = ckpi(ex[">3 mois"], ex["Total"], 0); ex["OT exécution 1mois< <3mois"] = ckpi(ex["1 mois < <3 mois"], ex["Total"], 0)
        la = pd.pivot_table(df[df["Statut OT"]=="LANC"], index="Poste travail princ.", columns="OT LANC ESTIME", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["OUI","NON"]: la[c] = la.get(c, 0)
        la["Total"] = la["OUI"]+la["NON"]; la["OT LANC ESTIME"] = ckpi(la["OUI"], la["Total"])
        pc = pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"], index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c] = pc.get(c, 0)
        pc["Total"] = pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"] = ckpi(pc["CARACTERISE"], pc["Total"])
        plc = pd.pivot_table(df[df["Statut OT"]=="LANC"], index="Poste travail princ.", columns="Backlog planification", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c] = plc.get(c, 0)
        plc["Total"] = plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"] = ckpi(plc["CARACTERISE"], plc["Total"])
        for kn, cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv = pd.pivot_table(df, index="Poste travail princ.", columns=cn, values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["OUI","NON"]: pv[c] = pv.get(c, 0)
            pv["Total"] = pv["OUI"]+pv["NON"]; pv[cn] = ckpi(pv["OUI"], pv["Total"]); res[kn.lower().replace(" ","_")] = pv
        avf = av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip()=="")].copy(); res['avf'] = avf
        tca = pd.pivot_table(avf, index="Poste travail princ.", columns="Statut utilisateur", values="Avis", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c] = tca.get(c, 0)
        tca["Total"] = tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"] = ckpi(tca["APRV"], tca["Total"])
        res['ckdf'] = pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT": an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois": pr["OT préparation <1 mois"],"OT préparation >3 mois": pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois": pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois": pl["OT planification <1 mois"],"OT planification >3 mois": pl["OT planification >3 mois"],"OT planification 1mois< <3mois": pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois": ex["OT exécution <1 mois"],"OT exécution >3 mois": ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois": ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé": tca["appel avis approuvé"],"OT LANC ESTIME": la["OT LANC ESTIME"],
            "Backlog préparation caractérisé": pc["Backlog préparation caractérisé"],"Backlog planification caractérisé": plc["Backlog planification caractérisé"],
            "OT CONFIME": res['ot_confime']["OT CONFIME"],"OT_COR_EGAL": res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res

    def ks(v, c):
        try: val = float(v)
        except: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c == "TAUX_REALISATION_CORRECTIF/PT": return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c == "appel avis approuvé": return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val = float(str(v).replace(' %','').strip())
        except: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val = int(v)
        except: return ""
        if val == 0: return "color:#cbd5e0"
        if val <= 3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val <= 10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k, a, t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k == "TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k == "appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        return 0
    def is_lb(k): return k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois","OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

    def html_table(rows, cols, tc, sc_col=None):
        h = '<table class="tw %s"><thead><tr>' % tc + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
        for r in rows:
            rc = "cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h += '<tr class="%s">' % rc
            for c in cols:
                v = r.get(c, "")
                if r.get("_t")=="cible": h += '<td>%s</td>' % v
                else:
                    s = cs(v) if sc_col and c in sc_col else ks(v, c)
                    h += '<td style="%s">%s</td>' % (s or "", v)
            h += '</tr>'
        return h + '</tbody></table>'

    def html_ano(rows, cols):
        h = '<table class="tw at"><thead><tr>' + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
        for r in rows:
            h += '<tr class="%s">' % ("tr" if r.get("_t")=="total" else "")
            for c in cols: v = r.get(c,""); h += '<td style="%s">%s</td>' % (kas(v) or "", v)
            h += '</tr>'
        return h + '</tbody></table>'

    def html_variance_table(rows, cols):
        h = '<table class="tw vt"><thead><tr>' + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
        for r in rows:
            rc = "tr" if r.get("_t")=="total" else ""
            h += '<tr class="%s">' % rc
            for c in cols:
                v = r.get(c, "")
                s = ""
                if "Δ" in str(c) or "Var%" in str(c):
                    try:
                        fv = float(str(v).replace('%','').replace('+','').strip())
                        if abs(fv) < 0.01: s = "v-zero"
                        else:
                            kpi_name = r.get("Indicateur", "")
                            if is_lb(kpi_name):
                                s = "v-pos" if fv < 0 else "v-neg"
                            else:
                                s = "v-pos" if fv > 0 else "v-neg"
                    except: pass
                h += '<td class="%s">%s</td>' % (s, v)
            h += '</tr>'
        return h + '</tbody></table>'

    def html_synth(kpi_list, actuals, targets, act_map, accent):
        h = ''
        for k in kpi_list:
            av, tv = actuals.get(k,0), targets.get(k,100)
            met = av <= tv if is_lb(k) else av >= tv
            sbg, sclr = ("#c6efce","#006100") if met else ("#ffc7ce","#9c0006")
            scbg = accent if met else "#e53e3e"
            act = "Objectif atteint" if met else act_map.get(k,"")
            h += '<div class="sr"><div class="sn">%s</div><div class="sc" style="background:%s">%.1f%%</div><div class="stg">Cible: %s%%</div><div class="sb" style="color:%s;background:%s">%s</div><div class="sa">%s</div></div>' % (k, scbg, av, tv, sclr, sbg, "ATTEINT" if met else "NON ATTEINT", act)
        return h

    def html_classement(scores, accent):
        sp = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        met_p, not_p = [(p,s) for p,s in sp if s>=80], [(p,s) for p,s in sp if s<80]
        t5, b5 = met_p[:5], not_p[-5:] if len(not_p)>5 else not_p
        h = '<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h += '<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (accent,i+1,p,cs("%.2f"%s),s)
        else: h += '<div style="padding:4px;font-size:8px;color:#718096">Aucun poste</div>'
        h += '</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h += '<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (len(b5)-i,p,cs("%.2f"%s),s)
        else: h += '<div style="padding:4px;font-size:8px;color:#38a169">Tous atteints</div>'
        h += '</div></div>'; return h

    def html_variance_ranking(top5, bot5):
        h = '<div class="cg"><div><div class="ct" style="color:#276749">🏆 Top 5 — Plus Grande Amelioration</div>'
        if top5:
            for i, item in enumerate(top5):
                p, var_val = item["poste"], item["variance"]
                kpi = item.get("kpi","")
                vc = "v-pos" if var_val > 0 else ("v-neg" if var_val < 0 else "v-zero")
                h += '<div class="cgr"><span class="rk" style="color:#276749">%s</span><span class="pn">%s</span><span class="pv" style="font-size:9px;color:#718096">%s</span><span class="ps" class="%s">%s%.1f pts</span></div>' % (i+1, p, kpi, vc, "+" if var_val>0 else "", var_val)
        else: h += '<div style="padding:4px;font-size:8px;color:#718096">Aucune variance positive</div>'
        h += '</div><div><div class="ct" style="color:#c53030">⚠️ Bottom 5 — Plus Grande Regression</div>'
        if bot5:
            for i, item in enumerate(reversed(bot5)):
                p, var_val = item["poste"], item["variance"]
                kpi = item.get("kpi","")
                vc = "v-pos" if var_val > 0 else ("v-neg" if var_val < 0 else "v-zero")
                h += '<div class="cgr"><span class="rk" style="color:#c53030">%s</span><span class="pn">%s</span><span class="pv" style="font-size:9px;color:#718096">%s</span><span class="ps" class="%s">%s%.1f pts</span></div>' % (len(bot5)-i, p, kpi, vc, "+" if var_val>0 else "", var_val)
        else: h += '<div style="padding:4px;font-size:8px;color:#38a169">Aucune regression</div>'
        h += '</div></div>'; return h

    def html_bars(data, title, color):
        h = '<div class="ca"><div class="ct" style="color:%s">%s</div>' % (color, title)
        for label, val in sorted(data, key=lambda x: x[1], reverse=True):
            bw = min(max(val,0),100)
            h += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>' % (label, bw, color, val)
        h += '</div>'; return h

    def html_grouped_bars(posts, pscores, qscores, title, var_scores=None):
        h = '<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>' % title
        has_var = var_scores is not None
        if has_var:
            h += '<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span><span><i style="background:linear-gradient(90deg,#e65100,#ff8f00)"></i> Variance</span></div>'
        else:
            h += '<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        sp2 = sorted(posts, key=lambda x: (pscores.get(x,0)+qscores.get(x,0))/2, reverse=True)
        for p in sp2:
            pv, qv = pscores.get(p,0), qscores.get(p,0)
            pw, qw = min(max(pv,0),100), min(max(qv,0),100)
            h += '<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div>' % (p, pw, pv, qw, qv)
            if has_var:
                vv = var_scores.get(p, 0)
                vw = min(max(abs(vv)*2, 0), 100)
                vc = "#276749" if vv >= 0 else "#c53030"
                h += '<div class="gbr-w"><div class="gbr-f" style="width:%s%%;background:%s"></div></div><div class="gbr-v" style="color:%s">%s%.1f</div>' % (vw, vc, vc, "+" if vv>0 else "", vv)
            h += '</div></div>'
        h += '</div>'; return h

    def anl_pie_chart(data, names_col, values_col, title, colors=None, mg=None):
        if data.empty: return None
        fig = px.pie(data, names=names_col, values=values_col, title=title,
                     color_discrete_sequence=colors or px.colors.qualitative.Set2)
        fig.update_traces(textposition='inside', textinfo='percent+label+value', textfont_size=9)
        fig.update_layout(margin=mg, height=450, autosize=True, title_font_size=11,
                          legend=dict(font_size=8, orientation="h", yanchor="bottom", y=-0.15))
        return fig

    def anl_html_table(df_out, pct_col=None, pct_thresh=(80,60)):
        h = '<table class="anl-tbl"><thead><tr>'
        for c in df_out.columns: h += '<th>%s</th>' % c
        h += '</tr></thead><tbody>'
        for idx, row in df_out.iterrows():
            is_tot = str(idx) == "TOTAL" or row.iloc[0] == "TOTAL"
            rc = "tot" if is_tot else ""
            h += '<tr class="%s">' % rc
            for c in df_out.columns:
                v = row[c]
                s = ""
                if pct_col and c == pct_col and not is_tot:
                    try:
                        pv = float(str(v).replace('%','').strip())
                        s = "g-green" if pv >= pct_thresh[0] else ("g-yellow" if pv >= pct_thresh[1] else "g-red")
                    except: pass
                if isinstance(v, float): v = round(v, 1)
                h += '<td class="%s">%s</td>' % (s, v)
            h += '</tr>'
        return h + '</tbody></table>'

    def export_btn(df, filename):
        buf = io.BytesIO()
        df.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)
        st.download_button("📥 Exporter Excel", data=buf, file_name=filename,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:18px;margin-bottom:2px">⚙️</div><div style="font-size:12px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:8px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""", unsafe_allow_html=True)
        st.markdown("---")
        unf = st.toggle("📁 Charger nouveaux fichiers", value=False, key="tf")
        ot_f = av_f = None; apm = []
        if unf:
            ot_f = st.file_uploader("Fichier OT", type=["xlsx"], key="uot")
            av_f = st.file_uploader("Fichier AVIS", type=["xlsx"], key="uav")
        else:
            if os.path.exists("ot.xlsx"):
                try:
                    _t = excr(pd.read_excel("ot.xlsx"))
                    apm = sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
                except: pass
            st.markdown("""<div style="background:rgba(255,255,255,.1);padding:5px 8px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:7px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:10px;color:white;font-weight:600;margin-top:1px">📅 %s</div></div>""" % fichier_date, unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("**🎯 Postes**")
        sp = st.multiselect("Poste", ["All"]+apm, ["All"], key="sp")
        st.markdown("**🏭 Atelier**")
        sa = st.multiselect("Atelier", ["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"], ["All"], key="sa")
        st.markdown("**🏢 Division**")
        sd = st.multiselect("Division", ["All","SF1","SF2"], ["All"], key="sd")
        st.markdown("---")
        st.markdown("**📅 Periode**")
        dr = st.date_input("Date debut planifiee", value=(datetime(2025,1,1).date(), datetime.today().date()), format="DD/MM/YYYY", key="dr")

    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf:
                raw_ot = pd.read_excel(ot_f); raw_av = pd.read_excel(av_f)
            else:
                raw_ot = pd.read_excel("ot.xlsx"); raw_av = pd.read_excel("avis.xlsx")
            raw_ot = excr(raw_ot); raw_av = excr(raw_av)
            for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c] = pd.to_datetime(raw_ot[c], errors="coerce")
            for c in ["Créé le","Début souhaité","Date de la clôture"]:
                if c in raw_av.columns: raw_av[c] = pd.to_datetime(raw_av[c], errors="coerce")
            if not apm: apm = sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp = apm
            if "All" in sa or not sa: sa = ["All"]
            if "All" in sd or not sd: sd = ["All"]
            sdt = pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt = pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p = str(poste).upper()
                if "All" not in sa:
                    m = False
                    if "Sulfurique (PS)" in sa and "PS" in p: m = True
                    if "Phosphorique (PP)" in sa and "PP" in p: m = True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m = True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m = True
                    if not m: return False
                if "All" not in sd:
                    m = False
                    if "SF1" in sd and "SF1" in p: m = True
                    if "SF2" in sd and "SF2" in p: m = True
                    if not m: return False
                return True

            vp = [p for p in apm if mf(p) and p in sp]

            df = raw_ot[(raw_ot["Poste travail princ."].isin(vp)) & (raw_ot["Date de début planifiée"].between(sdt, edt))].copy()
            avdf = raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df = excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)].drop_duplicates())
            avdf = excr(avdf[(avdf["Ordre"].isna())|(avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns: df["Statut OT"] = df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            df_dash = raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash = excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns: df_dash["Statut OT"] = df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now = pd.Timestamp.now()
            res = calc_kpis(df, avdf, now, vp)
            ckdf = res['ckdf']; dfp = res['dfp']
            res_d = calc_kpis(df_dash, avdf, now, vp)
            ckdf_d = res_d['ckdf']

            all_kpis = list(ckdf.columns)
            var_df = pd.DataFrame(index=ckdf.index)
            for k in all_kpis:
                var_df[k + "_periode"] = ckdf[k].round(2)
                var_df[k + "_reference"] = ckdf_d[k].round(2)
                var_df[k + "_delta"] = (ckdf[k] - ckdf_d[k]).round(2)
                var_df[k + "_var%"] = np.where(ckdf_d[k].abs() < 0.01, 0, ((ckdf[k] - ckdf_d[k]) / ckdf_d[k].abs() * 100)).round(1)

            poste_variance = {}
            for p in vp:
                if p in var_df.index:
                    total_var = 0
                    for k in all_kpis:
                        delta = var_df.loc[p, k + "_delta"]
                        if is_lb(k):
                            total_var -= abs(delta)
                        else:
                            total_var += delta
                    poste_variance[p] = round(total_var / len(all_kpis), 2)
                else:
                    poste_variance[p] = 0.0

            sorted_var = sorted(poste_variance.items(), key=lambda x: x[1], reverse=True)

            def get_top_kpi_var(poste):
                if poste not in var_df.index: return ""
                best_k, best_v = "", 0
                for k in all_kpis:
                    d = var_df.loc[poste, k + "_delta"]
                    if is_lb(k):
                        eff = -abs(d)
                    else:
                        eff = d
                    if abs(eff) > abs(best_v):
                        best_v = eff
                        best_k = k
                return best_k

            top5_var = [{"poste": p, "variance": v, "kpi": get_top_kpi_var(p)} for p, v in sorted_var[:5]]
            bot5_var = [{"poste": p, "variance": v, "kpi": get_top_kpi_var(p)} for p, v in sorted_var[-5:]]

            def build_variance_table(kpi_list, var_dataframe, label_type):
                cols = ["Poste de travail", "Indicateur", "Valeur Periode", "Valeur Reference", "Δ Absolu", "Var%"]
                rows = []
                for p in var_dataframe.index:
                    for k in kpi_list:
                        vp_v = var_dataframe.loc[p, k + "_periode"]
                        vr_v = var_dataframe.loc[p, k + "_reference"]
                        delta_v = var_dataframe.loc[p, k + "_delta"]
                        varp_v = var_dataframe.loc[p, k + "_var%"]
                        rows.append({
                            "Poste de travail": p, "Indicateur": k,
                            "Valeur Periode": vp_v, "Valeur Reference": vr_v,
                            "Δ Absolu": delta_v, "Var%": varp_v
                        })
                tot_row = {"Poste de travail": "TOTAL", "Indicateur": "", "_t": "total"}
                avg_per = np.mean([r["Valeur Periode"] for r in rows]) if rows else 0
                avg_ref = np.mean([r["Valeur Reference"] for r in rows]) if rows else 0
                avg_del = np.mean([r["Δ Absolu"] for r in rows]) if rows else 0
                tot_row["Valeur Periode"] = round(avg_per, 2)
                tot_row["Valeur Reference"] = round(avg_ref, 2)
                tot_row["Δ Absolu"] = round(avg_del, 2)
                tot_row["Var%"] = ""
                rows.append(tot_row)
                return cols, rows

            var_p_cols, var_p_rows = build_variance_table(qk, var_df, "P")
            var_q_cols, var_q_rows = build_variance_table(pk, var_df, "Q")

            pscores = {}; qscores = {}
            for poste in ckdf.index:
                r = ckdf.loc[poste]
                pscores[poste] = (sum(gscore(k,r[k],cible[k]) for k in qk if k in r.index)/len(qk)*100) if qk else 0
                qscores[poste] = (sum(gscore(k,r[k],cible[k]) for k in pk if k in r.index)/len(pk)*100) if pk else 0
            pa = {k: round(ckdf[k].mean(),2) for k in qk}
            qa = {k: round(ckdf[k].mean(),2) for k in pk}

            pscores_d = {}; qscores_d = {}
            for poste in ckdf_d.index:
                r = ckdf_d.loc[poste]
                pscores_d[poste] = (sum(gscore(k,r[k],cible[k]) for k in qk if k in r.index)/len(qk)*100) if qk else 0
                qscores_d[poste] = (sum(gscore(k,r[k],cible[k]) for k in pk if k in r.index)/len(pk)*100) if pk else 0

            all_ano = []
            sub_p = {"TAUX_REALISATION_CORRECTIF/PT":lambda d:d[(d["Nº appel pl.entret."].fillna(0)==0)&(~d["Statut OT"].isin(["CLOT","TCLO"]))],"OT préparation <1 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]!="<1 mois")],"OT préparation >3 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]==">3 mois")],"OT planification <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]!="<1 mois")],"OT planification >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]==">3 mois")],"OT exécution <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]!="<1 mois")],"OT exécution >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]==">3 mois")]}
            sub_q = {"OT LANC ESTIME":lambda d:d[(d["Statut OT"]=="LANC")&(d["OT LANC ESTIME"]=="NON")],"Backlog préparation caractérisé":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["Backlog preparation"]=="NON CARACTERISE")],"Backlog planification caractérisé":lambda d:d[(d["Statut OT"]=="LANC")&(d["Backlog planification"]=="NON CARACTERISE")],"OT CONFIME":lambda d:d[d["OT CONFIME"]=="NON"],"OT_COR_EGAL":lambda d:d[d["OT_COR_EGAL"]=="NON"]}
            for poste in vp:
                if poste not in dfp["Poste travail princ."].values: continue
                dp = dfp[dfp["Poste travail princ."]==poste]
                for kn, sf in sub_p.items():
                    vk = ckdf.loc[poste, kn] if poste in ckdf.index else 100
                    if pd.notna(vk) and vk < cible[kn]:
                        cnt = len(sf(dp))
                        if cnt > 0: all_ano.append({"Poste":poste,"KPI":kn,"Nb":cnt,"Type":"P"})
                for kn, sf in sub_q.items():
                    vk = ckdf.loc[poste, kn] if poste in ckdf.index else 100
                    if pd.notna(vk) and vk < cible[kn]:
                        cnt = len(sf(dp))
                        if cnt > 0: all_ano.append({"Poste":poste,"KPI":kn,"Nb":cnt,"Type":"Q"})
                vk_av = ckdf.loc[poste, "appel avis approuvé"] if poste in ckdf.index else 100
                if pd.notna(vk_av) and vk_av < cible["appel avis approuvé"]:
                    cnt = len(res['avf'][res['avf']["Poste travail princ."]==poste])
                    if cnt > 0: all_ano.append({"Poste":poste,"KPI":"appel avis approuvé","Nb":cnt,"Type":"Q"})

            def build_ano(ano_list, kpi_list):
                if not ano_list: return [], []
                adf = pd.DataFrame(ano_list)
                pv = adf.pivot_table(index="Poste", columns="KPI", values="Nb", aggfunc="sum", fill_value=0).astype(int)
                pv["Total"] = pv.sum(axis=1); tot = pv.sum()
                cols = [c for c in kpi_list if c in pv.columns] + ["Total"]; rows = []
                for idx in pv.index:
                    r = {"_t":"n","Poste de travail":idx}
                    for c in cols: r[c] = pv.loc[idx, c]
                    rows.append(r)
                tr = {"_t":"total","Poste de travail":"Total general"}
                for c in cols: tr[c] = int(tot[c])
                rows.append(tr); return ["Poste de travail"]+cols, rows

            ano_p_c, ano_p_r = build_ano([a for a in all_ano if a["Type"]=="P"], qk)
            ano_q_c, ano_q_r = build_ano([a for a in all_ano if a["Type"]=="Q"], pk)

            def build_kpi(kpi_list, scores, sname):
                sp2 = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)
                cols = ["Poste de travail"]+kpi_list+[sname]; rows = []
                cr = {"_t":"cible","Poste de travail":"CIBLE"}
                for k in kpi_list: cr[k] = cible[k]
                cr[sname] = "100.00 %"; rows.append(cr)
                for p in sp2:
                    r = {"_t":"n","Poste de travail":p}
                    for k in kpi_list: r[k] = round(ckdf.loc[p, k], 2) if p in ckdf.index else ""
                    r[sname] = "%.2f %%" % scores[p]; rows.append(r)
                tr = {"_t":"total","Poste de travail":"Total general"}
                for k in kpi_list: tr[k] = round(ckdf[k].mean(), 2)
                tr[sname] = "%.2f %%" % (np.mean(list(scores.values())) if scores else 0)
                rows.append(tr); return cols, rows

            pcols, prows = build_kpi(qk, pscores, "Score Performance")
            qcols, qrows = build_kpi(pk, qscores, "Score Qualite")

            save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_r, ano_p_c, ano_q_r, ano_q_c, var_p_r, var_p_c, var_q_r, var_q_c, fichier_date)

            df_sc_d = pd.DataFrame([{"Poste":p,"Perf":pscores_d[p],"Qual":qscores_d[p],"Metier":get_metier(p),"Atelier":get_atelier(p),"Division":get_division(p)} for p in vp if p in pscores_d])
            by_at = df_sc_d.groupby("Atelier")[["Perf","Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()
            by_mt = df_sc_d.groupby("Metier")[["Perf","Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()
            by_div = df_sc_d.groupby("Division")[["Perf","Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()

            total_ot = len(df); avg_p = np.mean(list(pscores.values())) if pscores else 0
            avg_q = np.mean(list(qscores.values())) if qscores else 0; total_ano = sum(a["Nb"] for a in all_ano)
            avg_var = np.mean(list(poste_variance.values())) if poste_variance else 0
            pos_var = sum(1 for v in poste_variance.values() if v > 0)
            neg_var = sum(1 for v in poste_variance.values() if v < 0)

            desig_col = None
            for cn in ["Désignation du travail","Designation du travail","Désignation","Designation","Description"]:
                if cn in dfp.columns: desig_col = cn; break

            bl_plan_data = dfp[dfp["Statut OT"]=="LANC"].copy()
            bl_plan_car = bl_plan_data[bl_plan_data["Backlog planification"]=="CARACTERISE"]
            bl_plan_ncar = bl_plan_data[bl_plan_data["Backlog planification"]=="NON CARACTERISE"]
            bl_plan_by_poste = bl_plan_data.groupby("Poste travail princ.").size().to_dict()
            bl_plan_car_by_poste = bl_plan_car.groupby("Poste travail princ.").size().to_dict()

            # VARIABLES DE MARGE ISOLEES POUR EVITER LE BUG D'EDITEUR
            mg_std = {"t": 40, "b": 10, "l": 10, "r": 10}
            mg_heat = {"t": 40, "b": 80, "l": 120, "r": 10}

            st.markdown('<div class="mh"><h1>📊 KPI Dashboard MC & FEED</h1><div class="db">📅 %s</div></div>' % fichier_date, unsafe_allow_html=True)
            st.markdown("""<div class="cr" style="grid-template-columns:repeat(6,1fr)">
            <div class="cc c1"><div class="cv">%s</div><div class="cl">Total OT Analyses</div></div>
            <div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div>
            <div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div>
            <div class="cc c4"><div class="cv">%s</div><div class="cl">Total Anomalies</div></div>
            <div class="cc c5"><div class="cv">%s%.1f</div><div class="cl">Variance Moyenne</div></div>
            <div class="cc c6"><div class="cv">%s / %s</div><div class="cl">Postes Amel. / Regr.</div></div>
            </div>""" % (total_ot, avg_p, avg_q, total_ano, "+" if avg_var>=0 else "", avg_var, pos_var, neg_var), unsafe_allow_html=True)

            tab0, tab1, tab2, tab3, tab4 = st.tabs(["📊 TABLEAU DE BORD", "📈 INDICATEURS PERFORMANCE", "✅ INDICATEUR QUALITE", "🔬 ANALYSE", "📉 VARIANCES & CHANGEMENT"])

            with tab0:
                st.markdown('<div class="stl c">📋 Analyse du Backlog Planification</div>', unsafe_allow_html=True)
                bp_total = len(bl_plan_data)
                bp_car_pct = (len(bl_plan_car)/bp_total*100) if bp_total>0 else 0
                bp_ncar_pct = (len(bl_plan_ncar)/bp_total*100) if bp_total>0 else 0

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown('<div class="cc c2" style="margin-bottom:4px"><div class="cv">%s</div><div class="cl">Total OT Lances</div></div>' % bp_total, unsafe_allow_html=True)
                    st.markdown('<div class="cc c2" style="margin-bottom:4px"><div class="cv">%.1f%%</div><div class="cl">Backlog Caracterise</div></div>' % bp_car_pct, unsafe_allow_html=True)
                    st.markdown('<div class="cc c4" style="margin-bottom:4px"><div class="cv">%.1f%%</div><div class="cl">Backlog Non Caracterise</div></div>' % bp_ncar_pct, unsafe_allow_html=True)
                with c2:
                    if bp_total > 0:
                        bp_pie_data = pd.DataFrame({
                            "Statut": ["Caracterise","Non Caracterise"],
                            "Nombre": [len(bl_plan_car), len(bl_plan_ncar)]
                        })
                        fig_bp = anl_pie_chart(bp_pie_data, "Statut", "Nombre", "Repartition Backlog Planification",
                                              colors=["#276749","#e53e3e"], mg=mg_std)
                        if fig_bp: st.plotly_chart(fig_bp, use_container_width=True)
                with c3:
                    if bl_plan_by_poste:
                        bp_bar_data = pd.DataFrame([
                            {"Poste": p, "Total": bl_plan_by_poste.get(p,0), "Caracterise": bl_plan_car_by_poste.get(p,0)}
                            for p in vp if p in bl_plan_by_poste
                        ]).sort_values("Total", ascending=True).tail(10)
                        fig_bpb = px.bar(bp_bar_data, x="Total", y="Poste", orientation="h",
                                         title="Top 10 Postes - Backlog Planification",
                                         color="Caracterise", color_discrete_sequence=["#276749","#e53e3e"])
                        fig_bpb.update_layout(
                            height=450,
                            autosize=True,
                            margin=mg_std,
                            title_font_size=11
                        )
                        st.plotly_chart(fig_bpb, use_container_width=True)

                st.markdown('<div class="stl p">📊 Vue d\'ensemble par poste (avec variance)</div>', unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp, pscores_d, qscores_d, "Performance & Qualite par Poste de Travail", poste_variance), unsafe_allow_html=True)

                st.markdown('<div class="stl v">📉 Classement des Variances par Poste</div>', unsafe_allow_html=True)
                st.markdown(html_variance_ranking(top5_var, bot5_var), unsafe_allow_html=True)

                st.markdown('<div class="stl p">🏭 Par Atelier</div>', unsafe_allow_html=True)
                if not by_at.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx,row["Perf"]) for idx,row in by_at.iterrows()], "Performance par Atelier", "#2b6cb0"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx,row["Qual"]) for idx,row in by_at.iterrows()], "Qualite par Atelier", "#276749"), unsafe_allow_html=True)

                st.markdown('<div class="stl p">🔧 Par Metier</div>', unsafe_allow_html=True)
                if not by_mt.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx,row["Perf"]) for idx,row in by_mt.iterrows()], "Performance par Metier", "#2b6cb0"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx,row["Qual"]) for idx,row in by_mt.iterrows()], "Qualite par Metier", "#27649"), unsafe_allow_html=True)

                st.markdown('<div class="stl p">🏢 Par Division</div>', unsafe_allow_html=True)
                if not by_div.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx,row["Perf"]) for idx,row in by_div.iterrows()], "Performance par Division", "#2b6cb0"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx,row["Qual"]) for idx,row in by_div.iterrows()], "Qualite par Division", "#276749"), unsafe_allow_html=True)

            with tab1:
                st.markdown('<div class="stl p">📈 Synthese Performance</div>', unsafe_allow_html=True)
                st.markdown(html_synth(qk, pa, cible, act_map, "#38a169"), unsafe_allow_html=True)
                st.markdown('<div class="stl p">📊 Indicateurs de Performance par Poste</div>', unsafe_allow_html=True)
                st.markdown(html_table(prows, pcols, "pt", ["Score Performance"]), unsafe_allow_html=True)
                st.markdown('<div class="stl p">🏆 Classement Performance</div>', unsafe_allow_html=True)
                st.markdown(html_classement(pscores, "#38a169"), unsafe_allow_html=True)
                if ano_p_r:
                    st.markdown('<div class="stl a">⚠️ Anomalies Performance</div>', unsafe_allow_html=True)
                    st.markdown(html_ano(ano_p_r, ano_p_c), unsafe_allow_html=True)

            with tab2:
                st.markdown('<div class="stl q">✅ Synthese Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_synth(pk, qa, cible, act_map, "#3182ce"), unsafe_allow_html=True)
                st.markdown('<div class="stl q">📊 Indicateurs de Qualite par Poste</div>', unsafe_allow_html=True)
                st.markdown(html_table(qrows, qcols, "qt", ["Score Qualite"]), unsafe_allow_html=True)
                st.markdown('<div class="stl q">🏆 Classement Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_classement(qscores, "#3182ce"), unsafe_allow_html=True)
                if ano_q_r:
                    st.markdown('<div class="stl a">⚠️ Anomalies Qualite</div>', unsafe_allow_html=True)
                    st.markdown(html_ano(ano_q_r, ano_q_c), unsafe_allow_html=True)

            with tab3:
                st.markdown('<div class="stl c">🔬 Analyse Detaillee</div>', unsafe_allow_html=True)

                st.markdown('<div class="stl p" style="margin-top:8px">📦 Analyse Backlog Preparation</div>', unsafe_allow_html=True)
                bl_prep_data = dfp[dfp["Statut OT"]=="CRÉÉ"].copy()
                bl_prep_car = bl_prep_data[bl_prep_data["Backlog preparation"]=="CARACTERISE"]
                bl_prep_ncar = bl_prep_data[bl_prep_data["Backlog preparation"]=="NON CARACTERISE"]
                bp_prep_total = len(bl_prep_data)
                if bp_prep_total > 0:
                    c1, c2 = st.columns(2)
                    with c1:
                        bp_prep_pie = pd.DataFrame({
                            "Statut": ["Caracterise (%d)" % len(bl_prep_car), "Non Caracterise (%d)" % len(bl_prep_ncar)],
                            "Nombre": [len(bl_prep_car), len(bl_prep_ncar)]
                        })
                        fig_pp = anl_pie_chart(bp_prep_pie, "Statut", "Nombre", "Backlog Preparation", ["#38a169","#e53e3e"], mg=mg_std)
                        if fig_pp: st.plotly_chart(fig_pp, use_container_width=True)
                    with c2:
                        prep_by_poste = bl_prep_data.groupby("Poste travail princ.").size().sort_values(ascending=False).head(10)
                        prep_car_by_poste = bl_prep_car.groupby("Poste travail princ.").size()
                        prep_df = pd.DataFrame({
                            "Poste": prep_by_poste.index,
                            "Total": prep_by_poste.values,
                            "Caracterise": [prep_car_by_poste.get(p,0) for p in prep_by_poste.index]
                        })
                        fig_pb = px.bar(prep_df, x="Total", y="Poste", orientation="h",
                                        title="Top 10 Postes - Backlog Preparation",
                                        color="Caracterise", color_discrete_sequence=["#38a169","#e53e3e"])
                        fig_pb.update_layout(
                            height=450,
                            autosize=True,
                            margin=mg_std,
                            title_font_size=11
                        )
                        st.plotly_chart(fig_pb, use_container_width=True)
                else:
                    st.markdown('<div class="es">Aucun OT en statut CRÉÉ pour cette periode</div>', unsafe_allow_html=True)

                st.markdown('<div class="stl p" style="margin-top:8px">⏳ Analyse Age des OT par Statut</div>', unsafe_allow_html=True)
                for statut_label, statut_filter, age_col, kpi_list_age in [
                    ("Preparation (CRÉÉ)", dfp["Statut OT"]=="CRÉÉ", "ap",
                     ["OT préparation <1 mois","OT préparation 1mois< <3mois","OT préparation >3 mois"]),
                    ("Planification (LANC sans SOPL)", (dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==0), "alp",
                     ["OT planification <1 mois","OT planification 1mois< <3mois","OT planification >3 mois"]),
                    ("Execution (LANC avec SOPL)", (dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==1), "aex",
                     ["OT exécution <1 mois","OT exécution 1mois< <3mois","OT exécution >3 mois"])
                ]:
                    sub = dfp[statut_filter]
                    if len(sub) > 0:
                        age_dist = sub[age_col].value_counts()
                        age_data = pd.DataFrame({"Categorie": age_dist.index, "Nombre": age_dist.values})
                        fig_age = px.pie(age_data, names="Categorie", values="Nombre",
                                         title=f"Repartition Age - {statut_label}",
                                         color_discrete_sequence=["#38a169","#ecc94b","#e53e3e"])
                        fig_age.update_traces(textposition='inside', textinfo='percent+label+value', textfont_size=9)
                        fig_age.update_layout(
                            height=450,
                            autosize=True,
                            margin=mg_std,
                            title_font_size=11,
                            legend=dict(font_size=8, orientation="h", yanchor="bottom", y=-0.15)
                        )
                        st.plotly_chart(fig_age, use_container_width=True)

                if all_ano:
                    st.markdown('<div class="stl a" style="margin-top:8px">🔍 Detail des Anomalies</div>', unsafe_allow_html=True)
                    ano_df = pd.DataFrame(all_ano)
                    if desig_col and desig_col in dfp.columns:
                        detail_rows = []
                        for _, ar in ano_df.iterrows():
                            poste = ar["Poste"]; kpi = ar["KPI"]
                            dp = dfp[dfp["Poste travail princ."]==poste]
                            sub_fn = sub_p.get(kpi) or sub_q.get(kpi)
                            if sub_fn:
                                anomalies = sub_fn(dp)
                                for _, arow in anomalies.head(5).iterrows():
                                    detail_rows.append({
                                        "Poste": poste, "KPI": kpi,
                                        "OT": arow.get("Ordre",""),
                                        "Designation": str(arow.get(desig_col,""))[:80],
                                        "Statut": arow.get("Statut OT","")
                                    })
                        if detail_rows:
                            ddf = pd.DataFrame(detail_rows)
                            st.markdown(anl_html_table(ddf), unsafe_allow_html=True)

            with tab4:
                st.markdown('<div class="stl v">📉 Analyse des Variances - Changement des Donnees OT & AVIS</div>', unsafe_allow_html=True)
                st.markdown("""<div style="background:#fff3e0;border:1px solid #ff8f00;border-radius:8px;padding:8px 12px;margin-bottom:8px;font-size:12px;color:#e65100">
                <strong>📋 Methode :</strong> Comparaison entre les KPIs de la <strong>periode selectionnee</strong> et les KPIs de <strong>reference (toutes les donnees)</strong>.
                La variance mesure l'impact du changement de filtrage des donnees OT et AVIS sur chaque indicateur.
                <strong>Vert</strong> = amelioration, <strong>Rouge</strong> = regression (inverse pour les indicateurs "plus bas c'est mieux").
                </div>""", unsafe_allow_html=True)

                st.markdown('<div class="stl v" style="margin-top:4px">📊 Resume Global des Variances par KPI</div>', unsafe_allow_html=True)
                kpi_var_rows = []
                for k in qk + pk:
                    ref_v = ckdf_d[k].mean()
                    per_v = ckdf[k].mean()
                    delta = per_v - ref_v
                    varp = ((per_v - ref_v) / ref_v.abs() * 100) if ref_v.abs() > 0.01 else 0
                    kpi_var_rows.append({
                        "Indicateur": k, "Valeur Periode": round(per_v,2),
                        "Valeur Reference": round(ref_v,2), "Δ Absolu": round(delta,2), "Var%": round(varp,1)
                    })
                tot_kpi = {"Indicateur": "TOTAL MOYEN", "Valeur Periode": round(np.mean([r["Valeur Periode"] for r in kpi_var_rows]),2),
                           "Valeur Reference": round(np.mean([r["Valeur Reference"] for r in kpi_var_rows]),2),
                           "Δ Absolu": round(np.mean([r["Δ Absolu"] for r in kpi_var_rows]),2), "Var%": "", "_t": "total"}
                kpi_var_rows.append(tot_kpi)
                st.markdown(html_variance_table(kpi_var_rows, ["Indicateur","Valeur Periode","Valeur Reference","Δ Absolu","Var%"]), unsafe_allow_html=True)

                st.markdown('<div class="stl v" style="margin-top:8px">🏆 Classement des Postes par Variance</div>', unsafe_allow_html=True)
                st.markdown(html_variance_ranking(top5_var, bot5_var), unsafe_allow_html=True)

                var_chart_data = pd.DataFrame([
                    {"Poste": p, "Variance": poste_variance.get(p,0), "Type": "Amelioration" if poste_variance.get(p,0)>=0 else "Regression"}
                    for p in vp
                ]).sort_values("Variance", ascending=True)
                fig_var = px.bar(var_chart_data, x="Variance", y="Poste", orientation="h",
                                 title="Variance Globale par Poste (Periode vs Reference)",
                                 color="Type", color_discrete_map={"Amelioration":"#276749","Regression":"#e53e3e"})
                fig_var.update_layout(
                    height=max(450, len(vp)*18),
                    autosize=True,
                    margin=mg_std,
                    title_font_size=11
                )
                fig_var.update_xaxes(zeroline=True, zerolinewidth=2, zerolinecolor="#718096")
                st.plotly_chart(fig_var, use_container_width=True)

                st.markdown('<div class="stl v" style="margin-top:8px">📈 Detail Variance - Indicateurs de Performance</div>', unsafe_allow_html=True)
                st.markdown(html_variance_table(var_p_rows, var_p_cols), unsafe_allow_html=True)

                st.markdown('<div class="stl v" style="margin-top:8px">✅ Detail Variance - Indicateurs de Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_variance_table(var_q_rows, var_q_cols), unsafe_allow_html=True)

                st.markdown('<div class="stl v" style="margin-top:8px">🗺️ Carte de Chaleur des Variances</div>', unsafe_allow_html=True)
                heat_data = []
                for p in vp:
                    row = {"Poste": p}
                    for k in qk + pk:
                        dk = k + "_delta"
                        row[k] = var_df.loc[p, dk] if p in var_df.index and dk in var_df.columns else 0
                    heat_data.append(row)
                if heat_data:
                    heat_df = pd.DataFrame(heat_data).set_index("Poste")
                    fig_heat = px.imshow(heat_df.values, labels=dict(x="Indicateur", y="Poste", color="Variance"),
                                         x=list(heat_df.columns), y=list(heat_df.index),
                                         title="Heatmap Variance (Vert=Amelioration, Rouge=Regression)",
                                         color_continuous_scale=["#c53030","#f7fafc","#276749"],
                                         aspect="auto", zmin=-50, zmax=50)
                    fig_heat.update_layout(
                        height=450,
                        autosize=True,
                        margin=mg_heat,
                        title_font_size=11
                    )
                    fig_heat.update_xaxes(tickangle=45, tickfont_size=8)
                    fig_heat.update_yaxes(tickfont_size=8)
                    st.plotly_chart(fig_heat, use_container_width=True)

                st.markdown('<div class="stl v" style="margin-top:8px">📥 Export</div>', unsafe_allow_html=True)
                var_export = pd.DataFrame(heat_data)
                if not var_export.empty:
                    export_btn(var_export, "variance_indicateurs.xlsx")

        except Exception as e:
            st.error(f"Erreur de chargement: {str(e)}")
            import traceback
            st.code(traceback.format_exc())

if __name__ == "__main__":
    main()
