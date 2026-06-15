# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
ALL_KPI = QK + PK
CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100}
ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois."}
LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]
MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

# ============================================================
def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste travail princ.") if "Poste travail princ." in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste travail princ.") if "Poste travail princ." in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste travail princ.") if "Poste travail princ." in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste travail princ.") if "Poste travail princ." in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    j["Sens"]=j.apply(lambda r:"Amelioration" if ((r["Tendance"]=="hausse" and r["KPI"] not in LOWER_BETTER) or (r["Tendance"]=="baisse" and r["KPI"] in LOWER_BETTER)) else "Degradation",axis=1)
    return j.sort_values(["Date actuelle","Sens","Ecart %"],ascending=[True,False,False])

def calculate_rankings_by_type(var_df, type_name):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    sub=var_df[var_df["Type"]==type_name].copy()
    if sub.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in sub["Poste"].unique():
        pv=sub[sub["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def get_caract_type(statut_user,keywords):
    s=str(statut_user).upper(); matched=[kw for kw in keywords if kw in s]
    return max(matched,key=len) if matched else "AUTRE"

# ============================================================
def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:6px 10px;background:#fff;border-radius:5px;margin-bottom:2px;border:1px solid var(--b);font-size:13px}
    .sr .sn{font-weight:700;color:var(--p);min-width:220px;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:3px 9px;border-radius:12px;font-weight:800;font-size:14px;min-width:50px;text-align:center;margin:0 8px;color:#fff}
    .sr .sa{color:#718096;font-size:12px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:4px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:24px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}
    .gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .trend-up{color:#276749;font-weight:800;font-size:16px}
    .trend-down{color:#c53030;font-weight:800;font-size:16px}
    .trend-stable{color:#718096;font-weight:800;font-size:16px}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:120px}.gbr-l{width:100px}}
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def contient_mot(t,lm):
        t=str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if a<=1: return "<1 mois"
        elif a>=3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
        return df

    def calc_kpis(df_i,av_i,now,posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
        df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
        for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc]=pd.to_datetime(df[dc],errors='coerce')
                df[am]=((now.year-df[dc].dt.year)*12+(now.month-df[dc].dt.month)).round(2)
                df[ac]=df[am].apply(cat_age)
            else: df[am]=np.nan; df[ac]="Inconnu"
        df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
        df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
        df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
        df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
        res['dfp']=df
        an=cpiv(df,df["Nº appel pl.entret."].fillna(0)==0,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["Total"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1); an["TAUX_REALISATION_CORRECTIF/PT"]=ckpi(an["TCLO"],an["Total"])
        pr=cpiv(df,df["Statut OT"]=="CRÉÉ","ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
        avf=av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip()=="")].copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res
    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "color:#cbd5e0"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        return 0
    def is_lb(k): return k in LOWER_BETTER

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                else: s=cs(v) if sc_col and c in sc_col else ks(v,c); h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
    def html_ano(rows,cols):
        h='<table class="tw at"><thead><tr>'+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            h+='<tr class="%s">'%("tr" if r.get("_t")=="total" else "")
            for c in cols: v=r.get(c,""); h+='<td style="%s">%s</td>'%(kas(v) or "",v)
            h+='</tr>'
        return h+'</tbody></table>'

    # ===== MODIFICATION : colonne "Indicateurs" au lieu de "KPI" =====
        # ===== Table Anomalies transposée (Indicateurs en lignes, Postes en colonnes) =====
    def html_ano_transpose(ano_rows, postes_list, type_label, color_header):
        if not ano_rows: return ""
        # Pivot : Indicateurs en lignes, Postes en colonnes
        indicators = []
        seen = set()
        for r in ano_rows:
            if r.get("_t")!="total" and r["Indicateurs"] not in seen:
                indicators.append(r["Indicateurs"])
                seen.add(r["Indicateurs"])
        
        cols = ["Indicateurs"] + postes_list + ["Total"]
        # Construire les données pivotées
        pivot_data = {}
        for ind in indicators:
            pivot_data[ind] = {"Indicateurs": ind}
            for p in postes_list:
                pivot_data[ind][p] = 0
            pivot_data[ind]["Total"] = 0
        
        for r in ano_rows:
            if r.get("_t")=="total": continue
            ind = r["Indicateurs"]
            pst = r["Poste travail princ."]
            cnt = r["Nb anomalies"]
            if ind in pivot_data:
                pivot_data[ind][pst] = cnt
                pivot_data[ind]["Total"] += cnt
        
        # Tri par Total décroissant
        indicators_sorted = sorted(indicators, key=lambda x: pivot_data[x]["Total"], reverse=True)
        
        h = '<table class="tw at"><thead><tr>'
        for c in cols:
            h += '<th>%s</th>' % c
        h += '</tr></thead><tbody>'
        
        grand_total = 0
        for ind in indicators_sorted:
            d = pivot_data[ind]
            grand_total += d["Total"]
            h += '<tr>'
            h += '<td style="font-weight:700;white-space:normal;max-width:280px">%s</td>' % ind
            for p in postes_list:
                v = d[p]
                h += '<td style="%s">%s</td>' % (kas(v) or "", v if v > 0 else "-")
            h += '<td style="font-weight:800;background:#2b6cb0;color:#fff">%s</td>' % d["Total"]
            h += '</tr>'
        
        # Ligne Total
        h += '<tr class="tr"><td style="font-weight:800">TOTAL</td>'
        for p in postes_list:
            s = sum(pivot_data[ind].get(p,0) for ind in indicators)
            h += '<td style="font-weight:800">%s</td>' % s
        h += '<td style="font-weight:900;font-size:13px">%s</td>' % grand_total
        h += '</tr></tbody></table>'
        return h

     # ===== Charts Anomalies =====
    def ano_charts(ano_rows, postes_list, type_label):
        if not ano_rows: return None, None
        indicators = []
        seen = set()
        for r in ano_rows:
            if r.get("_t")!="total" and r["Indicateurs"] not in seen:
                indicators.append(r["Indicateurs"])
                seen.add(r["Indicateurs"])
        pivot_data = {}
        for ind in indicators:
            pivot_data[ind] = {p: 0 for p in postes_list}
            pivot_data[ind]["Total"] = 0
        for r in ano_rows:
            if r.get("_t")=="total": continue
            ind = r["Indicateurs"]
            pst = r["Poste travail princ."]
            if ind in pivot_data and pst in pivot_data[ind]:
                pivot_data[ind][pst] = r["Nb anomalies"]
                pivot_data[ind]["Total"] += r["Nb anomalies"]
        indicators_sorted = sorted(indicators, key=lambda x: pivot_data[x]["Total"], reverse=True)

        # Chart 1 : Par Indicateur (barres horizontales)
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(
            y=indicators_sorted,
            x=[pivot_data[ind]["Total"] for ind in indicators_sorted],
            orientation='h',
            marker_color='#e53e3e',
            text=[pivot_data[ind]["Total"] for ind in indicators_sorted],
            textposition='outside',
            textfont_size=12,
            hovertemplate='%{y}: %{x} anomalies<extra></extra>'
        ))
        fig1.update_layout(
            title=dict(text='<b>Anomalies %s — Par Indicateur</b>'%type_label, font_size=14),
            height=max(300, len(indicators_sorted)*45 + 80),
            margin=dict(l=300, r=60, t=50, b=30),
            xaxis_title="Nombre d'anomalies",
            yaxis=dict(tickfont_size=11),
            template='plotly_white',
            showlegend=False
        )

        # Chart 2 : Total par Poste travail princ. (barres simples)
        totals_par_poste = []
        for p in postes_list:
            total_p = sum(pivot_data[ind].get(p, 0) for ind in indicators)
            totals_par_poste.append(total_p)
        # Tri par total décroissant
        poste_total = list(zip(postes_list, totals_par_poste))
        poste_total_sorted = sorted(poste_total, key=lambda x: x[1], reverse=True)
        postes_tri = [x[0] for x in poste_total_sorted]
        totaux_tri = [x[1] for x in poste_total_sorted]

        fig2 = go.Figure()
        fig2.add_trace(go.Bar(
            x=postes_tri,
            y=totaux_tri,
            marker_color='#e53e3e',
            text=totaux_tri,
            textposition='outside',
            textfont_size=12,
            hovertemplate='%{x}: %{y} anomalies<extra></extra>'
        ))
        fig2.update_layout(
            title=dict(text='<b>Anomalies %s — Total par Poste travail princ.</b>'%type_label, font_size=14),
            height=400,
            margin=dict(l=40, r=40, t=50, b=120),
            xaxis_title="Poste travail princ.",
            yaxis_title="Nombre total d'anomalies",
            xaxis_tickangle=-45,
            template='plotly_white',
            showlegend=False
        )

        return fig1, fig2
              # ===== Table Anomalies transposée (Indicateurs en lignes, Postes en colonnes) =====
    def html_ano_transpose(ano_rows, postes_list, type_label, color_header):
        if not ano_rows: return ""
        indicators = []
        seen = set()
        for r in ano_rows:
            if r.get("_t")!="total" and r["Indicateurs"] not in seen:
                indicators.append(r["Indicateurs"])
                seen.add(r["Indicateurs"])
        cols = ["Indicateurs"] + postes_list + ["Total"]
        pivot_data = {}
        for ind in indicators:
            pivot_data[ind] = {"Indicateurs": ind}
            for p in postes_list:
                pivot_data[ind][p] = 0
            pivot_data[ind]["Total"] = 0
        for r in ano_rows:
            if r.get("_t")=="total": continue
            ind = r["Indicateurs"]
            pst = r["Poste travail princ."]
            cnt = r["Nb anomalies"]
            if ind in pivot_data:
                pivot_data[ind][pst] = cnt
                pivot_data[ind]["Total"] += cnt
        indicators_sorted = sorted(indicators, key=lambda x: pivot_data[x]["Total"], reverse=True)
        h = '<table class="tw at"><thead><tr>'
        for c in cols:
            h += '<th>%s</th>' % c
        h += '</tr></thead><tbody>'
        grand_total = 0
        for ind in indicators_sorted:
            d = pivot_data[ind]
            grand_total += d["Total"]
            h += '<tr>'
            h += '<td style="font-weight:700;white-space:normal;max-width:280px">%s</td>' % ind
            for p in postes_list:
                v = d[p]
                h += '<td style="%s">%s</td>' % (kas(v) or "", v if v > 0 else "-")
            h += '<td style="font-weight:800;background:#2b6cb0;color:#fff">%s</td>' % d["Total"]
            h += '</tr>'
        h += '<tr class="tr"><td style="font-weight:800">TOTAL</td>'
        for p in postes_list:
            s = sum(pivot_data[ind].get(p,0) for ind in indicators)
            h += '<td style="font-weight:800">%s</td>' % s
        h += '<td style="font-weight:900;font-size:13px">%s</td>' % grand_total
        h += '</tr></tbody></table>'
        return h

    # ===== Charts Anomalies =====
    def ano_charts(ano_rows, postes_list, type_label):
        if not ano_rows: return None, None
        indicators = []
        seen = set()
        for r in ano_rows:
            if r.get("_t")!="total" and r["Indicateurs"] not in seen:
                indicators.append(r["Indicateurs"])
                seen.add(r["Indicateurs"])
        pivot_data = {}
        for ind in indicators:
            pivot_data[ind] = {p: 0 for p in postes_list}
            pivot_data[ind]["Total"] = 0
        for r in ano_rows:
            if r.get("_t")=="total": continue
            ind = r["Indicateurs"]
            pst = r["Poste travail princ."]
            if ind in pivot_data and pst in pivot_data[ind]:
                pivot_data[ind][pst] = r["Nb anomalies"]
                pivot_data[ind]["Total"] += r["Nb anomalies"]
        indicators_sorted = sorted(indicators, key=lambda x: pivot_data[x]["Total"], reverse=True)
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(
            y=indicators_sorted,
            x=[pivot_data[ind]["Total"] for ind in indicators_sorted],
            orientation='h',
            marker_color='#e53e3e',
            text=[pivot_data[ind]["Total"] for ind in indicators_sorted],
            textposition='outside',
            textfont_size=12,
            hovertemplate='%{y}: %{x} anomalies<extra></extra>'
        ))
        fig1.update_layout(
            title=dict(text='<b>Anomalies %s — Par Indicateur</b>'%type_label, font_size=14),
            height=max(300, len(indicators_sorted)*45 + 80),
            margin=dict(l=300, r=60, t=50, b=30),
            xaxis_title="Nombre d'anomalies",
            yaxis=dict(tickfont_size=11),
            template='plotly_white',
            showlegend=False
        )
        colors_seq = px.colors.qualitative.Set2
        fig2 = go.Figure()
        for i, ind in enumerate(indicators_sorted):
            vals = [pivot_data[ind].get(p, 0) for p in postes_list]
            if sum(vals) > 0:
                fig2.add_trace(go.Bar(
                    name=ind[:30],
                    x=postes_list,
                    y=vals,
                    marker_color=colors_seq[i % len(colors_seq)],
                    text=vals,
                    textposition='inside',
                    textfont_size=10
                ))
        fig2.update_layout(
            title=dict(text='<b>Anomalies %s — Par Poste travail princ.</b>'%type_label, font_size=14),
            barmode='stack',
            height=400,
            margin=dict(l=40, r=200, t=50, b=100),
            xaxis_title="Poste travail princ.",
            yaxis_title="Nombre d'anomalies",
            xaxis_tickangle=-45,
            legend=dict(font_size=9, orientation="v", yanchor="top", y=0.99, xanchor="left", x=1.01),
            template='plotly_white'
        )
        return fig1, fig2

    def html_actions_table(kpi_list,actuals,targets,act_map):
      def html_actions_table(kpi_list,actuals,targets,act_map):
        h='<table class="tw at"><thead><tr><th>Indicateurs</th><th>Valeur Actuelle</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action Recommandée</th></tr></thead><tbody>'
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            status="ATTEINT" if met else "NON ATTEINT"
            st_s="background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
            ec_clr="#276749" if met else "#c53030"
            action="Objectif atteint" if met else act_map.get(k,"")
            h+='<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td></tr>'%(k,av,tv,ec_clr,diff,st_s,status,action)
        return h+'</tbody></table>'

    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'; return h
    def html_kpi_bars(kpi_list,actuals,targets,title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,title)
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); met=av<=tv if is_lb(k) else av>=tv
            bw=min(max(av,0),100); bg=color_ok if met else color_fail
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(k,bw,bg,av)
        return h+'</div>'
    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        for p in sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True):
            pv,qv=pscores.get(p,0),qscores.get(p,0)
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>'%(p,min(max(pv,0),100),pv,min(max(qv,0),100),qv)
        return h+'</div>'

    # ===== Pie of Pie intelligent =====
    def anl_pie_chart(data, names_col, values_col, title, colors=None, min_pct=3.0):
        if data.empty: return None
        df = data[[names_col, values_col]].dropna().copy()
        df[values_col] = pd.to_numeric(df[values_col], errors='coerce').fillna(0)
        total = df[values_col].sum()
        if total == 0: return None
        df["_pct"] = df[values_col] / total * 100
        big = df[df["_pct"] >= min_pct].copy()
        small = df[df["_pct"] < min_pct].copy()
        has_small = len(small) >= 1 and small[values_col].sum() > 0
        if not has_small:
            fig = px.pie(df, names=names_col, values=values_col, title="<b>%s</b>"%title,
                         color_discrete_sequence=colors or px.colors.qualitative.Set2)
            fig.update_traces(textposition='inside', textinfo='percent+label+value', textfont_size=12, pull=[0.02]*len(df))
            fig.update_layout(margin=dict(t=60,b=50,l=20,r=20),height=480,autosize=True,title_font_size=15,
                legend=dict(font_size=11,orientation="h",yanchor="bottom",y=-0.12,title_text="Légende détaillée",title_font_size=12))
            return fig
        else:
            others_label = "Autres (%d secteurs)"%len(small)
            others_row = pd.DataFrame([{names_col: others_label, values_col: small[values_col].sum(), "_pct": small["_pct"].sum()}])
            main_df = pd.concat([big, others_row], ignore_index=True)
            sub_df = small.sort_values(values_col, ascending=False).copy()
            base_colors = colors or px.colors.qualitative.Set2
            main_colors = [base_colors[i % len(base_colors)] for i in range(len(main_df)-1)] + ["#CBD5E0"]
            sub_colors = [base_colors[(len(big)+i) % len(base_colors)] for i in range(len(sub_df))]
            fig = make_subplots(rows=1, cols=2, specs=[[{"type":"pie"},{"type":"pie"}]],
                subplot_titles=["<b>%s</b>"%title, "<b>Détail 'Autres' (%d secteurs)</b>"%len(small)], horizontal_spacing=0.08)
            fig.add_trace(go.Pie(labels=main_df[names_col].tolist(), values=main_df[values_col].tolist(),
                textinfo='percent+label+value', textposition='inside', textfont_size=12, marker_colors=main_colors,
                pull=[0.03 if i==len(main_df)-1 else 0.01 for i in range(len(main_df))]), row=1, col=1)
            fig.add_trace(go.Pie(labels=sub_df[names_col].tolist(), values=sub_df[values_col].tolist(),
                textinfo='percent+label+value', textposition='inside', textfont_size=11, marker_colors=sub_colors, hole=0.3), row=1, col=2)
            fig.update_layout(margin=dict(t=60,b=50,l=10,r=10),height=480,autosize=True,title_font_size=15,
                legend=dict(font_size=10,orientation="h",yanchor="bottom",y=-0.08,title_text="Légende détaillée",title_font_size=11),showlegend=True)
            return fig

    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",unsafe_allow_html=True)
        st.markdown("---")
        show_filters=st.checkbox("Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf")
            ot_f=av_f=None; apm=[]
            if unf:
                ot_f=st.file_uploader("Fichier OT",type=["xlsx"],key="uot")
                av_f=st.file_uploader("Fichier AVIS",type=["xlsx"],key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t=excr(pd.read_excel("ot.xlsx"))
                        apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            st.markdown("**🏭 Atelier**")
            sa=st.multiselect("Atelier",["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"],["All"],key="sa")
            st.markdown("**🏢 Division**")
            sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=av_f=None; apm=[]; sp=["All"]; sa=["All"]; sd=["All"]
            dr=(datetime(2025,1,1).date(),datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try:
                    _t=excr(pd.read_excel("ot.xlsx"))
                    apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception: pass

    # ===================== DATA LOADING =====================
    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf: raw_ot=pd.read_excel(ot_f); raw_av=pd.read_excel(av_f)
            else: raw_ot=pd.read_excel("ot.xlsx"); raw_av=pd.read_excel("avis.xlsx")
            raw_ot=excr(raw_ot); raw_av=excr(raw_av)
            for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
            for c in ["Créé le","Début souhaité","Date de la clôture"]:
                if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
            if not apm: apm=sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp=apm
            if "All" in sa or not sa: sa=["All"]
            if "All" in sd or not sd: sd=["All"]
            sdt=pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt=pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p=str(poste).upper()
                if "All" not in sa:
                    m=False
                    if "Sulfurique (PS)" in sa and "PS" in p: m=True
                    if "Phosphorique (PP)" in sa and "PP" in p: m=True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m=True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m=True
                    if not m: return False
                if "All" not in sd:
                    m=False
                    if "SF1" in sd and "SF1" in p: m=True
                    if "SF2" in sd and "SF2" in p: m=True
                    if not m: return False
                return True

            vp=[p for p in apm if mf(p) and p in sp]
            df=raw_ot[(raw_ot["Poste travail princ."].isin(vp))&(raw_ot["Date de début planifiée"].between(sdt,edt))].copy()
            avdf=raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df=excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            avdf=excr(avdf[(avdf["Ordre"].isna())|(avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
            df_dash=raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash=excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns: df_dash["Statut OT"]=df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now=pd.Timestamp.now()
            res=calc_kpis(df,avdf,now,vp); ckdf=res['ckdf']; dfp=res['dfp']
            res_d=calc_kpis(df_dash,avdf,now,vp); ckdf_d=res_d['ckdf']
            pa={k:round(ckdf[k].mean(),2) for k in QK}; qa={k:round(ckdf[k].mean(),2) for k in PK}
            pa_d={k:round(ckdf_d[k].mean(),2) for k in QK}; qa_d={k:round(ckdf_d[k].mean(),2) for k in PK}
            pscores={}; qscores={}
            for poste in ckdf.index:
                r=ckdf.loc[poste]
                pscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0
            pscores_d={}; qscores_d={}
            for poste in ckdf_d.index:
                r=ckdf_d.loc[poste]
                pscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0

            # ===== ANOMALIES (colonne "Poste travail princ.") =====
            all_ano=[]
            sub_p={"TAUX_REALISATION_CORRECTIF/PT":lambda d:d[(d["Nº appel pl.entret."].fillna(0)==0)&(~d["Statut OT"].isin(["CLOT","TCLO"]))],
                   "OT préparation <1 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]!="<1 mois")],
                   "OT préparation >3 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]==">3 mois")],
                   "OT planification <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]!="<1 mois")],
                   "OT planification >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]==">3 mois")],
                   "OT exécution <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]!="<1 mois")],
                   "OT exécution >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]==">3 mois")]}
            sub_q={"OT LANC ESTIME":lambda d:d[(d["Statut OT"]=="LANC")&(d["OT LANC ESTIME"]=="NON")],
                   "Backlog préparation caractérisé":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["Backlog preparation"]=="NON CARACTERISE")],
                   "Backlog planification caractérisé":lambda d:d[(d["Statut OT"]=="LANC")&(d["Backlog planification"]=="NON CARACTERISE")],
                   "appel avis approuvé":lambda d:d[d["Statut utilisateur"].str.contains("APRV|APRQ|REJT",na=False)],
                   "OT CONFIME":lambda d:d[(d["OT CONFIME"]=="NON")&(d["Statut OT"].isin(["CLOT","TCLO"]))],
                   "OT_COR_EGAL":lambda d:d[(d["OT_COR_EGAL"]=="NON")&(d["Statut OT"].isin(["CLOT","TCLO"]))&(d["Total coûts réels"].fillna(0)>0)]}

            ano_p_rows=[]; ano_q_rows=[]
            for k,fn in sub_p.items():
                try:
                    sdf=fn(dfp)
                    if sdf.empty: continue
                    for pst,g in sdf.groupby("Poste travail princ."):
                        ano_p_rows.append({"Poste travail princ.":pst,"Indicateurs":k,"Nb anomalies":len(g),"_t":""})
                except Exception: pass
            for k,fn in sub_q.items():
                try:
                    sdf=fn(dfp)
                    if sdf.empty: continue
                    for pst,g in sdf.groupby("Poste travail princ."):
                        ano_q_rows.append({"Poste travail princ.":pst,"Indicateurs":k,"Nb anomalies":len(g),"_t":""})
                except Exception: pass
            if ano_p_rows:
                tot_p=sum(r["Nb anomalies"] for r in ano_p_rows)
                ano_p_rows.append({"Poste travail princ.":"Total","Indicateurs":"","Nb anomalies":tot_p,"_t":"total"})
            if ano_q_rows:
                tot_q=sum(r["Nb anomalies"] for r in ano_q_rows)
                ano_q_rows.append({"Poste travail princ.":"Total","Indicateurs":"","Nb anomalies":tot_q,"_t":"total"})

            # ===== TABLES LIGNES =====
            pcols=["Poste travail princ."]+QK+["Score Performance"]
            prows=[{"Poste travail princ.":p,"_t":""} for p in ckdf.index]
            for r in prows:
                p_=r["Poste travail princ."]
                if p_ in ckdf.index:
                    for k in QK: r[k]=round(ckdf.loc[p_,k],2) if k in ckdf.columns else 0
                    r["Score Performance"]=round(pscores.get(p_,0),2)
            prows.append({"Poste travail princ.":"CIBLE","_t":"cible"})
            for k in QK: prows[-1][k]=CIBLE.get(k,"")
            prows[-1]["Score Performance"]="80%"

            qcols=["Poste travail princ."]+PK+["Score Qualite"]
            qrows=[{"Poste travail princ.":p,"_t":""} for p in ckdf.index]
            for r in qrows:
                p_=r["Poste travail princ."]
                if p_ in ckdf.index:
                    for k in PK: r[k]=round(ckdf.loc[p_,k],2) if k in ckdf.columns else 0
                    r["Score Qualite"]=round(qscores.get(p_,0),2)
            qrows.append({"Poste travail princ.":"CIBLE","_t":"cible"})
            for k in PK: qrows[-1][k]=CIBLE.get(k,"")
            qrows[-1]["Score Qualite"]="80%"

            save_kpis_to_excel(prows,pcols,qrows,qcols,
                               ano_p_rows,["Poste travail princ.","Indicateurs","Nb anomalies"],
                               ano_q_rows,["Poste travail princ.","Indicateurs","Nb anomalies"],
                               fichier_date)

            avg_p=round(np.mean(list(pscores.values())),2) if pscores else 0
            avg_q=round(np.mean(list(qscores.values())),2) if qscores else 0
            avg_p_d=round(np.mean(list(pscores_d.values())),2) if pscores_d else 0
            avg_q_d=round(np.mean(list(qscores_d.values())),2) if qscores_d else 0
            total_ot=len(df); total_ot_d=len(df_dash)
            nb_ano_p=sum(r["Nb anomalies"] for r in ano_p_rows if r.get("_t")!="total")
            nb_ano_q=sum(r["Nb anomalies"] for r in ano_q_rows if r.get("_t")!="total")

            # ===== HISTORIQUE =====
            hist_path=os.path.join("kpis","indicateurs_kpis.xlsx")
            hist_df=load_historical_kpis(hist_path)
            var_df=calculate_variations(hist_df)
            journal_df=generate_journal(var_df)

            # ===== Top/Bottom SEPARES Performance et Qualité =====
            top5_perf,bot5_perf=calculate_rankings_by_type(var_df,"Performance")
            top5_qual,bot5_qual=calculate_rankings_by_type(var_df,"Qualite")

            # ===== OMS + THERMO =====
            oms_df=dfp[dfp["Statut utilisateur"].str.contains("OMS",case=False,na=False)].copy() if "Statut utilisateur" in dfp.columns else pd.DataFrame()
            thermo_df=dfp[dfp["Statut utilisateur"].str.contains("THERM|THERMO",case=False,na=False)].copy() if "Statut utilisateur" in dfp.columns else pd.DataFrame()

            # ===================== INTERFACE =====================
            st.markdown('<div class="mh"><h1>📊 Dashboard KPI Maintenance</h1><span class="db">📅 %s</span></div>'%fichier_date,unsafe_allow_html=True)
            st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%s</div><div class="cl">OT (période)</div></div><div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div><div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualité</div></div><div class="cc c4"><div class="cv">%s</div><div class="cl">Anomalies</div></div></div>'%(total_ot,avg_p,avg_q,nb_ano_p+nb_ano_q),unsafe_allow_html=True)

            # ===== MODIFICATION : Synthèse & Actions EN PREMIER =====
            tabs=st.tabs(["📋 Synthèse & Actions","📊 Indicateurs de Performance","✅ Indicateurs de Qualité","🔍 Analyse OMS & Thermographie"])

                       # ===================== TAB 0 : SYNTHESE & ACTIONS (PREMIER) =====================
            with tabs[0]:
                st.markdown(html_grouped_bars(vp,pscores,qscores,"Scores par Poste travail princ."),unsafe_allow_html=True)

                # Classement Performance
                st.markdown('<div class="stl p" style="margin-top:10px">🏆 Classement — Indicateurs Performance</div>',unsafe_allow_html=True)
                st.markdown(html_classement(pscores,"#38a169"),unsafe_allow_html=True)

                # Classement Qualité
                st.markdown('<div class="stl q" style="margin-top:10px">🏆 Classement — Indicateurs Qualité</div>',unsafe_allow_html=True)
                st.markdown(html_classement(qscores,"#3182ce"),unsafe_allow_html=True)

                # Plan d'Actions Consolidé
                st.markdown('<div class="stl a" style="margin-top:10px">📋 Plan d\'Actions Consolidé — Tous les Indicateurs</div>',unsafe_allow_html=True)
                all_actuals={**pa,**qa}
                st.markdown(html_actions_table(ALL_KPI, all_actuals, CIBLE, ACT_MAP),unsafe_allow_html=True)

                # ===== Top/Bottom basés sur les SCORES ACTUELS =====
                def get_top_bottom(scores_dict, n=5):
                    ranked = sorted(scores_dict.items(), key=lambda x: x[1], reverse=True)
                    top = ranked[:n]
                    bottom = ranked[-n:][::-1] if len(ranked) > n else ranked[::-1]
                    return top, bottom

                top5_p, bot5_p = get_top_bottom(pscores)
                top5_q, bot5_q = get_top_bottom(qscores)

                # Top/Bottom Performance
                st.markdown('<div class="dgrid"><div>')
                st.markdown('<div class="stl p">🏆 Top 5 — Performance</div>',unsafe_allow_html=True)
                for i, (p, s) in enumerate(top5_p):
                    st.markdown('<div class="sr"><span class="sn">%s</span><span class="sc" style="background:#38a169">%.1f%%</span><span class="sa">Score Performance</span></div>'%(p,s),unsafe_allow_html=True)
                st.markdown('</div><div>')
                st.markdown('<div class="stl a">📉 Bottom 5 — Performance</div>',unsafe_allow_html=True)
                for i, (p, s) in enumerate(bot5_p):
                    st.markdown('<div class="sr"><span class="sn">%s</span><span class="sc" style="background:#e53e3e">%.1f%%</span><span class="sa">Score Performance</span></div>'%(p,s),unsafe_allow_html=True)
                st.markdown('</div></div>',unsafe_allow_html=True)

                # Top/Bottom Qualité
                st.markdown('<div class="dgrid" style="margin-top:6px"><div>')
                st.markdown('<div class="stl q">🏆 Top 5 — Qualité</div>',unsafe_allow_html=True)
                for i, (p, s) in enumerate(top5_q):
                    st.markdown('<div class="sr"><span class="sn">%s</span><span class="sc" style="background:#3182ce">%.1f%%</span><span class="sa">Score Qualité</span></div>'%(p,s),unsafe_allow_html=True)
                st.markdown('</div><div>')
                st.markdown('<div class="stl a">📉 Bottom 5 — Qualité</div>',unsafe_allow_html=True)
                for i, (p, s) in enumerate(bot5_q):
                    st.markdown('<div class="sr"><span class="sn">%s</span><span class="sc" style="background:#e53e3e">%.1f%%</span><span class="sa">Score Qualité</span></div>'%(p,s),unsafe_allow_html=True)
                st.markdown('</div></div>',unsafe_allow_html=True)

                # Journal des variations
                if not journal_df.empty:
                    st.markdown('<div class="stl c" style="margin-top:10px">📜 Journal des Variations Significatives (≥5%%)</div>',unsafe_allow_html=True)
                    jcols=["Date actuelle","Poste","Type","Indicateurs","Valeur precedente","Valeur actuelle","Ecart %%","Sens"]
                    jh='<table class="tw st"><thead><tr>'+''.join('<th>%s</th>'%c for c in jcols)+'</tr></thead><tbody>'
                    for _,r in journal_df.head(50).iterrows():
                        sens_clr="#276749" if r["Sens"]=="Amelioration" else "#c53030"
                        jh+='<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%.1f</td><td>%.1f</td><td>%.1f%%</td><td style="color:%s;font-weight:700">%s</td></tr>'%(r["Date actuelle"],r["Poste"],r["Type"],r["KPI"],r["Valeur precedente"],r["Valeur actuelle"],r["Ecart %"],sens_clr,r["Sens"])
                    jh+='</tbody></table>'
                    st.markdown(jh,unsafe_allow_html=True)
                      # ===================== TAB 1 : INDICATEURS DE PERFORMANCE =====================
            with tabs[1]:
                st.markdown('<div class="stl p">Indicateurs de Performance par Poste travail princ.</div>',unsafe_allow_html=True)
                st.markdown(html_table(prows,pcols,"pt",sc_col=set(QK+["Score Performance"])),unsafe_allow_html=True)

                st.markdown('<div class="stl a" style="margin-top:12px">🛠️ Actions — Performance</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table(QK, pa, CIBLE, ACT_MAP),unsafe_allow_html=True)
                st.markdown(html_kpi_bars(QK,pa,CIBLE,"Progression par Indicateurs — Performance","#38a169","#e53e3e"),unsafe_allow_html=True)

                if ano_p_rows:
                    st.markdown('<div class="stl a" style="margin-top:10px">⚠️ Anomalies Performance</div>',unsafe_allow_html=True)
                    st.markdown(html_ano_transpose(ano_p_rows, vp, "Performance", "#e53e3e"),unsafe_allow_html=True)
                    fig_ano_p1, fig_ano_p2 = ano_charts(ano_p_rows, vp, "Performance")
                    if fig_ano_p1 and fig_ano_p2:
                        col_c1, col_c2 = st.columns(2)
                        with col_c1:
                            st.plotly_chart(fig_ano_p1, use_container_width=True)
                        with col_c2:
                            st.plotly_chart(fig_ano_p2, use_container_width=True)
                       # ===================== TAB 2 : INDICATEURS DE QUALITE =====================
            with tabs[2]:
                st.markdown('<div class="stl q">Indicateurs de Qualité par Poste travail princ.</div>',unsafe_allow_html=True)
                st.markdown(html_table(qrows,qcols,"qt",sc_col=set(PK+["Score Qualite"])),unsafe_allow_html=True)

                st.markdown('<div class="stl a" style="margin-top:12px">🛠️ Actions — Qualité</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table(PK, qa, CIBLE, ACT_MAP),unsafe_allow_html=True)
                st.markdown(html_kpi_bars(PK,qa,CIBLE,"Progression par Indicateurs — Qualité","#3182ce","#e53e3e"),unsafe_allow_html=True)

                if ano_q_rows:
                    st.markdown('<div class="stl a" style="margin-top:10px">⚠️ Anomalies Qualité</div>',unsafe_allow_html=True)
                    st.markdown(html_ano_transpose(ano_q_rows, vp, "Qualité", "#e53e3e"),unsafe_allow_html=True)
                    fig_ano_q1, fig_ano_q2 = ano_charts(ano_q_rows, vp, "Qualité")
                    if fig_ano_q1 and fig_ano_q2:
                        col_c1, col_c2 = st.columns(2)
                        with col_c1:
                            st.plotly_chart(fig_ano_q1, use_container_width=True)
                        with col_c2:
                            st.plotly_chart(fig_ano_q2, use_container_width=True)
            # ===================== TAB 3 : OMS & THERMOGRAPHIE (MEME DASH) =====================
            with tabs[3]:
                st.markdown('<div class="dgrid">')

                # --- OMS ---
                st.markdown('<div>')
                st.markdown('<div class="stl q">🔍 Analyse OMS — Contrôle Conditionnel</div>',unsafe_allow_html=True)
                if oms_df.empty:
                    st.markdown('<div class="es">Aucune donnée OMS.</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div class="cr" style="grid-template-columns:repeat(3,1fr)"><div class="cc c1"><div class="cv">%s</div><div class="cl">OT OMS</div></div><div class="cc c2"><div class="cv">%s</div><div class="cl">Postes</div></div><div class="cc c3"><div class="cv">%s</div><div class="cl">Types</div></div></div>'%(len(oms_df),oms_df["Poste travail princ."].nunique(),oms_df["Statut utilisateur"].nunique()),unsafe_allow_html=True)
                    oms_bp=oms_df.groupby("Poste travail princ.").size().reset_index(name="Nombre")
                    fig1=anl_pie_chart(oms_bp,"Poste travail princ.","Nombre","OMS par Poste travail princ.")
                    if fig1: st.plotly_chart(fig1,use_container_width=True)
                    oms_bt=oms_df.groupby("Statut utilisateur").size().reset_index(name="Nombre")
                    fig2=anl_pie_chart(oms_bt,"Statut utilisateur","Nombre","OMS par Type")
                    if fig2: st.plotly_chart(fig2,use_container_width=True)
                    oms_det=oms_df[["Ordre","Poste travail princ.","Statut utilisateur","Description","Créé le","Date de début planifiée"]].copy()
                    oms_det["Créé le"]=oms_det["Créé le"].dt.strftime("%d/%m/%Y")
                    oms_det["Date de début planifiée"]=oms_det["Date de début planifiée"].dt.strftime("%d/%m/%Y")
                    st.dataframe(oms_det,use_container_width=True,height=300)
                    export_btn(oms_det,"analyse_oms.xlsx")
                st.markdown('</div>')

                # --- THERMOGRAPHIE ---
                st.markdown('<div>')
                st.markdown('<div class="stl q">🌡️ Analyse Thermographie — Inspection Thermique</div>',unsafe_allow_html=True)
                if thermo_df.empty:
                    st.markdown('<div class="es">Aucune donnée Thermographie.</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div class="cr" style="grid-template-columns:repeat(3,1fr)"><div class="cc c1"><div class="cv">%s</div><div class="cl">OT Thermo</div></div><div class="cc c2"><div class="cv">%s</div><div class="cl">Postes</div></div><div class="cc c3"><div class="cv">%s</div><div class="cl">Types</div></div></div>'%(len(thermo_df),thermo_df["Poste travail princ."].nunique(),thermo_df["Statut utilisateur"].nunique()),unsafe_allow_html=True)
                    th_bp=thermo_df.groupby("Poste travail princ.").size().reset_index(name="Nombre")
                    fig3=anl_pie_chart(th_bp,"Poste travail princ.","Nombre","Thermo par Poste travail princ.")
                    if fig3: st.plotly_chart(fig3,use_container_width=True)
                    th_bt=thermo_df.groupby("Statut utilisateur").size().reset_index(name="Nombre")
                    fig4=anl_pie_chart(th_bt,"Statut utilisateur","Nombre","Thermo par Type")
                    if fig4: st.plotly_chart(fig4,use_container_width=True)
                    th_det=thermo_df[["Ordre","Poste travail princ.","Statut utilisateur","Description","Créé le","Date de début planifiée"]].copy()
                    th_det["Créé le"]=th_det["Créé le"].dt.strftime("%d/%m/%Y")
                    th_det["Date de début planifiée"]=th_det["Date de début planifiée"].dt.strftime("%d/%m/%Y")
                    st.dataframe(th_det,use_container_width=True,height=300)
                    export_btn(th_det,"analyse_thermo.xlsx")
                st.markdown('</div>')

                st.markdown('</div>',unsafe_allow_html=True)

        except Exception as e:
            st.error(f"Erreur de chargement : {str(e)}")
    else:
        st.markdown('<div class="es" style="margin-top:80px">📁 Veuillez charger les fichiers OT et AVIS depuis la barre latérale.</div>',unsafe_allow_html=True)

if __name__ == "__main__":
    main()
