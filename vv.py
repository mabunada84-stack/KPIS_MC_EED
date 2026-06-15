# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
ALL_KPI = QK + PK
CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100}
ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois."}
LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]
MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

# ============================================================
def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    j["Sens"]=j.apply(lambda r:"Amelioration" if ((r["Tendance"]=="hausse" and r["KPI"] not in LOWER_BETTER) or (r["Tendance"]=="baisse" and r["KPI"] in LOWER_BETTER)) else "Degradation",axis=1)
    return j.sort_values(["Date actuelle","Sens","Ecart %"],ascending=[True,False,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def get_caract_type(statut_user,keywords):
    s=str(statut_user).upper(); matched=[kw for kw in keywords if kw in s]
    return max(matched,key=len) if matched else "AUTRE"

# ============================================================
# FONCTION TELECHARGEMENT EXCEL PAR KPI
# ============================================================
def create_kpi_anomaly_excel(kpi_name, ot_anomalies, avis_anomalies, action_text):
    """Cree un fichier Excel pour un KPI avec OT et AVIS en anomalie, KPI action en rouge."""
    wb=Workbook()
    sn=str(kpi_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","").replace("<","").replace(">","")[:31]
    ws=wb.active; ws.title=sn
    red_font=Font(color="FF0000",bold=True,size=11)
    header_font=Font(bold=True,color="FFFFFF",size=10)
    header_fill=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    title_font=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
    row=1

    # --- SECTION OT ---
    ot_cols=["Ordre","Désignation","Poste technique","Désignation du poste technique",
             "Équipement","Description de l'objet technique","Poste travail princ.",
             "Statut système","Statut utilisateur","Date de début planifiée","KPI action"]
    ws.cell(row=row,column=1,value="ORDRES DE TRAVAIL EN ANOMALIE").font=title_font
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=len(ot_cols))
    row+=1
    for j,col in enumerate(ot_cols,1):
        cl=ws.cell(row=row,column=j,value=col); cl.font=header_font; cl.fill=header_fill; cl.alignment=ca; cl.border=tb
    row+=1
    if ot_anomalies is not None and not ot_anomalies.empty:
        for _,r in ot_anomalies.iterrows():
            for j,col in enumerate(ot_cols,1):
                if col=="KPI action":
                    cl=ws.cell(row=row,column=j,value=action_text); cl.font=red_font
                elif col=="Date de début planifiée":
                    v=r.get(col,"")
                    if pd.notna(v) and hasattr(v,'strftime'): v=v.strftime("%d/%m/%Y")
                    cl=ws.cell(row=row,column=j,value=v)
                else:
                    cl=ws.cell(row=row,column=j,value=r.get(col,""))
                cl.border=tb; cl.alignment=ca
            row+=1
    else:
        ws.cell(row=row,column=1,value="Aucun OT en anomalie").font=Font(italic=True,color="718096")
        row+=1

    row+=2
    # --- SECTION AVIS ---
    avis_cols=["Avis","Ordre","Description","Poste technique","Désignation du poste technique",
               "Poste travail princ.","Statut système","Statut utilisateur","Créé le","KPI action"]
    ws.cell(row=row,column=1,value="AVIS EN ANOMALIE").font=title_font
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=len(avis_cols))
    row+=1
    for j,col in enumerate(avis_cols,1):
        cl=ws.cell(row=row,column=j,value=col); cl.font=header_font; cl.fill=header_fill; cl.alignment=ca; cl.border=tb
    row+=1
    if avis_anomalies is not None and not avis_anomalies.empty:
        for _,r in avis_anomalies.iterrows():
            for j,col in enumerate(avis_cols,1):
                if col=="KPI action":
                    cl=ws.cell(row=row,column=j,value=action_text); cl.font=red_font
                elif col=="Créé le":
                    v=r.get(col,"")
                    if pd.notna(v) and hasattr(v,'strftime'): v=v.strftime("%d/%m/%Y")
                    cl=ws.cell(row=row,column=j,value=v)
                else:
                    cl=ws.cell(row=row,column=j,value=r.get(col,""))
                cl.border=tb; cl.alignment=ca
            row+=1
    else:
        ws.cell(row=row,column=1,value="Aucun avis en anomalie").font=Font(italic=True,color="718096")
        row+=1

    for col_idx in range(1, max(len(ot_cols),len(avis_cols))+1):
        max_len=0
        for cell in ws[column_idx]:
            if cell.value: max_len=max(max_len,len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1,column=col_idx).column_letter].width=min(max_len+4,40)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ============================================================
def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:6px 10px;background:#fff;border-radius:5px;margin-bottom:2px;border:1px solid var(--b);font-size:13px}
    .sr .sn{font-weight:700;color:var(--p);min-width:220px;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:3px 9px;border-radius:12px;font-weight:800;font-size:14px;min-width:50px;text-align:center;margin:0 8px;color:#fff}
    .sr .sa{color:#718096;font-size:12px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .stg{font-size:11px;color:#718096;min-width:60px;text-align:center;white-space:nowrap}
    .sr .sb{font-size:11px;font-weight:700;padding:2px 8px;border-radius:3px;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:4px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:24px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}
    .gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .anl-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;margin:0}
    .anl-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:12px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .anl-tbl tbody td{padding:5px 8px;border-bottom:1px solid #edf2f7}
    .anl-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .anl-tbl tbody tr:hover td{background:#ebf8ff!important}
    .anl-tbl .tot td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .trend-up{color:#276749;font-weight:800;font-size:16px}
    .trend-down{color:#c53030;font-weight:800;font-size:16px}
    .trend-stable{color:#718096;font-weight:800;font-size:16px}
    .spark-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:8px}
    .spark-card{background:#fff;border-radius:var(--r);padding:10px 12px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .spark-card .sp-title{font-size:13px;font-weight:800;color:var(--p);margin-bottom:3px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .spark-card .sp-sub{font-size:11px;color:#718096;margin-bottom:5px}
    .rank-card{background:#fff;border-radius:var(--r);padding:12px 16px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .rank-card .rank-title{font-size:15px;font-weight:800;margin-bottom:8px;padding-bottom:5px;border-bottom:2px solid var(--b)}
    .rank-row{display:flex;align-items:center;padding:5px 0;font-size:13px;border-bottom:1px solid #f7fafc}
    .rank-row:last-child{border:none}
    .rank-row .rank-num{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:13px;color:#fff;margin-right:10px;flex-shrink:0}
    .rank-row .rank-name{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .rank-row .rank-score{font-weight:900;min-width:70px;text-align:right}
    .dl-row{display:flex;align-items:center;gap:8px;margin:4px 0}
    .dl-row .dl-label{flex:1;font-size:12px;font-weight:600;color:var(--p);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:120px}.gbr-l{width:100px}.spark-grid{grid-template-columns:1fr}}
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def contient_mot(t,lm):
        t=str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if pd.isna(a): return "Inconnu"
        if a<=1: return "<1 mois"
        elif a>=3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def cpiv_pt(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste technique",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
        return df

    def calc_kpis(df_i,av_i,now,posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
        df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
        for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc]=pd.to_datetime(df[dc],errors='coerce')
                df[am]=((now.year-df[dc].dt.year)*12+(now.month-df[dc].dt.month)).round(2)
                df[ac]=df[am].apply(cat_age)
            else: df[am]=np.nan; df[ac]="Inconnu"
        df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
        df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
        df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
        df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
        res['dfp']=df
        an=cpiv(df,df["Nº appel pl.entret."].fillna(0)==0,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["Total"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1); an["TAUX_REALISATION_CORRECTIF/PT"]=ckpi(an["TCLO"],an["Total"])
        pr=cpiv(df,df["Statut OT"]=="CRÉÉ","ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
        avf=av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip().eq(""))].copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "color:#cbd5e0"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        return 0
    def is_lb(k): return k in LOWER_BETTER

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                else: s=cs(v) if sc_col and c in sc_col else ks(v,c); h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
    def html_ano(rows,cols):
        h='<table class="tw at"><thead><tr>'+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            h+='<tr class="%s">'%("tr" if r.get("_t")=="total" else "")
            for c in cols: v=r.get(c,""); h+='<td style="%s">%s</td>'%(kas(v) or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
    def html_actions_table_sep(kpi_list,actuals,targets,act_map,section_title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,section_title)
        h+='<table class="tw at" style="margin:0"><thead><tr><th>KPI</th><th>Valeur Actuelle</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action Recommandée</th><th>Telecharger</th></tr></thead><tbody>'
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            status="ATTEINT" if met else "NON ATTEINT"
            st_s="background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
            ec_clr="#276749" if met else "#c53030"
            action="Objectif atteint" if met else act_map.get(k,"")
            h+='<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td>'%(k,av,tv,ec_clr,diff,st_s,status,action)
            if not met:
                safe_kpi=k.replace("/","-").replace("<","").replace(">","").replace(" ","_")[:30]
                h+='<td><button onclick="window.dispatchEvent(new CustomEvent(\'dl_kpi\',{detail:\'%s\'}))" style="background:#e53e3e;color:#fff;border:none;padding:3px 10px;border-radius:4px;cursor:pointer;font-weight:700;font-size:11px">📥 %s</button></td>'%(safe_kpi,safe_kpi)
            else:
                h+='<td></td>'
            h+='</tr>'
        h+='</tbody></table></div>'; return h
    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'; return h
    def html_kpi_bars(kpi_list,actuals,targets,title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,title)
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); met=av<=tv if is_lb(k) else av>=tv
            bw=min(max(av,0),100); bg=color_ok if met else color_fail
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(k,bw,bg,av)
        return h+'</div>'
    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        for p in sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True):
            pv,qv=pscores.get(p,0),qscores.get(p,0)
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>'%(p,min(max(pv,0),100),pv,min(max(qv,0),100),qv)
        return h+'</div>'
    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",unsafe_allow_html=True)
        st.markdown("---")
        show_filters=st.checkbox("Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf")
            ot_f=av_f=None; apm=[]
            if unf:
                ot_f=st.file_uploader("Fichier OT",type=["xlsx"],key="uot")
                av_f=st.file_uploader("Fichier AVIS",type=["xlsx"],key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t=excr(pd.read_excel("ot.xlsx"))
                        apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            # === MODIF: Atelier => Centrale / Utilitaires ===
            st.markdown("**🏭 Atelier**")
            sa=st.multiselect("Atelier",["All","Centrale","Utilitaires"],["All"],key="sa")
            st.markdown("**🏢 Division**")
            sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=av_f=None; apm=[]; sp=["All"]; sa=["All"]; sd=["All"]
            dr=(datetime(2025,1,1).date(),datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try:
                    _t=excr(pd.read_excel("ot.xlsx"))
                    apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception: pass

    # ===================== DATA LOADING =====================
    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf: raw_ot=pd.read_excel(ot_f); raw_av=pd.read_excel(av_f)
            else: raw_ot=pd.read_excel("ot.xlsx"); raw_av=pd.read_excel("avis.xlsx")
            raw_ot=excr(raw_ot); raw_av=excr(raw_av)
            for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
            for c in ["Créé le","Début souhaité","Date de la clôture"]:
                if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
            if not apm: apm=sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp=apm
            if "All" in sa or not sa: sa=["All"]
            if "All" in sd or not sd: sd=["All"]
            sdt=pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt=pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            # === MODIF: Filtre Atelier = Centrale / Utilitaires ===
            def mf(poste):
                p=str(poste).upper()
                if "All" not in sa:
                    m=False
                    if "Centrale" in sa and ("CENTRALE" in p or "CNTR" in p): m=True
                    if "Utilitaires" in sa and ("UTILITAIRE" in p or "UTIL" in p): m=True
                    if not m: return False
                if "All" not in sd:
                    m=False
                    if "SF1" in sd and "SF1" in p: m=True
                    if "SF2" in sd and "SF2" in p: m=True
                    if not m: return False
                return True

            vp=[p for p in apm if mf(p) and p in sp]
            df=raw_ot[(raw_ot["Poste travail princ."].isin(vp))&(raw_ot["Date de début planifiée"].between(sdt,edt))].copy()
            avdf=raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df=excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            avdf=excr(avdf[(avdf["Ordre"].isna())|(avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
            df_dash=raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash=excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns: df_dash["Statut OT"]=df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now=pd.Timestamp.now()
            res=calc_kpis(df,avdf,now,vp); ckdf=res['ckdf']; dfp=res['dfp']
            res_d=calc_kpis(df_dash,avdf,now,vp); ckdf_d=res_d['ckdf']
            pa={k:round(ckdf[k].mean(),2) for k in QK}; qa={k:round(ckdf[k].mean(),2) for k in PK}
            pa_d={k:round(ckdf_d[k].mean(),2) for k in QK}; qa_d={k:round(ckdf_d[k].mean(),2) for k in PK}
            pscores={}; qscores={}
            for poste in ckdf.index:
                r=ckdf.loc[poste]
                pscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0
            pscores_d={}; qscores_d={}
            for poste in ckdf_d.index:
                r=ckdf_d.loc[poste]
                pscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0

            # === ANOMALIES DETECTION ===
            sub_p={"TAUX_REALISATION_CORRECTIF/PT":lambda d:d[(d["Nº appel pl.entret."].fillna(0)==0)&(~d["Statut OT"].isin(["CLOT","TCLO"]))],
                   "OT préparation <1 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]!="<1 mois")],
                   "OT préparation >3 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]==">3 mois")],
                   "OT planification <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]!="<1 mois")],
                   "OT planification >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]==">3 mois")],
                   "OT exécution <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]!="<1 mois")],
                   "OT exécution >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]==">3 mois")]}
            sub_q={"OT LANC ESTIME":lambda d:d[(d["Statut OT"]=="LANC")&(d["OT LANC ESTIME"]=="NON")],
                   "Backlog préparation caractérisé":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["Backlog preparation"]=="NON CARACTERISE")],
                   "Backlog planification caractérisé":lambda d:d[(d["Statut OT"]=="LANC")&(d["Backlog planification"]=="NON CARACTERISE")],
                   "OT CONFIME":lambda d:d[(d["Statut OT"]=="TCLO")&(d["OT CONFIME"]=="NON")],
                   "OT_COR_EGAL":lambda d:d[(d["Statut OT"]=="TCLO")&(d["OT_COR_EGAL"]=="NON")],
                   "appel avis approuvé":lambda d:pd.DataFrame()}

            # Collecter anomalies OT par KPI
            ano_ot_data={}
            for kpi_name, filt_func in {**sub_p,**sub_q}.items():
                try:
                    ano_ot_data[kpi_name]=filt_func(dfp).copy()
                except Exception:
                    ano_ot_data[kpi_name]=pd.DataFrame()

            # Anomalies AVIS pour "appel avis approuvé"
            ano_avis_data={}
            avf=res.get('avf',pd.DataFrame())
            if not avf.empty:
                try:
                    ano_avis_data["appel avis approuvé"]=avf[~avf["Statut utilisateur"].str.contains("APRV",na=False)].copy()
                except Exception:
                    ano_avis_data["appel avis approuvé"]=pd.DataFrame()
            else:
                ano_avis_data["appel avis approuvé"]=pd.DataFrame()

            # Compter anomalies par poste
            ano_p_counts={}; ano_q_counts={}
            for kpi in QK:
                if kpi in ano_ot_data and not ano_ot_data[kpi].empty:
                    grp=ano_ot_data[kpi].groupby("Poste travail princ.")["Ordre"].count()
                    ano_p_counts[kpi]=grp.reindex(vp,fill_value=0).to_dict()
                else:
                    ano_p_counts[kpi]={p:0 for p in vp}
            for kpi in PK:
                if kpi=="appel avis approuvé":
                    if "appel avis approuvé" in ano_avis_data and not ano_avis_data["appel avis approuvé"].empty:
                        grp=ano_avis_data["appel avis approuvé"].groupby("Poste travail princ.")["Avis"].count()
                        ano_q_counts[kpi]=grp.reindex(vp,fill_value=0).to_dict()
                    else:
                        ano_q_counts[kpi]={p:0 for p in vp}
                elif kpi in ano_ot_data and not ano_ot_data[kpi].empty:
                    grp=ano_ot_data[kpi].groupby("Poste travail princ.")["Ordre"].count()
                    ano_q_counts[kpi]=grp.reindex(vp,fill_value=0).to_dict()
                else:
                    ano_q_counts[kpi]={p:0 for p in vp}

            # Total anomalies par poste
            tot_ano_p={p:sum(ano_p_counts[k].get(p,0) for k in QK) for p in vp}
            tot_ano_q={p:sum(ano_q_counts[k].get(p,0) for k in PK) for p in vp}

            # Scores globaux
            avg_p=round(np.mean(list(pscores.values())),2) if pscores else 0
            avg_q=round(np.mean(list(qscores.values())),2) if qscores else 0
            avg_p_d=round(np.mean(list(pscores_d.values())),2) if pscores_d else 0
            avg_q_d=round(np.mean(list(qscores_d.values())),2) if qscores_d else 0
            tot_ot=len(df); tot_av=len(avf)
            tot_ano_all=sum(tot_ano_p.values())+sum(tot_ano_q.values())

            # Sauvegarde Excel KPIs
            pcols=["Poste de travail"]+QK+["Score Performance"]
            qcols=["Poste de travail"]+PK+["Score Qualite"]
            prows=[{"Poste de travail":p,"_t":""} for p in vp]
            qrows=[{"Poste de travail":p,"_t":""} for p in vp]
            for i,p in enumerate(vp):
                for k in QK: prows[i][k]=round(ckdf.loc[p,k],2) if p in ckdf.index and k in ckdf.columns else 0
                prows[i]["Score Performance"]=round(pscores.get(p,0),2)
                for k in PK: qrows[i][k]=round(ckdf.loc[p,k],2) if p in ckdf.index and k in ckdf.columns else 0
                qrows[i]["Score Qualite"]=round(qscores.get(p,0),2)
            prows.append({"Poste de travail":"CIBLE","_t":"cible"})
            for k in QK: prows[-1][k]=CIBLE[k]
            prows[-1]["Score Performance"]=80
            prows.append({"Poste de travail":"Total general","_t":"total"})
            for k in QK: prows[-1][k]=round(ckdf[k].mean(),2) if k in ckdf.columns else 0
            prows[-1]["Score Performance"]=avg_p
            qrows.append({"Poste de travail":"CIBLE","_t":"cible"})
            for k in PK: qrows[-1][k]=CIBLE[k]
            qrows[-1]["Score Qualite"]=90
            qrows.append({"Poste de travail":"Total general","_t":"total"})
            for k in PK: qrows[-1][k]=round(ckdf[k].mean(),2) if k in ckdf.columns else 0
            qrows[-1]["Score Qualite"]=avg_q

            ano_p_c=["Poste de travail"]+QK+["Total"]
            ano_p_r=[{"Poste de travail":p,"_t":""} for p in vp]
            for i,p in enumerate(vp):
                for k in QK: ano_p_r[i][k]=ano_p_counts[k].get(p,0)
                ano_p_r[i]["Total"]=tot_ano_p[p]
            ano_p_r.append({"Poste de travail":"Total general","_t":"total"})
            for k in QK: ano_p_r[-1][k]=sum(ano_p_counts[k].values())
            ano_p_r[-1]["Total"]=sum(tot_ano_p.values())

            ano_q_c=["Poste de travail"]+PK+["Total"]
            ano_q_r=[{"Poste de travail":p,"_t":""} for p in vp]
            for i,p in enumerate(vp):
                for k in PK: ano_q_r[i][k]=ano_q_counts[k].get(p,0)
                ano_q_r[i]["Total"]=tot_ano_q[p]
            ano_q_r.append({"Poste de travail":"Total general","_t":"total"})
            for k in PK: ano_q_r[-1][k]=sum(ano_q_counts[k].values())
            ano_q_r[-1]["Total"]=sum(tot_ano_q.values())

            save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,fichier_date)

            # === Historique & Variations ===
            hist_path=os.path.join("kpis","indicateurs_kpis.xlsx")
            hist_df=load_historical_kpis(hist_path)
            var_df=calculate_variations(hist_df)
            journal_df=generate_journal(var_df)
            top5_df,bot5_df=calculate_rankings(var_df)

            # === Colonnes Poste technique pour OMS/Thermo ===
            pt_col="Poste technique" if "Poste technique" in dfp.columns else None
            pt_desc_col="Désignation du poste technique" if "Désignation du poste technique" in dfp.columns else None

            # ===================== RENDERING =====================
            st.markdown('<div class="mh"><h1>📊 Dashboard KPI Maintenance</h1><span class="db">📅 %s</span></div>'%fichier_date,unsafe_allow_html=True)
            st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%d</div><div class="cl">OT Total</div></div><div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div><div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div><div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies</div></div></div>'%(tot_ot,avg_p,avg_q,tot_ano_all),unsafe_allow_html=True)

            tab1,tab2,tab3,tab4,tab5,tab6,tab7,tab8=st.tabs(["📊 Dashboard","📈 Indicateur Performance","✅ Indicateur Qualite","🔍 Analyse OMS","🌡️ Analyse Thermographie","📋 Synthèse & Actions","📉 Variations","🏆 Classement"])

            # ============ TAB 1: DASHBOARD ============
            with tab1:
                st.markdown('<div class="stl p">Score Global par Poste de Travail</div>',unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp,pscores,qscores,"Comparaison Performance vs Qualite par Poste"),unsafe_allow_html=True)
                c1,c2=st.columns(2)
                with c1:
                    st.markdown('<div class="stl p">Indicateurs de Performance</div>',unsafe_allow_html=True)
                    st.markdown(html_kpi_bars(QK,pa,CIBLE,"Performance Globale","#38a169","#e53e3e"),unsafe_allow_html=True)
                with c2:
                    st.markdown('<div class="stl q">Indicateurs de Qualite</div>',unsafe_allow_html=True)
                    st.markdown(html_kpi_bars(PK,qa,CIBLE,"Qualite Globale","#3182ce","#e53e3e"),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Anomalies - Performance</div>',unsafe_allow_html=True)
                st.markdown(html_ano(ano_p_r,ano_p_c),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Anomalies - Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_ano(ano_q_r,ano_q_c),unsafe_allow_html=True)

            # ============ TAB 2: INDICATEUR PERFORMANCE ============
            with tab2:
                # === CHART BAR EN PREMIER ===
                st.markdown('<div class="stl p">Chart des Indicateurs de Performance</div>',unsafe_allow_html=True)
                fig_perf=go.Figure()
                kpi_labels_perf=[k.replace("OT ","OT\n").replace("<","&lt;").replace(">","&gt;") for k in QK]
                fig_perf.add_trace(go.Bar(name='Valeur Actuelle',x=QK,y=[pa.get(k,0) for k in QK],
                    marker_color=['#38a169' if (pa.get(k,0)<=CIBLE[k] if is_lb(k) else pa.get(k,0)>=CIBLE[k]) else '#e53e3e' for k in QK]))
                fig_perf.add_trace(go.Bar(name='Cible',x=QK,y=[CIBLE[k] for k in QK],
                    marker_color='rgba(200,200,200,0.5)',marker_line_color='rgba(150,150,150,0.8)',marker_line_width=1))
                fig_perf.update_layout(barmode='group',height=420,margin=dict(t=40,b=120,l=60,r=20),
                    xaxis_tickangle=-45,title_font_size=14,legend=dict(orientation="h",yanchor="bottom",y=1.02))
                st.plotly_chart(fig_perf,use_container_width=True)

                st.markdown('<div class="stl p">Tableau Detaille - Performance par Poste</div>',unsafe_allow_html=True)
                st.markdown(html_table(prows,pcols,"pt",sc_col=QK+["Score Performance"]),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Anomalies Performance</div>',unsafe_allow_html=True)
                st.markdown(html_ano(ano_p_r,ano_p_c),unsafe_allow_html=True)

                # === BOUTONS TELECHARGEMENT PAR KPI PERFORMANCE ===
                st.markdown('<div class="stl a">Telechargement des Anomalies par KPI (Performance)</div>',unsafe_allow_html=True)
                dl_cols_p=st.columns(3)
                for idx,kpi in enumerate(QK):
                    with dl_cols_p[idx%3]:
                        ot_ano=ano_ot_data.get(kpi,pd.DataFrame())
                        av_ano=pd.DataFrame()
                        action_txt=ACT_MAP.get(kpi,"")
                        if not ot_ano.empty:
                            buf=create_kpi_anomaly_excel(kpi,ot_ano,av_ano,action_txt)
                            safe_fn=kpi.replace("/","-").replace("<","").replace(">","").replace(" ","_")[:30]
                            st.download_button("📥 %s (%d OTs)"%(kpi,len(ot_ano)),data=buf,
                                file_name="%s_anomalies.xlsx"%safe_fn,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_p_%s"%safe_fn)

            # ============ TAB 3: INDICATEUR QUALITE ============
            with tab3:
                # === CHART BAR EN PREMIER ===
                st.markdown('<div class="stl q">Chart des Indicateurs de Qualite</div>',unsafe_allow_html=True)
                fig_qual=go.Figure()
                fig_qual.add_trace(go.Bar(name='Valeur Actuelle',x=PK,y=[qa.get(k,0) for k in PK],
                    marker_color=['#3182ce' if (qa.get(k,0)<=CIBLE[k] if is_lb(k) else qa.get(k,0)>=CIBLE[k]) else '#e53e3e' for k in PK]))
                fig_qual.add_trace(go.Bar(name='Cible',x=PK,y=[CIBLE[k] for k in PK],
                    marker_color='rgba(200,200,200,0.5)',marker_line_color='rgba(150,150,150,0.8)',marker_line_width=1))
                fig_qual.update_layout(barmode='group',height=420,margin=dict(t=40,b=120,l=60,r=20),
                    xaxis_tickangle=-45,title_font_size=14,legend=dict(orientation="h",yanchor="bottom",y=1.02))
                st.plotly_chart(fig_qual,use_container_width=True)

                st.markdown('<div class="stl q">Tableau Detaille - Qualite par Poste</div>',unsafe_allow_html=True)
                st.markdown(html_table(qrows,qcols,"qt",sc_col=PK+["Score Qualite"]),unsafe_allow_html=True)
                st.markdown('<div class="stl a">Anomalies Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_ano(ano_q_r,ano_q_c),unsafe_allow_html=True)

                # === BOUTONS TELECHARGEMENT PAR KPI QUALITE ===
                st.markdown('<div class="stl a">Telechargement des Anomalies par KPI (Qualite)</div>',unsafe_allow_html=True)
                dl_cols_q=st.columns(3)
                for idx,kpi in enumerate(PK):
                    with dl_cols_q[idx%3]:
                        ot_ano=ano_ot_data.get(kpi,pd.DataFrame())
                        av_ano=ano_avis_data.get(kpi,pd.DataFrame()) if kpi=="appel avis approuvé" else pd.DataFrame()
                        action_txt=ACT_MAP.get(kpi,"")
                        n_items=len(ot_ano)+len(av_ano)
                        if n_items>0:
                            buf=create_kpi_anomaly_excel(kpi,ot_ano,av_ano,action_txt)
                            safe_fn=kpi.replace("/","-").replace("<","").replace(">","").replace(" ","_")[:30]
                            st.download_button("📥 %s (%d)"%(kpi,n_items),data=buf,
                                file_name="%s_anomalies.xlsx"%safe_fn,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_q_%s"%safe_fn)

            # ============ TAB 4: ANALYSE OMS (par Poste technique) ============
            with tab4:
                st.markdown('<div class="stl c">Analyse OMS par Poste Technique</div>',unsafe_allow_html=True)
                if pt_col and pt_col in dfp.columns:
                    pt_list=dfp[pt_col].dropna().unique().tolist()
                    selected_pt_oms=st.multiselect("Filtrer par Poste technique (OMS)",["All"]+sorted(pt_list),["All"],key="pt_oms")
                    df_oms=dfp.copy()
                    if "All" not in selected_pt_oms and selected_pt_oms:
                        df_oms=df_oms[df_oms[pt_col].isin(selected_pt_oms)]

                    # Stats OMS par Poste technique
                    oms_pivot=pd.pivot_table(df_oms,index=pt_col,columns="Statut OT",values="Ordre",aggfunc="count",fill_value=0)
                    for c in ["CLOT","CRÉÉ","LANC","TCLO"]:
                        if c not in oms_pivot.columns: oms_pivot[c]=0
                    oms_pivot["Total"]=oms_pivot.sum(axis=1)
                    oms_pivot=oms_pivot.sort_values("Total",ascending=False)

                    st.markdown('<div class="ca"><div class="ct" style="color:#6b46c1">Distribution des OT par Poste Technique</div></div>',unsafe_allow_html=True)
                    fig_oms=go.Figure()
                    status_colors={"CRÉÉ":"#e53e3e","LANC":"#d69e2e","TCLO":"#38a169","CLOT":"#3182ce"}
                    for sc in ["CRÉÉ","LANC","TCLO","CLOT"]:
                        if sc in oms_pivot.columns:
                            fig_oms.add_trace(go.Bar(name=sc,x=oms_pivot.index.tolist(),y=oms_pivot[sc].tolist(),
                                marker_color=status_colors.get(sc,"#718096")))
                    fig_oms.update_layout(barmode='stack',height=450,margin=dict(t=30,b=150,l=200,r=20),
                        xaxis_tickangle=-60,yaxis_title="Nombre d'OT",legend=dict(orientation="h",yanchor="bottom",y=1.02))
                    st.plotly_chart(fig_oms,use_container_width=True)

                    # Caracterisation OMS par Poste technique
                    if "Backlog preparation" in df_oms.columns:
                        oms_car=pd.pivot_table(df_oms[df_oms["Statut OT"]=="CRÉÉ"],index=pt_col,columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0)
                        for c in ["CARACTERISE","NON CARACTERISE"]:
                            if c not in oms_car.columns: oms_car[c]=0
                        oms_car["Total"]=oms_car.sum(axis=1)
                        oms_car=oms_car.sort_values("Total",ascending=False).head(20)

                        oms_cols=["Poste technique","CARACTERISE","NON CARACTERISE","Total"]
                        oms_rows=[]
                        for pt_idx,row in oms_car.iterrows():
                            oms_rows.append({"Poste technique":str(pt_idx),"CARACTERISE":int(row.get("CARACTERISE",0)),"NON CARACTERISE":int(row.get("NON CARACTERISE",0)),"Total":int(row["Total"]),"_t":""})
                        oms_rows.append({"Poste technique":"Total","CARACTERISE":int(oms_car["CARACTERISE"].sum()),"NON CARACTERISE":int(oms_car["NON CARACTERISE"].sum()),"Total":int(oms_car["Total"].sum()),"_t":"total"})
                        st.markdown('<div class="stl c">Caracterisation Backlog Preparation par Poste Technique</div>',unsafe_allow_html=True)
                        st.markdown(html_table(oms_rows,oms_cols,"st"),unsafe_allow_html=True)

                    # Age OMS par Poste technique
                    if "amp" in df_oms.columns:
                        oms_age=pd.pivot_table(df_oms[df_oms["Statut OT"]=="CRÉÉ"],index=pt_col,columns="ap",values="Ordre",aggfunc="count",fill_value=0)
                        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]:
                            if c not in oms_age.columns: oms_age[c]=0
                        oms_age["Total"]=oms_age.sum(axis=1)
                        oms_age=oms_age.sort_values("Total",ascending=False).head(20)

                        age_cols=["Poste technique","<1 mois","1 mois < <3 mois",">3 mois","Inconnu","Total"]
                        age_rows=[]
                        for pt_idx,row in oms_age.iterrows():
                            age_rows.append({"Poste technique":str(pt_idx),"<1 mois":int(row.get("<1 mois",0)),"1 mois < <3 mois":int(row.get("1 mois < <3 mois",0)),">3 mois":int(row.get(">3 mois",0)),"Inconnu":int(row.get("Inconnu",0)),"Total":int(row["Total"]),"_t":""})
                        age_rows.append({"Poste technique":"Total","<1 mois":int(oms_age["<1 mois"].sum()),"1 mois < <3 mois":int(oms_age["1 mois < <3 mois"].sum()),">3 mois":int(oms_age[">3 mois"].sum()),"Inconnu":int(oms_age["Inconnu"].sum()),"Total":int(oms_age["Total"].sum()),"_t":"total"})
                        st.markdown('<div class="stl c">Age des OT par Poste Technique</div>',unsafe_allow_html=True)
                        st.markdown(html_table(age_rows,age_cols,"st"),unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">Colonne "Poste technique" non trouvee dans les donnees.</div>',unsafe_allow_html=True)

            # ============ TAB 5: ANALYSE THERMOGRAPHIE (par Poste technique) ============
            with tab5:
                st.markdown('<div class="stl s">Analyse Thermographie par Poste Technique</div>',unsafe_allow_html=True)
                if pt_col and pt_col in dfp.columns:
                    # Filtrer les OT liés à la thermographie (mots-clés dans désignation ou statut)
                    thermo_kw=["THERMO","THERM","INFRAROUGE","IR ","TEMPERATURE","CHAUDE","CHAUD","SURCHAUFFE"]
                    df_thermo=dfp[dfp.apply(lambda r:any(kw in str(r.get("Désignation","")).upper() or kw in str(r.get("Description de l'objet technique","")).upper() or kw in str(r.get("Statut utilisateur","")).upper() for kw in thermo_kw),axis=1)].copy()

                    selected_pt_th=st.multiselect("Filtrer par Poste technique (Thermo)",["All"]+sorted(pt_list),["All"],key="pt_th")
                    if "All" not in selected_pt_th and selected_pt_th:
                        df_thermo=df_thermo[df_thermo[pt_col].isin(selected_pt_th)]

                    if df_thermo.empty:
                        st.markdown('<div class="es">Aucun OT de thermographie trouve dans les donnees filtrees.<br>Recherche par mots-cles: %s</div>'%", ".join(thermo_kw),unsafe_allow_html=True)
                    else:
                        st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%d</div><div class="cl">OT Thermographie</div></div><div class="cc c2"><div class="cv">%d</div><div class="cl">Postes Techniques</div></div><div class="cc c3"><div class="cv">%d</div><div class="cl">En Cours</div></div><div class="cc c4"><div class="cv">%d</div><div class="cl">Clos</div></div></div>'%(
                            len(df_thermo),df_thermo[pt_col].nunique(),
                            len(df_thermo[df_thermo["Statut OT"].isin(["CRÉÉ","LANC"])]),
                            len(df_thermo[df_thermo["Statut OT"].isin(["CLOT","TCLO"])])),unsafe_allow_html=True)

                        # Chart thermographie par poste technique
                        th_pivot=pd.pivot_table(df_thermo,index=pt_col,columns="Statut OT",values="Ordre",aggfunc="count",fill_value=0)
                        for c in ["CLOT","CRÉÉ","LANC","TCLO"]:
                            if c not in th_pivot.columns: th_pivot[c]=0
                        th_pivot["Total"]=th_pivot.sum(axis=1)
                        th_pivot=th_pivot.sort_values("Total",ascending=False).head(15)

                        fig_th=go.Figure()
                        status_colors={"CRÉÉ":"#e53e3e","LANC":"#d69e2e","TCLO":"#38a169","CLOT":"#3182ce"}
                        for sc in ["CRÉÉ","LANC","TCLO","CLOT"]:
                            if sc in th_pivot.columns:
                                fig_th.add_trace(go.Bar(name=sc,x=th_pivot.index.tolist(),y=th_pivot[sc].tolist(),
                                    marker_color=status_colors.get(sc,"#718096")))
                        fig_th.update_layout(barmode='stack',height=450,margin=dict(t=30,b=150,l=200,r=20),
                            xaxis_tickangle=-60,yaxis_title="Nombre d'OT",legend=dict(orientation="h",yanchor="bottom",y=1.02))
                        st.plotly_chart(fig_th,use_container_width=True)

                        # Tableau detail thermographie
                        th_cols=[pt_col,"Statut OT","Statut utilisateur","Désignation","Créé le","Date de début planifiée"]
                        th_display_cols=["Poste technique","Statut OT","Statut utilisateur","Désignation","Créé le","Date de début planifiée"]
                        th_rows=[]
                        for _,r in df_thermo.sort_values("Créé le",ascending=False).head(50).iterrows():
                            row_dict={"_t":""}
                            for src,dst in zip(th_cols,th_display_cols):
                                v=r.get(src,"")
                                if pd.notna(v) and hasattr(v,'strftime'): v=v.strftime("%d/%m/%Y")
                                row_dict[dst]=str(v) if pd.notna(v) else ""
                            th_rows.append(row_dict)
                        st.markdown('<div class="stl s">Detail des OT Thermographie</div>',unsafe_allow_html=True)
                        st.markdown(html_table(th_rows,th_display_cols,"st"),unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">Colonne "Poste technique" non trouvee dans les donnees.</div>',unsafe_allow_html=True)

            # ============ TAB 6: SYNTHESE & ACTIONS (SEPARE Performance / Qualite) ============
            with tab6:
                # === SECTION PERFORMANCE ===
                st.markdown('<div class="stl p">Synthese & Actions — PERFORMANCE</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table_sep(QK,pa,CIBLE,ACT_MAP,"Actions Performance","#38a169","#e53e3e"),unsafe_allow_html=True)

                # Boutons telechargement Performance
                st.markdown('<div class="stl a">Telecharger Anomalies Performance</div>',unsafe_allow_html=True)
                dl_sp=st.columns(4)
                for idx,kpi in enumerate(QK):
                    with dl_sp[idx%4]:
                        ot_ano=ano_ot_data.get(kpi,pd.DataFrame())
                        action_txt=ACT_MAP.get(kpi,"")
                        if not ot_ano.empty:
                            buf=create_kpi_anomaly_excel(kpi,ot_ano,pd.DataFrame(),action_txt)
                            safe_fn=kpi.replace("/","-").replace("<","").replace(">","").replace(" ","_")[:30]
                            st.download_button("📥 %s (%d)"%(kpi[:25],len(ot_ano)),data=buf,
                                file_name="%s_anomalies.xlsx"%safe_fn,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_s_p_%s"%safe_fn)

                st.markdown("---",unsafe_allow_html=True)

                # === SECTION QUALITE ===
                st.markdown('<div class="stl q">Synthese & Actions — QUALITE</div>',unsafe_allow_html=True)
                st.markdown(html_actions_table_sep(PK,qa,CIBLE,ACT_MAP,"Actions Qualite","#3182ce","#e53e3e"),unsafe_allow_html=True)

                # Boutons telechargement Qualite
                st.markdown('<div class="stl a">Telecharger Anomalies Qualite</div>',unsafe_allow_html=True)
                dl_sq=st.columns(3)
                for idx,kpi in enumerate(PK):
                    with dl_sq[idx%3]:
                        ot_ano=ano_ot_data.get(kpi,pd.DataFrame())
                        av_ano=ano_avis_data.get(kpi,pd.DataFrame()) if kpi=="appel avis approuvé" else pd.DataFrame()
                        action_txt=ACT_MAP.get(kpi,"")
                        n_items=len(ot_ano)+len(av_ano)
                        if n_items>0:
                            buf=create_kpi_anomaly_excel(kpi,ot_ano,av_ano,action_txt)
                            safe_fn=kpi.replace("/","-").replace("<","").replace(">","").replace(" ","_")[:30]
                            st.download_button("📥 %s (%d)"%(kpi[:25],n_items),data=buf,
                                file_name="%s_anomalies.xlsx"%safe_fn,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_s_q_%s"%safe_fn)

            # ============ TAB 7: VARIATIONS ============
            with tab7:
                st.markdown('<div class="stl s">Journal des Variations Significatives</div>',unsafe_allow_html=True)
                if journal_df.empty:
                    st.markdown('<div class="es">Pas assez d\'historique pour calculer les variations (minimum 2 periodes requises).</div>',unsafe_allow_html=True)
                else:
                    jcols=["Date actuelle","Poste","Type","KPI","Valeur precedente","Valeur actuelle","Ecart %","Sens"]
                    jrows=[]
                    for _,r in journal_df.iterrows():
                        jrows.append({"Date actuelle":r["Date actuelle"],"Poste":r["Poste"],"Type":r["Type"],
                            "KPI":r["KPI"],"Valeur precedente":r["Valeur precedente"],"Valeur actuelle":r["Valeur actuelle"],
                            "Ecart %":round(r["Ecart %"],2),"Sens":r["Sens"],"_t":""})
                    jrows.sort(key=lambda x:x["Ecart %"] if x["Sens"]=="Degradation" else -x["Ecart %"],reverse=True)
                    st.markdown(html_table(jrows,jcols,"st",sc_col=["Ecart %"]),unsafe_allow_html=True)

                    if not top5_df.empty and not bot5_df.empty:
                        c1,c2=st.columns(2)
                        with c1:
                            st.markdown('<div class="stl p">Top 5 Amelioration</div>',unsafe_allow_html=True)
                            t5c=["Poste","Score variation"]
                            t5r=[{"Poste":r["Poste"],"Score variation":round(r["Score variation"],2),"_t":""} for _,r in top5_df.iterrows()]
                            st.markdown(html_table(t5r,t5c,"pt"),unsafe_allow_html=True)
                        with c2:
                            st.markdown('<div class="stl a">Top 5 Degradation</div>',unsafe_allow_html=True)
                            b5c=["Poste","Score variation"]
                            b5r=[{"Poste":r["Poste"],"Score variation":round(r["Score variation"],2),"_t":""} for _,r in bot5_df.iterrows()]
                            st.markdown(html_table(b5r,b5c,"at"),unsafe_allow_html=True)

            # ============ TAB 8: CLASSEMENT ============
            with tab8:
                st.markdown('<div class="stl p">Classement Performance</div>',unsafe_allow_html=True)
                st.markdown(html_classement(pscores,"#38a169"),unsafe_allow_html=True)
                st.markdown('<div class="stl q">Classement Qualite</div>',unsafe_allow_html=True)
                st.markdown(html_classement(qscores,"#3182ce"),unsafe_allow_html=True)

                # Classement combine
                st.markdown('<div class="stl c">Classement Combine</div>',unsafe_allow_html=True)
                combined={p:(pscores.get(p,0)+qscores.get(p,0))/2 for p in vp}
                st.markdown(html_classement(combined,"#6b46c1"),unsafe_allow_html=True)

                # Chart radar top 5
                top5_combined=sorted(combined.items(),key=lambda x:x[1],reverse=True)[:5]
                if top5_combined:
                    fig_radar=go.Figure()
                    for poste,score in top5_combined:
                        fig_radar.add_trace(go.Scatterpolar(
                            r=[pscores.get(poste,0),qscores.get(poste,0),score],
                            theta=["Performance","Qualite","Combine"],
                            fill='toself',name=poste))
                    fig_radar.update_layout(polar=dict(radialaxis=dict(visible=True,range=[0,100])),height=500,showlegend=True)
                    st.plotly_chart(fig_radar,use_container_width=True)

        except Exception as e:
            st.error("Erreur de chargement: %s"%str(e))
            import traceback; st.code(traceback.format_exc())
    else:
        st.markdown('<div class="es" style="padding:60px"><div style="font-size:64px;margin-bottom:20px">📁</div><h2 style="color:#1e3a5f;margin-bottom:10px">Aucun fichier charge</h2><p style="color:#718096">Veuillez activer "Charger nouveaux fichiers" dans le menu lateral ou placer les fichiers ot.xlsx et avis.xlsx dans le repertoire.</p></div>',unsafe_allow_html=True)

if __name__=="__main__":
    main()
