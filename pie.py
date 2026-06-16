# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os, json, hashlib
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
ALL_KPI = QK + PK
CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100}
ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois."}
LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]
MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

# ============================================================
# PALETTE PROFESSIONNELLE POUR PIE CHARTS
# ============================================================
PIE_COLORS = [
    "#1e3a5f","#2b6cb0","#3182ce","#4299e1","#63b3ed",
    "#276749","#38a169","#48bb78","#68d391","#9ae6b4",
    "#805ad5","#9f7aea","#b794f4","#d6bcfa","#e9d8fd",
    "#c53030","#e53e3e","#fc8181","#feb2b2","#fed7d7",
    "#d69e2e","#ecc94b","#f6e05e","#faf089","#fefcbf",
    "#744210","#975a16","#b7791f","#d69e2e","#ed8936",
    "#2d3748","#4a5568","#718096","#a0aec0","#cbd5e0"
]

# ============================================================
# CACHE SYSTEM - basé sur date.txt
# ============================================================
CACHE_FILE = ".dashboard_cache.pkl"

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def build_cache_key(fichier_date, sp, sa, sd, dr):
    """Construit une clé de cache unique basée sur date.txt + filtres"""
    raw = json.dumps({
        "date": fichier_date,
        "sp": sorted(sp),
        "sa": sorted(sa),
        "sd": sorted(sd),
        "dr": [str(dr[0]), str(dr[1])] if len(dr)==2 else []
    }, sort_keys=True)
    return hashlib.md5(raw.encode()).hexdigest()

def save_cache(key, data):
    """Sauvegarde le cache en JSON"""
    try:
        cache = {}
        if os.path.exists(CACHE_FILE.replace(".pkl",".json")):
            with open(CACHE_FILE.replace(".pkl",".json"),"r") as f:
                cache = json.load(f)
        # Convertir les DataFrames en dictionnaires pour le JSON
        serializable = {}
        for k, v in data.items():
            if isinstance(v, pd.DataFrame):
                serializable[k] = {"_type": "df", "data": v.to_dict(orient="split")}
            elif isinstance(v, dict):
                serializable[k] = {"_type": "dict", "data": v}
            elif isinstance(v, (int, float, str, bool, type(None))):
                serializable[k] = {"_type": "val", "data": v}
            elif isinstance(v, list):
                serializable[k] = {"_type": "list", "data": v}
            else:
                serializable[k] = {"_type": "val", "data": str(v)}
        cache[key] = serializable
        with open(CACHE_FILE.replace(".pkl",".json"),"w") as f:
            json.dump(cache, f)
    except Exception:
        pass

def load_cache(key):
    """Charge le cache depuis JSON"""
    try:
        if not os.path.exists(CACHE_FILE.replace(".pkl",".json")):
            return None
        with open(CACHE_FILE.replace(".pkl",".json"),"r") as f:
            cache = json.load(f)
        if key not in cache:
            return None
        raw = cache[key]
        result = {}
        for k, v in raw.items():
            t = v.get("_type","val")
            d = v.get("data")
            if t == "df":
                result[k] = pd.DataFrame(**d)
                if "index" in d and d["index"]:
                    result[k].index = d["index"]
                    if "columns" in d:
                        result[k] = result[k][d["columns"]]
            elif t == "dict":
                # Convertir les clés numériques
                result[k] = {}
                for dk, dv in d.items():
                    try: dk_conv = float(dk)
                    except: dk_conv = dk
                    result[k][dk_conv] = dv
            elif t == "list":
                result[k] = d
            else:
                result[k] = d
        return result
    except Exception:
        return None

# ============================================================
def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    j["Sens"]=j.apply(lambda r:"Amelioration" if ((r["Tendance"]=="hausse" and r["KPI"] not in LOWER_BETTER) or (r["Tendance"]=="baisse" and r["KPI"] in LOWER_BETTER)) else "Degradation",axis=1)
    return j.sort_values(["Date actuelle","Sens","Ecart %"],ascending=[True,False,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def get_caract_type(statut_user,keywords):
    s=str(statut_user).upper(); matched=[kw for kw in keywords if kw in s]
    return max(matched,key=len) if matched else "AUTRE"

# ============================================================
# PIE CHART PROFESSIONNEL - Style référence avec % + Nombre
# ============================================================
def create_professional_pie(labels, values, title="", colors=None, hole=0.45,
                             pull_small=0.12, small_threshold=5,
                             show_center_text=True, center_text="",
                             height=480, font_size_label=12, font_size_pct=13):
    """
    Crée un pie chart professionnel style référence :
    - Donut avec texte central
    - Secteurs petits tirés vers l'extérieur (pull)
    - Labels avec % ET nombre
    - Leader lines pour les petits secteurs
    - Palette professionnelle
    """
    total = sum(values)
    if total == 0:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", xref="paper", yref="paper",
                          x=0.5, y=0.5, showarrow=False, font=dict(size=18, color="#718096"))
        fig.update_layout(height=height, margin=dict(t=40,b=10,l=10,r=10))
        return fig

    n = len(labels)
    if colors is None:
        colors = [PIE_COLORS[i % len(PIE_COLORS)] for i in range(n)]

    # Calculer le pull pour chaque secteur (tirer les petits)
    pulls = []
    for v in values:
        pct = (v / total * 100) if total > 0 else 0
        if 0 < pct < small_threshold:
            pulls.append(pull_small)
        else:
            pulls.append(0)

    # Préparer les textes : "Label\nXX.X% (N)"
    text_labels = []
    for i, (lab, val) in enumerate(zip(labels, values)):
        pct = (val / total * 100) if total > 0 else 0
        if pct < 1 and val > 0:
            text_labels.append(f"{lab}<br>{pct:.1f}%<br>({int(val)})")
        elif pct >= 1:
            text_labels.append(f"{lab}<br>{pct:.1f}%<br>({int(val)})")
        else:
            text_labels.append("")

    # Position des textes : outside pour les petits, inside pour les grands
    text_positions = []
    for v in values:
        pct = (v / total * 100) if total > 0 else 0
        if pct < small_threshold:
            text_positions.append("outside")
        else:
            text_positions.append("inside")

    fig = go.Figure(go.Pie(
        labels=labels,
        values=values,
        hole=hole,
        pull=pulls,
        marker=dict(colors=colors,
                    line=dict(color='white', width=2.5)),
        text=text_labels,
        textposition=text_positions,
        textfont=dict(size=font_size_label, color="#1a202c", family="Inter, sans-serif"),
        hovertemplate='<b>%{label}</b><br>'
                      'Nombre: <b>%{value}</b><br>'
                      'Pourcentage: <b>%{percent}</b><br>'
                      'Total: <b>%{total}</b>'
                      '<extra></extra>',
        sort=False,
        direction='clockwise',
        rotation=0,
    ))

    # Texte central dans le trou du donut
    if show_center_text and hole > 0:
        if not center_text:
            center_text = f"Total<br><b>{int(total)}</b>"
        fig.add_annotation(
            text=center_text,
            x=0.5, y=0.5,
            xref="paper", yref="paper",
            showarrow=False,
            font=dict(size=20, color="#1e3a5f", family="Inter, sans-serif", weight="bold"),
            align="center"
        )

    fig.update_layout(
        title=dict(
            text=title,
            font=dict(size=16, color="#1e3a5f", family="Inter, sans-serif", weight="bold"),
            x=0.5, xanchor="center",
            y=0.97, yanchor="top",
            pad=dict(t=5, b=5)
        ),
        height=height,
        margin=dict(t=50, b=20, l=30, r=30),
        showlegend=True,
        legend=dict(
            font=dict(size=11, color="#4a5568", family="Inter, sans-serif"),
            orientation="h",
            yanchor="bottom", y=-0.08,
            xanchor="center", x=0.5,
            itemwidth=30,
            itemsel="none"
        ),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family="Inter, sans-serif"),
    )

    return fig


def create_status_pie_chart(df, status_col, title="", colors_map=None, height=480):
    """
    Crée un pie chart pour la répartition par statut avec % + nombre.
    Gère automatiquement les petits secteurs.
    """
    if df.empty or status_col not in df.columns:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", xref="paper", yref="paper",
                          x=0.5, y=0.5, showarrow=False, font=dict(size=18, color="#718096"))
        fig.update_layout(height=height)
        return fig

    counts = df[status_col].value_counts()
    labels = counts.index.tolist()
    values = counts.values.tolist()

    # Palette par défaut pour les statuts
    default_colors = {
        "CLOT": "#276749", "TCLO": "#38a169", "CRÉÉ": "#2b6cb0",
        "LANC": "#d69e2e", "ENCO": "#805ad5", "LIBE": "#4299e1",
        "CARACTERISE": "#276749", "NON CARACTERISE": "#c53030",
        "OUI": "#276749", "NON": "#c53030",
        "APRV": "#276749", "APRQ": "#2b6cb0", "REJT": "#c53030",
        "<1 mois": "#276749", ">3 mois": "#c53030", "1 mois < <3 mois": "#d69e2e",
        "APRV AVAU": "#38a169",
    }

    colors = []
    for lab in labels:
        lab_s = str(lab).strip()
        if colors_map and lab_s in colors_map:
            colors.append(colors_map[lab_s])
        elif lab_s in default_colors:
            colors.append(default_colors[lab_s])
        else:
            colors.append(PIE_COLORS[len(colors) % len(PIE_COLORS)])

    total = sum(values)
    center = f"Total<br><b>{int(total)}</b>"

    return create_professional_pie(
        labels=labels, values=values, title=title,
        colors=colors, hole=0.42, pull_small=0.15,
        small_threshold=6, show_center_text=True,
        center_text=center, height=height,
        font_size_label=11, font_size_pct=12
    )


def create_age_pie_chart(df, age_col, title="", height=480):
    """Pie chart spécialisé pour la répartition par âge des OT"""
    age_colors = {"<1 mois": "#276749", "1 mois < <3 mois": "#d69e2e", ">3 mois": "#c53030", "Inconnu": "#a0aec0"}
    return create_status_pie_chart(df, age_col, title=title, colors_map=age_colors, height=height)


def create_kpi_pie_by_poste(ckdf, kpi_name, title="", height=500):
    """
    Pie chart montrant la répartition d'un KPI par poste.
    Affiche les postes qui ont des valeurs != 0, avec % et nombre.
    """
    if kpi_name not in ckdf.columns:
        fig = go.Figure()
        fig.add_annotation(text="KPI non disponible", xref="paper", yref="paper",
                          x=0.5, y=0.5, showarrow=False, font=dict(size=18, color="#718096"))
        fig.update_layout(height=height)
        return fig

    vals = ckdf[kpi_name].dropna()
    vals = vals[vals != 0]

    if vals.empty:
        fig = go.Figure()
        fig.add_annotation(text="Aucune valeur non-nulle", xref="paper", yref="paper",
                          x=0.5, y=0.5, showarrow=False, font=dict(size=18, color="#718096"))
        fig.update_layout(height=height)
        return fig

    # Pour les KPI en %, on montre la contribution de chaque poste
    total = vals.sum()
    if total == 0:
        fig = go.Figure()
        fig.add_annotation(text="Total = 0", xref="paper", yref="paper",
                          x=0.5, y=0.5, showarrow=False, font=dict(size=18, color="#718096"))
        fig.update_layout(height=height)
        return fig

    labels = [str(idx) for idx in vals.index]
    values = vals.values.tolist()
    n = len(labels)
    colors = [PIE_COLORS[i % len(PIE_COLORS)] for i in range(n)]

    return create_professional_pie(
        labels=labels, values=values, title=title,
        colors=colors, hole=0.40, pull_small=0.18,
        small_threshold=4, show_center_text=True,
        center_text=f"{kpi_name[:25]}<br>Moy: <b>{total/max(n,1):.1f}%</b>",
        height=height, font_size_label=10, font_size_pct=11
    )


# ============================================================
def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--p);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:6px 10px;background:#fff;border-radius:5px;margin-bottom:2px;border:1px solid var(--b);font-size:13px}
    .sr .sn{font-weight:700;color:var(--p);min-width:220px;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:3px 9px;border-radius:12px;font-weight:800;font-size:14px;min-width:50px;text-align:center;margin:0 8px;color:#fff}
    .sr .sa{color:#718096;font-size:12px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .stg{font-size:11px;color:#718096;min-width:60px;text-align:center;white-space:nowrap}
    .sr .sb{font-size:11px;font-weight:700;padding:2px 8px;border-radius:3px;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:4px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:24px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}
    .gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .anl-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;margin:0}
    .anl-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:12px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .anl-tbl tbody td{padding:5px 8px;border-bottom:1px solid #edf2f7}
    .anl-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .anl-tbl tbody tr:hover td{background:#ebf8ff!important}
    .anl-tbl .tot td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .trend-up{color:#276749;font-weight:800;font-size:16px}
    .trend-down{color:#c53030;font-weight:800;font-size:16px}
    .trend-stable{color:#718096;font-weight:800;font-size:16px}
    .spark-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:8px}
    .spark-card{background:#fff;border-radius:var(--r);padding:10px 12px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .spark-card .sp-title{font-size:13px;font-weight:800;color:var(--p);margin-bottom:3px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .spark-card .sp-sub{font-size:11px;color:#718096;margin-bottom:5px}
    .rank-card{background:#fff;border-radius:var(--r);padding:12px 16px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .rank-card .rank-title{font-size:15px;font-weight:800;margin-bottom:8px;padding-bottom:5px;border-bottom:2px solid var(--b)}
    .rank-row{display:flex;align-items:center;padding:5px 0;font-size:13px;border-bottom:1px solid #f7fafc}
    .rank-row:last-child{border:none}
    .rank-row .rank-num{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:13px;color:#fff;margin-right:10px;flex-shrink:0}
    .rank-row .rank-name{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .rank-row .rank-score{font-weight:900;min-width:70px;text-align:right}
    .pie-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(420px,1fr));gap:8px}
    .pie-container{background:#fff;border-radius:var(--r);padding:12px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:120px}.gbr-l{width:100px}.spark-grid{grid-template-columns:1fr}.pie-grid{grid-template-columns:1fr}}
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def contient_mot(t,lm):
        t=str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if a<=1: return "<1 mois"
        elif a>=3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
        return df
    def get_metier(p):
        p=str(p).upper()
        if "E" in p: return "Electrique"
        if "M" in p: return "Mecanique"
        if "R" in p: return "Instrumentation"
        if "G" in p: return "Genie Civil"
        return "Autre"
    def get_atelier(p):
        p=str(p).upper()
        if "PS" in p: return "Sulfurique"
        if "PP" in p: return "Phosphorique"
        if "TSP" in p or "REX" in p: return "Engrais"
        if "MCP" in p or "DCP" in p: return "Feed"
        return "Autre"
    def get_division(p):
        p=str(p).upper()
        if "SF1" in p: return "SF1"
        if "SF2" in p: return "SF2"
        return "Autre"

    def calc_kpis(df_i,av_i,now,posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
        df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
        for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc]=pd.to_datetime(df[dc],errors='coerce')
                df[am]=((now.year-df[dc].dt.year)*12+(now.month-df[dc].dt.month)).round(2)
                df[ac]=df[am].apply(cat_age)
            else: df[am]=np.nan; df[ac]="Inconnu"
        df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
        df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
        df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
        df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
        res['dfp']=df
        an=cpiv(df,df["Nº appel pl.entret."].fillna(0)==0,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["Total"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1); an["TAUX_REALISATION_CORRECTIF/PT"]=ckpi(an["TCLO"],an["Total"])
        pr=cpiv(df,df["Statut OT"]=="CRÉÉ","ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
        avf=av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip()=="")].copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "color:#cbd5e0"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        return 0
    def is_lb(k): return k in LOWER_BETTER

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                else: s=cs(v) if sc_col and c in sc_col else ks(v,c); h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
    def html_ano(rows,cols):
        h='<table class="tw at"><thead><tr>'+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            h+='<tr class="%s">'%("tr" if r.get("_t")=="total" else "")
            for c in cols: v=r.get(c,""); h+='<td style="%s">%s</td>'%(kas(v) or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
    def html_actions_table(kpi_list,actuals,targets,act_map):
        h='<table class="tw at"><thead><tr><th>KPI</th><th>Valeur Actuelle</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action Recommandee</th></tr></thead><tbody>'
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            status="ATTEINT" if met else "NON ATTEINT"
            st_s="background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
            ec_clr="#276749" if met else "#c53030"
            action="Objectif atteint" if met else act_map.get(k,"")
            h+='<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td></tr>'%(k,av,tv,ec_clr,diff,st_s,status,action)
        return h+'</tbody></table>'
    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 - Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">Bottom 5 - Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'; return h
    def html_kpi_bars(kpi_list,actuals,targets,title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,title)
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); met=av<=tv if is_lb(k) else av>=tv
            bw=min(max(av,0),100); bg=color_ok if met else color_fail
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(k,bw,bg,av)
        return h+'</div>'
    def html_bars(data,title,color):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color,title)
        for label,val in sorted(data,key=lambda x:x[1],reverse=True):
            bw=min(max(val,0),100)
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(label,bw,color,val)
        return h+'</div>'
    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        for p in sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True):
            pv,qv=pscores.get(p,0),qscores.get(p,0)
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>'%(p,min(max(pv,0),100),pv,min(max(qv,0),100),qv)
        return h+'</div>'
    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",unsafe_allow_html=True)
        st.markdown("---")
        show_filters=st.checkbox("Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf")
            ot_f=av_f=None; apm=[]
            if unf:
                ot_f=st.file_uploader("Fichier OT",type=["xlsx"],key="uot")
                av_f=st.file_uploader("Fichier AVIS",type=["xlsx"],key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t=excr(pd.read_excel("ot.xlsx"))
                        apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            st.markdown("**🏭 Atelier**")
            sa=st.multiselect("Atelier",["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"],["All"],key="sa")
            st.markdown("**🏢 Division**")
            sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=av_f=None; apm=[]; sp=["All"]; sa=["All"]; sd=["All"]
            dr=(datetime(2025,1,1).date(),datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try:
                    _t=excr(pd.read_excel("ot.xlsx"))
                    apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception: pass

    # ===================== DATA LOADING AVEC CACHE =====================
    if not unf or (ot_f is not None and av_f is not None):
        # Construire la clé de cache
        cache_key = None
        if not unf:
            # Mode fichiers locaux : clé basée sur date.txt + filtres
            cache_key = build_cache_key(fichier_date, sp, sa, sd, dr)

        # Tentative de chargement du cache (uniquement en mode fichiers locaux)
        cached_data = None
        if cache_key:
            cached_data = load_cache(cache_key)
            if cached_data is not None:
                # Restaurer toutes les variables depuis le cache
                ckdf = cached_data.get('ckdf')
                dfp = cached_data.get('dfp')
                avf = cached_data.get('avf')
                pa = cached_data.get('pa', {})
                qa = cached_data.get('qa', {})
                pa_d = cached_data.get('pa_d', {})
                qa_d = cached_data.get('qa_d', {})
                pscores = cached_data.get('pscores', {})
                qscores = cached_data.get('qscores', {})
                pscores_d = cached_data.get('pscores_d', {})
                qscores_d = cached_data.get('qscores_d', {})
                vp = cached_data.get('vp', [])
                df_dash = cached_data.get('df_dash')
                all_ano = cached_data.get('all_ano', [])
                ano_data = cached_data.get('ano_data', {})
                raw_ot = cached_data.get('raw_ot')
                raw_av = cached_data.get('raw_av')
                _cache_hit = True
            else:
                _cache_hit = False
        else:
            _cache_hit = False

        if not _cache_hit:
            # ============ CALCUL COMPLET (pas de cache ou fichiers uploadés) ============
            try:
                if unf: raw_ot=pd.read_excel(ot_f); raw_av=pd.read_excel(av_f)
                else: raw_ot=pd.read_excel("ot.xlsx"); raw_av=pd.read_excel("avis.xlsx")
                raw_ot=excr(raw_ot); raw_av=excr(raw_av)
                for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                    if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
                for c in ["Créé le","Début souhaité","Date de la clôture"]:
                    if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
                if not apm: apm=sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                if "All" in sp or not sp: sp=apm
                if "All" in sa or not sa: sa=["All"]
                if "All" in sd or not sd: sd=["All"]
                sdt=pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
                edt=pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

                def mf(poste):
                    p=str(poste).upper()
                    if "All" not in sa:
                        m=False
                        if "Sulfurique (PS)" in sa and "PS" in p: m=True
                        if "Phosphorique (PP)" in sa and "PP" in p: m=True
                        if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m=True
                        if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m=True
                        if not m: return False
                    if "All" not in sd:
                        m=False
                        if "SF1" in sd and "SF1" in p: m=True
                        if "SF2" in sd and "SF2" in p: m=True
                        if not m: return False
                    return True

                vp=[p for p in apm if mf(p) and p in sp]
                df=raw_ot[(raw_ot["Poste travail princ."].isin(vp))&(raw_ot["Date de début planifiée"].between(sdt,edt))].copy()
                avdf=raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
                df=excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
                avdf=excr(avdf[(avdf["Ordre"].isna())|(avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
                if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
                df_dash=raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
                df_dash=excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
                if "Statut système" in df_dash.columns: df_dash["Statut OT"]=df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

                now=pd.Timestamp.now()
                res=calc_kpis(df,avdf,now,vp); ckdf=res['ckdf']; dfp=res['dfp']; avf=res['avf']
                res_d=calc_kpis(df_dash,avdf,now,vp); ckdf_d=res_d['ckdf']
                pa={k:round(ckdf[k].mean(),2) for k in QK}; qa={k:round(ckdf[k].mean(),2) for k in PK}
                pa_d={k:round(ckdf_d[k].mean(),2) for k in QK}; qa_d={k:round(ckdf_d[k].mean(),2) for k in PK}
                pscores={}; qscores={}
                for poste in ckdf.index:
                    r=ckdf.loc[poste]
                    pscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                    qscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0
                pscores_d={}; qscores_d={}
                for poste in ckdf_d.index:
                    r=ckdf_d.loc[poste]
                    pscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                    qscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0

                all_ano=[]
                sub_p={"TAUX_REALISATION_CORRECTIF/PT":lambda d:d[(d["Nº appel pl.entret."].fillna(0)==0)&(~d["Statut OT"].isin(["CLOT","TCLO"]))],"OT préparation <1 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]!="<1 mois")],"OT préparation >3 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]==">3 mois")],"OT planification <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]!="<1 mois")],"OT planification >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]==">3 mois")],"OT exécution <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]!="<1 mois")],"OT exécution >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]==">3 mois")],"OT préparation 1mois< <3mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]=="1 mois < <3 mois")],"OT planification 1mois< <3mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]=="1 mois < <3 mois")],"OT exécution 1mois< <3mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]=="1 mois < <3 mois")]}
                sub_q={"OT LANC ESTIME":lambda d:d[(d["Statut OT"]=="LANC")&(d["OT LANC ESTIME"]=="NON")],"Backlog préparation caractérisé":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["Backlog preparation"]=="NON CARACTERISE")],"Backlog planification caractérisé":lambda d:d[(d["Statut OT"]=="LANC")&(d["Backlog planification"]=="NON CARACTERISE")],"OT COR Egal":lambda d:d[(d["OT COR EGAL"]=="NON")],"OT CONFIME":lambda d:d[(d["OT CONFIME"]=="NON")&(d["Statut OT"].isin(["CLOT","TCLO"]))],"appel avis approuvé":lambda d:d[(d["Statut utilisateur"].isin(["APRQ","REJT"]))]}

                ano_data={}
                for kn,fn in sub_p.items():
                    try:
                        sd2=fn(dfp); cnt=len(sd2)
                        if cnt>0:
                            grp=sd2.groupby("Poste travail princ.")["Ordre"].count().to_dict()
                            all_ano.extend([{"KPI":kn,"Poste":p,"Nb anomalies":n} for p,n in grp.items() if n>0])
                            ano_data[kn]=cnt
                    except Exception: pass
                for kn,fn in sub_q.items():
                    try:
                        sd2=fn(dfp); cnt=len(sd2)
                        if cnt>0:
                            grp=sd2.groupby("Poste travail princ.")["Ordre"].count().to_dict()
                            all_ano.extend([{"KPI":kn,"Poste":p,"Nb anomalies":n} for p,n in grp.items() if n>0])
                            ano_data[kn]=cnt
                    except Exception: pass

                # Sauvegarder dans le cache si mode fichiers locaux
                if cache_key:
                    cache_data = {
                        'ckdf': ckdf, 'dfp': dfp, 'avf': avf,
                        'pa': pa, 'qa': qa, 'pa_d': pa_d, 'qa_d': qa_d,
                        'pscores': pscores, 'qscores': qscores,
                        'pscores_d': pscores_d, 'qscores_d': qscores_d,
                        'vp': vp, 'df_dash': df_dash, 'all_ano': all_ano,
                        'ano_data': ano_data,
                        'raw_ot': raw_ot.head(0), 'raw_av': raw_av.head(0)
                    }
                    save_cache(cache_key, cache_data)

            except Exception as e:
                st.error(f"Erreur de chargement: {str(e)}")
                st.stop()
        else:
            # Cache hit - afficher indicateur de performance
            pass

        # ===================== DASHBOARD DISPLAY =====================
        p_score=round(np.mean(list(pscores.values())),2) if pscores else 0
        q_score=round(np.mean(list(qscores.values())),2) if qscores else 0
        total_ot=len(dfp)
        total_anom=len(all_ano)
        p_score_d=round(np.mean(list(pscores_d.values())),2) if pscores_d else 0
        q_score_d=round(np.mean(list(qscores_d.values())),2) if qscores_d else 0
        total_ot_d=len(df_dash) if df_dash is not None else 0

        st.markdown('<div class="mh"><h1>📊 DASHBOARD KPI - SUIVI MAINTENANCE</h1><span class="db">📅 %s</span></div>'%fichier_date,unsafe_allow_html=True)
        st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%d</div><div class="cl">OT (Periode)</div></div><div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div><div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div><div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies</div></div></div>'%(total_ot,p_score,q_score,total_anom),unsafe_allow_html=True)
        st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%d</div><div class="cl">OT (Total)</div></div><div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Perf. (Total)</div></div><div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Qual. (Total)</div></div><div class="cc c4"><div class="cv">%d</div><div class="cl">Postes</div></div></div>'%(total_ot_d,p_score_d,q_score_d,len(vp)),unsafe_allow_html=True)

        # ===================== TABS =====================
        tab1,tab2,tab3,tab4,tab5,tab6=st.tabs(["📋 Performance","🎯 Qualite","⚠️ Anomalies","📊 Graphiques","📈 Tendances","💾 Export"])

        with tab1:
            st.markdown('<div class="stl p">INDICATEURS DE PERFORMANCE PAR POSTE</div>',unsafe_allow_html=True)
            pcols=["Poste de travail"]+QK+["Score Performance"]
            prows=[]
            for poste in ckdf.index:
                r=ckdf.loc[poste]; row={"Poste de travail":poste,"_t":""}
                for k in QK: row[k]="%.1f"%r[k] if k in r.index and not pd.isna(r[k]) else "N/A"
                row["Score Performance"]="%.2f"%pscores.get(poste,0); prows.append(row)
            cible_row={"Poste de travail":"CIBLE","_t":"cible"}
            for k in QK: cible_row[k]=CIBLE.get(k,"")
            cible_row["Score Performance"]="100.00"; prows.insert(0,cible_row)
            tot_row={"Poste de travail":"Moyenne","_t":"total"}
            for k in QK: tot_row[k]="%.1f"%pa.get(k,0)
            tot_row["Score Performance"]="%.2f"%p_score; prows.append(tot_row)
            st.markdown(html_table(prows,pcols,"pt",sc_col={"Score Performance"}),unsafe_allow_html=True)
            st.markdown('<div class="stl p" style="margin-top:8px">BAREMES PERFORMANCE</div>',unsafe_allow_html=True)
            st.markdown(html_kpi_bars(QK,pa,CIBLE,"Performance Globale","#38a169","#e53e3e"),unsafe_allow_html=True)
            st.markdown('<div class="stl c" style="margin-top:8px">CLASSEMENT POSTES - PERFORMANCE</div>',unsafe_allow_html=True)
            st.markdown(html_classement(pscores,"#276749"),unsafe_allow_html=True)

        with tab2:
            st.markdown('<div class="stl q">INDICATEURS DE QUALITE PAR POSTE</div>',unsafe_allow_html=True)
            qcols=["Poste de travail"]+PK+["Score Qualite"]
            qrows=[]
            for poste in ckdf.index:
                r=ckdf.loc[poste]; row={"Poste de travail":poste,"_t":""}
                for k in PK: row[k]="%.1f"%r[k] if k in r.index and not pd.isna(r[k]) else "N/A"
                row["Score Qualite"]="%.2f"%qscores.get(poste,0); qrows.append(row)
            cible_row2={"Poste de travail":"CIBLE","_t":"cible"}
            for k in PK: cible_row2[k]=CIBLE.get(k,"")
            cible_row2["Score Qualite"]="100.00"; qrows.insert(0,cible_row2)
            tot_row2={"Poste de travail":"Moyenne","_t":"total"}
            for k in PK: tot_row2[k]="%.1f"%qa.get(k,0)
            tot_row2["Score Qualite"]="%.2f"%q_score; qrows.append(tot_row2)
            st.markdown(html_table(qrows,qcols,"qt",sc_col={"Score Qualite"}),unsafe_allow_html=True)
            st.markdown('<div class="stl q" style="margin-top:8px">BAREMES QUALITE</div>',unsafe_allow_html=True)
            st.markdown(html_kpi_bars(PK,qa,CIBLE,"Qualite Globale","#3182ce","#e53e3e"),unsafe_allow_html=True)
            st.markdown('<div class="stl c" style="margin-top:8px">CLASSEMENT POSTES - QUALITE</div>',unsafe_allow_html=True)
            st.markdown(html_classement(qscores,"#2b6cb0"),unsafe_allow_html=True)

        with tab3:
            st.markdown('<div class="stl a">ANOMALIES DETECTEES</div>',unsafe_allow_html=True)
            if all_ano:
                ano_df=pd.DataFrame(all_ano)
                ano_grp=ano_df.groupby("KPI")["Nb anomalies"].sum().sort_values(ascending=False).reset_index()
                ano_grp.columns=["KPI","Total"]
                ano_pivot=ano_df.pivot_table(index="Poste",columns="KPI",values="Nb anomalies",aggfunc="sum",fill_value=0)
                ano_pivot["Total"]=ano_pivot.sum(axis=1)
                ano_pivot=ano_pivot.sort_values("Total",ascending=False)
                acols=["KPI"]+ano_pivot.columns.tolist()
                arows=[{"KPI":k,"_t":"total","Total":int(ano_grp[ano_grp["KPI"]==k]["Total"].values[0]) if len(ano_grp[ano_grp["KPI"]==k])>0 else 0} for k in ano_pivot.columns if k!="Total"]
                arows=[{"KPI":"Total","_t":"total","Total":int(ano_grp["Total"].sum())}]+arows
                for poste in ano_pivot.index:
                    row={"KPI":poste,"_t":""}; 
                    for c in ano_pivot.columns: row[c]=int(ano_pivot.loc[poste,c])
                    arows.append(row)
                st.markdown(html_ano(arows,acols),unsafe_allow_html=True)
                st.markdown('<div class="stl a" style="margin-top:8px">ACTIONS CORRECTIVES RECOMMANDEES</div>',unsafe_allow_html=True)
                all_kpi_actuals={**pa,**qa}
                st.markdown(html_actions_table(list(ano_data.keys()),all_kpi_actuals,CIBLE,ACT_MAP),unsafe_allow_html=True)
            else:
                st.markdown('<div class="es">✅ Aucune anomalie detectee</div>',unsafe_allow_html=True)

        with tab4:
            # ===================== GRAPHIQUES AVEC PIE CHARTS PROFESSIONNELS =====================
            st.markdown('<div class="stl c">REPARTITION PAR STATUT OT - PIE CHARTS</div>',unsafe_allow_html=True)

            # Grille de pie charts
            col_p1, col_p2 = st.columns(2)

            with col_p1:
                # Pie 1: Répartition globale par statut OT
                fig_statut = create_status_pie_chart(dfp, "Statut OT",
                    title="Répartition par Statut OT", height=460)
                st.plotly_chart(fig_statut, use_container_width=True, config={"displayModeBar": False})

                # Pie 2: Répartition par âge préparation
                if "ap" in dfp.columns:
                    crees = dfp[dfp["Statut OT"]=="CRÉÉ"]
                    fig_age_prep = create_age_pie_chart(crees, "ap",
                        title="Age Preparation (OT Créés)", height=460)
                    st.plotly_chart(fig_age_prep, use_container_width=True, config={"displayModeBar": False})

                # Pie 3: Backlog préparation caractérisé
                if "Backlog preparation" in dfp.columns:
                    crees2 = dfp[dfp["Statut OT"]=="CRÉÉ"]
                    fig_bp = create_status_pie_chart(crees2, "Backlog preparation",
                        title="Backlog Préparation Caractérisé", height=460)
                    st.plotly_chart(fig_bp, use_container_width=True, config={"displayModeBar": False})

            with col_p2:
                # Pie 4: Répartition par âge planification
                if "alp" in dfp.columns:
                    lanc_sopl0 = dfp[(dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==0)]
                    fig_age_plan = create_age_pie_chart(lanc_sopl0, "alp",
                        title="Age Planification (OT Lancés)", height=460)
                    st.plotly_chart(fig_age_plan, use_container_width=True, config={"displayModeBar": False})

                # Pie 5: Répartition par âge exécution
                if "aex" in dfp.columns:
                    lanc_sopl1 = dfp[(dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==1)]
                    fig_age_exec = create_age_pie_chart(lanc_sopl1, "aex",
                        title="Age Exécution (OT en Cours)", height=460)
                    st.plotly_chart(fig_age_exec, use_container_width=True, config={"displayModeBar": False})

                # Pie 6: Backlog planification caractérisé
                if "Backlog planification" in dfp.columns:
                    lanc2 = dfp[dfp["Statut OT"]=="LANC"]
                    fig_bpl = create_status_pie_chart(lanc2, "Backlog planification",
                        title="Backlog Planification Caractérisé", height=460)
                    st.plotly_chart(fig_bpl, use_container_width=True, config={"displayModeBar": False})

            # Deuxième rangée : Pie charts par poste pour les KPIs clés
            st.markdown('<div class="stl s" style="margin-top:10px">REPARTITION PAR POSTE - KPIs CLES</div>',unsafe_allow_html=True)

            kpi_pie_selection = st.selectbox(
                "Sélectionner un KPI pour voir la répartition par poste",
                ALL_KPI, index=0, key="kpi_pie_sel"
            )

            col_kp1, col_kp2 = st.columns(2)
            with col_kp1:
                fig_kpi_poste = create_kpi_pie_by_poste(ckdf, kpi_pie_selection,
                    title=f"Répartition: {kpi_pie_selection}", height=480)
                st.plotly_chart(fig_kpi_poste, use_container_width=True, config={"displayModeBar": False})

            with col_kp2:
                # Pie: OT Confirmés vs Non confirmés
                if "OT CONFIME" in dfp.columns:
                    fig_conf = create_status_pie_chart(dfp, "OT CONFIME",
                        title="OT Confirmés (CLOT+CONF)", height=480)
                    st.plotly_chart(fig_conf, use_container_width=True, config={"displayModeBar": False})

            # Troisième rangée : Distribution par atelier et division
            st.markdown('<div class="stl p" style="margin-top:10px">REPARTITION PAR ATELIER & DIVISION</div>',unsafe_allow_html=True)

            dfp_copy = dfp.copy()
            dfp_copy["Atelier"] = dfp_copy["Poste travail princ."].apply(get_atelier)
            dfp_copy["Division"] = dfp_copy["Poste travail princ."].apply(get_division)
            dfp_copy["Metier"] = dfp_copy["Poste travail princ."].apply(get_metier)

            col_a1, col_a2, col_a3 = st.columns(3)
            with col_a1:
                fig_atel = create_professional_pie(
                    labels=dfp_copy["Atelier"].value_counts().index.tolist(),
                    values=dfp_copy["Atelier"].value_counts().values.tolist(),
                    title="Répartition par Atelier",
                    colors=["#276749","#2b6cb0","#d69e2e","#805ad5","#a0aec0"],
                    hole=0.40, height=420
                )
                st.plotly_chart(fig_atel, use_container_width=True, config={"displayModeBar": False})

            with col_a2:
                fig_div = create_professional_pie(
                    labels=dfp_copy["Division"].value_counts().index.tolist(),
                    values=dfp_copy["Division"].value_counts().values.tolist(),
                    title="Répartition par Division",
                    colors=["#1e3a5f","#4299e1","#a0aec0"],
                    hole=0.40, height=420
                )
                st.plotly_chart(fig_div, use_container_width=True, config={"displayModeBar": False})

            with col_a3:
                fig_met = create_professional_pie(
                    labels=dfp_copy["Metier"].value_counts().index.tolist(),
                    values=dfp_copy["Metier"].value_counts().values.tolist(),
                    title="Répartition par Metier",
                    colors=["#e53e3e","#2b6cb0","#805ad5","#d69e2e","#38a169"],
                    hole=0.40, height=420
                )
                st.plotly_chart(fig_met, use_container_width=True, config={"displayModeBar": False})

            # Barres groupées
            st.markdown('<div class="stl p" style="margin-top:10px">SCORES PAR POSTE</div>',unsafe_allow_html=True)
            st.markdown(html_grouped_bars(vp,pscores,qscores,"Comparaison Performance vs Qualite par Poste"),unsafe_allow_html=True)

            # OT LANC ESTIME
            if "OT LANC ESTIME" in dfp.columns:
                lanc3 = dfp[dfp["Statut OT"]=="LANC"]
                col_e1, col_e2 = st.columns(2)
                with col_e1:
                    fig_est = create_status_pie_chart(lanc3, "OT LANC ESTIME",
                        title="OT Lancés Estimés", height=420)
                    st.plotly_chart(fig_est, use_container_width=True, config={"displayModeBar": False})
                with col_e2:
                    fig_cor = create_status_pie_chart(dfp, "OT_COR_EGAL",
                        title="OT Coûts Réels = Budgétés", height=420)
                    st.plotly_chart(fig_cor, use_container_width=True, config={"displayModeBar": False})

            # Appels avis
            if avf is not None and not avf.empty:
                st.markdown('<div class="stl q" style="margin-top:10px">APPELS AVIS</div>',unsafe_allow_html=True)
                col_av1, col_av2 = st.columns(2)
                with col_av1:
                    fig_av = create_status_pie_chart(avf, "Statut utilisateur",
                        title="Statut Appels Avis", height=420)
                    st.plotly_chart(fig_av, use_container_width=True, config={"displayModeBar": False})
                with col_av2:
                    # Pie par poste pour les avis
                    av_poste = avf["Poste travail princ."].value_counts()
                    fig_avp = create_professional_pie(
                        labels=av_poste.index.tolist(),
                        values=av_poste.values.tolist(),
                        title="Appels Avis par Poste",
                        hole=0.40, height=420
                    )
                    st.plotly_chart(fig_avp, use_container_width=True, config={"displayModeBar": False})

        with tab5:
            st.markdown('<div class="stl s">TENDANCES & EVOLUTIONS</div>',unsafe_allow_html=True)
            hist_path=os.path.join("kpis","indicateurs_kpis.xlsx")
            if os.path.exists(hist_path):
                hist_df=load_historical_kpis(hist_path)
                var_df=calculate_variations(hist_df)
                if not var_df.empty:
                    jrn=generate_journal(var_df)
                    if not jrn.empty:
                        st.markdown('<div class="ca"><div class="ct">Journal des Variations Significatives (|e|>=5%)</div>',unsafe_allow_html=True)
                        for _,row in jrn.iterrows():
                            sens_clr="#276749" if row["Sens"]=="Amelioration" else "#c53030"
                            sens_icon="▲" if row["Sens"]=="Amelioration" else "▼"
                            st.markdown('<div class="sr"><span class="sn">%s - %s</span><span class="sc" style="background:%s">%s %.1f%%</span><span class="sa">%s: %.1f → %.1f</span><span class="stg">%s → %s</span></div>'%(row["Poste"],row["Type"],sens_clr,sens_icon,row["Ecart %"],row["KPI"],row["Valeur precedente"],row["Valeur actuelle"],row["Date precedente"],row["Date actuelle"]),unsafe_allow_html=True)
                        st.markdown('</div>',unsafe_allow_html=True)
                    top5,bot5=calculate_rankings(var_df)
                    if not top5.empty:
                        st.markdown('<div class="dgrid">',unsafe_allow_html=True)
                        st.markdown('<div class="rank-card"><div class="rank-title" style="color:#276749;border-bottom-color:#38a169">🏆 Top 5 Amelioration</div>',unsafe_allow_html=True)
                        for i,(_,row) in enumerate(top5.iterrows()):
                            st.markdown('<div class="rank-row"><span class="rank-num" style="background:#276749">%s</span><span class="rank-name">%s</span><span class="rank-score" style="color:#276749">%+.1f</span></div>'%(i+1,row["Poste"],row["Score variation"]),unsafe_allow_html=True)
                        st.markdown('</div>',unsafe_allow_html=True)
                        st.markdown('<div class="rank-card"><div class="rank-title" style="color:#c53030;border-bottom-color:#e53e3e">⚠️ Top 5 Degradation</div>',unsafe_allow_html=True)
                        for i,(_,row) in enumerate(bot5.iterrows()):
                            st.markdown('<div class="rank-row"><span class="rank-num" style="background:#c53030">%s</span><span class="rank-name">%s</span><span class="rank-score" style="color:#c53030">%+.1f</span></div>'%(i+1,row["Poste"],row["Score variation"]),unsafe_allow_html=True)
                        st.markdown('</div>',unsafe_allow_html=True)
                        st.markdown('</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">Pas assez de donnees historiques pour calculer les tendances (minimum 2 periodes requises)</div>',unsafe_allow_html=True)
            else:
                st.markdown('<div class="es">Aucun fichier historique trouve dans kpis/indicateurs_kpis.xlsx</div>',unsafe_allow_html=True)

        with tab6:
            st.markdown('<div class="stl s">EXPORT DES DONNEES</div>',unsafe_allow_html=True)
            col_e1,col_e2=st.columns(2)
            with col_e1:
                st.markdown("**Indicateurs de Performance**")
                pcols_exp=["Poste de travail"]+QK+["Score Performance"]
                pdf_exp=pd.DataFrame(prows)
                if not pdf_exp.empty: pdf_exp=pdf_exp[pcols_exp]
                export_btn(pdf_exp,"performance_kpis.xlsx")
                st.markdown("**Indicateurs de Qualite**")
                qcols_exp=["Poste de travail"]+PK+["Score Qualite"]
                qdf_exp=pd.DataFrame(qrows)
                if not qdf_exp.empty: qdf_exp=qdf_exp[qcols_exp]
                export_btn(qdf_exp,"qualite_kpis.xlsx")
            with col_e2:
                st.markdown("**Anomalies**")
                if all_ano:
                    export_btn(pd.DataFrame(all_ano),"anomalies.xlsx")
                else:
                    st.info("Aucune anomalie")
                st.markdown("**Sauvegarde KPIs historiques**")
                if st.button("💾 Sauvegarder dans Excel historique",key="save_hist"):
                    pcols_h=["Poste de travail"]+QK+["Score Performance"]
                    qcols_h=["Poste de travail"]+PK+["Score Qualite"]
                    pr_h=[{k:r[k] for k in pcols_h if k in r} for r in prows if r.get("_t")!="cible"]
                    qr_h=[{k:r[k] for k in qcols_h if k in r} for r in qrows if r.get("_t")!="cible"]
                    ano_p_h=[]; ano_q_h=[]
                    if all_ano:
                        adf=pd.DataFrame(all_ano)
                        for kn in QK:
                            sub=adf[adf["KPI"]==kn]
                            if not sub.empty:
                                ano_p_h.append({"KPI":kn,"Poste":p,"Nb":int(n)} for _,r in sub.iterrows() for p,n in r.items() if p!="KPI" and p!="Nb")
                        for kn in PK:
                            sub=adf[adf["KPI"]==kn]
                            if not sub.empty:
                                ano_q_h.append({"KPI":kn,"Poste":p,"Nb":int(n)} for _,r in sub.iterrows() for p,n in r.items() if p!="KPI" and p!="Nb")
                    save_kpis_to_excel(pr_h,pcols_h,qr_h,qcols_h,ano_p_h,["KPI","Poste","Nb"] if ano_p_h else [],ano_q_h,["KPI","Poste","Nb"] if ano_q_h else [],fichier_date)
                    st.success("✅ Sauvegarde effectuee avec succes!")
    else:
        st.markdown('<div class="es" style="margin-top:100px">📁 Veuillez charger les fichiers OT et AVIS depuis la barre laterale</div>',unsafe_allow_html=True)

if __name__=="__main__":
    main()
