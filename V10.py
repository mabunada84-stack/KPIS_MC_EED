# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
ALL_KPI = QK + PK
CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100}
ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois."}
LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]
MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

# ============================================================
def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    j["Sens"]=j.apply(lambda r:"Amelioration" if ((r["Tendance"]=="hausse" and r["KPI"] not in LOWER_BETTER) or (r["Tendance"]=="baisse" and r["KPI"] in LOWER_BETTER)) else "Degradation",axis=1)
    return j.sort_values(["Date actuelle","Sens","Ecart %"],ascending=[True,False,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

# ============================================================
def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:.5rem!important;padding-right:.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:8px 0 4px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}.stl.bl{border-left-color:#dd6b20}.stl.om{border-left-color:#b83280}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw.bl thead th{background:linear-gradient(135deg,#c05621,#dd6b20)}
    .tw.om thead th{background:linear-gradient(135deg,#97266d,#b83280)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:4px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:24px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}
    .gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-ok{background:linear-gradient(90deg,#276749,#48bb78)}.gb-warn{background:linear-gradient(90deg,#d69e2e,#f6e05e)}.gb-ko{background:linear-gradient(90deg,#c53030,#fc8181)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700;flex-wrap:wrap}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .rank-card{background:#fff;border-radius:var(--r);padding:12px 16px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .rank-card .rank-title{font-size:15px;font-weight:800;margin-bottom:8px;padding-bottom:5px;border-bottom:2px solid var(--b)}
    .rank-row{display:flex;align-items:center;padding:5px 0;font-size:13px;border-bottom:1px solid #f7fafc}
    .rank-row:last-child{border:none}
    .rank-row .rank-num{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:13px;color:#fff;margin-right:10px;flex-shrink:0}
    .rank-row .rank-name{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .rank-row .rank-score{font-weight:900;min-width:70px;text-align:right}
    .synth-box{background:linear-gradient(135deg,#f7fafc,#edf2f7);border-radius:var(--r);padding:14px 18px;margin-top:8px;border:1px solid #cbd5e0}
    .synth-box .synth-title{font-size:16px;font-weight:800;color:var(--p);margin-bottom:8px;display:flex;align-items:center;gap:8px}
    .synth-kpi-row{display:flex;align-items:center;padding:5px 0;border-bottom:1px solid #e2e8f0;font-size:12px;gap:8px}
    .synth-kpi-row:last-child{border:none}
    .synth-kpi-name{width:280px;font-weight:700;color:#2d3748;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .synth-kpi-val{min-width:65px;text-align:center;font-weight:800;font-size:13px;padding:2px 8px;border-radius:4px}
    .synth-kpi-target{min-width:55px;text-align:center;color:#718096;font-weight:600;font-size:11px}
    .synth-kpi-ecart{min-width:70px;text-align:center;font-weight:800;font-size:12px}
    .synth-kpi-badge{min-width:90px;text-align:center;padding:3px 8px;border-radius:12px;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px}
    .synth-kpi-action{flex:1;color:#4a5568;font-size:11px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .badge-ok{background:#c6efce;color:#006100}.badge-ko{background:#ffc7ce;color:#9c0006}.badge-warn{background:#ffeb9c;color:#9c6500}
    .synth-resume{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:10px}
    .synth-resume-card{background:#fff;border-radius:8px;padding:10px 14px;text-align:center;border:1px solid var(--b)}
    .synth-resume-card .src-val{font-size:28px;font-weight:900}
    .synth-resume-card .src-lbl{font-size:11px;color:#718096;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .action-list{margin-top:6px}
    .action-item{display:flex;align-items:flex-start;padding:6px 10px;margin-bottom:4px;background:#fff;border-radius:6px;border-left:4px solid;font-size:12px;gap:8px}
    .action-item.act-ko{border-left-color:#e53e3e}.action-item.act-warn{border-left-color:#d69e2e}
    .action-item .act-kpi{font-weight:800;color:#2d3748;min-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .action-item .act-ecart{font-weight:700;min-width:60px;text-align:center}
    .action-item .act-text{flex:1;color:#4a5568;font-size:11px;line-height:1.4}
    .hbar-chart{background:#fff;border-radius:var(--r);padding:14px 18px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04);margin-bottom:8px}
    .hbar-chart .hbar-title{font-size:15px;font-weight:800;color:var(--p);margin-bottom:4px;padding-bottom:6px;border-bottom:2px solid var(--b)}
    .hbar-chart .hbar-sub{font-size:11px;color:#718096;margin-bottom:10px}
    .hbar-row{display:flex;align-items:center;padding:4px 0;gap:8px}
    .hbar-label{width:260px;font-size:11px;font-weight:700;color:#2d3748;text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .hbar-track{flex:1;position:relative;height:26px;background:#edf2f7;border-radius:4px;overflow:visible}
    .hbar-fill{height:100%;border-radius:4px;position:relative;min-width:2px;transition:width .5s ease}
    .hbar-fill.ok{background:linear-gradient(90deg,#276749,#48bb78)}.hbar-fill.ko{background:linear-gradient(90deg,#c53030,#fc8181)}.hbar-fill.warn{background:linear-gradient(90deg,#d69e2e,#f6e05e)}
    .hbar-val{position:absolute;right:-58px;top:50%;transform:translateY(-50%);font-size:12px;font-weight:900;color:#1a202c;white-space:nowrap}
    .hbar-target{position:absolute;top:50%;transform:translateY(-50%);width:2px;height:20px;background:#e53e3e;z-index:2}
    .hbar-target-label{position:absolute;top:-14px;transform:translateX(-50%);font-size:9px;color:#e53e3e;font-weight:700;white-space:nowrap}
    .hbar-legend{display:flex;gap:16px;margin-top:10px;padding-top:8px;border-top:1px solid #edf2f7;font-size:11px;font-weight:600;color:#4a5568;flex-wrap:wrap}
    .hbar-legend span{display:flex;align-items:center;gap:5px}
    .hbar-legend .lg-swatch{width:14px;height:14px;border-radius:3px;display:inline-block}
    .toggle-bar{display:flex;align-items:center;gap:0;margin:6px 0 4px 0;border-radius:8px;overflow:hidden;border:1px solid var(--b);width:fit-content}
    .toggle-btn{padding:8px 20px;font-size:13px;font-weight:700;cursor:pointer;border:none;background:#f7fafc;color:#718096;transition:all .2s}
    .toggle-btn.active-p{background:linear-gradient(135deg,#276749,#38a169);color:#fff}
    .toggle-btn.active-q{background:linear-gradient(135deg,#2b6cb0,#3182ce);color:#fff}
    .toggle-btn:hover{background:#e2e8f0}
    .pie-card{background:#fff;border-radius:var(--r);padding:14px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .pie-card .pie-title{font-size:14px;font-weight:800;color:var(--p);margin-bottom:8px;text-align:center}
    .pie-legend{display:flex;flex-wrap:wrap;gap:6px 14px;justify-content:center;margin-top:8px}
    .pie-legend span{display:flex;align-items:center;gap:5px;font-size:11px;font-weight:600;color:#2d3748}
    .pie-legend .pl-dot{width:12px;height:12px;border-radius:2px;flex-shrink:0}
    .pie-sub{margin-top:6px;padding-top:6px;border-top:1px solid #edf2f7}
    .pie-sub-title{font-size:11px;font-weight:700;color:#718096;margin-bottom:4px;text-align:center}
    .backlog-section{background:#fff;border-radius:var(--r);padding:14px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04);margin-top:8px}
    .backlog-section .bs-title{font-size:15px;font-weight:800;margin-bottom:8px;padding-bottom:6px;border-bottom:2px solid var(--b)}
    .backlog-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-top:8px}
    @media(max-width:900px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg,.dgrid,.synth-resume,.backlog-grid{grid-template-columns:1fr}.hbar-label{width:140px}.synth-kpi-name{width:140px}.action-item .act-kpi{min-width:140px}.gbr-l{width:110px}}
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def contient_mot(t,lm):
        t=str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if a<=1: return "<1 mois"
        elif a>=3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
        return df
    def get_atelier(p):
        p=str(p).upper()
        if "PS" in p: return "Sulfurique"
        if "PP" in p: return "Phosphorique"
        if "TSP" in p or "REX" in p: return "Engrais"
        if "MCP" in p or "DCP" in p: return "Feed"
        return "Centrale et Utilitaires"
    def get_metier(p):
        p=str(p).upper()
        if "E" in p: return "Electrique"
        if "M" in p: return "Mecanique"
        if "R" in p: return "Instrumentation"
        if "G" in p: return "Genie Civil"
        return "Autre"
    def get_division(p):
        p=str(p).upper()
        if "SF1" in p: return "SF1"
        if "SF2" in p: return "SF2"
        return "Autre"
    def get_caract_type_detail(statut_user,keywords):
        s=str(statut_user).upper(); matched=[kw for kw in keywords if kw in s]
        return max(matched,key=len) if matched else "NON CARACTERISE"

    def calc_kpis(df_i,av_i,now,posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
        df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
        df["Type caract preparation"]=df["Statut utilisateur"].apply(lambda x:get_caract_type_detail(x,MP_KW))
        df["Type caract planification"]=df["Statut utilisateur"].apply(lambda x:get_caract_type_detail(x,MPLAN_KW))
        for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc]=pd.to_datetime(df[dc],errors='coerce')
                df[am]=((now.year-df[dc].dt.year)*12+(now.month-df[dc].dt.month)).round(2)
                df[ac]=df[am].apply(cat_age)
            else: df[am]=np.nan; df[ac]="Inconnu"
        df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
        df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
        df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
        df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
        res['dfp']=df
        an=cpiv(df,df["Nº appel pl.entret."].fillna(0)==0,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["Total"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1); an["TAUX_REALISATION_CORRECTIF/PT"]=ckpi(an["TCLO"],an["Total"])
        pr=cpiv(df,df["Statut OT"]=="CRÉÉ","ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
        avf=av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip()=="")].copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res

    # ===== STYLE HELPERS =====
    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "color:#cbd5e0"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def kan(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "background:#c6efce;color:#006100;font-weight:600"
        if val<=2: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        return "background:#ffc7ce;color:#9c0006;font-weight:600"
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        return 0
    def is_lb(k): return k in LOWER_BETTER

    # ===== HTML GENERATORS =====
    def html_table(rows,cols,tc,sc_col=None,style_fn=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                elif style_fn and c not in ["Poste de travail"]: s=style_fn(v,c); h+='<td style="%s">%s</td>'%(s or "",v)
                elif sc_col and c in sc_col: s=cs(v); h+='<td style="%s">%s</td>'%(s or "",v)
                else: h+='<td>%s</td>'%v
            h+='</tr>'
        return h+'</tbody></table>'
    def html_ano(rows,cols):
        h='<table class="tw at"><thead><tr>'+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            h+='<tr class="%s">'%("tr" if r.get("_t")=="total" else "")
            for c in cols: v=r.get(c,""); h+='<td style="%s">%s</td>'%(kas(v) or "",v)
            h+='</tr>'
        return h+'</tbody></table>'

    def html_synthese_actions(kpi_list, actuals, targets, act_map, section_type):
        color_main = "#276749" if section_type == "perf" else "#2b6cb0"
        label = "PERFORMANCE" if section_type == "perf" else "QUALITE"
        icon = "📊" if section_type == "perf" else "✅"
        nb_total=len(kpi_list); nb_ok=0; nb_ko=0; nb_warn=0; details_html=""
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            if is_lb(k): is_warn=(not met) and (av<=tv*1.5)
            else: wt=tv-5; is_warn=(not met) and (av>=wt)
            if met: nb_ok+=1; badge_class="badge-ok"; badge_txt="ATTEINT"; val_style="background:#c6efce;color:#006100"
            elif is_warn: nb_warn+=1; badge_class="badge-warn"; badge_txt="ATTENTION"; val_style="background:#ffeb9c;color:#9c6500"
            else: nb_ko+=1; badge_class="badge-ko"; badge_txt="NON ATTEINT"; val_style="background:#ffc7ce;color:#9c0006"
            ec_color="#276749" if met else ("#9c6500" if is_warn else "#c53030")
            action="✅ Objectif atteint" if met else act_map.get(k,"")
            details_html+='<div class="synth-kpi-row"><div class="synth-kpi-name">%s</div><div class="synth-kpi-val" style="%s">%.1f%%</div><div class="synth-kpi-target">Cible: %.0f%%</div><div class="synth-kpi-ecart" style="color:%s">%+.1f%%</div><div class="synth-kpi-badge %s">%s</div><div class="synth-kpi-action">%s</div></div>'%(k,val_style,av,tv,ec_color,diff,badge_class,badge_txt,action)
        actions_html=""
        ko_kpis=[]; warn_kpis=[]
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv; met=av<=tv if is_lb(k) else av>=tv
            if not met:
                if is_lb(k): is_w=av<=tv*1.5
                else: is_w=av>=tv-5
                (warn_kpis if is_w else ko_kpis).append((k,diff))
        if ko_kpis or warn_kpis:
            actions_html='<div class="action-list"><div style="font-size:13px;font-weight:800;color:#c53030;margin-bottom:4px">🔴 Actions Correctives Prioritaires</div>'
            for k,diff in ko_kpis: actions_html+='<div class="action-item act-ko"><div class="act-kpi">%s</div><div class="act-ecart" style="color:#c53030">%+.1f%%</div><div class="act-text">%s</div></div>'%(k,diff,act_map.get(k,""))
            if warn_kpis:
                actions_html+='<div style="font-size:13px;font-weight:800;color:#d69e2e;margin:6px 0 4px 0">🟡 Actions d\'Amélioration</div>'
                for k,diff in warn_kpis: actions_html+='<div class="action-item act-warn"><div class="act-kpi">%s</div><div class="act-ecart" style="color:#9c6500">%+.1f%%</div><div class="act-text">%s</div></div>'%(k,diff,act_map.get(k,""))
            actions_html+='</div>'
        h='<div class="synth-box"><div class="synth-title"><span style="font-size:20px">%s</span> Synthèse & Actions — %s</div>'%(icon,label)
        h+='<div class="synth-resume"><div class="synth-resume-card"><div class="src-val" style="color:#276749">%d</div><div class="src-lbl">Atteints</div></div><div class="synth-resume-card"><div class="src-val" style="color:#d69e2e">%d</div><div class="src-lbl">Attention</div></div><div class="synth-resume-card"><div class="src-val" style="color:#c53030">%d</div><div class="src-lbl">Non Atteints</div></div></div>'
        h+=details_html
        if actions_html: h+=actions_html
        h+='</div>'; return h

    def html_hbar_chart(kpi_list, actuals, targets, total_general, section_type, title):
        h='<div class="hbar-chart"><div class="hbar-title">%s</div>'%title
        h+='<div class="hbar-sub">Comparaison KPI vs Total Général (%.1f%%)</div>'%total_general
        all_vals=[actuals.get(k,0) for k in kpi_list]+[total_general]+[targets.get(k,100) for k in kpi_list]
        max_val=max(all_vals)*1.15; 
        if max_val==0: max_val=100
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); met=av<=tv if is_lb(k) else av>=tv
            if met: fill_class="ok"
            else:
                if is_lb(k): fill_class="warn" if av<=tv*1.5 else "ko"
                else: fill_class="warn" if av>=tv-5 else "ko"
            bar_width=max((av/max_val)*100,0.5); target_pos=min((tv/max_val)*100,100); total_pos=min((total_general/max_val)*100,100)
            h+='<div class="hbar-row"><div class="hbar-label">%s</div><div class="hbar-track"><div class="hbar-fill %s" style="width:%.1f%%"><div class="hbar-val">%.1f%%</div></div><div class="hbar-target" style="left:%.1f%%"><div class="hbar-target-label">C:%.0f</div></div><div style="position:absolute;left:%.1f%%;top:50%%;transform:translateY(-50%%);width:2px;height:26px;border-left:2px dashed #805ad5;z-index:1;opacity:.7"></div></div></div>'%(k,fill_class,bar_width,av,target_pos,tv,total_pos)
        h+='<div class="hbar-legend"><span><span class="lg-swatch" style="background:linear-gradient(90deg,#276749,#48bb78)"></span> Atteint</span><span><span class="lg-swatch" style="background:linear-gradient(90deg,#d69e2e,#f6e05e)"></span> Attention</span><span><span class="lg-swatch" style="background:linear-gradient(90deg,#c53030,#fc8181)"></span> Non Atteint</span><span><span class="lg-swatch" style="background:transparent;border-left:2px solid #e53e3e;height:14px"></span> Cible</span><span><span class="lg-swatch" style="background:transparent;border-left:2px dashed #805ad5;height:14px"></span> Total Général</span></div></div>'
        return h

    def html_pie_chart(data_dict, title, colors=None, threshold=5):
        if not data_dict: return ""
        total=sum(data_dict.values())
        if total==0: return '<div class="es">Aucune donnée</div>'
        if colors is None: colors=["#3182ce","#38a169","#e53e3e","#d69e2e","#805ad5","#ed8936","#4299e1","#48bb78","#fc8181","#f6e05e","#b83280","#dd6b20"]
        # Separate main segments and "others"
        main_items=[]; other_items=[]
        for label,val in data_dict.items():
            pct=val/total*100
            if pct>=threshold or len(data_dict)<=6: main_items.append((label,val))
            else: other_items.append((label,val))
        if other_items:
            main_items.append(("Autres",sum(v for _,v in other_items)))
        # Draw main pie
        svg_s=260; cx=svg_s/2; cy=svg_s/2; r=95
        paths=""; start_angle=-90; legend_main=[]
        for i,(label,val) in enumerate(main_items):
            pct=val/total*100
            if pct==0: continue
            angle=pct/100*360; end_angle=start_angle+angle; color=colors[i%len(colors)]
            sa_r=np.radians(start_angle); ea_r=np.radians(end_angle)
            x1=cx+r*np.cos(sa_r); y1=cy+r*np.sin(sa_r); x2=cx+r*np.cos(ea_r); y2=cy+r*np.sin(ea_r)
            la=1 if angle>180 else 0
            if abs(angle-360)<0.01: d="M %f %f L %f %f A %f %f 0 1 1 %f %f Z"%(cx,cy,x1,y1,r,r,x2,y2)
            else: d="M %f %f L %f %f A %f %f 0 %d 1 %f %f Z"%(cx,cy,x1,y1,r,r,la,x2,y2)
            paths+='<path d="%s" fill="%s" stroke="#fff" stroke-width="2" opacity="0.9"/>'%(d,color)
            mid_a=np.radians(start_angle+angle/2); lx=cx+r*0.6*np.cos(mid_a); ly=cy+r*0.6*np.sin(mid_a)
            if pct>4: paths+='<text x="%f" y="%f" text-anchor="middle" dominant-baseline="central" font-size="11" font-weight="700" fill="#fff">%.1f%%</text>'%(lx,ly,pct)
            if pct<=4 and pct>0: paths+='<text x="%f" y="%f" text-anchor="start" dominant-baseline="central" font-size="9" font-weight="700" fill="#1a202c">%.1f%%</text>'%(cx+r+5,ly,pct)
            legend_main.append('<span><span class="pl-dot" style="background:%s"></span>%s (%d — %.1f%%)</span>'%(color,label,val,pct))
            start_angle=end_angle
        h='<div class="pie-card"><div class="pie-title">%s</div>'%title
        h+='<svg width="%d" height="%d" viewBox="0 0 %d %d" style="display:block;margin:0 auto">%s</svg>'%(svg_s,svg_s,svg_s,svg_s,paths)
        h+='<div class="pie-legend">%s</div>'% ''.join(legend_main)
        # Sub-pie for "Autres" if needed
        if other_items:
            h+='<div class="pie-sub"><div class="pie-sub-title">📊 Détail « Autres » (%.1f%%)</div>'%(sum(v for _,v in other_items)/total*100)
            sub_s=160; scx=sub_s/2; scy=sub_s/2; sr=55; sub_paths=""; sa=-90
            for i,(label,val) in enumerate(other_items):
                pct=val/total*100; oth_total=sum(v for _,v in other_items)
                opct=val/oth_total*100; angle=opct/100*360; ea=sa+angle; color=colors[(len(main_items)-1+i)%len(colors)]
                sa_r=np.radians(sa); ea_r=np.radians(ea)
                x1=scx+sr*np.cos(sa_r); y1=scy+sr*np.sin(sa_r); x2=scx+sr*np.cos(ea_r); y2=scy+sr*np.sin(ea_r)
                la=1 if angle>180 else 0
                if abs(angle-360)<0.01: dd="M %f %f L %f %f A %f %f 0 1 1 %f %f Z"%(scx,scy,x1,y1,sr,sr,x2,y2)
                else: dd="M %f %f L %f %f A %f %f 0 %d 1 %f %f Z"%(scx,scy,x1,y1,sr,sr,la,x2,y2)
                sub_paths+='<path d="%s" fill="%s" stroke="#fff" stroke-width="1.5"/>'%(dd,color)
                mid_a=np.radians(sa+angle/2); lx=scx+sr*0.55*np.cos(mid_a); ly=scy+sr*0.55*np.sin(mid_a)
                if opct>8: sub_paths+='<text x="%f" y="%f" text-anchor="middle" dominant-baseline="central" font-size="9" font-weight="700" fill="#fff">%.0f%%</text>'%(lx,ly,opct)
                sa=ea
            h+='<svg width="%d" height="%d" viewBox="0 0 %d %d" style="display:block;margin:0 auto">%s</svg>'%(sub_s,sub_s,sub_s,sub_s,sub_paths)
            sub_legend=''
            for i,(label,val) in enumerate(other_items):
                color=colors[(len(main_items)-1+i)%len(colors)]
                sub_legend+='<span><span class="pl-dot" style="background:%s"></span>%s (%d)</span>'%(color,label,val)
            h+='<div class="pie-legend">%s</div>'%sub_legend
            h+='</div>'
        h+='</div>'; return h

    def html_grouped_bars_colored(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Perf (≥80%)</span><span><i style="background:linear-gradient(90deg,#d69e2e,#f6e05e)"></i> Perf (60-80%)</span><span><i style="background:linear-gradient(90deg,#c53030,#fc8181)"></i> Perf (<60%)</span><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Qual (≥80%)</span><span><i style="background:linear-gradient(90deg,#9b59b6,#af7ac5)"></i> Qual (60-80%)</span><span><i style="background:linear-gradient(90deg,#c0392b,#e74c3c)"></i> Qual (<60%)</span></div>'
        for p in sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True):
            pv=pscores.get(p,0); qv=qscores.get(p,0)
            pc="gb-ok" if pv>=80 else ("gb-warn" if pv>=60 else "gb-ko")
            qc="gb-ok" if qv>=80 else ("gb-warn" if qv>=60 else "gb-ko")
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f %s" style="width:%s%%"></div></div><div class="gbr-v" style="%s">%.1f%%</div><div class="gbr-w"><div class="gbr-f %s" style="width:%s%%"></div></div><div class="gbr-v" style="%s">%.1f%%</div></div></div>'%(p,min(max(pv,0),100),pc,cs("%.1f"%pv),pv,min(max(qv,0),100),qc,cs("%.1f"%qv),qv)
        h+='</div>'; return h

    def html_classement(scores,accent,detail_label):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">🏆 Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">⚠️ Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'
        h+='<div style="font-size:11px;color:#718096;margin-top:4px;text-align:center">📊 Classement basé sur : %s</div>'%detail_label
        return h

    def html_backlog_section(dfp, posts, backlog_col, type_col, section_title, table_tc):
        bl_df=dfp[dfp["Statut OT"]=="CRÉÉ"].copy() if "Préparation" in section_title else dfp[dfp["Statut OT"]=="LANC"].copy()
        # Table par poste
        pivot=pd.pivot_table(bl_df,index="Poste travail princ.",columns=backlog_col,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pivot[c]=pivot.get(c,0)
        pivot["Total"]=pivot.sum(axis=1)
        pivot["% Caractérisé"]=np.where(pivot["Total"]==0,"—","%.1f%%"%(pivot.get("CARACTERISE",0)/pivot["Total"]*100))
        rows=[]; cols=["Poste de travail","Total","CARACTERISE","NON CARACTERISE","% Caractérisé"]
        for p in pivot.index:
            rows.append({"Poste de travail":p,"Total":int(pivot.loc[p,"Total"]),"CARACTERISE":int(pivot.loc[p,"CARACTERISE"]),"NON CARACTERISE":int(pivot.loc[p,"NON CARACTERISE"]),"% Caractérisé":pivot.loc[p,"% Caractérisé"],"_t":""})
        tot_c=int(pivot["Total"].sum()); tot_car=int(pivot.get("CARACTERISE",pd.Series(0,index=pivot.index)).sum()); tot_nc=int(pivot.get("NON CARACTERISE",pd.Series(0,index=pivot.index)).sum())
        tot_pct="%.1f%%"%(tot_car/tot_c*100) if tot_c>0 else "—"
        rows.append({"Poste de travail":"Total général","Total":tot_c,"CARACTERISE":tot_car,"NON CARACTERISE":tot_nc,"% Caractérisé":tot_pct,"_t":"total"})
        # Pie 1: Caractérisé vs Non
        pie1_data={"Caractérisé":tot_car,"Non Caractérisé":tot_nc}
        # Pie 2: Types de caractérisation
        type_counts=bl_df[bl_df[backlog_col]=="CARACTERISE"][type_col].value_counts().to_dict()
        if not type_counts: type_counts={"Aucun":0}
        h='<div class="backlog-section"><div class="bs-title" style="color:#dd6b20">📦 %s</div>'%section_title
        h+=html_table(rows,cols,table_tc)
        h+='<div class="backlog-grid">'
        h+=html_pie_chart(pie1_data,"Taux de Caractérisation",["#38a169","#e53e3e"])
        h+=html_pie_chart(type_counts,"Répartition Types de Caractérisation",["#3182ce","#805ad5","#d69e2e","#e53e3e","#48bb78","#ed8936","#b83280","#dd6b20"])
        h+='</div></div>'; return h

    def html_special_analysis(dfp, posts, keyword, section_title, table_tc, icon):
        desc_col="Description" if "Description" in dfp.columns else None
        if desc_col is None:
            return '<div class="backlog-section"><div class="bs-title" style="color:#b83280">%s %s</div><div class="es">Colonne Description non trouvée</div></div>'%(icon,section_title)
        spec_df=dfp[dfp[desc_col].str.contains(keyword,case=False,na=False)].copy()
        pivot=pd.pivot_table(spec_df,index="Poste travail princ.",columns="Statut OT",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        statuts=sorted(pivot.columns.tolist())
        rows=[]; cols=["Poste de travail"]+statuts+["Total"]
        for p in pivot.index:
            rd={"Poste de travail":p,"_t":""}; t=0
            for s in statuts: rd[s]=int(pivot.loc[p,s]); t+=rd[s]
            rd["Total"]=t; rows.append(rd)
        tot_r={"Poste de travail":"Total général","_t":"total"}; gt=0
        for s in statuts: tot_r[s]=int(pivot[s].sum()); gt+=tot_r[s]
        tot_r["Total"]=gt; rows.append(tot_r)
        # Pie par statut
        stat_data=pivot.sum().to_dict()
        h='<div class="backlog-section"><div class="bs-title" style="color:#b83280">%s %s</div>'%(icon,section_title)
        h+=html_table(rows,cols,table_tc)
        if stat_data:
            h+='<div style="max-width:350px;margin:8px auto 0">'
            h+=html_pie_chart(stat_data,"Répartition %s par Statut"%keyword,["#3182ce","#38a169","#e53e3e","#d69e2e","#805ad5","#ed8936"])
            h+='</div>'
        h+='</div>'; return h

    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",unsafe_allow_html=True)
        st.markdown("---")
        show_filters=st.checkbox("Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf")
            ot_f=av_f=None; apm=[]
            if unf:
                ot_f=st.file_uploader("Fichier OT",type=["xlsx"],key="uot")
                av_f=st.file_uploader("Fichier AVIS",type=["xlsx"],key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t=excr(pd.read_excel("ot.xlsx"))
                        apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            st.markdown("**🏭 Atelier**")
            sa=st.multiselect("Atelier",["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"],["All"],key="sa")
            st.markdown("**🏢 Division**")
            sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=av_f=None; apm=[]; sp=["All"]; sa=["All"]; sd=["All"]
            dr=(datetime(2025,1,1).date(),datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try:
                    _t=excr(pd.read_excel("ot.xlsx"))
                    apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception: pass

    # ===================== DATA LOADING =====================
    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf: raw_ot=pd.read_excel(ot_f); raw_av=pd.read_excel(av_f)
            else: raw_ot=pd.read_excel("ot.xlsx"); raw_av=pd.read_excel("avis.xlsx")
            raw_ot=excr(raw_ot); raw_av=excr(raw_av)
            for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
            for c in ["Créé le","Début souhaité","Date de la clôture"]:
                if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
            if not apm: apm=sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp=apm
            if "All" in sa or not sa: sa=["All"]
            if "All" in sd or not sd: sd=["All"]
            sdt=pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt=pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p=str(poste).upper()
                if "All" not in sa:
                    m=False
                    if "Sulfurique (PS)" in sa and "PS" in p: m=True
                    if "Phosphorique (PP)" in sa and "PP" in p: m=True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m=True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m=True
                    if not m: return False
                if "All" not in sd:
                    m=False
                    if "SF1" in sd and "SF1" in p: m=True
                    if "SF2" in sd and "SF2" in p: m=True
                    if not m: return False
                return True

            vp=[p for p in apm if mf(p) and p in sp]
            df=raw_ot[(raw_ot["Poste travail princ."].isin(vp))&(raw_ot["Date de début planifiée"].between(sdt,edt))].copy()
            avdf=raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df=excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            avdf=excr(avdf[(avdf["Ordre"].isna())|(avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
            df_dash=raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash=excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns: df_dash["Statut OT"]=df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now=pd.Timestamp.now()
            res=calc_kpis(df,avdf,now,vp); ckdf=res['ckdf']; dfp=res['dfp']
            res_d=calc_kpis(df_dash,avdf,now,vp); ckdf_d=res_d['ckdf']
            pa={k:round(ckdf[k].mean(),2) for k in QK}; qa={k:round(ckdf[k].mean(),2) for k in PK}
            pa_d={k:round(ckdf_d[k].mean(),2) for k in QK}; qa_d={k:round(ckdf_d[k].mean(),2) for k in PK}

            # Scores par poste
            pscores={}; qscores={}
            for poste in ckdf.index:
                r=ckdf.loc[poste]
                pscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0
            sp_avg=np.mean(list(pscores.values())) if pscores else 0
            sq_avg=np.mean(list(qscores.values())) if qscores else 0
            sg_avg=(sp_avg+sq_avg)/2
            total_perf=np.mean([pa[k] for k in QK]) if QK else 0
            total_qual=np.mean([qa[k] for k in PK]) if PK else 0

            # ===== ANOMALY COUNTS PER POSTE =====
            sub_p={"TAUX_REALISATION_CORRECTIF/PT":lambda d:d[(d["Nº appel pl.entret."].fillna(0)==0)&(~d["Statut OT"].isin(["CLOT","TCLO"]))],
                   "OT préparation <1 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]!="<1 mois")],
                   "OT préparation >3 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]==">3 mois")],
                   "OT planification <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]!="<1 mois")],
                   "OT planification >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]==">3 mois")],
                   "OT exécution <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]!="<1 mois")],
                   "OT exécution >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]==">3 mois")]}
            sub_q={"OT LANC ESTIME":lambda d:d[(d["Statut OT"]=="LANC")&(d["OT LANC ESTIME"]=="NON")],
                   "Backlog préparation caractérisé":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["Backlog preparation"]=="NON CARACTERISE")],
                   "Backlog planification caractérisé":lambda d:d[(d["Statut OT"]=="LANC")&(d["Backlog planification"]=="NON CARACTERISE")],
                   "OT CONFIME":lambda d:d[(d["Statut système"].str.contains("CLO",na=False))&(~d["Statut système"].str.contains("CONF",na=False))],
                   "OT_COR_EGAL":lambda d:d[(d["OT_COR_EGAL"]=="NON")&(d["Total coûts réels"].fillna(0)>0)]}

            # Compute anomaly counts per poste
            ano_p_counts={}
            for k in QK:
                if k in sub_p:
                    try:
                        fd=sub_p[k](dfp)
                        cnt=fd.groupby("Poste travail princ.").size()
                        ano_p_counts[k]=cnt.reindex(vp,fill_value=0)
                    except: ano_p_counts[k]=pd.Series(0,index=vp)
                else: ano_p_counts[k]=pd.Series(0,index=vp)
            ano_p_cnt_df=pd.DataFrame(ano_p_counts,index=vp)
            ano_p_cnt_df["Total"]=ano_p_cnt_df.sum(axis=1)

            ano_q_counts={}
            for k in PK:
                if k in sub_q:
                    try:
                        fd=sub_q[k](dfp)
                        cnt=fd.groupby("Poste travail princ.").size()
                        ano_q_counts[k]=cnt.reindex(vp,fill_value=0)
                    except: ano_q_counts[k]=pd.Series(0,index=vp)
                else: ano_q_counts[k]=pd.Series(0,index=vp)
            ano_q_cnt_df=pd.DataFrame(ano_q_counts,index=vp)
            ano_q_cnt_df["Total"]=ano_q_cnt_df.sum(axis=1)

            # ===== ANOMALIES DETAIL =====
            ano_p_data=[]; ano_q_data=[]
            for kn,filt in sub_p.items():
                try:
                    fd=filt(dfp)
                    for _,rw in fd.iterrows():
                        ano_p_data.append({"Poste":rw.get("Poste travail princ.",""),"KPI":kn,"OT":rw.get("Ordre",""),"Description":str(rw.get("Description",""))[:60]})
                except: pass
            for kn,filt in sub_q.items():
                try:
                    fd=filt(dfp)
                    for _,rw in fd.iterrows():
                        ano_q_data.append({"Poste":rw.get("Poste travail princ.",""),"KPI":kn,"OT":rw.get("Ordre",""),"Description":str(rw.get("Description",""))[:60]})
                except: pass
            ano_p_df=pd.DataFrame(ano_p_data) if ano_p_data else pd.DataFrame(columns=["Poste","KPI","OT","Description"])
            ano_q_df=pd.DataFrame(ano_q_data) if ano_q_data else pd.DataFrame(columns=["Poste","KPI","OT","Description"])

            # ===== HISTORIQUE =====
            hist_path=os.path.join("kpis","indicateurs_kpis.xlsx") if os.path.exists(os.path.join("kpis","indicateurs_kpis.xlsx")) else None
            hist_df=load_historical_kpis(hist_path) if hist_path else pd.DataFrame()
            var_df=calculate_variations(hist_df)
            journal_df=generate_journal(var_df)
            top5_df,bot5_df=calculate_rankings(var_df)

            # ===== BUILD TABLES =====
            # Performance indicators table
            pcols=["Poste de travail"]+QK+["Score Performance"]
            prows=[{"Poste de travail":p,"_t":""} for p in ckdf.index]
            for i,p in enumerate(ckdf.index):
                for k in QK: prows[i][k]="%.1f"%(ckdf.loc[p,k]) if k in ckdf.columns else "0.0"
                prows[i]["Score Performance"]="%.2f"%(pscores[p])
            tg={"Poste de travail":"Total général","_t":"total"}
            for k in QK: tg[k]="%.1f"%(ckdf[k].mean()) if k in ckdf.columns else "0.0"
            tg["Score Performance"]="%.2f"%total_perf; prows.append(tg)
            cible_row={"Poste de travail":"CIBLE","_t":"cible"}
            for k in QK: cible_row[k]=CIBLE.get(k,"")
            cible_row["Score Performance"]="80.00"; prows.append(cible_row)

            # Performance anomalies count table
            pa_cols=["Poste de travail"]+QK+["Total"]
            pa_rows=[]
            for p in vp:
                rd={"Poste de travail":p,"_t":""}
                for k in QK: rd[k]=int(ano_p_cnt_df.loc[p,k]) if p in ano_p_cnt_df.index else 0
                rd["Total"]=int(ano_p_cnt_df.loc[p,"Total"]) if p in ano_p_cnt_df.index else 0
                pa_rows.append(rd)
            pa_tot={"Poste de travail":"Total général","_t":"total"}
            for k in QK: pa_tot[k]=int(ano_p_cnt_df[k].sum())
            pa_tot["Total"]=int(ano_p_cnt_df["Total"].sum()); pa_rows.append(pa_tot)

            # Qualite indicators table
            qcols=["Poste de travail"]+PK+["Score Qualite"]
            qrows=[{"Poste de travail":p,"_t":""} for p in ckdf.index]
            for i,p in enumerate(ckdf.index):
                for k in PK: qrows[i][k]="%.1f"%(ckdf.loc[p,k]) if k in ckdf.columns else "0.0"
                qrows[i]["Score Qualite"]="%.2f"%(qscores[p])
            tgq={"Poste de travail":"Total général","_t":"total"}
            for k in PK: tgq[k]="%.1f"%(ckdf[k].mean()) if k in ckdf.columns else "0.0"
            tgq["Score Qualite"]="%.2f"%total_qual; qrows.append(tgq)
            cible_row_q={"Poste de travail":"CIBLE","_t":"cible"}
            for k in PK: cible_row_q[k]=CIBLE.get(k,"")
            cible_row_q["Score Qualite"]="80.00"; qrows.append(cible_row_q)

            # Qualite anomalies count table
            qa_cols=["Poste de travail"]+PK+["Total"]
            qa_rows=[]
            for p in vp:
                rd={"Poste de travail":p,"_t":""}
                for k in PK: rd[k]=int(ano_q_cnt_df.loc[p,k]) if p in ano_q_cnt_df.index else 0
                rd["Total"]=int(ano_q_cnt_df.loc[p,"Total"]) if p in ano_q_cnt_df.index else 0
                qa_rows.append(rd)
            qa_tot={"Poste de travail":"Total général","_t":"total"}
            for k in PK: qa_tot[k]=int(ano_q_cnt_df[k].sum())
            qa_tot["Total"]=int(ano_q_cnt_df["Total"].sum()); qa_rows.append(qa_tot)

            # Anomalies detail tables
            ano_p_tbl_rows=[]
            for _,rw in ano_p_df.iterrows():
                ano_p_tbl_rows.append({"Poste":rw["Poste"],"KPI":rw["KPI"],"OT":rw["OT"],"Description":rw["Description"],"_t":""})
            ano_p_tbl_rows.append({"Poste":"TOTAL","KPI":"","OT":str(len(ano_p_df)),"Description":"","_t":"total"})
            ano_q_tbl_rows=[]
            for _,rw in ano_q_df.iterrows():
                ano_q_tbl_rows.append({"Poste":rw["Poste"],"KPI":rw["KPI"],"OT":rw["OT"],"Description":rw["Description"],"_t":""})
            ano_q_tbl_rows.append({"Poste":"TOTAL","KPI":"","OT":str(len(ano_q_df)),"Description":"","_t":"total"})

            save_kpis_to_excel(prows,pcols,qrows,qcols,
                ano_p_tbl_rows if ano_p_tbl_rows else [],["Poste","KPI","OT","Description"],
                ano_q_tbl_rows if ano_q_tbl_rows else [],["Poste","KPI","OT","Description"],
                fichier_date)

            # ===================== AFFICHAGE =====================
            nb_ot=len(df); nb_ano_p=len(ano_p_df); nb_ano_q=len(ano_q_df)
            st.markdown('<div class="mh"><h1>📊 Dashboard KPI — Maintenance</h1><div class="db">📅 %s</div></div>'%fichier_date,unsafe_allow_html=True)
            st.markdown('<div class="cr"><div class="cc c1"><div class="cv">%d</div><div class="cl">OT Traites</div></div><div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Global</div></div><div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div><div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies</div></div></div>'%(nb_ot,sg_avg,sp_avg,nb_ano_p+nb_ano_q),unsafe_allow_html=True)

            tab1,tab2,tab3,tab4=st.tabs(["📋 Performance","✅ Qualité","⚠️ Anomalies","📊 Présentation"])

            # ===== TAB 1 : PERFORMANCE =====
            with tab1:
                # Chart Bar Horizontal AVANT le tableau
                st.markdown(html_hbar_chart(QK,pa,CIBLE,total_perf,"perf","📊 Performance — KPI vs Total Général"),unsafe_allow_html=True)
                # Toggle
                st.markdown('<div class="toggle-bar"><button class="toggle-btn active-p" onclick="this.parentElement.querySelectorAll(\'.toggle-btn\').forEach(b=>b.className=\'toggle-btn\');this.className=\'toggle-btn active-p\';document.getElementById(\'perf-ind\').style.display=\'\';document.getElementById(\'perf-ano\').style.display=\'none\'">📊 Indicateurs de Performance</button><button class="toggle-btn" onclick="this.parentElement.querySelectorAll(\'.toggle-btn\').forEach(b=>b.className=\'toggle-btn\');this.className=\'toggle-btn active-p\';document.getElementById(\'perf-ind\').style.display=\'none\';document.getElementById(\'perf-ano\').style.display=\'\'">⚠️ Nombre d\'Anomalies</button></div>',unsafe_allow_html=True)
                st.markdown('<div id="perf-ind">',unsafe_allow_html=True)
                st.markdown('<div class="stl p">Indicateurs de Performance</div>',unsafe_allow_html=True)
                st.markdown(html_table(prows,pcols,"pt",["Score Performance"],ks),unsafe_allow_html=True)
                st.markdown('</div><div id="perf-ano" style="display:none">',unsafe_allow_html=True)
                st.markdown('<div class="stl a">Nombre d\'Anomalies par KPI — Performance</div>',unsafe_allow_html=True)
                st.markdown(html_table(pa_rows,pa_cols,"at",style_fn=kan),unsafe_allow_html=True)
                st.markdown('</div>',unsafe_allow_html=True)
                st.markdown(html_synthese_actions(QK,pa,CIBLE,ACT_MAP,"perf"),unsafe_allow_html=True)

            # ===== TAB 2 : QUALITE =====
            with tab2:
                st.markdown(html_hbar_chart(PK,qa,CIBLE,total_qual,"qual","📊 Qualité — KPI vs Total Général"),unsafe_allow_html=True)
                st.markdown('<div class="toggle-bar"><button class="toggle-btn active-q" onclick="this.parentElement.querySelectorAll(\'.toggle-btn\').forEach(b=>b.className=\'toggle-btn\');this.className=\'toggle-btn active-q\';document.getElementById(\'qual-ind\').style.display=\'\';document.getElementById(\'qual-ano\').style.display=\'none\'">✅ Indicateurs de Qualité</button><button class="toggle-btn" onclick="this.parentElement.querySelectorAll(\'.toggle-btn\').forEach(b=>b.className=\'toggle-btn\');this.className=\'toggle-btn active-q\';document.getElementById(\'qual-ind\').style.display=\'none\';document.getElementById(\'qual-ano\').style.display=\'\'">⚠️ Nombre d\'Anomalies</button></div>',unsafe_allow_html=True)
                st.markdown('<div id="qual-ind">',unsafe_allow_html=True)
                st.markdown('<div class="stl q">Indicateurs de Qualité</div>',unsafe_allow_html=True)
                st.markdown(html_table(qrows,qcols,"qt",["Score Qualite"],ks),unsafe_allow_html=True)
                st.markdown('</div><div id="qual-ano" style="display:none">',unsafe_allow_html=True)
                st.markdown('<div class="stl a">Nombre d\'Anomalies par KPI — Qualité</div>',unsafe_allow_html=True)
                st.markdown(html_table(qa_rows,qa_cols,"at",style_fn=kan),unsafe_allow_html=True)
                st.markdown('</div>',unsafe_allow_html=True)
                st.markdown(html_synthese_actions(PK,qa,CIBLE,ACT_MAP,"qual"),unsafe_allow_html=True)

            # ===== TAB 3 : ANOMALIES =====
            with tab3:
                st.markdown('<div class="stl a">Anomalies Performance (%d)</div>'%nb_ano_p,unsafe_allow_html=True)
                if ano_p_tbl_rows: st.markdown(html_ano(ano_p_tbl_rows,["Poste","KPI","OT","Description"]),unsafe_allow_html=True)
                else: st.markdown('<div class="es">✅ Aucune anomalie performance</div>',unsafe_allow_html=True)
                st.markdown('<div class="stl a">Anomalies Qualité (%d)</div>'%nb_ano_q,unsafe_allow_html=True)
                if ano_q_tbl_rows: st.markdown(html_ano(ano_q_tbl_rows,["Poste","KPI","OT","Description"]),unsafe_allow_html=True)
                else: st.markdown('<div class="es">✅ Aucune anomalie qualité</div>',unsafe_allow_html=True)

            # ===== TAB 4 : PRÉSENTATION =====
            with tab4:
                # Grouped bars colorés par objectif
                st.markdown('<div class="stl c">Performance vs Qualité par Poste — Coloré par Objectif</div>',unsafe_allow_html=True)
                if vp: st.markdown(html_grouped_bars_colored(vp,pscores,qscores,"Comparaison Performance / Qualité"),unsafe_allow_html=True)

                # Répartition par Atelier
                st.markdown('<div class="stl c">Répartition par Atelier</div>',unsafe_allow_html=True)
                atelier_data={}
                for p in vp: a=get_atelier(p); atelier_data[a]=atelier_data.get(a,0)+1
                if atelier_data:
                    c1,c2,c3=st.columns(3)
                    with c1: st.markdown(html_pie_chart(atelier_data,"Répartition par Atelier",["#3182ce","#38a169","#e53e3e","#d69e2e","#805ad5"]),unsafe_allow_html=True)

                # Répartition par Division (NOUVEAU)
                st.markdown('<div class="stl c">Répartition par Division</div>',unsafe_allow_html=True)
                div_data={}
                for p in vp: d=get_division(p); div_data[d]=div_data.get(d,0)+1
                if div_data:
                    with c2: st.markdown(html_pie_chart(div_data,"Répartition par Division",["#1e3a5f","#2c5282","#4a5568"]),unsafe_allow_html=True)

                # Répartition par Métier
                st.markdown('<div class="stl c">Répartition par Métier</div>',unsafe_allow_html=True)
                metier_data={}
                for p in vp: m=get_metier(p); metier_data[m]=metier_data.get(m,0)+1
                if metier_data:
                    with c3: st.markdown(html_pie_chart(metier_data,"Répartition par Métier",["#2b6cb0","#276749","#c53030","#d69e2e","#6b46c1"]),unsafe_allow_html=True)

                # ===== CLASSEMENT =====
                st.markdown('<div class="stl s">Classement des Postes</div>',unsafe_allow_html=True)
                if pscores:
                    st.markdown('<div style="font-size:13px;font-weight:700;color:#276749;margin-bottom:4px">📊 Classement Performance</div>',unsafe_allow_html=True)
                    st.markdown(html_classement(pscores,"#38a169","Pourcentage de KPI Performance atteignant leur cible respective (seuils : ≥75% pour âge, ≥80% pour taux, ≥95% pour conformité)"),unsafe_allow_html=True)
                if qscores:
                    st.markdown('<div style="font-size:13px;font-weight:700;color:#2b6cb0;margin-bottom:4px;margin-top:8px">📊 Classement Qualité</div>',unsafe_allow_html=True)
                    st.markdown(html_classement(qscores,"#3182ce","Pourcentage de KPI Qualité atteignant leur cible respective (seuils : ≥90% pour appels avis, ≥95% pour conformité)"),unsafe_allow_html=True)
                if pscores and qscores:
                    gs={p:(pscores.get(p,0)+qscores.get(p,0))/2 for p in vp}
                    st.markdown('<div style="font-size:13px;font-weight:700;color:#805ad5;margin-bottom:4px;margin-top:8px">📊 Classement Global Combiné</div>',unsafe_allow_html=True)
                    st.markdown(html_classement(gs,"#805ad5","Moyenne des scores Performance et Qualité — classement global de la maintenance"),unsafe_allow_html=True)

                # Top/Bottom historique
                if not top5_df.empty and not bot5_df.empty:
                    st.markdown('<div class="stl s">Évolution Historique — Top & Bottom 5</div>',unsafe_allow_html=True)
                    c1,c2=st.columns(2)
                    with c1:
                        st.markdown('<div class="rank-card"><div class="rank-title" style="color:#38a169">📈 Top 5 Progression</div>',unsafe_allow_html=True)
                        for i,(_,rw) in enumerate(top5_df.iterrows()):
                            st.markdown('<div class="rank-row"><div class="rank-num" style="background:#38a169">%d</div><div class="rank-name">%s</div><div class="rank-score" style="color:#38a169">%.1f</div></div>'%(i+1,rw["Poste"],rw["Score variation"]),unsafe_allow_html=True)
                        st.markdown('</div>',unsafe_allow_html=True)
                    with c2:
                        st.markdown('<div class="rank-card"><div class="rank-title" style="color:#e53e3e">📉 Bottom 5 Régression</div>',unsafe_allow_html=True)
                        for i,(_,rw) in enumerate(bot5_df.iterrows()):
                            st.markdown('<div class="rank-row"><div class="rank-num" style="background:#e53e3e">%d</div><div class="rank-name">%s</div><div class="rank-score" style="color:#e53e3e">%.1f</div></div>'%(i+1,rw["Poste"],rw["Score variation"]),unsafe_allow_html=True)
                        st.markdown('</div>',unsafe_allow_html=True)

                # Journal des variations
                if not journal_df.empty:
                    st.markdown('<div class="stl s">Journal des Variations Significatives</div>',unsafe_allow_html=True)
                    jcols=list(journal_df.columns); jrows=[]
                    for _,rw in journal_df.head(30).iterrows():
                        rd=dict(rw); rd["_t"]=""; jrows.append(rd)
                    st.markdown(html_table(jrows,jcols,"st"),unsafe_allow_html=True)
                    export_btn(journal_df,"journal_variations.xlsx")
                else:
                    st.markdown('<div class="es">Aucune variation significative (nécessite au moins 2 périodes historiques)</div>',unsafe_allow_html=True)

                # ===== ANALYSE BACKLOG PRÉPARATION =====
                st.markdown('<div class="stl bl"></div>',unsafe_allow_html=True)
                st.markdown(html_backlog_section(dfp,vp,"Backlog preparation","Type caract preparation","Analyse Backlog Préparation","bl"),unsafe_allow_html=True)

                # ===== ANALYSE BACKLOG PLANIFICATION =====
                st.markdown(html_backlog_section(dfp,vp,"Backlog planification","Type caract planification","Analyse Backlog Planification","bl"),unsafe_allow_html=True)

                # ===== ANALYSE OMS =====
                st.markdown('<div class="stl om"></div>',unsafe_allow_html=True)
                st.markdown(html_special_analysis(dfp,vp,"OMS","Analyse OMS","om","🔬"),unsafe_allow_html=True)

                # ===== ANALYSE THERMOGRAPHIE =====
                st.markdown(html_special_analysis(dfp,vp,"Thermographi","Analyse Thermographie","om","🌡️"),unsafe_allow_html=True)

            # EXPORTS
            with st.sidebar:
                st.markdown("---")
                st.markdown("**📥 Exports**")
                if not ano_p_df.empty: export_btn(ano_p_df,"anomalies_performance.xlsx")
                if not ano_q_df.empty: export_btn(ano_q_df,"anomalies_qualite.xlsx")

        except Exception as e:
            st.error(f"Erreur de chargement : {str(e)}")
            st.markdown('<div class="es">Vérifiez que ot.xlsx et avis.xlsx sont dans le répertoire du script.</div>',unsafe_allow_html=True)
    else:
        if unf:
            st.markdown('<div class="es">📁 Chargez les fichiers OT et AVIS dans la barre latérale.</div>',unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">📁 Placez ot.xlsx et avis.xlsx dans le répertoire du script.</div>',unsafe_allow_html=True)

if __name__ == "__main__":
    main()
