# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os, inspect, textwrap, pprint
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
ALL_KPI = QK + PK
CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100}
ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois."}
LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]
MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

# ============================================================
# FONCTIONS DE CALCUL (niveau module pour inspection dynamique)
# ============================================================

def contient_mot(t, lm):
    """Verifie si le texte t contient au moins un mot cle de la liste lm."""
    t = str(t)
    return any(m in t for l in lm for m in l.split())

def cat_age(a):
    """Categorise l'age en mois en 3 classes."""
    if a <= 1: return "<1 mois"
    elif a >= 3: return ">3 mois"
    return "1 mois < <3 mois"

def ckpi(n, d, sz=100):
    """Calcule le pourcentage KPI: (n/d)*100, avec securite contre division par zero."""
    return np.where(d == 0, sz, (n / d) * 100)

def cpiv(df, f, c, p):
    """Cree un pivot table count par poste et colonne c, reindexe sur la liste p."""
    return pd.pivot_table(df[f], index="Poste travail princ.", columns=c,
                          values="Ordre", aggfunc="count", fill_value=0).reindex(p, fill_value=0)

def excr(df):
    """Exclut les lignes contenant 'cresseur' dans le poste de travail."""
    if "Poste travail princ." in df.columns:
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur", case=False, na=False)].copy()
    return df

def get_metier(p):
    """Determine le metier a partir du code poste."""
    p = str(p).upper()
    if "E" in p: return "Electrique"
    if "M" in p: return "Mecanique"
    if "R" in p: return "Instrumentation"
    if "G" in p: return "Genie Civil"
    return "Autre"

def get_atelier(p):
    """Determine l'atelier a partir du code poste."""
    p = str(p).upper()
    if "PS" in p: return "Sulfurique"
    if "PP" in p: return "Phosphorique"
    if "TSP" in p or "REX" in p: return "Engrais"
    if "MCP" in p or "DCP" in p: return "Feed"
    return "Autre"

def get_division(p):
    """Determine la division a partir du code poste."""
    p = str(p).upper()
    if "SF1" in p: return "SF1"
    if "SF2" in p: return "SF2"
    return "Autre"

def get_caract_type(statut_user, keywords):
    """Determine le type de caracterisation a partir du statut utilisateur et des mots cles."""
    s = str(statut_user).upper()
    matched = [kw for kw in keywords if kw in s]
    return max(matched, key=len) if matched else "AUTRE"

# ----- Colonnes ajoutees -----
def creer_colonnes_additionnelles(df, now):
    """
    Cree toutes les colonnes additionnelles utilisees pour les calculs KPI.
    Colonnes creees:
      - Backlog preparation : CARACTERISE / NON CARACTERISE
      - Backlog planification : CARACTERISE / NON CARACTERISE
      - amp / ap : age en mois preparation / categorie
      - amlp / alp : age en mois planification / categorie
      - amex / aex : age en mois execution / categorie
      - OT CONFIME : OUI / NON
      - Contient SOPL : 1 / 0
      - OT LANC ESTIME : OUI / NON
      - OT_COR_EGAL : OUI / NON
    """
    df = df.copy()
    # Backlog preparation
    df["Backlog preparation"] = np.where(
        df["Statut utilisateur"].apply(lambda x: contient_mot(x, MP_KW)),
        "CARACTERISE", "NON CARACTERISE")
    # Backlog planification
    df["Backlog planification"] = np.where(
        df["Statut utilisateur"].apply(lambda x: contient_mot(x, MPLAN_KW)),
        "CARACTERISE", "NON CARACTERISE")
    # Ages et categories
    for dc, am, ac in [('Créé le', "amp", "ap"),
                        ('Date de début planifiée', "amlp", "alp"),
                        ('Date de début planifiée', "amex", "aex")]:
        if dc in df.columns:
            df[dc] = pd.to_datetime(df[dc], errors='coerce')
            df[am] = ((now.year - df[dc].dt.year) * 12 + (now.month - df[dc].dt.month)).round(2)
            df[ac] = df[am].apply(cat_age)
        else:
            df[am] = np.nan
            df[ac] = "Inconnu"
    # OT CONFIME
    df["OT CONFIME"] = np.where(
        df["Statut système"].str.contains("CLO", na=False) &
        df["Statut système"].str.contains("CONF", na=False),
        "OUI", "NON")
    # Contient SOPL
    df["Contient SOPL"] = df["Statut utilisateur"].str.contains("SOPL", na=False).map({True: 1, False: 0})
    # OT LANC ESTIME
    df["OT LANC ESTIME"] = np.where(df["Total coûts budgétés"].fillna(0) == 0, "NON", "OUI")
    # OT_COR_EGAL
    df["OT_COR_EGAL"] = np.where(
        (df["Total coûts budgétés"].fillna(0) - df["Total coûts réels"].fillna(0)) == 0,
        "OUI", "NON")
    return df

# ----- Calcul complet des KPI -----
def calc_kpis(df_i, av_i, now, posts):
    """
    Calcule tous les KPI par poste de travail.
    Retourne un dictionnaire contenant:
      - 'dfp' : DataFrame OT avec colonnes additionnelles
      - 'avf' : DataFrame avis filtres
      - 'ckdf' : DataFrame pivot des valeurs KPI par poste
      - 'ot_confime' : pivot OT CONFIME
      - 'ot_cor_egal' : pivot OT_COR_EGAL
    """
    res = {}
    df = creer_colonnes_additionnelles(df_i, now)
    av = av_i.copy()
    res['dfp'] = df

    # 1. TAUX_REALISATION_CORRECTIF/PT
    an = cpiv(df, df["Nº appel pl.entret."].fillna(0) == 0, "Statut OT", posts)
    for c in ["CLOT", "CRÉÉ", "LANC", "TCLO"]: an[c] = an.get(c, 0)
    an["Total"] = an[["CLOT", "CRÉÉ", "LANC", "TCLO"]].sum(axis=1)
    an["TAUX_REALISATION_CORRECTIF/PT"] = ckpi(an["TCLO"], an["Total"])

    # 2. Preparation
    pr = cpiv(df, df["Statut OT"] == "CRÉÉ", "ap", posts)
    for c in ["<1 mois", ">3 mois", "1 mois < <3 mois"]: pr[c] = pr.get(c, 0)
    pr["Total"] = pr[["<1 mois", "1 mois < <3 mois", ">3 mois"]].sum(axis=1)
    pr["OT préparation <1 mois"] = ckpi(pr["<1 mois"], pr["Total"])
    pr["OT préparation >3 mois"] = ckpi(pr[">3 mois"], pr["Total"], 0)
    pr["OT préparation 1mois< <3mois"] = ckpi(pr["1 mois < <3 mois"], pr["Total"], 0)

    # 3. Planification (hors SOPL)
    pl = cpiv(df, (df["Statut OT"] == "LANC") & (df["Contient SOPL"] == 0), "alp", posts)
    for c in ["<1 mois", ">3 mois", "1 mois < <3 mois"]: pl[c] = pl.get(c, 0)
    pl["Total"] = pl[["<1 mois", "1 mois < <3 mois", ">3 mois"]].sum(axis=1)
    pl["OT planification <1 mois"] = ckpi(pl["<1 mois"], pl["Total"])
    pl["OT planification >3 mois"] = ckpi(pl[">3 mois"], pl["Total"], 0)
    pl["OT planification 1mois< <3mois"] = ckpi(pl["1 mois < <3 mois"], pl["Total"], 0)

    # 4. Execution (avec SOPL)
    ex = cpiv(df, (df["Statut OT"] == "LANC") & (df["Contient SOPL"] == 1), "aex", posts)
    for c in ["<1 mois", ">3 mois", "1 mois < <3 mois"]: ex[c] = ex.get(c, 0)
    ex["Total"] = ex[["<1 mois", "1 mois < <3 mois", ">3 mois"]].sum(axis=1)
    ex["OT exécution <1 mois"] = ckpi(ex["<1 mois"], ex["Total"])
    ex["OT exécution >3 mois"] = ckpi(ex[">3 mois"], ex["Total"], 0)
    ex["OT exécution 1mois< <3mois"] = ckpi(ex["1 mois < <3 mois"], ex["Total"], 0)

    # 5. OT LANC ESTIME
    la = pd.pivot_table(df[df["Statut OT"] == "LANC"], index="Poste travail princ.",
                        columns="OT LANC ESTIME", values="Ordre", aggfunc="count",
                        fill_value=0).reindex(posts, fill_value=0)
    for c in ["OUI", "NON"]: la[c] = la.get(c, 0)
    la["Total"] = la["OUI"] + la["NON"]
    la["OT LANC ESTIME"] = ckpi(la["OUI"], la["Total"])

    # 6. Backlog préparation caractérisé
    pc = pd.pivot_table(df[df["Statut OT"] == "CRÉÉ"], index="Poste travail princ.",
                        columns="Backlog preparation", values="Ordre", aggfunc="count",
                        fill_value=0).reindex(posts, fill_value=0)
    for c in ["CARACTERISE", "NON CARACTERISE"]: pc[c] = pc.get(c, 0)
    pc["Total"] = pc["CARACTERISE"] + pc["NON CARACTERISE"]
    pc["Backlog préparation caractérisé"] = ckpi(pc["CARACTERISE"], pc["Total"])

    # 7. Backlog planification caractérisé
    plc = pd.pivot_table(df[df["Statut OT"] == "LANC"], index="Poste travail princ.",
                         columns="Backlog planification", values="Ordre", aggfunc="count",
                         fill_value=0).reindex(posts, fill_value=0)
    for c in ["CARACTERISE", "NON CARACTERISE"]: plc[c] = plc.get(c, 0)
    plc["Total"] = plc["CARACTERISE"] + plc["NON CARACTERISE"]
    plc["Backlog planification caractérisé"] = ckpi(plc["CARACTERISE"], plc["Total"])

    # 8. OT CONFIME & OT_COR_EGAL
    for kn, cn in [("OT CONFIME", "OT CONFIME"), ("OT_COR_EGAL", "OT_COR_EGAL")]:
        pv = pd.pivot_table(df, index="Poste travail princ.", columns=cn,
                            values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["OUI", "NON"]: pv[c] = pv.get(c, 0)
        pv["Total"] = pv["OUI"] + pv["NON"]
        pv[cn] = ckpi(pv["OUI"], pv["Total"])
        res[kn.lower().replace(" ", "_")] = pv

    # 9. Appel avis approuvé
    avf = av[(av["Ordre"].isna()) | (av["Ordre"].astype(str).str.strip() == "")].copy()
    res['avf'] = avf
    tca = pd.pivot_table(avf, index="Poste travail princ.", columns="Statut utilisateur",
                         values="Avis", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
    for c in ["APRQ", "APRV", "APRV AVAU", "REJT"]: tca[c] = tca.get(c, 0)
    tca["Total"] = tca[["APRQ", "APRV", "APRV AVAU", "REJT"]].sum(axis=1)
    tca["appel avis approuvé"] = ckpi(tca["APRV"], tca["Total"])

    # Assemblage final
    res['ckdf'] = pd.DataFrame({
        "TAUX_REALISATION_CORRECTIF/PT": an["TAUX_REALISATION_CORRECTIF/PT"],
        "OT préparation <1 mois": pr["OT préparation <1 mois"],
        "OT préparation >3 mois": pr["OT préparation >3 mois"],
        "OT préparation 1mois< <3mois": pr["OT préparation 1mois< <3mois"],
        "OT planification <1 mois": pl["OT planification <1 mois"],
        "OT planification >3 mois": pl["OT planification >3 mois"],
        "OT planification 1mois< <3mois": pl["OT planification 1mois< <3mois"],
        "OT exécution <1 mois": ex["OT exécution <1 mois"],
        "OT exécution >3 mois": ex["OT exécution >3 mois"],
        "OT exécution 1mois< <3mois": ex["OT exécution 1mois< <3mois"],
        "appel avis approuvé": tca["appel avis approuvé"],
        "OT LANC ESTIME": la["OT LANC ESTIME"],
        "Backlog préparation caractérisé": pc["Backlog préparation caractérisé"],
        "Backlog planification caractérisé": plc["Backlog planification caractérisé"],
        "OT CONFIME": res['ot_confime']["OT CONFIME"],
        "OT_COR_EGAL": res['ot_cor_egal']["OT_COR_EGAL"]
    })
    return res

# ----- Seuils de couleur KPI -----
def ks(v, c):
    """Retourne le style CSS selon la valeur v et le KPI c."""
    try: val = float(v)
    except Exception: return ""
    if c in ["OT préparation <1 mois", "OT planification <1 mois", "OT exécution <1 mois"]:
        return "background:#c6efce;color:#006100;font-weight:600" if val >= 80 else (
            "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 75 else
            "background:#ffc7ce;color:#9c0006;font-weight:600")
    if c in ["OT préparation 1mois< <3mois", "OT planification 1mois< <3mois", "OT exécution 1mois< <3mois"]:
        return "background:#c6efce;color:#006100;font-weight:600" if val <= 15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
    if c in ["OT préparation >3 mois", "OT planification >3 mois", "OT exécution >3 mois"]:
        return "background:#c6efce;color:#006100;font-weight:600" if val <= 5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
    if c == "TAUX_REALISATION_CORRECTIF/PT":
        return "background:#c6efce;color:#006100;font-weight:600" if val >= 85 else (
            "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 80 else
            "background:#ffc7ce;color:#9c0006;font-weight:600")
    if c == "appel avis approuvé":
        return "background:#c6efce;color:#006100;font-weight:600" if val >= 95 else (
            "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 90 else
            "background:#ffc7ce;color:#9c0006;font-weight:600")
    if c in ["OT LANC ESTIME", "Backlog préparation caractérisé", "Backlog planification caractérisé",
             "OT CONFIME", "OT_COR_EGAL"]:
        return "background:#c6efce;color:#006100;font-weight:600" if val >= 100 else (
            "background:#ffeb9c;color:#9c6500;font-weight:600" if val >= 95 else
            "background:#ffc7ce;color:#9c0006;font-weight:600")
    return ""

# ----- Seuil de couleur Score -----
def cs(v):
    """Retourne le style CSS du score global."""
    try: val = float(str(v).replace(' %', '').strip())
    except Exception: return ""
    return "background:#c6efce;color:#006100;font-weight:700" if val >= 90 else (
        "background:#ffeb9c;color:#9c6500;font-weight:700" if val >= 80 else
        "background:#ffc7ce;color:#9c0006;font-weight:700")

# ----- Seuil de couleur Anomalies -----
def kas(v):
    """Retourne le style CSS du nombre d'anomalies."""
    try: val = int(v)
    except Exception: return ""
    if val == 0: return "color:#cbd5e0"
    if val <= 3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
    if val <= 10: return "background:#fed7d7;color:#c53030;font-weight:600"
    return "background:#fc8181;color:#742a2a;font-weight:800"

# ----- Calcul du score -----
def gscore(k, a, t):
    """
    Calcule le score individuel d'un KPI (0 ou 1).
    k: nom du KPI, a: valeur actuelle, t: cible.
    Retourne 1 si l'objectif est atteint, 0 sinon.
    """
    if pd.isna(a) or pd.isna(t): return 0
    if k in ["OT préparation <1 mois", "OT planification <1 mois", "OT exécution <1 mois"]:
        return 1 if a >= 75 else 0
    if k in ["OT préparation 1mois< <3mois", "OT planification 1mois< <3mois", "OT exécution 1mois< <3mois"]:
        return 1 if a <= 15 else 0
    if k in ["OT préparation >3 mois", "OT planification >3 mois", "OT exécution >3 mois"]:
        return 1 if a <= 5 else 0
    if k == "TAUX_REALISATION_CORRECTIF/PT": return 1 if a >= 80 else 0
    if k == "appel avis approuvé": return 1 if a >= 90 else 0
    if k in ["OT LANC ESTIME", "Backlog préparation caractérisé", "Backlog planification caractérisé",
             "OT CONFIME", "OT_COR_EGAL"]: return 1 if a >= 95 else 0
    return 0

def is_lb(k):
    """Indique si le KPI est de type 'plus bas c'est mieux'."""
    return k in LOWER_BETTER

# ----- Filtres d'anomalies -----
def get_anomalie_filters():
    """Retourne les dictionnaires de filtres d'anomalies pour Performance et Qualite."""
    sub_p = {
        "TAUX_REALISATION_CORRECTIF/PT": lambda d: d[(d["Nº appel pl.entret."].fillna(0) == 0) & (~d["Statut OT"].isin(["CLOT", "TCLO"]))],
        "OT préparation <1 mois": lambda d: d[(d["Statut OT"] == "CRÉÉ") & (d["ap"] != "<1 mois")],
        "OT préparation >3 mois": lambda d: d[(d["Statut OT"] == "CRÉÉ") & (d["ap"] == ">3 mois")],
        "OT planification <1 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 0) & (d["alp"] != "<1 mois")],
        "OT planification >3 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 0) & (d["alp"] == ">3 mois")],
        "OT exécution <1 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 1) & (d["aex"] != "<1 mois")],
        "OT exécution >3 mois": lambda d: d[(d["Statut OT"] == "LANC") & (d["Contient SOPL"] == 1) & (d["aex"] == ">3 mois")]
    }
    sub_q = {
        "OT LANC ESTIME": lambda d: d[(d["Statut OT"] == "LANC") & (d["OT LANC ESTIME"] == "NON")],
        "Backlog préparation caractérisé": lambda d: d[(d["Statut OT"] == "CRÉÉ") & (d["Backlog preparation"] == "NON CARACTERISE")],
        "Backlog planification caractérisé": lambda d: d[(d["Statut OT"] == "LANC") & (d["Backlog planification"] == "NON CARACTERISE")],
        "OT CONFIME": lambda d: d[(d["OT CONFIME"] == "NON")],
        "OT_COR_EGAL": lambda d: d[(d["OT_COR_EGAL"] == "NON")]
    }
    return sub_p, sub_q

# ============================================================
# FONCTIONS UTILITAIRES
# ============================================================

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt", "r", encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_r, ano_p_c, ano_q_r, ano_q_c, sheet_name):
    kpis_dir = "kpis"; os.makedirs(kpis_dir, exist_ok=True)
    filepath = os.path.join(kpis_dir, "indicateurs_kpis.xlsx")
    sn = str(sheet_name).replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")[:31]
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfl = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    tf = Font(bold=True, size=12, color="1E3A5F")
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    try: wb = load_workbook(filepath)
    except Exception: wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws = wb.create_sheet(sn); rn = 1
    def ws_sec(title, cols, rows, sr):
        ws.cell(row=sr, column=1, value=title).font = tf; sr += 1
        for j, c in enumerate(cols, 1):
            cl = ws.cell(row=sr, column=j, value=c); cl.font = hf; cl.fill = hfl
            cl.alignment = Alignment(horizontal='center'); cl.border = tb
        sr += 1
        for r in rows:
            for j, c in enumerate(cols, 1):
                cl = ws.cell(row=sr, column=j, value=r.get(c, "")); cl.border = tb
                cl.alignment = Alignment(horizontal='center')
            sr += 1
        return sr + 1
    rn = ws_sec("INDICATEURS DE PERFORMANCE", pcols, prows, rn)
    if ano_p_c and ano_p_r: rn = ws_sec("ANOMALIES PERFORMANCE", ano_p_c, ano_p_r, rn)
    rn = ws_sec("INDICATEURS DE QUALITE", qcols, qrows, rn)
    if ano_q_c and ano_q_r: rn = ws_sec("ANOMALIES QUALITE", ano_q_c, ano_q_r, rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception: return pd.DataFrame()
    records = []; section = None; headers = None
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]; rows_data = list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0 = str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section = "perf"; headers = None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section = "qual"; headers = None; continue
                elif "ANOMALIES" in cell0.upper(): section = None; continue
                if section and headers is None and cell0:
                    headers = [str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE", "Total general", ""):
                    entry = {"Date": sheet_name}
                    for j, h in enumerate(headers):
                        if j < len(row): entry[h] = row[j]
                    entry["_section"] = section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df = pd.DataFrame(records)
    df["Date_parsed"] = pd.to_datetime(df["Date"].str.replace("-", "/"), format="%d/%m/%Y", errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates = sorted(hist_df["Date"].unique())
    if len(dates) < 2: return pd.DataFrame()
    perf_df = hist_df[hist_df["_section"] == "perf"].copy()
    qual_df = hist_df[hist_df["_section"] == "qual"].copy()
    variations = []
    for i in range(1, len(dates)):
        prev_date, curr_date = dates[i - 1], dates[i]
        prev_perf = perf_df[perf_df["Date"] == prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf = perf_df[perf_df["Date"] == curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual = qual_df[qual_df["Date"] == prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual = qual_df[qual_df["Date"] == curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name, prev_d, curr_d, kpi_list in [("Performance", prev_perf, curr_perf, QK + ["Score Performance"]),
                                                      ("Qualite", prev_qual, curr_qual, PK + ["Score Qualite"])]:
            for poste in set(prev_d.index) & set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv = float(prev_d.loc[poste, kpi])
                    except Exception: continue
                    try: cv = float(curr_d.loc[poste, kpi])
                    except Exception: continue
                    diff = cv - pv; pct = (diff / pv * 100) if pv != 0 else (100 if cv != 0 else 0)
                    if abs(diff) <= 0.5: trend = "stabilite"
                    elif diff > 0.5: trend = "hausse"
                    else: trend = "baisse"
                    variations.append({"Date precedente": prev_date, "Date actuelle": curr_date, "Poste": poste,
                                       "Type": sec_name, "KPI": kpi, "Valeur precedente": round(pv, 2),
                                       "Valeur actuelle": round(cv, 2), "Ecart": round(diff, 2),
                                       "Ecart %": round(pct, 2), "Tendance": trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j = var_df.copy(); j["Significatif"] = j["Ecart %"].abs() >= 5
    j = j[j["Significatif"]].copy()
    j["Sens"] = j.apply(lambda r: "Amelioration" if ((r["Tendance"] == "hausse" and r["KPI"] not in LOWER_BETTER) or
                                                       (r["Tendance"] == "baisse" and r["KPI"] in LOWER_BETTER))
                         else "Degradation", axis=1)
    return j.sort_values(["Date actuelle", "Sens", "Ecart %"], ascending=[True, False, False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(), pd.DataFrame()
    scores = {}
    for poste in var_df["Poste"].unique():
        pv = var_df[var_df["Poste"] == poste].copy()
        scores[poste] = sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _, r in pv.iterrows())
    ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    return pd.DataFrame(ranked[:5], columns=["Poste", "Score variation"]), pd.DataFrame(ranked[-5:][::-1], columns=["Poste", "Score variation"])

# ============================================================
# CSS
# ============================================================

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:6px 10px;background:#fff;border-radius:5px;margin-bottom:2px;border:1px solid var(--b);font-size:13px}
    .sr .sn{font-weight:700;color:var(--p);min-width:220px;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:3px 9px;border-radius:12px;font-weight:800;font-size:14px;min-width:50px;text-align:center;margin:0 8px;color:#fff}
    .sr .sa{color:#718096;font-size:12px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .stg{font-size:11px;color:#718096;min-width:60px;text-align:center;white-space:nowrap}
    .sr .sb{font-size:11px;font-weight:700;padding:2px 8px;border-radius:3px;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:4px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:24px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}
    .gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .anl-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;margin:0}
    .anl-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:12px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .anl-tbl tbody td{padding:5px 8px;border-bottom:1px solid #edf2f7}
    .anl-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .anl-tbl tbody tr:hover td{background:#ebf8ff!important}
    .anl-tbl .tot td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .trend-up{color:#276749;font-weight:800;font-size:16px}
    .trend-down{color:#c53030;font-weight:800;font-size:16px}
    .trend-stable{color:#718096;font-weight:800;font-size:16px}
    .spark-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:8px}
    .spark-card{background:#fff;border-radius:var(--r);padding:10px 12px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .spark-card .sp-title{font-size:13px;font-weight:800;color:var(--p);margin-bottom:3px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .spark-card .sp-sub{font-size:11px;color:#718096;margin-bottom:5px}
    .rank-card{background:#fff;border-radius:var(--r);padding:12px 16px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .rank-card .rank-title{font-size:15px;font-weight:800;margin-bottom:8px;padding-bottom:5px;border-bottom:2px solid var(--b)}
    .rank-row{display:flex;align-items:center;padding:5px 0;font-size:13px;border-bottom:1px solid #f7fafc}
    .rank-row:last-child{border:none}
    .rank-row .rank-num{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:13px;color:#fff;margin-right:10px;flex-shrink:0}
    .rank-row .rank-name{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .rank-row .rank-score{font-weight:900;min-width:70px;text-align:right}
    .code-page-section{background:#fff;border-radius:var(--r);padding:16px;margin-bottom:10px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .code-page-section h3{font-size:16px;font-weight:800;color:var(--p);margin:0 0 10px 0;padding-bottom:6px;border-bottom:2px solid var(--pl)}
    .code-page-section h4{font-size:13px;font-weight:700;color:var(--pl);margin:12px 0 4px 0}
    .code-page-desc{font-size:12px;color:#4a5568;margin-bottom:8px;line-height:1.6}
    .lock-overlay{min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:120px}.gbr-l{width:100px}.spark-grid{grid-template-columns:1fr}}
    </style>""", unsafe_allow_html=True)

# ============================================================
# PAGE "CODE DE CALCUL" — extraction dynamique via inspect
# ============================================================

CODE_AUTORISATION = "96800221"

def page_code_de_calcul():
    """Affiche la page Code de calcul avec authentification."""
    inject_custom_css()

    # Gestion authentification
    if "code_auth_ok" not in st.session_state:
        st.session_state.code_auth_ok = False

    if not st.session_state.code_auth_ok:
        st.markdown("""<div class="lock-overlay">
        <div style="font-size:64px;margin-bottom:20px">🔐</div>
        <h1 style="text-align:center;font-size:36px;color:#fff;font-weight:900;margin:0 0 8px 0">Code de calcul</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:18px;margin:0 0 30px 0">Acces restreint — Saisissez le code d'autorisation</p>
        </div>""", unsafe_allow_html=True)

        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            code_saisi = st.text_input("Code d'autorisation", type="password",
                                       key="code_input_auth", label_visibility="collapsed",
                                       placeholder="Entrez le code a 8 chiffres")
            if st.button("🔓 Acceder", type="primary", use_container_width=True):
                if code_saisi.strip() == CODE_AUTORISATION:
                    st.session_state.code_auth_ok = True
                    st.rerun()
                else:
                    st.error("❌ Code incorrect. Acces refuse.")
        return

    # --- Authentifié : affichage du code ---
    st.markdown("""<div class="mh">
    <h1>📜 Code de calcul</h1>
    <span class="db">Extraction dynamique — Mise a jour automatique</span>
    </div>""", unsafe_allow_html=True)

    st.markdown("""<div style="background:#ebf8ff;border:1px solid #bee3f8;border-radius:8px;padding:10px 14px;margin-bottom:10px;font-size:13px;color:#2c5282">
    ℹ️ <strong>Remarque :</strong> Cette page affiche le code source reel des fonctions de calcul en temps reel.
    Toute modification apportee aux fonctions dans le code de l'application sera automatiquement reflétee ici.
    </div>""", unsafe_allow_html=True)

    if st.button("🔒 Deconnecter", key="btn_deco_code"):
        st.session_state.code_auth_ok = False
        st.rerun()

    # ---- SECTION 1 : Constantes et configurations ----
    st.markdown("""<div class="code-page-section">
    <h3>1️⃣ Constantes et configurations</h3>
    <p class="code-page-desc">Listes de KPI, cibles, actions recommandees et mots cles utilises dans les calculs.</p>
    <h4>Liste des KPI de Qualite (QK)</h4>""", unsafe_allow_html=True)
    st.code("QK = " + pprint.pformat(QK, width=120), language="python")

    st.markdown("<h4>Liste des KPI de Performance (PK)</h4>", unsafe_allow_html=True)
    st.code("PK = " + pprint.pformat(PK, width=120), language="python")

    st.markdown("<h4>Cibles par KPI (CIBLE)</h4>", unsafe_allow_html=True)
    st.code("CIBLE = " + pprint.pformat(CIBLE, width=120), language="python")

    st.markdown("<h4>KPI ou 'plus bas = mieux' (LOWER_BETTER)</h4>", unsafe_allow_html=True)
    st.code("LOWER_BETTER = " + pprint.pformat(LOWER_BETTER, width=120), language="python")

    st.markdown("<h4>Actions recommandees par KPI (ACT_MAP)</h4>", unsafe_allow_html=True)
    st.code("ACT_MAP = " + pprint.pformat(ACT_MAP, width=120), language="python")

    st.markdown("<h4>Mots cles preparation (MP_KW)</h4>", unsafe_allow_html=True)
    st.code("MP_KW = " + pprint.pformat(MP_KW, width=120), language="python")

    st.markdown("<h4>Mots cles planification (MPLAN_KW)</h4>", unsafe_allow_html=True)
    st.code("MPLAN_KW = " + pprint.pformat(MPLAN_KW, width=120), language="python")

    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 2 : Code des colonnes ajoutees ----
    st.markdown("""<div class="code-page-section">
    <h3>2️⃣ Code des colonnes ajoutees</h3>
    <p class="code-page-desc">Fonction <code>creer_colonnes_additionnelles(df, now)</code> — cree toutes les colonnes derivees utilisees pour les calculs KPI.</p>
    <p class="code-page-desc"><strong>Colonnes generees :</strong><br>
    • <code>Backlog preparation</code> : CARACTERISE / NON CARACTERISE (selon MP_KW)<br>
    • <code>Backlog planification</code> : CARACTERISE / NON CARACTERISE (selon MPLAN_KW)<br>
    • <code>amp</code> / <code>ap</code> : age en mois preparation / categorie age<br>
    • <code>amlp</code> / <code>alp</code> : age en mois planification / categorie age<br>
    • <code>amex</code> / <code>aex</code> : age en mois execution / categorie age<br>
    • <code>OT CONFIME</code> : OUI / NON (CLO + CONF dans Statut systeme)<br>
    • <code>Contient SOPL</code> : 1 / 0 (SOPL dans Statut utilisateur)<br>
    • <code>OT LANC ESTIME</code> : OUI / NON (Total couts budgetes != 0)<br>
    • <code>OT_COR_EGAL</code> : OUI / NON (couts budgetes == couts reels)</p>""", unsafe_allow_html=True)
    try:
        src = inspect.getsource(creer_colonnes_additionnelles)
        st.code(src, language="python")
    except Exception as e:
        st.warning(f"Impossible d'extraire le source: {e}")
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 3 : Fonctions utilitaires de calcul ----
    st.markdown("""<div class="code-page-section">
    <h3>3️⃣ Fonctions utilitaires de calcul</h3>
    <p class="code-page-desc">Fonctions de base utilisees dans les formules KPI.</p>
    <h4>contient_mot(t, lm)</h4>
    <p class="code-page-desc">Verifie si le texte contient un mot cle de la liste.</p>""", unsafe_allow_html=True)
    try: st.code(inspect.getsource(contient_mot), language="python")
    except Exception: pass

    st.markdown("<h4>cat_age(a)</h4><p class=\"code-page-desc\">Categorise l'age en 3 classes.</p>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(cat_age), language="python")
    except Exception: pass

    st.markdown("<h4>ckpi(n, d, sz=100)</h4><p class=\"code-page-desc\">Calcul du pourcentage KPI avec protection division par zero.</p>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(ckpi), language="python")
    except Exception: pass

    st.markdown("<h4>cpiv(df, f, c, p)</h4><p class=\"code-page-desc\">Cree un pivot table count par poste et colonne.</p>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(cpiv), language="python")
    except Exception: pass

    st.markdown("<h4>excr(df)</h4><p class=\"code-page-desc\">Exclut les lignes 'cresseur'.</p>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(excr), language="python")
    except Exception: pass

    st.markdown("<h4>get_caract_type(statut_user, keywords)</h4><p class=\"code-page-desc\">Determine le type de caracterisation.</p>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(get_caract_type), language="python")
    except Exception: pass

    st.markdown("<h4>is_lb(k)</h4><p class=\"code-page-desc\">Indique si le KPI est 'plus bas c'est mieux'.</p>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(is_lb), language="python")
    except Exception: pass

    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 4 : Formules de calcul des KPI ----
    st.markdown("""<div class="code-page-section">
    <h3>4️⃣ Formules de calcul des KPI</h3>
    <p class="code-page-desc">Fonction principale <code>calc_kpis(df_i, av_i, now, posts)</code> — calcule l'ensemble des 16 KPI par poste de travail.</p>
    <p class="code-page-desc"><strong>Detail des formules :</strong><br><br>
    <table style="width:100%;font-size:12px;border-collapse:collapse">
    <tr style="background:#1e3a5f;color:#fff"><th style="padding:6px;text-align:left">KPI</th><th style="padding:6px;text-align:left">Formule</th><th style="padding:6px;text-align:left">Filtre</th></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">TAUX_REALISATION_CORRECTIF/PT</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(TCLO, Total) = (TCLO/Total)*100</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">Appel=0</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT preparation &lt;1 mois</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(&lt;1 mois, Total prep)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">Statut OT=CREÉ</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT preparation &gt;3 mois</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(&gt;3 mois, Total prep, 0)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">Statut OT=CREÉ</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT preparation 1m&lt;&lt;3m</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(1m&lt;&lt;3m, Total prep, 0)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">Statut OT=CREÉ</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT planification &lt;1 mois</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(&lt;1 mois, Total planif)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">LANC, SOPL=0</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT planification &gt;3 mois</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(&gt;3 mois, Total planif, 0)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">LANC, SOPL=0</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT planification 1m&lt;&lt;3m</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(1m&lt;&lt;3m, Total planif, 0)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">LANC, SOPL=0</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT execution &lt;1 mois</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(&lt;1 mois, Total exec)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">LANC, SOPL=1</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT execution &gt;3 mois</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(&gt;3 mois, Total exec, 0)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">LANC, SOPL=1</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT execution 1m&lt;&lt;3m</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(1m&lt;&lt;3m, Total exec, 0)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">LANC, SOPL=1</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">appel avis approuve</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(APRV, Total avis)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">Avis sans Ordre</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT LANC ESTIME</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(OUI, Total LANC)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">LANC, Couts budgetes>0</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">Backlog prep. caracterise</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(CARACTERISE, Total CREÉ)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">Statut OT=CREÉ</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">Backlog planif. caracterise</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(CARACTERISE, Total LANC)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">Statut OT=LANC</td></tr>
    <tr><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0;font-weight:600">OT CONFIME</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">ckpi(OUI, Total)</td><td style="padding:4px 6px;border-bottom:1px solid #e2e8f0">CLO+CONF</td></tr>
    <tr><td style="padding:4px 6px;font-weight:600">OT_COR_EGAL</td><td style="padding:4px 6px">ckpi(OUI, Total)</td><td style="padding:4px 6px">Couts budgetes == reels</td></tr>
    </table></p>""", unsafe_allow_html=True)
    try:
        src = inspect.getsource(calc_kpis)
        st.code(src, language="python")
    except Exception as e:
        st.warning(f"Impossible d'extraire le source: {e}")
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 5 : Methode de calcul du score ----
    st.markdown("""<div class="code-page-section">
    <h3>5️⃣ Methode de calcul du score</h3>
    <p class="code-page-desc">
    <strong>Principe :</strong> Chaque KPI est evalue individuellement par la fonction <code>gscore(k, a, t)</code> qui retourne <strong>1</strong> si l'objectif est atteint, <strong>0</strong> sinon.<br><br>
    <strong>Score Performance</strong> = (somme des gscore pour les 10 KPI de QK) / 10 × 100<br>
    <strong>Score Qualite</strong> = (somme des gscore pour les 6 KPI de PK) / 6 × 100<br><br>
    <strong>Seuils d'atteinte par KPI (dans gscore) :</strong><br>
    • OT preparation/planification/execution &lt;1 mois : atteint si ≥ 75%<br>
    • OT preparation/planification/execution 1m&lt;&lt;3m : atteint si ≤ 15%<br>
    • OT preparation/planification/execution &gt;3 mois : atteint si ≤ 5%<br>
    • TAUX_REALISATION_CORRECTIF/PT : atteint si ≥ 80%<br>
    • appel avis approuvé : atteint si ≥ 90%<br>
    • OT LANC ESTIME, Backlog prep/planif, OT CONFIME, OT_COR_EGAL : atteint si ≥ 95%<br><br>
    <strong>Couleurs du score global (dans cs) :</strong><br>
    • Vert : ≥ 90%<br>
    • Jaune : ≥ 80% et &lt; 90%<br>
    • Rouge : &lt; 80%
    </p>
    <h4>gscore(k, a, t) — Score individuel par KPI</h4>""", unsafe_allow_html=True)
    try: st.code(inspect.getsource(gscore), language="python")
    except Exception: pass

    st.markdown("<h4>cs(v) — Couleur du score global</h4>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(cs), language="python")
    except Exception: pass

    st.markdown("""<h4>Calcul du score par poste (extrait du dashboard)</h4>
    <p class="code-page-desc">Pour chaque poste, le score est calcule ainsi :</p>""", unsafe_allow_html=True)
    st.code("""# Score Performance par poste
pscores[poste] = (
    sum(gscore(k, ckdf.loc[poste, k], CIBLE[k]) for k in QK if k in ckdf.columns)
    / len(QK) * 100
)

# Score Qualite par poste
qscores[poste] = (
    sum(gscore(k, ckdf.loc[poste, k], CIBLE[k]) for k in PK if k in ckdf.columns)
    / len(PK) * 100
)""", language="python")

    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 6 : Seuils de couleur KPI ----
    st.markdown("""<div class="code-page-section">
    <h3>6️⃣ Seuils de couleur par KPI</h3>
    <p class="code-page-desc">Fonction <code>ks(v, c)</code> — determine la couleur de la cellule selon la valeur et le KPI.</p>
    <p class="code-page-desc"><strong>Regles :</strong><br>
    • OT prep/planif/exec &lt;1 mois : 🟢 ≥80, 🟡 ≥75, 🔴 &lt;75<br>
    • OT prep/planif/exec 1m&lt;&lt;3m : 🟢 ≤15, 🔴 &gt;15<br>
    • OT prep/planif/exec &gt;3 mois : 🟢 ≤5, 🔴 &gt;5<br>
    • TAUX_REALISATION : 🟢 ≥85, 🟡 ≥80, 🔴 &lt;80<br>
    • appel avis approuvé : 🟢 ≥95, 🟡 ≥90, 🔴 &lt;90<br>
    • OT LANC ESTIME, Backlog, CONFIME, COR_EGAL : 🟢 ≥100, 🟡 ≥95, 🔴 &lt;95
    </p>""", unsafe_allow_html=True)
    try: st.code(inspect.getsource(ks), language="python")
    except Exception: pass

    st.markdown("<h4>kas(v) — Couleur des anomalies</h4>", unsafe_allow_html=True)
    try: st.code(inspect.getsource(kas), language="python")
    except Exception: pass

    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 7 : Filtres d'anomalies ----
    st.markdown("""<div class="code-page-section">
    <h3>7️⃣ Filtres d'anomalies</h3>
    <p class="code-page-desc">Fonction <code>get_anomalie_filters()</code> — retourne les filtres lambda utilises pour identifier les OT en anomalie pour chaque KPI.</p>""", unsafe_allow_html=True)
    try: st.code(inspect.getsource(get_anomalie_filters), language="python")
    except Exception: pass
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 8 : Fonctions de classification ----
    st.markdown("""<div class="code-page-section">
    <h3>8️⃣ Fonctions de classification (Metier, Atelier, Division)</h3>
    <p class="code-page-desc">Fonctions utilisees pour les filtres et l'analyse par categorie.</p>""", unsafe_allow_html=True)
    for fn in [get_metier, get_atelier, get_division]:
        st.markdown(f"<h4>{fn.__name__}(p)</h4>", unsafe_allow_html=True)
        try: st.code(inspect.getsource(fn), language="python")
        except Exception: pass
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- SECTION 9 : Calcul de l'age ----
    st.markdown("""<div class="code-page-section">
    <h3>9️⃣ Formule de calcul de l'age</h3>
    <p class="code-page-desc">L'age est calcule en mois ecoules entre la date de reference et la date de l'evenement :</p>""", unsafe_allow_html=True)
    st.code("""# Formule d'age en mois
age_mois = (now.year - date_evenement.year) * 12 + (now.month - date_evenement.month)

# Categorisation
# si age <= 1 mois  => "<1 mois"
# si age >= 3 mois  => ">3 mois"
# sinon             => "1 mois < <3 mois"

# Les 3 dates utilisees :
# - Preparation : colonne "Créé le"
# - Planification : colonne "Date de début planifiée" (filtre SOPL=0)
# - Execution : colonne "Date de début planifiée" (filtre SOPL=1)""", language="python")
    st.markdown("</div>", unsafe_allow_html=True)

    # ---- Pied de page ----
    st.markdown("""<div style="text-align:center;padding:16px;color:#718096;font-size:12px;border-top:1px solid #e2e8f0;margin-top:10px">
    📜 Code de calcul — Extraction dynamique via <code>inspect.getsource()</code><br>
    Derniere extraction : %s — Toute modification du code source est automatiquement reflétée.
    </div>""" % datetime.now().strftime("%d/%m/%Y %H:%M:%S"), unsafe_allow_html=True)


# ============================================================
# FONCTIONS HTML POUR LE DASHBOARD
# ============================================================

def html_table(rows, cols, tc, sc_col=None):
    h = '<table class="tw %s"><thead><tr>' % tc + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
    for r in rows:
        rc = "cb" if r.get("_t") == "cible" else ("tr" if r.get("_t") == "total" else "")
        h += '<tr class="%s">' % rc
        for c in cols:
            v = r.get(c, "")
            if r.get("_t") == "cible":
                h += '<td>%s</td>' % v
            else:
                s = cs(v) if sc_col and c in sc_col else ks(v, c)
                h += '<td style="%s">%s</td>' % (s or "", v)
        h += '</tr>'
    return h + '</tbody></table>'

def html_ano(rows, cols):
    h = '<table class="tw at"><thead><tr>' + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
    for r in rows:
        h += '<tr class="%s">' % ("tr" if r.get("_t") == "total" else "")
        for c in cols:
            v = r.get(c, "")
            h += '<td style="%s">%s</td>' % (kas(v) or "", v)
        h += '</tr>'
    return h + '</tbody></table>'

def html_actions_table(kpi_list, actuals, targets, act_map):
    h = '<table class="tw at"><thead><tr><th>KPI</th><th>Valeur Actuelle</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action Recommandée</th></tr></thead><tbody>'
    for k in kpi_list:
        av = actuals.get(k, 0); tv = targets.get(k, 100); diff = av - tv
        met = av <= tv if is_lb(k) else av >= tv
        status = "ATTEINT" if met else "NON ATTEINT"
        st_s = "background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
        ec_clr = "#276749" if met else "#c53030"
        action = "Objectif atteint" if met else act_map.get(k, "")
        h += '<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td></tr>' % (k, av, tv, ec_clr, diff, st_s, status, action)
    return h + '</tbody></table>'

def html_classement(scores, accent):
    sp = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    met_p = [(p, s) for p, s in sp if s >= 80]; not_p = [(p, s) for p, s in sp if s < 80]
    t5 = met_p[:5]; b5 = not_p[-5:] if len(not_p) > 5 else not_p
    h = '<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
    if t5:
        for i, (p, s) in enumerate(t5):
            h += '<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (accent, i + 1, p, cs("%.2f" % s), s)
    else:
        h += '<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
    h += '</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
    if b5:
        for i, (p, s) in enumerate(reversed(b5)):
            h += '<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (len(b5) - i, p, cs("%.2f" % s), s)
    else:
        h += '<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
    h += '</div></div>'; return h

def html_kpi_bars(kpi_list, actuals, targets, title, color_ok, color_fail):
    h = '<div class="ca"><div class="ct" style="color:%s">%s</div>' % (color_ok, title)
    for k in kpi_list:
        av = actuals.get(k, 0); tv = targets.get(k, 100); met = av <= tv if is_lb(k) else av >= tv
        bw = min(max(av, 0), 100); bg = color_ok if met else color_fail
        h += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>' % (k, bw, bg, av)
    return h + '</div>'

def html_bars(data, title, color):
    h = '<div class="ca"><div class="ct" style="color:%s">%s</div>' % (color, title)
    for label, val in sorted(data, key=lambda x: x[1], reverse=True):
        bw = min(max(val, 0), 100)
        h += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>' % (label, bw, color, val)
    return h + '</div>'

def html_grouped_bars(posts, pscores, qscores, title):
    h = '<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>' % title
    h += '<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
    for p in sorted(posts, key=lambda x: (pscores.get(x, 0) + qscores.get(x, 0)) / 2, reverse=True):
        pv, qv = pscores.get(p, 0), qscores.get(p, 0)
        h += '<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>' % (p, min(max(pv, 0), 100), pv, min(max(qv, 0), 100), qv)
    return h + '</div>'

def anl_pie_chart(data, names_col, values_col, title, colors=None):
    if data.empty: return None
    fig = px.pie(data, names=names_col, values=values_col, title=title,
                 color_discrete_sequence=colors or px.colors.qualitative.Set2)
    fig.update_traces(textposition='inside', textinfo='percent+label+value', textfont_size=12)
    fig.update_layout(margin=dict(t=50, b=20, l=20, r=20), height=450, autosize=True,
                      title_font_size=15, legend=dict(font_size=12, orientation="h", yanchor="bottom", y=-0.15))
    return fig

def export_btn(df, filename):
    buf = io.BytesIO(); df.to_excel(buf, index=False, engine='openpyxl'); buf.seek(0)
    st.download_button("📥 Exporter Excel", data=buf, file_name=filename,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# MAIN — DASHBOARD
# ============================================================

def main():
    try: locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL, 'fr_FR')
        except Exception: pass

    # Navigation
    if "current_page" not in st.session_state:
        st.session_state.current_page = "dashboard"

    # Ecran HSE
    if "hse_affiche" not in st.session_state:
        st.session_state.hse_affiche = False
    if not st.session_state.hse_affiche:
        c = random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>""" % c, unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche = True; st.rerun(); st.stop()

    # Si la page "Code de calcul" est demandee
    if st.session_state.current_page == "code_calcul":
        page_code_de_calcul()
        return

    # ===== DASHBOARD =====
    inject_custom_css()
    fichier_date = get_date_from_file()

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""", unsafe_allow_html=True)

        # Bouton Code de calcul
        st.markdown("---")
        if st.button("📜 Code de calcul", key="btn_go_code", use_container_width=True,
                     type="secondary"):
            st.session_state.current_page = "code_calcul"
            st.rerun()

        st.markdown("---")
        show_filters = st.checkbox("Afficher les filtres", value=True, key="show_filters")
        if show_filters:
            unf = st.toggle("📁 Charger nouveaux fichiers", value=False, key="tf")
            ot_f = av_f = None; apm = []
            if unf:
                ot_f = st.file_uploader("Fichier OT", type=["xlsx"], key="uot")
                av_f = st.file_uploader("Fichier AVIS", type=["xlsx"], key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t = excr(pd.read_excel("ot.xlsx"))
                        apm = sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1", "SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>""" % fichier_date, unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp = st.multiselect("Poste", ["All"] + apm, ["All"], key="sp")
            st.markdown("**🏭 Atelier**")
            sa = st.multiselect("Atelier", ["All", "Sulfurique (PS)", "Phosphorique (PP)", "Engrais (TSP/REX)", "Feed (MCP/DCP)"], ["All"], key="sa")
            st.markdown("**🏢 Division**")
            sd = st.multiselect("Division", ["All", "SF1", "SF2"], ["All"], key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr = st.date_input("Date debut planifiee", value=(datetime(2025, 1, 1).date(), datetime.today().date()), format="DD/MM/YYYY", key="dr")
        else:
            unf = False; ot_f = av_f = None; apm = []; sp = ["All"]; sa = ["All"]; sd = ["All"]
            dr = (datetime(2025, 1, 1).date(), datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try:
                    _t = excr(pd.read_excel("ot.xlsx"))
                    apm = sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1", "SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception: pass

    # ===================== DATA LOADING =====================
    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf: raw_ot = pd.read_excel(ot_f); raw_av = pd.read_excel(av_f)
            else: raw_ot = pd.read_excel("ot.xlsx"); raw_av = pd.read_excel("avis.xlsx")
            raw_ot = excr(raw_ot); raw_av = excr(raw_av)
            for c in ["Créé le", "Date de début planifiée", "Date de clôture", "Début réel", "Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c] = pd.to_datetime(raw_ot[c], errors="coerce")
            for c in ["Créé le", "Début souhaité", "Date de la clôture"]:
                if c in raw_av.columns: raw_av[c] = pd.to_datetime(raw_av[c], errors="coerce")
            if not apm:
                apm = sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1", "SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp = apm
            if "All" in sa or not sa: sa = ["All"]
            if "All" in sd or not sd: sd = ["All"]
            sdt = pd.to_datetime(dr[0]) if len(dr) == 2 else pd.to_datetime(datetime(2025, 1, 1))
            edt = pd.to_datetime(dr[1]) if len(dr) == 2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p = str(poste).upper()
                if "All" not in sa:
                    m = False
                    if "Sulfurique (PS)" in sa and "PS" in p: m = True
                    if "Phosphorique (PP)" in sa and "PP" in p: m = True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m = True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m = True
                    if not m: return False
                if "All" not in sd:
                    m = False
                    if "SF1" in sd and "SF1" in p: m = True
                    if "SF2" in sd and "SF2" in p: m = True
                    if not m: return False
                return True

            vp = [p for p in apm if mf(p) and p in sp]
            df = raw_ot[(raw_ot["Poste travail princ."].isin(vp)) & (raw_ot["Date de début planifiée"].between(sdt, edt))].copy()
            avdf = raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df = excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1", "SF2"), na=False)].drop_duplicates())
            avdf = excr(avdf[(avdf["Ordre"].isna()) | (avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns:
                df["Statut OT"] = df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
            df_dash = raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash = excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1", "SF2"), na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns:
                df_dash["Statut OT"] = df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now = pd.Timestamp.now()
            res = calc_kpis(df, avdf, now, vp); ckdf = res['ckdf']; dfp = res['dfp']
            res_d = calc_kpis(df_dash, avdf, now, vp); ckdf_d = res_d['ckdf']
            pa = {k: round(ckdf[k].mean(), 2) for k in QK}; qa = {k: round(ckdf[k].mean(), 2) for k in PK}
            pa_d = {k: round(ckdf_d[k].mean(), 2) for k in QK}; qa_d = {k: round(ckdf_d[k].mean(), 2) for k in PK}
            pscores = {}; qscores = {}
            for poste in ckdf.index:
                r = ckdf.loc[poste]
                pscores[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in QK if k in r.index) / len(QK) * 100) if QK else 0
                qscores[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in PK if k in r.index) / len(PK) * 100) if PK else 0
            pscores_d = {}; qscores_d = {}
            for poste in ckdf_d.index:
                r = ckdf_d.loc[poste]
                pscores_d[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in QK if k in r.index) / len(QK) * 100) if QK else 0
                qscores_d[poste] = (sum(gscore(k, r[k], CIBLE[k]) for k in PK if k in r.index) / len(PK) * 100) if PK else 0

            # Anomalies
            sub_p, sub_q = get_anomalie_filters()
            all_ano = []
            ano_p_data = {}; ano_q_data = {}
            for k, filt in sub_p.items():
                try:
                    fd = filt(dfp)
                    gp = fd.groupby("Poste travail princ.")["Ordre"].count().reindex(vp, fill_value=0)
                    ano_p_data[k] = gp
                    all_ano.append({"KPI": k, "Nb anomalies": int(gp.sum()), "Type": "Performance"})
                except Exception: pass
            for k, filt in sub_q.items():
                try:
                    fd = filt(dfp)
                    gp = fd.groupby("Poste travail princ.")["Ordre"].count().reindex(vp, fill_value=0)
                    ano_q_data[k] = gp
                    all_ano.append({"KPI": k, "Nb anomalies": int(gp.sum()), "Type": "Qualite"})
                except Exception: pass

            # Construction tables
            pcols = ["Poste de travail"] + QK + ["Score Performance"]
            prows = []
            for p in ckdf.index:
                r = {"Poste de travail": p, "_t": ""}
                for k in QK: r[k] = round(ckdf.loc[p, k], 1) if p in ckdf.index and k in ckdf.columns else 0
                r["Score Performance"] = round(pscores.get(p, 0), 2)
                prows.append(r)
            prows.append({"Poste de travail": "CIBLE", "_t": "cible",
                          **{k: CIBLE.get(k, "") for k in QK}, "Score Performance": "≥80%"})
            tot_r = {"Poste de travail": "Total general", "_t": "total"}
            for k in QK: tot_r[k] = round(pa.get(k, 0), 1)
            tot_r["Score Performance"] = round(sum(pscores.values()) / len(pscores) * 100, 2) if pscores else 0
            prows.append(tot_r)

            qcols = ["Poste de travail"] + PK + ["Score Qualite"]
            qrows = []
            for p in ckdf.index:
                r = {"Poste de travail": p, "_t": ""}
                for k in PK: r[k] = round(ckdf.loc[p, k], 1) if p in ckdf.index and k in ckdf.columns else 0
                r["Score Qualite"] = round(qscores.get(p, 0), 2)
                qrows.append(r)
            qrows.append({"Poste de travail": "CIBLE", "_t": "cible",
                          **{k: CIBLE.get(k, "") for k in PK}, "Score Qualite": "≥80%"})
            tot_q = {"Poste de travail": "Total general", "_t": "total"}
            for k in PK: tot_q[k] = round(qa.get(k, 0), 1)
            tot_q["Score Qualite"] = round(sum(qscores.values()) / len(qscores) * 100, 2) if qscores else 0
            qrows.append(tot_q)

            # Tables anomalies
            ano_p_cols = ["Poste de travail"] + list(ano_p_data.keys()) + ["Total"]
            ano_p_rows = []
            for p in vp:
                r = {"Poste de travail": p, "_t": ""}
                t = 0
                for k in ano_p_data: r[k] = int(ano_p_data[k].get(p, 0)); t += r[k]
                r["Total"] = t; ano_p_rows.append(r)
            tot_ap = {"Poste de travail": "Total", "_t": "total"}
            ta = 0
            for k in ano_p_data: tot_ap[k] = int(ano_p_data[k].sum()); ta += tot_ap[k]
            tot_ap["Total"] = ta; ano_p_rows.append(tot_ap)

            ano_q_cols = ["Poste de travail"] + list(ano_q_data.keys()) + ["Total"]
            ano_q_rows = []
            for p in vp:
                r = {"Poste de travail": p, "_t": ""}
                t = 0
                for k in ano_q_data: r[k] = int(ano_q_data[k].get(p, 0)); t += r[k]
                r["Total"] = t; ano_q_rows.append(r)
            tot_aq = {"Poste de travail": "Total", "_t": "total"}
            ta = 0
            for k in ano_q_data: tot_aq[k] = int(ano_q_data[k].sum()); ta += tot_aq[k]
            tot_aq["Total"] = ta; ano_q_rows.append(tot_aq)

            # Sauvegarde Excel
            save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_rows, ano_p_cols, ano_q_rows, ano_q_cols, fichier_date)

            # Historique et tendances
            hist_path = os.path.join("kpis", "indicateurs_kpis.xlsx")
            hist_df = load_historical_kpis(hist_path)
            var_df = calculate_variations(hist_df)
            journal_df = generate_journal(var_df)
            top5_df, bot5_df = calculate_rankings(var_df)

            # ===== AFFICHAGE DASHBOARD =====
            avg_ps = round(sum(pscores.values()) / len(pscores) * 100, 2) if pscores else 0
            avg_qs = round(sum(qscores.values()) / len(qscores) * 100, 2) if qscores else 0
            total_ano = sum(a["Nb anomalies"] for a in all_ano)
            nb_postes = len(vp)

            st.markdown("""<div class="mh"><h1>📊 Dashboard KPI — Maintenance</h1><span class="db">📅 %s</span></div>""" % fichier_date, unsafe_allow_html=True)
            st.markdown("""<div class="cr">
            <div class="cc c1"><div class="cv">%s%%</div><div class="cl">Score Performance</div></div>
            <div class="cc c2"><div class="cv">%s%%</div><div class="cl">Score Qualite</div></div>
            <div class="cc c3"><div class="cv">%s</div><div class="cl">Postes</div></div>
            <div class="cc c4"><div class="cv">%s</div><div class="cl">Anomalies</div></div>
            </div>""" % (avg_ps, avg_qs, nb_postes, total_ano), unsafe_allow_html=True)

            tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
                "📊 Vue d'ensemble", "⚡ Performance", "✅ Qualite",
                "🚨 Anomalies", "🏆 Classement", "📋 Actions", "📈 Tendances"])

            with tab1:
                st.markdown('<div class="stl p">Barres de progression — Performance</div>', unsafe_allow_html=True)
                st.markdown(html_kpi_bars(QK, pa, CIBLE, "KPI de Performance (moyenne globale)", "#2b6cb0", "#e53e3e"), unsafe_allow_html=True)
                st.markdown('<div class="stl q">Barres de progression — Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_kpi_bars(PK, qa, CIBLE, "KPI de Qualite (moyenne globale)", "#276749", "#e53e3e"), unsafe_allow_html=True)
                st.markdown('<div class="stl s">Comparaison par poste</div>', unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp, pscores, qscores, "Score Performance vs Qualite par poste"), unsafe_allow_html=True)

                # Graphiques
                if not dfp.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st_pie = dfp[dfp["Nº appel pl.entret."].fillna(0) == 0].groupby("Statut OT")["Ordre"].count().reset_index()
                        st_pie.columns = ["Statut", "Nombre"]
                        if not st_pie.empty:
                            fig = anl_pie_chart(st_pie, "Statut", "Nombre", "Repartition OT Correctif/PT")
                            if fig: st.plotly_chart(fig, use_container_width=True)
                    with c2:
                        metier_data = dfp["Poste travail princ."].apply(get_metier).value_counts().reset_index()
                        metier_data.columns = ["Metier", "Nombre"]
                        if not metier_data.empty:
                            fig2 = anl_pie_chart(metier_data, "Metier", "Nombre", "Repartition par Metier")
                            if fig2: st.plotly_chart(fig2, use_container_width=True)

            with tab2:
                st.markdown('<div class="stl p">Indicateurs de Performance</div>', unsafe_allow_html=True)
                st.markdown(html_table(prows, pcols, "pt", sc_col={"Score Performance"}), unsafe_allow_html=True)
                export_btn(pd.DataFrame(prows).drop(columns=["_t"], errors="ignore"), "performance_kpi.xlsx")

            with tab3:
                st.markdown('<div class="stl q">Indicateurs de Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_table(qrows, qcols, "qt", sc_col={"Score Qualite"}), unsafe_allow_html=True)
                export_btn(pd.DataFrame(qrows).drop(columns=["_t"], errors="ignore"), "qualite_kpi.xlsx")

            with tab4:
                st.markdown('<div class="stl a">Anomalies Performance</div>', unsafe_allow_html=True)
                st.markdown(html_ano(ano_p_rows, ano_p_cols), unsafe_allow_html=True)
                st.markdown('<div class="stl a">Anomalies Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_ano(ano_q_rows, ano_q_cols), unsafe_allow_html=True)

            with tab5:
                st.markdown('<div class="stl c">Classement Performance</div>', unsafe_allow_html=True)
                st.markdown(html_classement(pscores, "#2b6cb0"), unsafe_allow_html=True)
                st.markdown('<div class="stl c">Classement Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_classement(qscores, "#276749"), unsafe_allow_html=True)
                # Top/Bottom 5 variations
                if not top5_df.empty:
                    st.markdown('<div class="stl s">Top 5 — Meilleures variations</div>', unsafe_allow_html=True)
                    st.dataframe(top5_df, use_container_width=True, hide_index=True)
                if not bot5_df.empty:
                    st.markdown('<div class="stl a">Bottom 5 — Variations negatives</div>', unsafe_allow_html=True)
                    st.dataframe(bot5_df, use_container_width=True, hide_index=True)

            with tab6:
                st.markdown('<div class="stl p">Actions Performance</div>', unsafe_allow_html=True)
                st.markdown(html_actions_table(QK, pa, CIBLE, ACT_MAP), unsafe_allow_html=True)
                st.markdown('<div class="stl q">Actions Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_actions_table(PK, qa, CIBLE, ACT_MAP), unsafe_allow_html=True)

            with tab7:
                if not journal_df.empty:
                    st.markdown('<div class="stl s">Journal des variations significatives (≥5%%)</div>', unsafe_allow_html=True)
                    jdisp = journal_df[["Date precedente", "Date actuelle", "Poste", "Type", "KPI",
                                        "Valeur precedente", "Valeur actuelle", "Ecart %", "Sens"]].copy()
                    st.dataframe(jdisp, use_container_width=True, hide_index=True)
                    export_btn(jdisp, "journal_variations.xlsx")
                else:
                    st.markdown('<div class="es">Aucune variation significative detectee (historique insuffisant).</div>', unsafe_allow_html=True)

                # Sparklines par KPI
                if not var_df.empty:
                    st.markdown('<div class="stl c">Evolution par KPI</div>', unsafe_allow_html=True)
                    spark_html = '<div class="spark-grid">'
                    for kpi in ALL_KPI:
                        kdf = var_df[var_df["KPI"] == kpi].sort_values("Date actuelle")
                        if len(kdf) >= 2:
                            vals = kdf["Valeur actuelle"].tolist()
                            dates = kdf["Date actuelle"].tolist()
                            fig_sp = go.Figure()
                            fig_sp.add_trace(go.Scatter(x=dates, y=vals, mode='lines+markers',
                                                        line=dict(width=2, color="#2b6cb0"),
                                                        marker=dict(size=5)))
                            if kpi in CIBLE:
                                fig_sp.add_hline(y=CIBLE[kpi], line_dash="dash", line_color="#e53e3e",
                                                 annotation_text="Cible: %s%%" % CIBLE[kpi])
                            fig_sp.update_layout(margin=dict(t=30, b=30, l=40, r=10), height=200,
                                                 font=dict(size=10), title=dict(text=kpi, font=dict(size=12)))
                            spark_html += '<div class="spark-card">'
                        else:
                            continue
                    spark_html += '</div>'
                    # Utiliser plotly directement pour chaque KPI
                    n_cols = 3
                    kpi_with_data = [k for k in ALL_KPI if len(var_df[var_df["KPI"] == k]) >= 2]
                    for i in range(0, len(kpi_with_data), n_cols):
                        cols = st.columns(n_cols)
                        for j in range(n_cols):
                            if i + j < len(kpi_with_data):
                                kpi = kpi_with_data[i + j]
                                kdf = var_df[var_df["KPI"] == kpi].sort_values("Date actuelle")
                                with cols[j]:
                                    fig_sp = go.Figure()
                                    fig_sp.add_trace(go.Scatter(x=kdf["Date actuelle"], y=kdf["Valeur actuelle"],
                                                                mode='lines+markers', line=dict(width=2, color="#2b6cb0"),
                                                                marker=dict(size=5)))
                                    if kpi in CIBLE:
                                        fig_sp.add_hline(y=CIBLE[kpi], line_dash="dash", line_color="#e53e3e",
                                                         annotation_text="Cible: %s%%" % CIBLE[kpi])
                                    fig_sp.update_layout(margin=dict(t=30, b=30, l=40, r=10), height=200,
                                                         font=dict(size=10),
                                                         title=dict(text=kpi, font=dict(size=11, color="#1e3a5f")))
                                    st.plotly_chart(fig_sp, use_container_width=True)

        except Exception as e:
            st.error(f"Erreur de chargement des donnees : {e}")
            st.exception(e)
    else:
        inject_custom_css()
        st.markdown("""<div class="mh"><h1>📊 Dashboard KPI — Maintenance</h1></div>""", unsafe_allow_html=True)
        st.markdown('<div class="es">Veuillez charger les fichiers OT et AVIS dans la barre laterale pour demarrer.</div>', unsafe_allow_html=True)


# ============================================================
if __name__ == "__main__":
    main()
