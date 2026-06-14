# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os, gc, re
from datetime import datetime
import plotly.express as px
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# -------- PAGE CONFIG --------
st.set_page_config(
    layout="wide",
    page_title="Dashboard KPI"
)

# -------- utilitaires de filesystem / date ----------
def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt", "r", encoding="utf-8") as f:
                return f.read().strip()
        except: pass
    return datetime.now().strftime("%d/%m/%Y")

# -------- sauvegarde excel -------------
def save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_r, ano_p_c, ano_q_r, ano_q_c, sheet_name):
    kpis_dir = "kpis"
    os.makedirs(kpis_dir, exist_ok=True)
    filepath = os.path.join(kpis_dir, "indicateurs_kpis.xlsx")
    sn = str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfl = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    tf = Font(bold=True, size=12, color="1E3A5F")
    tb = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb = load_workbook(filepath)
    except:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws = wb.create_sheet(sn)
    rn = 1
    def ws_section(title, cols, rows, sr):
        ws.cell(row=sr, column=1, value=title).font = tf; sr += 1
        for j, c in enumerate(cols, 1):
            cl = ws.cell(row=sr, column=j, value=c); cl.font = hf; cl.fill = hfl; cl.alignment = Alignment(horizontal='center'); cl.border = tb
        sr += 1
        for r in rows:
            for j, c in enumerate(cols, 1):
                cl = ws.cell(row=sr, column=j, value=r.get(c, "")); cl.border = tb; cl.alignment = Alignment(horizontal='center')
            sr += 1
        return sr + 1
    rn = ws_section("INDICATEURS DE PERFORMANCE", pcols, prows, rn)
    if ano_p_c and ano_p_r: rn = ws_section("ANOMALIES PERFORMANCE", ano_p_c, ano_p_r, rn)
    rn = ws_section("INDICATEURS DE QUALITE", qcols, qrows, rn)
    if ano_q_c and ano_q_r: rn = ws_section("ANOMALIES QUALITE", ano_q_c, ano_q_r, rn)
    try: wb.save(filepath)
    except: pass

# -------- CSS injection --------
def inject_custom_css():
    st.markdown("""<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{
        max-width: 100% !important;
        width: 100% !important;
        padding-left: 0.5rem !important;
        padding-right: 0.5rem !important;
        padding-top:.6rem;
        padding-bottom:.6rem;
    }
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:10px 16px;border-radius:var(--r);margin-bottom:4px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:16px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:2px 10px;border-radius:14px;color:#fff;font-size:10px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:4px;margin-bottom:4px}
    .cc{background:#fff;border-radius:var(--r);padding:8px 10px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:22px;font-weight:900;line-height:1}
    .cc .cl{font-size:7px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:1px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:11px;font-weight:700;color:var(--p);margin:4px 0 1px 0;padding-left:8px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:8px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:7px;text-transform:uppercase;letter-spacing:.3px;padding:3px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw tbody td{padding:2px 3px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:8px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:8px!important}
    .stTabs [data-baseweb="tab-list"]{gap:2px;background:#e2e8f0;padding:2px;border-radius:6px;margin-bottom:3px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:5px 10px;font-weight:600;font-size:10px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:4px 8px;background:#fff;border-radius:5px;margin-bottom:1px;border:1px solid var(--b);font-size:9px}
    .sr .sn{font-weight:700;color:var(--p);min-width:200px;font-size:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:2px 7px;border-radius:12px;font-weight:800;font-size:10px;min-width:40px;text-align:center;margin:0 6px;color:#fff}
    .sr .sa{color:#718096;font-size:8px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .stg{font-size:7px;color:#718096;min-width:50px;text-align:center;white-space:nowrap}
    .sr .sb{font-size:7px;font-weight:700;padding:1px 5px;border-radius:3px;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:8px;margin-top:2px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:10px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:3px;font-size:8px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:160px;font-weight:600;color:var(--p);text-align:right;padding-right:6px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:22px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:8px;font-weight:800;color:#1a202c;min-width:50px;text-align:right;padding-left:4px}
    .gbr{display:flex;align-items:center;padding:2px 0;font-size:8px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:140px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:7px}
    .gbr-g{display:flex;align-items:center;gap:3px;flex:1}
    .gbr-w{flex:1;height:18px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}
    .gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:7px;font-weight:800;min-width:42px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:12px;margin-bottom:4px;font-size:8px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:4px}
    .gbr-legend i{display:inline-block;width:12px;height:12px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:4px}
    .cg>div{background:#fff;border-radius:var(--r);padding:6px 8px;border:1px solid var(--b)}
    .cg .ct{font-size:9px;font-weight:700;margin-bottom:2px;padding-bottom:2px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:2px 0;font-size:8px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:14px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:45px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:4px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:6px 12px;font-weight:700;font-size:11px;width:100%}
    ::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:2px}
    section[data-testid="stSidebar"] { width: 250px !important; min-width: 250px !important; }
    section[data-testid="stSidebar"][aria-expanded="false"] { width: 0px !important; min-width: 0px !important; }
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:9px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:2px 6px;margin-bottom:2px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    div[data-testid="stSidebar"] .stCheckbox div[data-testid="stWidget"]{ background:rgba(255,255,255,.15)!important; border:1px solid rgba(255,255,255,.25)!important; }
    .es{text-align:center;padding:10px;color:#718096;font-size:10px}
    .anl-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:9px;margin:0}
    .anl-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:8px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0}
    .anl-tbl tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7}
    .anl-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .anl-tbl tbody tr:hover td{background:#ebf8ff!important}
    .anl-tbl .tot td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:13px}.mh .db{float:none;display:block;margin-top:2px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:100px}.gbr-l{width:90px}}
    </style>""", unsafe_allow_html=True)

# ============================================================
# -------- Lecture robuste des fichiers (local ou upload) --------
# ============================================================

def is_xlsx_buffer(f):
    """Vérifie si un file-like commence par la signature ZIP/PK (xlsx)."""
    try:
        pos = f.tell()
    except Exception:
        pos = None
    try:
        b = f.read(4)
        f.seek(0)
    except Exception:
        return False
    return b == b'PK\x03\x04'

def read_any_excel(source, usecols=None, parse_dates=None):
    """
    Lit un fichier Excel/CSV depuis un chemin (str) ou un UploadedFile/BytesIO.
    Gère les fichiers .xls renommés en .xlsx, CSV renommés, etc.
    """
    last_err = None

    # ---- Cas 1 : chemin fichier (string) ----
    if isinstance(source, str):
        ext = os.path.splitext(source)[1].lower()
        
        # 1) Si l'extension est .csv ou .txt, essayer CSV en priorité
        if ext in (".csv", ".txt"):
            for csv_sep in [";", ",", "\t", "|"]:
                try:
                    return pd.read_csv(source, usecols=usecols, parse_dates=parse_dates, sep=csv_sep, on_bad_lines='skip')
                except Exception as e:
                    last_err = e

        # 2) Moteurs Excel selon l'extension
        if ext == ".xls":
            engines = ["xlrd", "openpyxl"]
        elif ext == ".xlsx":
            engines = ["openpyxl", "xlrd"]   # openpyxl en premier, xlrd en secours si renommé
        else:
            engines = ["openpyxl", "xlrd"]   # extension inconnue, tenter les deux

        for eng in engines:
            try:
                return pd.read_excel(source, usecols=usecols, parse_dates=parse_dates, engine=eng)
            except Exception as e:
                last_err = e

        # 3) Si tous les moteurs Excel ont échoué, tenter les séparateurs CSV courants
        for csv_sep in [";", ",", "\t", "|"]:
            try:
                df = pd.read_csv(source, usecols=usecols, parse_dates=parse_dates, sep=csv_sep, on_bad_lines='skip')
                if df.shape[1] > 1:
                    return df
            except Exception:
                continue

        raise ValueError(f"Impossible de lire le fichier '{source}': {last_err}")

    # ---- Cas 2 : file-like (Streamlit UploadedFile, BytesIO, etc.) ----
    try:
        source.seek(0)
    except Exception:
        pass

    # Essai 1 : xlsx (openpyxl) si signature PK..
    if is_xlsx_buffer(source):
        try:
            source.seek(0)
            return pd.read_excel(source, usecols=usecols, parse_dates=parse_dates, engine="openpyxl")
        except Exception as e:
            last_err = e
            try:
                source.seek(0)
            except Exception:
                pass

    # Essai 2 : xls legacy (xlrd)
    try:
        source.seek(0)
        return pd.read_excel(source, usecols=usecols, parse_dates=parse_dates, engine="xlrd")
    except Exception as e:
        last_err = e
        try:
            source.seek(0)
        except Exception:
            pass

    # Essai 3 : csv avec séparateurs explicites (évite 'bad delimiter value')
    for csv_sep in [";", ",", "\t", "|"]:
        try:
            source.seek(0)
            df = pd.read_csv(source, usecols=usecols, parse_dates=parse_dates, sep=csv_sep, on_bad_lines='skip')
            if df.shape[1] > 1:
                return df
        except Exception:
            continue

    # Essai 4 : dernier recours, forcer openpyxl
    try:
        source.seek(0)
        return pd.read_excel(source, usecols=usecols, parse_dates=parse_dates, engine="openpyxl")
    except Exception as e:
        last_err = e

    raise ValueError(f"Aucun moteur ne peut lire ce fichier. Derniere erreur: {last_err}")


# ============================================================
# -------- Suite du programme --------
# ============================================================

def main():
    try: locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
    except:
        try: locale.setlocale(locale.LC_ALL, 'fr_FR')
        except: pass
    inject_custom_css()
    fichier_date = get_date_from_file()

    consignes = ["Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.","Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.","Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.","Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.","Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.","Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.","Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.","Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.","Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.","Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.","Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.","Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.","Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.","Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.","Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.","Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.","Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.","Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.","Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.","Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.","Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.","Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.","Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.","Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.","Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.","Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.","Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.","Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.","La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.","Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche = False
    if not st.session_state.hse_affiche:
        c = random.choice(consignes)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:42px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:18px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:28px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:32px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>""" % c, unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche = True; st.rerun(); st.stop()

    def contient_mot(t, lm):
        t = str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if pd.isna(a): return "Inconnu"
        if a <= 1: return "<1 mois"
        elif a >= 3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n, d, sz=100): return np.where(d == 0, sz, (n / d) * 100)
    def cpiv(df, f, c, p):
        return pd.pivot_table(df[f], index="Poste travail princ.", columns=c, values="Ordre", aggfunc="count", fill_value=0).reindex(p, fill_value=0)
    def excr(df):
        return df[~df["Poste travail princ."].astype(str).str.contains("cresseur", case=False, na=False)].copy() if "Poste travail princ." in df.columns else df
    def get_metier(p):
        p = str(p).upper()
        if "E" in p: return "Electrique"
        if "M" in p: return "Mecanique"
        if "R" in p: return "Instrumentation"
        if "G" in p: return "Genie Civil"
        return "Autre"
    def get_atelier(p):
        p = str(p).upper()
        if "PS" in p: return "Sulfurique"
        if "PP" in p: return "Phosphorique"
        if "TSP" in p or "REX" in p: return "Engrais"
        if "MCP" in p or "DCP" in p: return "Feed"
        return "Autre"
    def get_division(p):
        p = str(p).upper()
        if "SF1" in p: return "SF1"
        if "SF2" in p: return "SF2"
        return "Autre"

    def calc_kpis(df_i, av_i, now, posts):
        res = {}; df = df_i.copy(); av = av_i.copy()
        mp = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
        df["Backlog preparation"] = np.where(df["Statut utilisateur"].apply(lambda x: contient_mot(x, mp)), "CARACTERISE", "NON CARACTERISE")
        mplan = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
        df["Backlog planification"] = np.where(df["Statut utilisateur"].apply(lambda x: contient_mot(x, mplan)), "CARACTERISE", "NON CARACTERISE")
        for dc, am, ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc] = pd.to_datetime(df[dc], errors='coerce')
                df[am] = ((now.year - df[dc].dt.year)*12 + (now.month - df[dc].dt.month)).round(2)
                df[ac] = df[am].apply(cat_age)
            else: df[am] = np.nan; df[ac] = "Inconnu"
        df["OT CONFIME"] = np.where(df["Statut système"].str.contains("CLO", na=False) & df["Statut système"].str.contains("CONF", na=False), "OUI", "NON")
        df["Contient SOPL"] = df["Statut utilisateur"].str.contains("SOPL", na=False).map({True:1, False:0})
        df["OT LANC ESTIME"] = np.where(df["Total coûts budgétés"].fillna(0) == 0, "NON", "OUI")
        df["OT_COR_EGAL"] = np.where((df["Total coûts budgétés"].fillna(0) - df["Total coûts réels"].fillna(0)) == 0, "OUI", "NON")
        res['dfp'] = df
        an = cpiv(df, df["Nº appel pl.entret."].fillna(0)==0, "Statut OT", posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c] = an.get(c, 0)
        an["Total"] = an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1); an["TAUX_REALISATION_CORRECTIF/PT"] = ckpi(an["TCLO"], an["Total"])
        pr = cpiv(df, df["Statut OT"]=="CRÉÉ", "ap", posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c] = pr.get(c, 0)
        pr["Total"] = pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"] = ckpi(pr["<1 mois"], pr["Total"]); pr["OT préparation >3 mois"] = ckpi(pr[">3 mois"], pr["Total"], 0); pr["OT préparation 1mois< <3mois"] = ckpi(pr["1 mois < <3 mois"], pr["Total"], 0)
        pl = cpiv(df, (df["Statut OT"]=="LANC") & (df["Contient SOPL"]==0), "alp", posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c] = pl.get(c, 0)
        pl["Total"] = pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"] = ckpi(pl["<1 mois"], pl["Total"]); pl["OT planification >3 mois"] = ckpi(pl[">3 mois"], pl["Total"], 0); pl["OT planification 1mois< <3mois"] = ckpi(pl["1 mois < <3 mois"], pl["Total"], 0)
        ex = cpiv(df, (df["Statut OT"]=="LANC") & (df["Contient SOPL"]==1), "aex", posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c] = ex.get(c, 0)
        ex["Total"] = ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"] = ckpi(ex["<1 mois"], ex["Total"]); ex["OT exécution >3 mois"] = ckpi(ex[">3 mois"], ex["Total"], 0); ex["OT exécution 1mois< <3mois"] = ckpi(ex["1 mois < <3 mois"], ex["Total"], 0)
        la = pd.pivot_table(df[df["Statut OT"]=="LANC"], index="Poste travail princ.", columns="OT LANC ESTIME", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["OUI","NON"]: la[c] = la.get(c, 0)
        la["Total"] = la["OUI"]+la["NON"]; la["OT LANC ESTIME"] = ckpi(la["OUI"], la["Total"])
        pc = pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"], index="Poste travail princ.", columns="Backlog preparation", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c] = pc.get(c, 0)
        pc["Total"] = pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"] = ckpi(pc["CARACTERISE"], pc["Total"])
        plc = pd.pivot_table(df[df["Statut OT"]=="LANC"], index="Poste travail princ.", columns="Backlog planification", values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c] = plc.get(c, 0)
        plc["Total"] = plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"] = ckpi(plc["CARACTERISE"], plc["Total"])
        for kn, cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv = pd.pivot_table(df, index="Poste travail princ.", columns=cn, values="Ordre", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
            for c in ["OUI","NON"]: pv[c] = pv.get(c, 0)
            pv["Total"] = pv["OUI"]+pv["NON"]; pv[cn] = ckpi(pv["OUI"], pv["Total"]); res[kn.lower().replace(" ","_")] = pv
        avf = av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip()=="")].copy(); res['avf'] = avf
        tca = pd.pivot_table(avf, index="Poste travail princ.", columns="Statut utilisateur", values="Avis", aggfunc="count", fill_value=0).reindex(posts, fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c] = tca.get(c, 0)
        tca["Total"] = tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"] = ckpi(tca["APRV"], tca["Total"])
        res['ckdf'] = pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT": an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois": pr["OT préparation <1 mois"],"OT préparation >3 mois": pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois": pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois": pl["OT planification <1 mois"],"OT planification >3 mois": pl["OT planification >3 mois"],"OT planification 1mois< <3mois": pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois": ex["OT exécution <1 mois"],"OT exécution >3 mois": ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois": ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé": tca["appel avis approuvé"],"OT LANC ESTIME": la["OT LANC ESTIME"],
            "Backlog préparation caractérisé": pc["Backlog préparation caractérisé"],"Backlog planification caractérisé": plc["Backlog planification caractérisé"],
            "OT CONFIME": res['ot_confime']["OT CONFIME"],"OT_COR_EGAL": res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res

    def ks(v, c):
        try: val = float(v)
        except: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c == "TAUX_REALISATION_CORRECTIF/PT": return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c == "appel avis approuvé": return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val = float(str(v).replace(' %','').strip())
        except: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val = int(v)
        except: return ""
        if val == 0: return "color:#cbd5e0"
        if val <= 3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val <= 10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k, a, t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k == "TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k == "appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        return 0
    def is_lb(k): return k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois","OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

    def html_table(rows, cols, tc, sc_col=None):
        h = '<table class="tw %s"><thead><tr>' % tc + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
        for r in rows:
            rc = "cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h += '<tr class="%s">' % rc
            for c in cols:
                v = r.get(c, "")
                if r.get("_t")=="cible": h += '<td>%s</td>' % v
                else:
                    s = cs(v) if sc_col and c in sc_col else ks(v, c)
                    h += '<td style="%s">%s</td>' % (s or "", v)
            h += '</tr>'
        return h + '</tbody></table>'

    def html_ano(rows, cols):
        h = '<table class="tw at"><thead><tr>' + ''.join('<th>%s</th>' % c for c in cols) + '</tr></thead><tbody>'
        for r in rows:
            h += '<tr class="%s">' % ("tr" if r.get("_t")=="total" else "")
            for c in cols: v = r.get(c,""); h += '<td style="%s">%s</td>' % (kas(v) or "", v)
            h += '</tr>'
        return h + '</tbody></table>'

    def html_synth(kpi_list, actuals, targets, act_map, accent):
        h = ''
        for k in kpi_list:
            av, tv = actuals.get(k,0), targets.get(k,100)
            met = av <= tv if is_lb(k) else av >= tv
            sbg, sclr = ("#c6efce","#006100") if met else ("#ffc7ce","#9c0006")
            scbg = accent if met else "#e53e3e"
            act = "Objectif atteint" if met else act_map.get(k,"")
            h += '<div class="sr"><div class="sn">%s</div><div class="sc" style="background:%s">%.1f%%</div><div class="stg">Cible: %s%%</div><div class="sb" style="color:%s;background:%s">%s</div><div class="sa">%s</div></div>' % (k, scbg, av, tv, sclr, sbg, "ATTEINT" if met else "NON ATTEINT", act)
        return h

    def html_classement(scores, accent):
        sp = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        met_p, not_p = [(p,s) for p,s in sp if s>=80], [(p,s) for p,s in sp if s<80]
        t5, b5 = met_p[:5], not_p[-5:] if len(not_p)>5 else not_p
        h = '<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h += '<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (accent,i+1,p,cs("%.2f"%s),s)
        else: h += '<div style="padding:4px;font-size:8px;color:#718096">Aucun poste</div>'
        h += '</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h += '<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>' % (len(b5)-i,p,cs("%.2f"%s),s)
        else: h += '<div style="padding:4px;font-size:8px;color:#38a169">Tous atteints</div>'
        h += '</div></div>'; return h

    def html_kpi_bars(kpi_list, actuals, targets, title, color_ok, color_fail):
        h = '<div class="ca"><div class="ct" style="color:%s">%s</div>' % (color_ok, title)
        for k in kpi_list:
            av, tv = actuals.get(k,0), targets.get(k,100)
            met = av <= tv if is_lb(k) else av >= tv
            bw = min(max(av,0),100); bg = color_ok if met else color_fail
            h += '<div class="car"><div class="cal" style="width:240px">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>' % (k, bw, bg, av)
        h += '</div>'; return h

    def html_bars(data, title, color):
        h = '<div class="ca"><div class="ct" style="color:%s">%s</div>' % (color, title)
        for label, val in sorted(data, key=lambda x: x[1], reverse=True):
            bw = min(max(val,0),100)
            h += '<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>' % (label, bw, color, val)
        h += '</div>'; return h

    def html_grouped_bars(posts, pscores, qscores, title):
        h = '<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>' % title
        h += '<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        sp2 = sorted(posts, key=lambda x: (pscores.get(x,0)+qscores.get(x,0))/2, reverse=True)
        for p in sp2:
            pv, qv = pscores.get(p,0), qscores.get(p,0)
            pw, qw = min(max(pv,0),100), min(max(qv,0),100)
            h += '<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>' % (p, pw, pv, qw, qv)
        h += '</div>'; return h

    def anl_pie_chart(data, names_col, values_col, title, colors=None):
        if data.empty: return None
        fig = px.pie(data, names=names_col, values=values_col, title=title,
                     color_discrete_sequence=colors or px.colors.qualitative.Set2)
        fig.update_traces(textposition='inside', textinfo='percent+label+value', textfont_size=9)
        fig.update_layout(height=450, autosize=True, margin=dict(t=40,b=10,l=10,r=10), title_font_size=11,
                          legend=dict(font_size=8, orientation="h", yanchor="bottom", y=-0.15))
        return fig

    def anl_html_table(df_out, pct_col=None, pct_thresh=(80,60)):
        h = '<table class="anl-tbl"><thead><tr>'
        for c in df_out.columns: h += '<th>%s</th>' % c
        h += '</tr></thead><tbody>'
        for idx, row in df_out.iterrows():
            is_tot = str(idx) == "TOTAL" or row.iloc[0] == "TOTAL"
            rc = "tot" if is_tot else ""
            h += '<tr class="%s">' % rc
            for c in df_out.columns:
                v = row[c]; s = ""
                if pct_col and c == pct_col and not is_tot:
                    try:
                        pv = float(str(v).replace('%','').strip())
                        s = "g-green" if pv >= pct_thresh[0] else ("g-yellow" if pv >= pct_thresh[1] else "g-red")
                    except: pass
                if isinstance(v, float): v = round(v, 1)
                h += '<td class="%s">%s</td>' % (s, v)
            h += '</tr>'
        return h + '</tbody></table>'

    def export_btn(df, filename):
        buf = io.BytesIO()
        df.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)
        st.download_button("📥 Exporter Excel", data=buf, file_name=filename,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:18px;margin-bottom:2px">⚙️</div><div style="font-size:12px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:8px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""", unsafe_allow_html=True)
        show_filters = st.checkbox("Afficher les filtres", value=True, key="show_filters")
        if show_filters:
            st.markdown("---")
            unf = st.toggle("📁 Charger nouveaux fichiers", value=False, key="tf")
            ot_f = av_f = None; apm = []
            if unf:
                ot_f = st.file_uploader("Fichier OT", type=["xlsx","xls","csv"], key="uot")
                av_f = st.file_uploader("Fichier AVIS", type=["xlsx","xls","csv"], key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t = excr(read_any_excel("ot.xlsx"))
                        apm = sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:5px 8px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:7px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:10px;color:white;font-weight:600;margin-top:1px">📅 %s</div></div>""" % fichier_date, unsafe_allow_html=True)
            st.markdown("---")
            st.markdown("**🎯 Postes**")
            sp = st.multiselect("Poste", ["All"]+apm, ["All"], key="sp")
            st.markdown("**🏭 Atelier**")
            sa = st.multiselect("Atelier", ["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"], ["All"], key="sa")
            st.markdown("**🏢 Division**")
            sd = st.multiselect("Division", ["All","SF1","SF2"], ["All"], key="sd")
            st.markdown("---")
            st.markdown("**📅 Periode**")
            dr = st.date_input("Date debut planifiee", value=(datetime(2025,1,1).date(), datetime.today().date()), format="DD/MM/YYYY", key="dr")
        else:
            unf = False; ot_f = None; av_f = None; apm = []
            if os.path.exists("ot.xlsx"):
                try:
                    _t = excr(read_any_excel("ot.xlsx"))
                    apm = sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
                except: pass
            sp = apm; sa = ["All"]; sd = ["All"]
            dr = (datetime(2025,1,1).date(), datetime.today().date())

    if not unf or (ot_f is not None and av_f is not None):
        try:
            # ========== LECTURE ROBUSTE ==========
            if unf:
                raw_ot = read_any_excel(ot_f)
                raw_av = read_any_excel(av_f)
            else:
                raw_ot = read_any_excel("ot.xlsx")
                raw_av = read_any_excel("avis.xlsx")
            # =======================================

            raw_ot = excr(raw_ot); raw_av = excr(raw_av)
            for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c] = pd.to_datetime(raw_ot[c], errors="coerce")
            for c in ["Créé le","Début souhaité","Date de la clôture"]:
                if c in raw_av.columns: raw_av[c] = pd.to_datetime(raw_av[c], errors="coerce")
            if not apm: apm = sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp = apm
            if "All" in sa or not sa: sa = ["All"]
            if "All" in sd or not sd: sd = ["All"]
            sdt = pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt = pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p = str(poste).upper()
                if "All" not in sa:
                    m = False
                    if "Sulfurique (PS)" in sa and "PS" in p: m = True
                    if "Phosphorique (PP)" in sa and "PP" in p: m = True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m = True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m = True
                    if not m: return False
                if "All" not in sd:
                    m = False
                    if "SF1" in sd and "SF1" in p: m = True
                    if "SF2" in sd and "SF2" in p: m = True
                    if not m: return False
                return True

            vp = [p for p in apm if mf(p) and p in sp]

            df = raw_ot[(raw_ot["Poste travail princ."].isin(vp)) & (raw_ot["Date de début planifiée"].between(sdt, edt))].copy()
            avdf = raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df = excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)].drop_duplicates())
            avdf = excr(avdf[(avdf["Ordre"].isna())|(avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns: df["Statut OT"] = df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            df_dash = raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash = excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"), na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns: df_dash["Statut OT"] = df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now = pd.Timestamp.now()
            res = calc_kpis(df, avdf, now, vp)
            ckdf = res['ckdf']; dfp = res['dfp']
            res_d = calc_kpis(df_dash, avdf, now, vp)
            ckdf_d = res_d['ckdf']

            qk = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois","OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois","OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois","OT exécution 1mois< <3mois"]
            pk = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
            cible = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,"OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,"OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,"OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,"Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,"OT CONFIME":100,"OT_COR_EGAL":100}
            act_map = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.","OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).","OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.","OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).","OT planification >3 mois":"Traiter les OT avec planification > 3 mois.","OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).","OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.","OT LANC ESTIME":"Estimer les couts des OT lances.","Backlog préparation caractérisé":"Caracteriser le backlog de preparation.","Backlog planification caractérisé":"Caracteriser le backlog de planification.","OT CONFIME":"Confirmer les OT termines.","OT_COR_EGAL":"Rapprocher les couts reels et budgetes.","appel avis approuvé":"Creer un OT pour les avis sans ordre."}

            pscores = {}; qscores = {}
            for poste in ckdf.index:
                r = ckdf.loc[poste]
                pscores[poste] = (sum(gscore(k,r[k],cible[k]) for k in qk if k in r.index)/len(qk)*100) if qk else 0
                qscores[poste] = (sum(gscore(k,r[k],cible[k]) for k in pk if k in r.index)/len(pk)*100) if pk else 0
            pa = {k: round(ckdf[k].mean(),2) for k in qk}
            qa = {k: round(ckdf[k].mean(),2) for k in pk}

            pscores_d = {}; qscores_d = {}
            for poste in ckdf_d.index:
                r = ckdf_d.loc[poste]
                pscores_d[poste] = (sum(gscore(k,r[k],cible[k]) for k in qk if k in r.index)/len(qk)*100) if qk else 0
                qscores_d[poste] = (sum(gscore(k,r[k],cible[k]) for k in pk if k in r.index)/len(pk)*100) if pk else 0
            pa_d = {k: round(ckdf_d[k].mean(),2) for k in qk}
            qa_d = {k: round(ckdf_d[k].mean(),2) for k in pk}

            # ===== ANOMALIES =====
            all_ano = []
            sub_p = {"TAUX_REALISATION_CORRECTIF/PT":lambda d:d[(d["Nº appel pl.entret."].fillna(0)==0)&(~d["Statut OT"].isin(["CLOT","TCLO"]))],"OT préparation <1 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]!="<1 mois")],"OT préparation >3 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]==">3 mois")],"OT planification <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]!="<1 mois")],"OT planification >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]==">3 mois")],"OT exécution <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]!="<1 mois")],"OT exécution >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]==">3 mois")]}
            sub_q = {"OT LANC ESTIME":lambda d:d[(d["Statut OT"]=="LANC")&(d["OT LANC ESTIME"]=="NON")],"Backlog préparation caractérisé":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["Backlog preparation"]=="NON CARACTERISE")],"Backlog planification caractérisé":lambda d:d[(d["Statut OT"]=="LANC")&(d["Backlog planification"]=="NON CARACTERISE")],"OT CONFIME":lambda d:d[d["OT CONFIME"]=="NON"],"OT_COR_EGAL":lambda d:d[d["OT_COR_EGAL"]=="NON"]}
            for poste in vp:
                if poste not in dfp["Poste travail princ."].values: continue
                dp = dfp[dfp["Poste travail princ."]==poste]
                for kn, sf in sub_p.items():
                    vk = ckdf.loc[poste, kn] if poste in ckdf.index else 100
                    if pd.notna(vk) and vk < cible[kn]:
                        cnt = len(sf(dp))
                        if cnt > 0: all_ano.append({"Poste":poste,"KPI":kn,"Nb":cnt,"Type":"P"})
                for kn, sf in sub_q.items():
                    vk = ckdf.loc[poste, kn] if poste in ckdf.index else 100
                    if pd.notna(vk) and vk < cible[kn]:
                        cnt = len(sf(dp))
                        if cnt > 0: all_ano.append({"Poste":poste,"KPI":kn,"Nb":cnt,"Type":"Q"})
                vk_av = ckdf.loc[poste, "appel avis approuvé"] if poste in ckdf.index else 100
                if pd.notna(vk_av) and vk_av < cible["appel avis approuvé"]:
                    cnt = len(res['avf'][res['avf']["Poste travail princ."]==poste])
                    if cnt > 0: all_ano.append({"Poste":poste,"KPI":"appel avis approuvé","Nb":cnt,"Type":"Q"})

            def build_ano(ano_list, kpi_list):
                if not ano_list: return [], []
                adf = pd.DataFrame(ano_list)
                pv = adf.pivot_table(index="Poste", columns="KPI", values="Nb", aggfunc="sum", fill_value=0).astype(int)
                pv["Total"] = pv.sum(axis=1); tot = pv.sum()
                cols = [c for c in kpi_list if c in pv.columns] + ["Total"]; rows = []
                for idx in pv.index:
                    r = {"_t":"n","Poste de travail":idx}
                    for c in cols: r[c] = pv.loc[idx, c]
                    rows.append(r)
                tr = {"_t":"total","Poste de travail":"Total general"}
                for c in cols: tr[c] = int(tot[c])
                rows.append(tr); return ["Poste de travail"]+cols, rows

            ano_p_c, ano_p_r = build_ano([a for a in all_ano if a["Type"]=="P"], qk)
            ano_q_c, ano_q_r = build_ano([a for a in all_ano if a["Type"]=="Q"], pk)

            def build_kpi(kpi_list, scores, sname):
                sp2 = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)
                cols = ["Poste de travail"]+kpi_list+[sname]; rows = []
                cr = {"_t":"cible","Poste de travail":"CIBLE"}
                for k in kpi_list: cr[k] = cible[k]
                cr[sname] = "100.00 %"; rows.append(cr)
                for p in sp2:
                    r = {"_t":"n","Poste de travail":p}
                    for k in kpi_list: r[k] = round(ckdf.loc[p, k], 2) if p in ckdf.index else ""
                    r[sname] = "%.2f %%" % scores[p]; rows.append(r)
                tr = {"_t":"total","Poste de travail":"Total general"}
                for k in kpi_list: tr[k] = round(ckdf[k].mean(), 2)
                tr[sname] = "%.2f %%" % (np.mean(list(scores.values())) if scores else 0)
                rows.append(tr); return cols, rows

            pcols, prows = build_kpi(qk, pscores, "Score Performance")
            qcols, qrows = build_kpi(pk, qscores, "Score Qualite")

            save_kpis_to_excel(prows, pcols, qrows, qcols, ano_p_r, ano_p_c, ano_q_r, ano_q_c, fichier_date)

            df_sc_d = pd.DataFrame([{"Poste":p,"Perf":pscores_d[p],"Qual":qscores_d[p],"Metier":get_metier(p),"Atelier":get_atelier(p),"Division":get_division(p)} for p in vp if p in pscores_d])
            by_at = df_sc_d.groupby("Atelier")[["Perf","Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()
            by_mt = df_sc_d.groupby("Metier")[["Perf","Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()
            by_div = df_sc_d.groupby("Division")[["Perf","Qual"]].mean().round(1) if not df_sc_d.empty else pd.DataFrame()

            total_ot = len(df); avg_p = np.mean(list(pscores.values())) if pscores else 0
            avg_q = np.mean(list(qscores.values())) if qscores else 0; total_ano = sum(a["Nb"] for a in all_ano)

            desig_col = None
            for cn in ["Désignation du travail","Designation du travail","Désignation","Designation","Description"]:
                if cn in dfp.columns: desig_col = cn; break

            # ==================== RENDER ====================
            st.markdown('<div class="mh"><h1>📊 KPI Dashboard MC & FEED</h1><div class="db">📅 %s</div></div>' % fichier_date, unsafe_allow_html=True)
            st.markdown("""<div class="cr">
            <div class="cc c1"><div class="cv">%s</div><div class="cl">Total OT Analyses</div></div>
            <div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div>
            <div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div>
            <div class="cc c4"><div class="cv">%s</div><div class="cl">Total Anomalies</div></div>
            </div>""" % (total_ot, avg_p, avg_q, total_ano), unsafe_allow_html=True)

            tab0, tab1, tab2, tab3 = st.tabs(["📊 TABLEAU DE BORD", "📈 INDICATEURS PERFORMANCE", "✅ INDICATEUR QUALITE", "🔬 ANALYSE"])

            # ==================== TAB0: DASHBOARD ====================
            with tab0:
                st.markdown('<div class="stl p">📊 Vue d\'ensemble par poste</div>', unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp, pscores_d, qscores_d, "Performance & Qualite par Poste de Travail"), unsafe_allow_html=True)

                st.markdown('<div class="stl p">🏭 Par Atelier</div>', unsafe_allow_html=True)
                if not by_at.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx,row["Perf"]) for idx,row in by_at.iterrows()], "Performance par Atelier", "#2b6cb0"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx,row["Qual"]) for idx,row in by_at.iterrows()], "Qualite par Atelier", "#276749"), unsafe_allow_html=True)

                st.markdown('<div class="stl c">🔧 Par Metier</div>', unsafe_allow_html=True)
                if not by_mt.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx,row["Perf"]) for idx,row in by_mt.iterrows()], "Performance par Metier", "#6b46c1"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx,row["Qual"]) for idx,row in by_mt.iterrows()], "Qualite par Metier", "#2b6cb0"), unsafe_allow_html=True)

                st.markdown('<div class="stl q">🏢 Par Division</div>', unsafe_allow_html=True)
                if not by_div.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(html_bars([(idx,row["Perf"]) for idx,row in by_div.iterrows()], "Performance par Division", "#276749"), unsafe_allow_html=True)
                    with c2:
                        st.markdown(html_bars([(idx,row["Qual"]) for idx,row in by_div.iterrows()], "Qualite par Division", "#6b46c1"), unsafe_allow_html=True)

                st.markdown('<div class="stl a">📋 Synthese Globale</div>', unsafe_allow_html=True)
                global_perf = np.mean(list(pscores_d.values())) if pscores_d else 0
                global_qual = np.mean(list(qscores_d.values())) if qscores_d else 0
                met_p = "ATTEINT" if global_perf >= 80 else "NON ATTEINT"
                met_q = "ATTEINT" if global_qual >= 80 else "NON ATTEINT"
                clr_p = "#38a169" if global_perf >= 80 else "#e53e3e"
                clr_q = "#38a169" if global_qual >= 80 else "#e53e3e"
                bg_p = "#c6efce" if global_perf >= 80 else "#ffc7ce"
                bg_q = "#c6efce" if global_qual >= 80 else "#ffc7ce"
                st.markdown("""<div style="display:grid;grid-template-columns:1fr 1fr;gap:4px">
                <div style="background:#fff;border-radius:10px;padding:12px;border:1px solid #e2e8f0;text-align:center">
                    <div style="font-size:9px;font-weight:700;color:#718096;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">Score Performance Global</div>
                    <div style="font-size:32px;font-weight:900;color:%s">%.1f%%</div>
                    <div style="display:inline-block;padding:3px 12px;border-radius:12px;font-size:10px;font-weight:800;color:%s;background:%s;margin-top:4px">%s</div>
                </div>
                <div style="background:#fff;border-radius:10px;padding:12px;border:1px solid #e2e8f0;text-align:center">
                    <div style="font-size:9px;font-weight:700;color:#718096;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">Score Qualite Global</div>
                    <div style="font-size:32px;font-weight:900;color:%s">%.1f%%</div>
                    <div style="display:inline-block;padding:3px 12px;border-radius:12px;font-size:10px;font-weight:800;color:%s;background:%s;margin-top:4px">%s</div>
                </div>
                </div>""" % (clr_p, global_perf, clr_p, bg_p, met_p, clr_q, global_qual, clr_q, bg_q, met_q), unsafe_allow_html=True)

            # ==================== TAB1: PERFORMANCE ====================
            with tab1:
                st.markdown('<div class="stl p">📈 Synthese Performance</div>', unsafe_allow_html=True)
                st.markdown(html_synth(qk, pa, cible, act_map, "#276749"), unsafe_allow_html=True)
                st.markdown('<div class="stl p">📊 Tableau Detaille Performance</div>', unsafe_allow_html=True)
                st.markdown(html_table(prows, pcols, "pt", ["Score Performance"]), unsafe_allow_html=True)
                if ano_p_r:
                    st.markdown('<div class="stl a">⚠️ Anomalies Performance</div>', unsafe_allow_html=True)
                    st.markdown(html_ano(ano_p_r, ano_p_c), unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">✅ Aucune anomalie performance detectee</div>', unsafe_allow_html=True)
                st.markdown('<div class="stl c">🏆 Classement Performance</div>', unsafe_allow_html=True)
                st.markdown(html_classement(pscores, "#276749"), unsafe_allow_html=True)
                st.markdown('<div class="stl p">📊 Barres de Progression Performance</div>', unsafe_allow_html=True)
                st.markdown(html_kpi_bars(qk, pa, cible, "Indicateurs Performance — Moyenne Globale", "#276749", "#e53e3e"), unsafe_allow_html=True)
                exp_df_p = pd.DataFrame(prows)
                st.markdown("---")
                export_btn(exp_df_p, "performance_kpis_%s.xlsx" % fichier_date.replace("/","-"))

            # ==================== TAB2: QUALITE ====================
            with tab2:
                st.markdown('<div class="stl q">✅ Synthese Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_synth(pk, qa, cible, act_map, "#2b6cb0"), unsafe_allow_html=True)
                st.markdown('<div class="stl q">📊 Tableau Detaille Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_table(qrows, qcols, "qt", ["Score Qualite"]), unsafe_allow_html=True)
                if ano_q_r:
                    st.markdown('<div class="stl a">⚠️ Anomalies Qualite</div>', unsafe_allow_html=True)
                    st.markdown(html_ano(ano_q_r, ano_q_c), unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">✅ Aucune anomalie qualite detectee</div>', unsafe_allow_html=True)
                st.markdown('<div class="stl c">🏆 Classement Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_classement(qscores, "#2b6cb0"), unsafe_allow_html=True)
                st.markdown('<div class="stl q">📊 Barres de Progression Qualite</div>', unsafe_allow_html=True)
                st.markdown(html_kpi_bars(pk, qa, cible, "Indicateurs Qualite — Moyenne Globale", "#2b6cb0", "#e53e3e"), unsafe_allow_html=True)
                exp_df_q = pd.DataFrame(qrows)
                st.markdown("---")
                export_btn(exp_df_q, "qualite_kpis_%s.xlsx" % fichier_date.replace("/","-"))

            # ==================== TAB3: ANALYSE ====================
            with tab3:
                st.markdown('<div class="stl p">📊 Distribution par Statut OT</div>', unsafe_allow_html=True)
                if "Statut OT" in df_dash.columns and not df_dash.empty:
                    stat_dist = df_dash["Statut OT"].value_counts().reset_index()
                    stat_dist.columns = ["Statut", "Nombre"]
                    fig_stat = anl_pie_chart(stat_dist, "Statut", "Nombre", "Repartition des OT par Statut",
                                             ["#276749","#2b6cb0","#805ad5","#e53e3e","#ed8936","#38a169","#4299e1"])
                    if fig_stat: st.plotly_chart(fig_stat, use_container_width=True)

                st.markdown('<div class="stl q">🏭 Distribution par Atelier</div>', unsafe_allow_html=True)
                if not df_sc_d.empty:
                    at_counts = df_sc_d["Atelier"].value_counts().reset_index()
                    at_counts.columns = ["Atelier", "Nb Postes"]
                    fig_at = anl_pie_chart(at_counts, "Atelier", "Nb Postes", "Postes par Atelier",
                                           ["#2b6cb0","#276749","#805ad5","#e53e3e","#ed8936"])
                    if fig_at: st.plotly_chart(fig_at, use_container_width=True)

                st.markdown('<div class="stl c">🔧 Analyse par Metier</div>', unsafe_allow_html=True)
                if not by_mt.empty:
                    mt_df = by_mt.reset_index(); mt_df.columns = ["Metier","Performance","Qualite"]
                    mt_df["Moyenne"] = ((mt_df["Performance"]+mt_df["Qualite"])/2).round(1)
                    tot_row = pd.DataFrame([{"Metier":"TOTAL","Performance":mt_df["Performance"].mean(),"Qualite":mt_df["Qualite"].mean(),"Moyenne":mt_df["Moyenne"].mean()}])
                    mt_df = pd.concat([mt_df, tot_row], ignore_index=True)
                    st.markdown(anl_html_table(mt_df, pct_col=None), unsafe_allow_html=True)

                st.markdown('<div class="stl q">🏭 Analyse par Atelier</div>', unsafe_allow_html=True)
                if not by_at.empty:
                    at_df = by_at.reset_index(); at_df.columns = ["Atelier","Performance","Qualite"]
                    at_df["Moyenne"] = ((at_df["Performance"]+at_df["Qualite"])/2).round(1)
                    tot_row = pd.DataFrame([{"Atelier":"TOTAL","Performance":at_df["Performance"].mean(),"Qualite":at_df["Qualite"].mean(),"Moyenne":at_df["Moyenne"].mean()}])
                    at_df = pd.concat([at_df, tot_row], ignore_index=True)
                    st.markdown(anl_html_table(at_df, pct_col=None), unsafe_allow_html=True)

                st.markdown('<div class="stl p">🏢 Analyse par Division</div>', unsafe_allow_html=True)
                if not by_div.empty:
                    dv_df = by_div.reset_index(); dv_df.columns = ["Division","Performance","Qualite"]
                    dv_df["Moyenne"] = ((dv_df["Performance"]+dv_df["Qualite"])/2).round(1)
                    tot_row = pd.DataFrame([{"Division":"TOTAL","Performance":dv_df["Performance"].mean(),"Qualite":dv_df["Qualite"].mean(),"Moyenne":dv_df["Moyenne"].mean()}])
                    dv_df = pd.concat([dv_df, tot_row], ignore_index=True)
                    st.markdown(anl_html_table(dv_df, pct_col=None), unsafe_allow_html=True)

                st.markdown('<div class="stl a">⚠️ Detail des Anomalies par KPI</div>', unsafe_allow_html=True)
                if all_ano:
                    ano_df = pd.DataFrame(all_ano)
                    ano_summary = ano_df.groupby("KPI")["Nb"].agg(["sum","count"]).reset_index()
                    ano_summary.columns = ["KPI","Total Anomalies","Nb Postes Concernes"]
                    ano_summary = ano_summary.sort_values("Total Anomalies", ascending=False)
                    tot_ano_row = pd.DataFrame([{"KPI":"TOTAL","Total Anomalies":ano_summary["Total Anomalies"].sum(),"Nb Postes Concernes":ano_summary["Nb Postes Concernes"].sum()}])
                    ano_summary = pd.concat([ano_summary, tot_ano_row], ignore_index=True)
                    st.markdown(anl_html_table(ano_summary, pct_col=None), unsafe_allow_html=True)

                    st.markdown('<div class="stl a">📋 Detail par Poste</div>', unsafe_allow_html=True)
                    ano_poste = ano_df.groupby("Poste").agg(Total=("Nb","sum"), KPIs=("KPI",lambda x:", ".join(x.unique()))).reset_index()
                    ano_poste = ano_poste.sort_values("Total", ascending=False)
                    st.markdown(anl_html_table(ano_poste, pct_col=None), unsafe_allow_html=True)
                else:
                    st.markdown('<div class="es">✅ Aucune anomalie detectee sur la periode</div>', unsafe_allow_html=True)

                st.markdown('<div class="stl a">🕐 OT avec Preparation > 3 mois</div>', unsafe_allow_html=True)
                if dfp is not None and not dfp.empty and "ap" in dfp.columns:
                    old_ot = dfp[(dfp["Statut OT"]=="CRÉÉ") & (dfp["ap"]==">3 mois")].copy()
                    if not old_ot.empty:
                        show_cols = ["Ordre","Poste travail princ.","Statut OT","Créé le","ap"]
                        if desig_col: show_cols.insert(2, desig_col)
                        show_cols = [c for c in show_cols if c in old_ot.columns]
                        old_show = old_ot[show_cols].sort_values("Créé le", ascending=True).head(50).copy()
                        old_show["Créé le"] = old_show["Créé le"].dt.strftime("%d/%m/%Y")
                        st.markdown(anl_html_table(old_show, pct_col=None), unsafe_allow_html=True)
                        export_btn(old_show, "ot_preparation_anciens_%s.xlsx" % fichier_date.replace("/","-"))
                    else:
                        st.markdown('<div class="es">✅ Aucun OT avec preparation > 3 mois</div>', unsafe_allow_html=True)

                st.markdown('<div class="stl a">💰 OT Lances Non Estimes</div>', unsafe_allow_html=True)
                if dfp is not None and not dfp.empty:
                    no_est = dfp[(dfp["Statut OT"]=="LANC") & (dfp["OT LANC ESTIME"]=="NON")].copy()
                    if not no_est.empty:
                        show_cols2 = ["Ordre","Poste travail princ.","Statut OT","Total coûts budgétés","Total coûts réels"]
                        if desig_col: show_cols2.insert(2, desig_col)
                        show_cols2 = [c for c in show_cols2 if c in no_est.columns]
                        no_est_show = no_est[show_cols2].head(50).copy()
                        st.markdown(anl_html_table(no_est_show, pct_col=None), unsafe_allow_html=True)
                        export_btn(no_est_show, "ot_non_estimes_%s.xlsx" % fichier_date.replace("/","-"))
                    else:
                        st.markdown('<div class="es">✅ Tous les OT lances sont estimes</div>', unsafe_allow_html=True)

                st.markdown("---")
                all_export = pd.DataFrame(prows + qrows)
                export_btn(all_export, "kpi_complet_%s.xlsx" % fichier_date.replace("/","-"))

        except Exception as e:
            st.error("Erreur lors du chargement des donnees: %s" % str(e))
            st.markdown('<div class="es">Veuillez verifier que les fichiers ot.xlsx et avis.xlsx sont presents a la racine, ou uploader les fichiers via le panneau lateral.<br><br><b>Detail :</b> %s</div>' % str(e).replace("<","&lt;").replace(">","&gt;"), unsafe_allow_html=True)
    else:
        if unf:
            st.markdown('<div class="es">📁 Veuillez uploader les fichiers OT et AVIS pour continuer.</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">📂 Aucun fichier de donnees detecte. Placez ot.xlsx et avis.xlsx a la racine du projet ou activez le chargement manuel.</div>', unsafe_allow_html=True)

    gc.collect()

if __name__ == "__main__":
    main()
