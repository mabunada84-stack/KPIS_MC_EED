# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]
ALL_KPI = QK + PK
CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100}
ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois."}
ACT_DETAIL = {
    "TAUX_REALISATION_CORRECTIF/PT": {"prio":"HAUTE","icon":"🎯","detail":"Identifier les OT Correctif/PT non realises et accelerer leur cloture. Prioriser les OT les plus anciens et allouer les ressources necessaires."},
    "OT préparation <1 mois": {"prio":"HAUTE","icon":"⏱️","detail":"Mettre en place un suivi hebdomadaire de l'age des OT en preparation. Objectif : 80% des OT prepares en moins d'un mois."},
    "OT préparation >3 mois": {"prio":"CRITIQUE","icon":"🚨","detail":"Traiter en urgence les OT avec un delai de preparation superieur a 3 mois. Analyser les causes de retard et mettre en place des actions correctives."},
    "OT planification <1 mois": {"prio":"HAUTE","icon":"📅","detail":"Optimiser le processus de planification pour reduire le delai a moins d'un mois pour 80% des OT."},
    "OT planification >3 mois": {"prio":"CRITIQUE","icon":"🚨","detail":"Revue hebdomadaire des OT non planifies depuis plus de 3 mois. Identifier les goulots d'etranglement et mobiliser les planificateurs."},
    "OT exécution <1 mois": {"prio":"HAUTE","icon":"⚡","detail":"Suivi quotidien de l'execution des OT lances. Objectif : 80% des OT executes en moins d'un mois apres lancement."},
    "OT exécution >3 mois": {"prio":"CRITIQUE","icon":"🚨","detail":"Analyser les OT en execution depuis plus de 3 mois. Identifier les blocages (materiaux, main d'oeuvre, autorisations) et lever les obstacles."},
    "OT LANC ESTIME": {"prio":"MOYENNE","icon":"💰","detail":"Exiger l'estimation des couts avant le lancement de tout OT. Mettre en place un controle a la passation de l'ordre de lancement."},
    "Backlog préparation caractérisé": {"prio":"HAUTE","icon":"📋","detail":"Caracteriser 100% du backlog de preparation avec les informations techniques necessaires (type d'intervention, ressources, delais)."},
    "Backlog planification caractérisé": {"prio":"HAUTE","icon":"📋","detail":"Caracteriser 100% du backlog de planification. Assurer la disponibilite des ressources avant planification."},
    "OT CONFIME": {"prio":"MOYENNE","icon":"✅","detail":"Verifier et confirmer systematiquement les OT apres cloture. Mettre en place un circuit de validation automatique."},
    "OT_COR_EGAL": {"prio":"MOYENNE","icon":"📊","detail":"Rapprocher les couts reels et budgetes pour chaque OT. Analyser les ecarts significatifs et ajuster les estimations futures."},
    "appel avis approuvé": {"prio":"HAUTE","icon":"🔔","detail":"Creer un OT pour chaque avis approuve sans ordre associe. Suivi quotidien des avis en attente de creation d'OT."},
    "OT préparation 1mois< <3mois": {"prio":"MOYENNE","icon":"📉","detail":"Reducire la proportion d'OT en preparation entre 1 et 3 mois. Cible : maintenir ce ratio en dessous de 15%."},
    "OT planification 1mois< <3mois": {"prio":"MOYENNE","icon":"📉","detail":"Reducire la proportion d'OT en planification entre 1 et 3 mois. Cible : maintenir ce ratio en dessous de 15%."},
    "OT exécution 1mois< <3mois": {"prio":"MOYENNE","icon":"📉","detail":"Reducire la proportion d'OT en execution entre 1 et 3 mois. Cible : maintenir ce ratio en dessous de 15%."}
}
LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]
MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]
CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

# ============================================================
def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    j["Sens"]=j.apply(lambda r:"Amelioration" if ((r["Tendance"]=="hausse" and r["KPI"] not in LOWER_BETTER) or (r["Tendance"]=="baisse" and r["KPI"] in LOWER_BETTER)) else "Degradation",axis=1)
    return j.sort_values(["Date actuelle","Sens","Ecart %"],ascending=[True,False,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def get_caract_type(statut_user,keywords):
    s=str(statut_user).upper(); matched=[kw for kw in keywords if kw in s]
    return max(matched,key=len) if matched else "AUTRE"

# ============================================================
def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .sr{display:flex;align-items:center;padding:6px 10px;background:#fff;border-radius:5px;margin-bottom:2px;border:1px solid var(--b);font-size:13px}
    .sr .sn{font-weight:700;color:var(--p);min-width:220px;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .sc{padding:3px 9px;border-radius:12px;font-weight:800;font-size:14px;min-width:50px;text-align:center;margin:0 8px;color:#fff}
    .sr .sa{color:#718096;font-size:12px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .sr .stg{font-size:11px;color:#718096;min-width:60px;text-align:center;white-space:nowrap}
    .sr .sb{font-size:11px;font-weight:700;padding:2px 8px;border-radius:3px;white-space:nowrap}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:4px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:24px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}
    .gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .anl-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:13px;margin:0}
    .anl-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:12px;padding:6px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .anl-tbl tbody td{padding:5px 8px;border-bottom:1px solid #edf2f7}
    .anl-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .anl-tbl tbody tr:hover td{background:#ebf8ff!important}
    .anl-tbl .tot td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important}
    .g-green{background:#c6efce;color:#006100;font-weight:600}
    .g-yellow{background:#ffeb9c;color:#9c6500;font-weight:600}
    .g-red{background:#ffc7ce;color:#9c0006;font-weight:600}
    .trend-up{color:#276749;font-weight:800;font-size:16px}
    .trend-down{color:#c53030;font-weight:800;font-size:16px}
    .trend-stable{color:#718096;font-weight:800;font-size:16px}
    .spark-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:8px}
    .spark-card{background:#fff;border-radius:var(--r);padding:10px 12px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .spark-card .sp-title{font-size:13px;font-weight:800;color:var(--p);margin-bottom:3px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .spark-card .sp-sub{font-size:11px;color:#718096;margin-bottom:5px}
    .rank-card{background:#fff;border-radius:var(--r);padding:12px 16px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .rank-card .rank-title{font-size:15px;font-weight:800;margin-bottom:8px;padding-bottom:5px;border-bottom:2px solid var(--b)}
    .rank-row{display:flex;align-items:center;padding:5px 0;font-size:13px;border-bottom:1px solid #f7fafc}
    .rank-row:last-child{border:none}
    .rank-row .rank-num{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:900;font-size:13px;color:#fff;margin-right:10px;flex-shrink:0}
    .rank-row .rank-name{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .rank-row .rank-score{font-weight:900;min-width:70px;text-align:right}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg,.dgrid{grid-template-columns:1fr}.car .cal{width:120px}.gbr-l{width:100px}.spark-grid{grid-template-columns:1fr}}

    /* === PROFESSIONAL ACTION CARDS === */
    .act-panel{display:grid;grid-template-columns:repeat(auto-fill,minmax(420px,1fr));gap:10px}
    .act-card{background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.06);border:1px solid var(--b);transition:transform .2s,box-shadow .2s}
    .act-card:hover{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,0,0,.1)}
    .act-card-head{padding:10px 14px;display:flex;align-items:center;gap:10px;border-bottom:1px solid #edf2f7}
    .act-card-icon{font-size:24px;flex-shrink:0}
    .act-card-kpi{flex:1;font-size:13px;font-weight:800;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .act-card-badge{padding:3px 10px;border-radius:20px;font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.5px;flex-shrink:0}
    .badge-critique{background:linear-gradient(135deg,#e53e3e,#c53030);color:#fff}
    .badge-haute{background:linear-gradient(135deg,#ed8936,#dd6b20);color:#fff}
    .badge-moyenne{background:linear-gradient(135deg,#ecc94b,#d69e2e);color:#744210}
    .badge-atteint{background:linear-gradient(135deg,#48bb78,#38a169);color:#fff}
    .act-card-body{padding:10px 14px}
    .act-metrics{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:8px}
    .act-metric{text-align:center;padding:6px 4px;background:#f7fafc;border-radius:8px}
    .act-metric-val{font-size:18px;font-weight:900;line-height:1.2}
    .act-metric-lbl{font-size:10px;color:#718096;font-weight:600;text-transform:uppercase;letter-spacing:.3px;margin-top:1px}
    .act-bar-wrap{margin-bottom:8px}
    .act-bar-label{display:flex;justify-content:space-between;font-size:11px;font-weight:700;margin-bottom:3px}
    .act-bar-track{height:8px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .act-bar-fill{height:100%;border-radius:4px;transition:width .5s ease}
    .act-bar-fill.green{background:linear-gradient(90deg,#48bb78,#38a169)}
    .act-bar-fill.red{background:linear-gradient(90deg,#fc8181,#e53e3e)}
    .act-bar-fill.orange{background:linear-gradient(90deg,#fbd38d,#ed8936)}
    .act-card-detail{font-size:12px;color:#4a5568;line-height:1.5;padding:8px 10px;background:#f7fafc;border-radius:8px;border-left:3px solid var(--pl)}
    .act-card-footer{padding:6px 14px 10px;display:flex;align-items:center;gap:6px}
    .act-ecart{font-size:12px;font-weight:800;padding:2px 8px;border-radius:4px}
    .act-ecart.neg{background:#fff5f5;color:#c53030}
    .act-ecart.pos{background:#f0fff4;color:#276749}
    .act-ecart.zero{background:#f7fafc;color:#718096}
    .act-section-title{font-size:16px;font-weight:800;color:var(--p);margin:10px 0 6px 0;padding:8px 14px;background:linear-gradient(135deg,#1e3a5f,#2c5282);color:#fff;border-radius:8px;display:flex;align-items:center;gap:8px}
    .act-section-title .act-count{background:rgba(255,255,255,.2);padding:2px 10px;border-radius:12px;font-size:12px;font-weight:700}
    .act-summary-row{display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-bottom:10px}
    .act-summary-item{background:#fff;border-radius:10px;padding:12px;text-align:center;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.03)}
    .act-summary-item .asv{font-size:28px;font-weight:900}
    .act-summary-item .asl{font-size:10px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .act-synth-col{background:#fff;border-radius:12px;padding:14px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.04)}
    .act-synth-col .asct{font-size:15px;font-weight:800;padding:8px 12px;border-radius:8px;color:#fff;margin-bottom:10px;display:flex;align-items:center;gap:8px}
    .act-synth-col .asct.perf{background:linear-gradient(135deg,#276749,#38a169)}
    .act-synth-col .asct.qual{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .act-synth-kpi-row{display:flex;align-items:center;padding:5px 0;border-bottom:1px solid #f7fafc;font-size:12px}
    .act-synth-kpi-row:last-child{border:none}
    .act-synth-kpi-name{width:220px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .act-synth-kpi-bar{flex:1;height:16px;background:#edf2f7;border-radius:3px;overflow:hidden;margin:0 8px}
    .act-synth-kpi-fill{height:100%;border-radius:3px}
    .act-synth-kpi-val{font-weight:800;min-width:48px;text-align:right;font-size:12px}
    .act-synth-kpi-status{min-width:24px;text-align:center;font-size:14px}
    @media(max-width:900px){.act-panel{grid-template-columns:1fr}.act-summary-row{grid-template-columns:1fr}}
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    def contient_mot(t,lm):
        t=str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if a<=1: return "<1 mois"
        elif a>=3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
        return df
    def get_metier(p):
        p=str(p).upper()
        if "E" in p: return "Electrique"
        if "M" in p: return "Mecanique"
        if "R" in p: return "Instrumentation"
        if "G" in p: return "Genie Civil"
        return "Autre"
    def get_atelier(p):
        p=str(p).upper()
        if "PS" in p: return "Sulfurique"
        if "PP" in p: return "Phosphorique"
        if "TSP" in p or "REX" in p: return "Engrais"
        if "MCP" in p or "DCP" in p: return "Feed"
        return "Autre"
    def get_division(p):
        p=str(p).upper()
        if "SF1" in p: return "SF1"
        if "SF2" in p: return "SF2"
        return "Autre"

        def calc_kpis(df_i,av_i,now,posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        # --- Conversion des colonnes numeriques potentiellement en texte ---
        for col_num in ["Total coûts budgétés","Total coûts réels","Nº appel pl.entret."]:
            if col_num in df.columns:
                df[col_num]=pd.to_numeric(df[col_num],errors='coerce').fillna(0)
        for col_num in ["Avis"]:
            if col_num in av.columns:
                av[col_num]=pd.to_numeric(av[col_num],errors='coerce').fillna(0)

        df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
        df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
        for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc]=pd.to_datetime(df[dc],errors='coerce')
                df[am]=((now.year-df[dc].dt.year)*12+(now.month-df[dc].dt.month)).round(2)
                df[ac]=df[am].apply(cat_age)
            else: df[am]=np.nan; df[ac]="Inconnu"
        df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
        df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
        df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"]==0,"NON","OUI")
        df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"]-df["Total coûts réels"])==0,"OUI","NON")
        res['dfp']=df
        an=cpiv(df,df["Nº appel pl.entret."]==0,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["Total"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1); an["TAUX_REALISATION_CORRECTIF/PT"]=ckpi(an["TCLO"],an["Total"])
        pr=cpiv(df,df["Statut OT"]=="CRÉÉ","ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==0),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
        avf=av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip().eq(""))].copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"]
        })
        return res

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "color:#cbd5e0"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        return 0
    def is_lb(k): return k in LOWER_BETTER

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                else: s=cs(v) if sc_col and c in sc_col else ks(v,c); h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'
    def html_ano(rows,cols):
        h='<table class="tw at"><thead><tr>'+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            h+='<tr class="%s">'%("tr" if r.get("_t")=="total" else "")
            for c in cols: v=r.get(c,""); h+='<td style="%s">%s</td>'%(kas(v) or "",v)
            h+='</tr>'
        return h+'</tbody></table>'

    # ========== AMELIOREE : Pie of Pie avec legende detaillee ==========
    def anl_pie_of_pie_chart(data, names_col, values_col, title, colors=None, threshold=5):
        if data.empty: return None
        df_pie = data[[names_col, values_col]].dropna()
        df_pie[values_col] = pd.to_numeric(df_pie[values_col], errors='coerce').fillna(0)
        total = df_pie[values_col].sum()
        if total == 0: return None
        df_pie['pct'] = (df_pie[values_col] / total * 100)
        small = df_pie[df_pie['pct'] < threshold]
        big = df_pie[df_pie['pct'] >= threshold].copy()
        if not small.empty:
            other_row = pd.DataFrame({names_col: ['Autres'], values_col: [small[values_col].sum()], 'pct': [small['pct'].sum()]})
            main_df = pd.concat([big, other_row], ignore_index=True)
        else:
            main_df = big.copy()
            small = pd.DataFrame(columns=df_pie.columns)

        clr = colors or px.colors.qualitative.Set2
        fig = make_subplots(
            rows=1, cols=2,
            specs=[[{"type": "domain"}, {"type": "domain"}]],
            subplot_titles=("", "Detail des petits secteurs"),
            column_widths=[0.6, 0.4],
            horizontal_spacing=0.05
        )
        main_colors = clr[:len(main_df)]
        fig.add_trace(go.Pie(
            labels=main_df[names_col],
            values=main_df[values_col],
            textinfo='percent+value',
            texttemplate='%{label}<br>%{value} (%{percent})',
            textfont_size=11,
            marker_colors=main_colors,
            hole=0.35,
            name="Principal"
        ), row=1, col=1)

        if not small.empty:
            small_colors = clr[len(big):len(big)+len(small)]
            while len(small_colors) < len(small):
                small_colors = small_colors + clr
            small_colors = small_colors[:len(small)]
            fig.add_trace(go.Pie(
                labels=small[names_col],
                values=small[values_col],
                textinfo='percent+value',
                texttemplate='%{label}<br>%{value} (%{percent})',
                textfont_size=10,
                marker_colors=small_colors,
                hole=0.25,
                name="Detail"
            ), row=1, col=2)
        else:
            fig.add_trace(go.Pie(
                labels=["Aucun petit secteur"],
                values=[1],
                textinfo='none',
                marker_colors=['#e2e8f0'],
                hole=0.25,
                name="Detail"
            ), row=1, col=2)

        legend_texts = []
        for _, row in main_df.iterrows():
            if row[names_col] == 'Autres':
                legend_texts.append(f"<b>Autres</b> ({row['pct']:.1f}%) = {int(row[values_col])}")
            else:
                legend_texts.append(f"<b>{row[names_col]}</b> ({row['pct']:.1f}%) = {int(row[values_col])}")
        legend_detail = "<br>".join(legend_texts)
        if not small.empty:
            legend_detail += "<br><br><b>--- Detail Autres ---</b><br>"
            for _, row in small.iterrows():
                legend_detail += f"• {row[names_col]} ({row['pct']:.1f}%) = {int(row[values_col])}<br>"

        fig.update_layout(
            title=dict(text=f"<b>{title}</b><br><span style='font-size:10px;color:#718096'>Total: {int(total)}</span>",
                       font_size=14, x=0.5, xanchor='center'),
            margin=dict(t=70, b=20, l=20, r=20),
            height=480,
            autosize=True,
            showlegend=True,
            legend=dict(
                font_size=10,
                orientation="v",
                xanchor="left",
                x=1.02,
                y=0.5,
                bgcolor="rgba(255,255,255,0.9)",
                bordercolor="#e2e8f0",
                borderwidth=1
            ),
            annotations=[
                dict(text=f"<span style='font-size:10px'>{legend_detail}</span>",
                     x=1.02, y=-0.02, xanchor='left', yanchor='top',
                     showarrow=False, font_size=9, align='left')
            ]
        )
        return fig

    # ========== PROFESSIONNELLE : Actions Recommandees en cartes ==========
    def html_professional_actions(kpi_list, actuals, targets, act_map, act_detail, section_type):
        cards_by_prio = {"CRITIQUE": [], "HAUTE": [], "MOYENNE": [], "ATTEINT": []}
        for k in kpi_list:
            av = actuals.get(k, 0)
            tv = targets.get(k, 100)
            diff = av - tv
            met = av <= tv if is_lb(k) else av >= tv
            detail_info = act_detail.get(k, {"prio": "MOYENNE", "icon": "📌", "detail": act_map.get(k, "")})
            prio = "ATTEINT" if met else detail_info["prio"]
            icon = detail_info["icon"]
            detail_text = detail_info["detail"] if not met else "Objectif atteint. Maintenir le niveau actuel et surveiller les tendances."
            cards_by_prio[prio].append({
                "kpi": k, "av": av, "tv": tv, "diff": diff, "met": met,
                "prio": prio, "icon": icon, "detail": detail_text
            })

        h = ""
        total_kpi = len(kpi_list)
        n_crit = len(cards_by_prio["CRITIQUE"])
        n_haute = len(cards_by_prio["HAUTE"])
        n_moy = len(cards_by_prio["MOYENNE"])
        n_att = len(cards_by_prio["ATTEINT"])

        type_color = "#276749" if section_type == "perf" else "#2b6cb0"
        type_label = "PERFORMANCE" if section_type == "perf" else "QUALITE"
        type_icon = "⚡" if section_type == "perf" else "🎯"

        h += f'<div class="act-summary-row">'
        h += f'<div class="act-summary-item"><div class="asv" style="color:#c53030">{n_crit}</div><div class="asl">Critique</div></div>'
        h += f'<div class="act-summary-item"><div class="asv" style="color:#dd6b20">{n_haute}</div><div class="asl">Haute Priorite</div></div>'
        h += f'<div class="act-summary-item"><div class="asv" style="color:#38a169">{n_att}/{total_kpi}</div><div class="asl">Atteints</div></div>'
        h += f'</div>'

        for prio_key, prio_label, prio_color in [
            ("CRITIQUE", "Actions Critiques — Intervention Immediate Requise", "#c53030"),
            ("HAUTE", "Actions Haute Priorite — Planification Urgente", "#dd6b20"),
            ("MOYENNE", "Actions Moyenne Priorite — Amelioration Continue", "#d69e2e"),
            ("ATTEINT", "Objectifs Atteints — Maintien & Surveillance", "#38a169")
        ]:
            cards = cards_by_prio[prio_key]
            if not cards:
                continue
            h += f'<div class="act-section-title" style="background:linear-gradient(135deg,{prio_color},{"#fff" if prio_key=="ATTEINT" else "#2d3748"})">'
            h += f'<span>{prio_label}</span><span class="act-count">{len(cards)}</span></div>'
            h += '<div class="act-panel">'
            for card in cards:
                badge_class = f"badge-{prio_key.lower()}"
                bw = min(max(card["av"], 0), 100)
                if card["met"]:
                    bar_class = "green"
                elif abs(card["diff"]) > 20:
                    bar_class = "red"
                else:
                    bar_class = "orange"
                ecart_class = "zero" if card["met"] else ("pos" if card["diff"] > 0 and not is_lb(card["kpi"]) else "neg")
                ecart_sign = "+" if card["diff"] > 0 else ""
                h += f'<div class="act-card">'
                h += f'<div class="act-card-head">'
                h += f'<span class="act-card-icon">{card["icon"]}</span>'
                h += f'<span class="act-card-kpi">{card["kpi"]}</span>'
                h += f'<span class="act-card-badge {badge_class}">{prio_key if prio_key != "ATTEINT" else "ATTEINT"}</span>'
                h += f'</div>'
                h += f'<div class="act-card-body">'
                h += f'<div class="act-metrics">'
                h += f'<div class="act-metric"><div class="act-metric-val" style="color:{type_color}">{card["av"]:.1f}%</div><div class="act-metric-lbl">Actuel</div></div>'
                h += f'<div class="act-metric"><div class="act-metric-val" style="color:#4a5568">{card["tv"]:.0f}%</div><div class="act-metric-lbl">Cible</div></div>'
                h += f'<div class="act-metric"><div class="act-metric-val" style="color:{"#276749" if card["met"] else "#c53030"}">{ecart_sign}{card["diff"]:.1f}%</div><div class="act-metric-lbl">Ecart</div></div>'
                h += f'</div>'
                h += f'<div class="act-bar-wrap">'
                h += f'<div class="act-bar-label"><span>Progression vers la cible</span><span>{min(card["av"],card["tv"]):.0f}% / {card["tv"]:.0f}%</span></div>'
                h += f'<div class="act-bar-track"><div class="act-bar-fill {bar_class}" style="width:{bw}%"></div></div>'
                h += f'</div>'
                h += f'<div class="act-card-detail">{card["detail"]}</div>'
                h += f'</div>'
                h += f'<div class="act-card-footer">'
                h += f'<span class="act-ecart {ecart_class}">Ecart: {ecart_sign}{card["diff"]:.1f} pts</span>'
                h += f'<span style="font-size:11px;color:#718096;flex:1">Type: {type_label}</span>'
                h += f'<span style="font-size:14px">{"✅" if card["met"] else "⚠️"}</span>'
                h += f'</div>'
                h += f'</div>'
            h += '</div>'
        return h

    # ========== SYNTHESE & ACTIONS SANS BOUTON ==========
    def html_synthese_actions(pa, qa, pscores, qscores, targets):
        p_score_avg = np.mean(list(pscores.values())) if pscores else 0
        q_score_avg = np.mean(list(qscores.values())) if qscores else 0
        p_met = sum(1 for k in QK if (pa.get(k,0) <= targets.get(k,100) if is_lb(k) else pa.get(k,0) >= targets.get(k,100)))
        q_met = sum(1 for k in PK if (qa.get(k,0) <= targets.get(k,100) if is_lb(k) else qa.get(k,0) >= targets.get(k,100)))

        h = '<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:8px">'
        h += '<div class="act-synth-col">'
        h += '<div class="asct perf">⚡ Synthese Performance <span style="background:rgba(255,255,255,.2);padding:2px 10px;border-radius:12px;font-size:12px;font-weight:700">%d/%d atteints</span></div>' % (p_met, len(QK))
        h += '<div style="text-align:center;margin-bottom:10px">'
        sc_clr = "#276749" if p_score_avg >= 80 else ("#dd6b20" if p_score_avg >= 60 else "#c53030")
        h += '<div style="font-size:42px;font-weight:900;color:%s;line-height:1">%.1f%%</div>' % (p_score_avg, sc_clr)
        h += '<div style="font-size:11px;color:#718096;font-weight:600;text-transform:uppercase;letter-spacing:1px">Score Global Performance</div>'
        h += '</div>'
        for k in QK:
            av = pa.get(k, 0); tv = targets.get(k, 100)
            met = av <= tv if is_lb(k) else av >= tv
            bw = min(max(av, 0), 100)
            bg = "#48bb78" if met else "#fc8181"
            h += '<div class="act-synth-kpi-row">'
            h += '<span class="act-synth-kpi-name">%s</span>' % k
            h += '<span class="act-synth-kpi-bar"><span class="act-synth-kpi-fill" style="width:%s%%;background:%s"></span></span>' % (bw, bg)
            h += '<span class="act-synth-kpi-val" style="color:%s">%.1f%%</span>' % ("#276749" if met else "#c53030", av)
            h += '<span class="act-synth-kpi-status">%s</span>' % ("✅" if met else "❌")
            h += '</div>'
        h += '</div>'

        h += '<div class="act-synth-col">'
        h += '<div class="asct qual">🎯 Synthese Qualite <span style="background:rgba(255,255,255,.2);padding:2px 10px;border-radius:12px;font-size:12px;font-weight:700">%d/%d atteints</span></div>' % (q_met, len(PK))
        h += '<div style="text-align:center;margin-bottom:10px">'
        sc_clr = "#2b6cb0" if q_score_avg >= 80 else ("#dd6b20" if q_score_avg >= 60 else "#c53030")
        h += '<div style="font-size:42px;font-weight:900;color:%s;line-height:1">%.1f%%</div>' % (q_score_avg, sc_clr)
        h += '<div style="font-size:11px;color:#718096;font-weight:600;text-transform:uppercase;letter-spacing:1px">Score Global Qualite</div>'
        h += '</div>'
        for k in PK:
            av = qa.get(k, 0); tv = targets.get(k, 100)
            met = av <= tv if is_lb(k) else av >= tv
            bw = min(max(av, 0), 100)
            bg = "#48bb78" if met else "#fc8181"
            h += '<div class="act-synth-kpi-row">'
            h += '<span class="act-synth-kpi-name">%s</span>' % k
            h += '<span class="act-synth-kpi-bar"><span class="act-synth-kpi-fill" style="width:%s%%;background:%s"></span></span>' % (bw, bg)
            h += '<span class="act-synth-kpi-val" style="color:%s">%.1f%%</span>' % ("#276749" if met else "#c53030", av)
            h += '<span class="act-synth-kpi-status">%s</span>' % ("✅" if met else "❌")
            h += '</div>'
        h += '</div>'
        h += '</div>'
        return h

    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'; return h
    def html_kpi_bars(kpi_list,actuals,targets,title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,title)
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); met=av<=tv if is_lb(k) else av>=tv
            bw=min(max(av,0),100); bg=color_ok if met else color_fail
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(k,bw,bg,av)
        return h+'</div>'
    def html_bars(data,title,color):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color,title)
        for label,val in sorted(data,key=lambda x:x[1],reverse=True):
            bw=min(max(val,0),100)
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(label,bw,color,val)
        return h+'</div>'
    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        for p in sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True):
            pv,qv=pscores.get(p,0),qscores.get(p,0)
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>'%(p,min(max(pv,0),100),pv,min(max(qv,0),100),qv)
        return h+'</div>'
    def export_btn(df,filename):
        buf=io.BytesIO(); df.to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
        st.download_button("📥 Exporter Excel",data=buf,file_name=filename,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ===================== SIDEBAR =====================
    with st.sidebar:
        st.markdown("""<div style="padding:10px 0 4px 0"><div style="font-size:22px;margin-bottom:2px">⚙️</div><div style="font-size:14px;font-weight:800;color:white">Filtres & Parametres</div><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Configuration</div></div>""",unsafe_allow_html=True)
        st.markdown("---")
        show_filters=st.checkbox("Afficher les filtres",value=True,key="show_filters")
        if show_filters:
            unf=st.toggle("📁 Charger nouveaux fichiers",value=False,key="tf")
            ot_f=av_f=None; apm=[]
            if unf:
                ot_f=st.file_uploader("Fichier OT",type=["xlsx"],key="uot")
                av_f=st.file_uploader("Fichier AVIS",type=["xlsx"],key="uav")
            else:
                if os.path.exists("ot.xlsx"):
                    try:
                        _t=excr(pd.read_excel("ot.xlsx"))
                        apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                    except Exception: pass
                st.markdown("""<div style="background:rgba(255,255,255,.1);padding:6px 10px;border-radius:6px;border:1px solid rgba(255,255,255,.15)"><div style="font-size:11px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:1px">Donnees</div><div style="font-size:14px;color:white;font-weight:600;margin-top:2px">📅 %s</div></div>"""%fichier_date,unsafe_allow_html=True)
            st.markdown("---"); st.markdown("**🎯 Postes**")
            sp=st.multiselect("Poste",["All"]+apm,["All"],key="sp")
            st.markdown("**🏭 Atelier**")
            sa=st.multiselect("Atelier",["All","Sulfurique (PS)","Phosphorique (PP)","Engrais (TSP/REX)","Feed (MCP/DCP)"],["All"],key="sa")
            st.markdown("**🏢 Division**")
            sd=st.multiselect("Division",["All","SF1","SF2"],["All"],key="sd")
            st.markdown("---"); st.markdown("**📅 Periode**")
            dr=st.date_input("Date debut planifiee",value=(datetime(2025,1,1).date(),datetime.today().date()),format="DD/MM/YYYY",key="dr")
        else:
            unf=False; ot_f=av_f=None; apm=[]; sp=["All"]; sa=["All"]; sd=["All"]
            dr=(datetime(2025,1,1).date(),datetime.today().date())
            if os.path.exists("ot.xlsx"):
                try:
                    _t=excr(pd.read_excel("ot.xlsx"))
                    apm=sorted(_t[_t["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
                except Exception: pass

    # ===================== DATA LOADING =====================
    if not unf or (ot_f is not None and av_f is not None):
        try:
            if unf: raw_ot=pd.read_excel(ot_f); raw_av=pd.read_excel(av_f)
            else: raw_ot=pd.read_excel("ot.xlsx"); raw_av=pd.read_excel("avis.xlsx")
            raw_ot=excr(raw_ot); raw_av=excr(raw_av)
            for c in ["Créé le","Date de début planifiée","Date de clôture","Début réel","Fin réelle"]:
                if c in raw_ot.columns: raw_ot[c]=pd.to_datetime(raw_ot[c],errors="coerce")
            for c in ["Créé le","Début souhaité","Date de la clôture"]:
                if c in raw_av.columns: raw_av[c]=pd.to_datetime(raw_av[c],errors="coerce")
            if not apm: apm=sorted(raw_ot[raw_ot["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)]["Poste travail princ."].dropna().unique().tolist())
            if "All" in sp or not sp: sp=apm
            if "All" in sa or not sa: sa=["All"]
            if "All" in sd or not sd: sd=["All"]
            sdt=pd.to_datetime(dr[0]) if len(dr)==2 else pd.to_datetime(datetime(2025,1,1))
            edt=pd.to_datetime(dr[1]) if len(dr)==2 else pd.to_datetime(datetime.today())

            def mf(poste):
                p=str(poste).upper()
                if "All" not in sa:
                    m=False
                    if "Sulfurique (PS)" in sa and "PS" in p: m=True
                    if "Phosphorique (PP)" in sa and "PP" in p: m=True
                    if "Engrais (TSP/REX)" in sa and ("TSP" in p or "REX" in p): m=True
                    if "Feed (MCP/DCP)" in sa and ("MCP" in p or "DCP" in p): m=True
                    if not m: return False
                if "All" not in sd:
                    m=False
                    if "SF1" in sd and "SF1" in p: m=True
                    if "SF2" in sd and "SF2" in p: m=True
                    if not m: return False
                return True

            vp=[p for p in apm if mf(p) and p in sp]
            df=raw_ot[(raw_ot["Poste travail princ."].isin(vp))&(raw_ot["Date de début planifiée"].between(sdt,edt))].copy()
            avdf=raw_av[raw_av["Poste travail princ."].isin(vp)].copy()
            df=excr(df[df["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            avdf=excr(avdf[(avdf["Ordre"].isna())|(avdf["Ordre"].astype(str).str.strip().eq(""))].drop_duplicates())
            if "Statut système" in df.columns: df["Statut OT"]=df["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]
            df_dash=raw_ot[raw_ot["Poste travail princ."].isin(vp)].copy()
            df_dash=excr(df_dash[df_dash["Poste travail princ."].astype(str).str.startswith(("SF1","SF2"),na=False)].drop_duplicates())
            if "Statut système" in df_dash.columns: df_dash["Statut OT"]=df_dash["Statut système"].fillna("").astype(str).str.strip().str.split().str[0]

            now=pd.Timestamp.now()
            res=calc_kpis(df,avdf,now,vp); ckdf=res['ckdf']; dfp=res['dfp']
            res_d=calc_kpis(df_dash,avdf,now,vp); ckdf_d=res_d['ckdf']
            pa={k:round(ckdf[k].mean(),2) for k in QK}; qa={k:round(ckdf[k].mean(),2) for k in PK}
            pa_d={k:round(ckdf_d[k].mean(),2) for k in QK}; qa_d={k:round(ckdf_d[k].mean(),2) for k in PK}
            pscores={}; qscores={}
            for poste in ckdf.index:
                r=ckdf.loc[poste]
                pscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0
            pscores_d={}; qscores_d={}
            for poste in ckdf_d.index:
                r=ckdf_d.loc[poste]
                pscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in QK if k in r.index)/len(QK)*100) if QK else 0
                qscores_d[poste]=(sum(gscore(k,r[k],CIBLE[k]) for k in PK if k in r.index)/len(PK)*100) if PK else 0

            # ===================== ANOMALIES =====================
            all_ano=[]
            sub_p={"TAUX_REALISATION_CORRECTIF/PT":lambda d:d[(d["Nº appel pl.entret."].fillna(0)==0)&(~d["Statut OT"].isin(["CLOT","TCLO"]))],
                   "OT préparation <1 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]!="<1 mois")],
                   "OT préparation >3 mois":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["ap"]==">3 mois")],
                   "OT planification <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]!="<1 mois")],
                   "OT planification >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==0)&(d["alp"]==">3 mois")],
                   "OT exécution <1 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]!="<1 mois")],
                   "OT exécution >3 mois":lambda d:d[(d["Statut OT"]=="LANC")&(d["Contient SOPL"]==1)&(d["aex"]==">3 mois")]}
            sub_q={"OT LANC ESTIME":lambda d:d[(d["Statut OT"]=="LANC")&(d["OT LANC ESTIME"]=="NON")],
                   "Backlog préparation caractérisé":lambda d:d[(d["Statut OT"]=="CRÉÉ")&(d["Backlog preparation"]=="NON CARACTERISE")],
                   "Backlog planification caractérisé":lambda d:d[(d["Statut OT"]=="LANC")&(d["Backlog planification"]=="NON CARACTERISE")],
                   "OT CONFIME":lambda d:d[(d["Statut OT"].isin(["CLOT","TCLO"]))&(d["OT CONFIME"]=="NON")],
                   "OT_COR_EGAL":lambda d:d[(d["OT_COR_EGAL"]=="NON")],
                   "appel avis approuvé":lambda d:d[~d["Statut utilisateur"].str.contains("APRV",na=False)]}
            ano_p_rows=[]; ano_q_rows=[]
            for k,filt in sub_p.items():
                sub=filt(dfp); cnt=len(sub)
                ano_p_rows.append({"KPI":k,"Nombre anomalies":cnt,"_t":""})
                all_ano.extend(sub.to_dict("records"))
            ano_p_rows.append({"KPI":"Total general","Nombre anomalies":sum(r["Nombre anomalies"] for r in ano_p_rows),"_t":"total"})
            ano_p_cols=["KPI","Nombre anomalies"]

            for k,filt in sub_q.items():
                sub=filt(dfp); cnt=len(sub)
                ano_q_rows.append({"KPI":k,"Nombre anomalies":cnt,"_t":""})
                all_ano.extend(sub.to_dict("records"))
            ano_q_rows.append({"KPI":"Total general","Nombre anomalies":sum(r["Nombre anomalies"] for r in ano_q_rows),"_t":"total"})
            ano_q_cols=["KPI","Nombre anomalies"]
            ano_df=pd.DataFrame(all_ano) if all_ano else pd.DataFrame()

            # ===================== PERFORMANCE TABLE =====================
            pcols=["Poste de travail"]+QK+["Score Performance"]
            prows=[]
            for poste in ckdf.index:
                r=ckdf.loc[poste]; row={"Poste de travail":poste}
                for k in QK: row[k]=round(r[k],1) if k in r.index else 0
                row["Score Performance"]=round(pscores[poste],2)
                prows.append(row)
            prows.append({"Poste de travail":"MOYENNE","_t":"total"})
            for k in QK: prows[-1][k]=round(pa.get(k,0),1)
            prows[-1]["Score Performance"]=round(np.mean(list(pscores.values())),2)
            prows.append({"Poste de travail":"CIBLE","_t":"cible"})
            for k in QK: prows[-1][k]=CIBLE.get(k,"-")
            prows[-1]["Score Performance"]=100

            # ===================== QUALITE TABLE =====================
            qcols=["Poste de travail"]+PK+["Score Qualite"]
            qrows=[]
            for poste in ckdf.index:
                r=ckdf.loc[poste]; row={"Poste de travail":poste}
                for k in PK: row[k]=round(r[k],1) if k in r.index else 0
                row["Score Qualite"]=round(qscores[poste],2)
                qrows.append(row)
            qrows.append({"Poste de travail":"MOYENNE","_t":"total"})
            for k in PK: qrows[-1][k]=round(qa.get(k,0),1)
            qrows[-1]["Score Qualite"]=round(np.mean(list(qscores.values())),2)
            qrows.append({"Poste de travail":"CIBLE","_t":"cible"})
            for k in PK: qrows[-1][k]=CIBLE.get(k,"-")
            qrows[-1]["Score Qualite"]=100

            # ===================== PIE CHART DATA =====================
            statut_pie=dfp[dfp["Statut OT"].isin(["CRÉÉ","LANC","CLOT","TCLO"])]["Statut OT"].value_counts().reset_index()
            statut_pie.columns=["Statut","Nombre"]
            age_prep_pie=dfp[dfp["Statut OT"]=="CRÉÉ"]["ap"].value_counts().reset_index()
            age_prep_pie.columns=["Age","Nombre"]
            age_plan_pie=dfp[(dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==0)]["alp"].value_counts().reset_index()
            age_plan_pie.columns=["Age","Nombre"]
            age_exec_pie=dfp[(dfp["Statut OT"]=="LANC")&(dfp["Contient SOPL"]==1)]["aex"].value_counts().reset_index()
            age_exec_pie.columns=["Age","Nombre"]
            atelier_pie=dfp.copy()
            atelier_pie["Atelier"]=atelier_pie["Poste travail princ."].apply(get_atelier)
            atelier_pie=atelier_pie["Atelier"].value_counts().reset_index()
            atelier_pie.columns=["Atelier","Nombre"]
            metier_pie=dfp.copy()
            metier_pie["Metier"]=metier_pie["Poste travail princ."].apply(get_metier)
            metier_pie=metier_pie["Metier"].value_counts().reset_index()
            metier_pie.columns=["Metier","Nombre"]

            # ===================== SAVE TO EXCEL =====================
            save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_rows,ano_p_cols,ano_q_rows,ano_q_cols,fichier_date)

            # ===================== HISTORICAL =====================
            hist_path=os.path.join("kpis","indicateurs_kpis.xlsx")
            hist_df=load_historical_kpis(hist_path)
            var_df=calculate_variations(hist_df)
            journal_df=generate_journal(var_df)
            top5_df,bottom5_df=calculate_rankings(var_df)

            # ===================== TABS =====================
            tabs=st.tabs(["📊 Tableau de Bord","⚡ Performance","🎯 Qualité","📋 Synthèse & Actions","📈 Analyse","🚨 Anomalies"])
            p_score_avg=round(np.mean(list(pscores.values())),1) if pscores else 0
            q_score_avg=round(np.mean(list(qscores.values())),1) if qscores else 0
            total_ot=len(dfp)

            # ========== TAB 1: TABLEAU DE BORD ==========
            with tabs[0]:
                st.markdown('<div class="mh"><h1>📊 Tableau de Bord KPI — Maintenance</h1><span class="db">📅 %s</span></div>'%fichier_date,unsafe_allow_html=True)
                st.markdown('<div class="cr">' +
                    '<div class="cc c1"><div class="cv">%d</div><div class="cl">Total OT</div></div>'%total_ot +
                    '<div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div>'%p_score_avg +
                    '<div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Qualite</div></div>'%q_score_avg +
                    '<div class="cc c4"><div class="cv">%d</div><div class="cl">Postes Suivis</div></div>'%len(vp) +
                    '</div>',unsafe_allow_html=True)
                st.markdown('<div class="dgrid">',unsafe_allow_html=True)
                c1,c2=st.columns(2)
                with c1: st.markdown(html_kpi_bars(QK,pa,CIBLE,"Indicateurs de Performance","#38a169","#e53e3e"),unsafe_allow_html=True)
                with c2: st.markdown(html_kpi_bars(PK,qa,CIBLE,"Indicateurs de Qualite","#3182ce","#e53e3e"),unsafe_allow_html=True)
                st.markdown('</div>',unsafe_allow_html=True)
                st.markdown(html_grouped_bars(vp,pscores,qscores,"Scores par Poste de Travail"),unsafe_allow_html=True)
                st.markdown('<div class="dgrid">',unsafe_allow_html=True)
                pc1,pc2=st.columns(2)
                with pc1:
                    fig1=anl_pie_of_pie_chart(statut_pie,"Statut","Nombre","Repartition par Statut OT",px.colors.qualitative.Set2,threshold=5)
                    if fig1: st.plotly_chart(fig1,use_container_width=True)
                with pc2:
                    fig2=anl_pie_of_pie_chart(atelier_pie,"Atelier","Nombre","Repartition par Atelier",px.colors.qualitative.Pastel,threshold=8)
                    if fig2: st.plotly_chart(fig2,use_container_width=True)
                st.markdown('</div>',unsafe_allow_html=True)
                st.markdown('<div class="dgrid">',unsafe_allow_html=True)
                pc3,pc4=st.columns(2)
                with pc3:
                    fig3=anl_pie_of_pie_chart(metier_pie,"Metier","Nombre","Repartition par Metier",px.colors.qualitative.Dark2,threshold=8)
                    if fig3: st.plotly_chart(fig3,use_container_width=True)
                with pc4:
                    if not age_prep_pie.empty:
                        fig4=anl_pie_of_pie_chart(age_prep_pie,"Age","Nombre","Age Preparation OT (CRÉÉ)",["#48bb78","#ecc94b","#fc8181"],threshold=5)
                        if fig4: st.plotly_chart(fig4,use_container_width=True)
                st.markdown('</div>',unsafe_allow_html=True)

            # ========== TAB 2: PERFORMANCE ==========
            with tabs[1]:
                st.markdown('<div class="mh"><h1>⚡ Indicateurs de Performance</h1><span class="db">%.1f%%</span></div>'%p_score_avg,unsafe_allow_html=True)
                st.markdown('<div class="cr">' +
                    '<div class="cc c1"><div class="cv">%.1f%%</div><div class="cl">Taux Realisation</div></div>'%pa.get("TAUX_REALISATION_CORRECTIF/PT",0) +
                    '<div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">Prep &lt;1 mois</div></div>'%pa.get("OT préparation <1 mois",0) +
                    '<div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Plan &lt;1 mois</div></div>'%pa.get("OT planification <1 mois",0) +
                    '<div class="cc c4"><div class="cv">%.1f%%</div><div class="cl">Exec &lt;1 mois</div></div>'%pa.get("OT exécution <1 mois",0) +
                    '</div>',unsafe_allow_html=True)
                st.markdown(html_table(prows,pcols,"pt",{"Score Performance"}),unsafe_allow_html=True)
                st.markdown(html_classement(pscores,"#38a169"),unsafe_allow_html=True)
                st.markdown('<div class="dgrid">',unsafe_allow_html=True)
                apc1,apc2=st.columns(2)
                with apc1:
                    if not age_plan_pie.empty:
                        fig5=anl_pie_of_pie_chart(age_plan_pie,"Age","Nombre","Age Planification OT",["#48bb78","#ecc94b","#fc8181"],threshold=5)
                        if fig5: st.plotly_chart(fig5,use_container_width=True)
                with apc2:
                    if not age_exec_pie.empty:
                        fig6=anl_pie_of_pie_chart(age_exec_pie,"Age","Nombre","Age Execution OT",["#48bb78","#ecc94b","#fc8181"],threshold=5)
                        if fig6: st.plotly_chart(fig6,use_container_width=True)
                st.markdown('</div>',unsafe_allow_html=True)

            # ========== TAB 3: QUALITE ==========
            with tabs[2]:
                st.markdown('<div class="mh"><h1>🎯 Indicateurs de Qualite</h1><span class="db">%.1f%%</span></div>'%q_score_avg,unsafe_allow_html=True)
                st.markdown('<div class="cr">' +
                    '<div class="cc c1"><div class="cv">%.1f%%</div><div class="cl">Avis Approuves</div></div>'%qa.get("appel avis approuvé",0) +
                    '<div class="cc c2"><div class="cv">%.1f%%</div><div class="cl">OT Estimes</div></div>'%qa.get("OT LANC ESTIME",0) +
                    '<div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Backlog Prep</div></div>'%qa.get("Backlog préparation caractérisé",0) +
                    '<div class="cc c4"><div class="cv">%.1f%%</div><div class="cl">OT Confirmes</div></div>'%qa.get("OT CONFIME",0) +
                    '</div>',unsafe_allow_html=True)
                st.markdown(html_table(qrows,qcols,"qt",{"Score Qualite"}),unsafe_allow_html=True)
                st.markdown(html_classement(qscores,"#3182ce"),unsafe_allow_html=True)

            # ========== TAB 4: SYNTHESE & ACTIONS (SANS BOUTON) ==========
            with tabs[3]:
                st.markdown('<div class="mh"><h1>📋 Synthese & Actions Recommandees</h1><span class="db">📅 %s</span></div>'%fichier_date,unsafe_allow_html=True)

                # --- Synthese cote a cote (sans bouton) ---
                st.markdown('<p class="stl s">Synthese Globale — Performance & Qualite</p>',unsafe_allow_html=True)
                st.markdown(html_synthese_actions(pa, qa, pscores, qscores, CIBLE),unsafe_allow_html=True)

                # --- Actions Recommandees Professionnelles ---
                st.markdown('<p class="stl a">Actions Recommandees — Performance</p>',unsafe_allow_html=True)
                st.markdown(html_professional_actions(QK, pa, CIBLE, ACT_MAP, ACT_DETAIL, "perf"),unsafe_allow_html=True)

                st.markdown('<p class="stl a">Actions Recommandees — Qualite</p>',unsafe_allow_html=True)
                st.markdown(html_professional_actions(PK, qa, CIBLE, ACT_MAP, ACT_DETAIL, "qual"),unsafe_allow_html=True)

            # ========== TAB 5: ANALYSE ==========
            with tabs[4]:
                st.markdown('<div class="mh"><h1>📈 Analyse des Tendances</h1></div>',unsafe_allow_html=True)
                if journal_df.empty:
                    st.markdown('<div class="es">📊 Aucune donnee historique suffisante pour l\'analyse des tendances.<br>Les variations seront disponibles a partir de la deuxieme extraction.</div>',unsafe_allow_html=True)
                else:
                    st.markdown('<p class="stl s">Journal des Variations Significatives</p>',unsafe_allow_html=True)
                    jcols=["Date precedente","Date actuelle","Poste","Type","KPI","Valeur precedente","Valeur actuelle","Ecart","Ecart %","Tendance","Sens"]
                    jrows=[{c:r[c] for c in jcols} for _,r in journal_df.iterrows()]
                    st.markdown(html_table(jrows,jcols,"st"),unsafe_allow_html=True)
                    if not top5_df.empty:
                        st.markdown('<p class="stl p">Top 5 — Amelioration</p>',unsafe_allow_html=True)
                        tcols=["Poste","Score variation"]; trows=[{c:r[c] for c in tcols} for _,r in top5_df.iterrows()]
                        st.markdown(html_table(trows,tcols,"pt"),unsafe_allow_html=True)
                    if not bottom5_df.empty:
                        st.markdown('<p class="stl a">Bottom 5 — Degradation</p>',unsafe_allow_html=True)
                        bcols=["Poste","Score variation"]; brows=[{c:r[c] for c in bcols} for _,r in bottom5_df.iterrows()]
                        st.markdown(html_table(brows,bcols,"at"),unsafe_allow_html=True)

            # ========== TAB 6: ANOMALIES ==========
            with tabs[5]:
                st.markdown('<div class="mh"><h1>🚨 Detail des Anomalies</h1></div>',unsafe_allow_html=True)
                st.markdown('<p class="stl a">Anomalies Performance</p>',unsafe_allow_html=True)
                st.markdown(html_ano(ano_p_rows,ano_p_cols),unsafe_allow_html=True)
                st.markdown('<p class="stl a">Anomalies Qualite</p>',unsafe_allow_html=True)
                st.markdown(html_ano(ano_q_rows,ano_q_cols),unsafe_allow_html=True)
                if not ano_df.empty:
                    st.markdown('<p class="stl c">Detail des OT Anormaux (%d)</p>'%len(ano_df),unsafe_allow_html=True)
                    acols=["Ordre","Poste travail princ.","Statut OT","Statut utilisateur","Description"]
                    display_cols=[c for c in acols if c in ano_df.columns]
                    if display_cols:
                        arows=[{c:str(r[c])[:50] for c in display_cols} for _,r in ano_df.head(200).iterrows()]
                        st.markdown(html_table(arows,display_cols,"at"),unsafe_allow_html=True)
                    if len(ano_df)>200:
                        st.markdown('<div class="es">... et %d autres lignes</div>'%(len(ano_df)-200),unsafe_allow_html=True)
                    buf=io.BytesIO(); ano_df.head(2000).to_excel(buf,index=False,engine='openpyxl'); buf.seek(0)
                    st.download_button("📥 Exporter les Anomalies",data=buf,file_name="anomalies.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.error(f"Erreur de chargement : {str(e)}")
            import traceback; st.code(traceback.format_exc())
    else:
        if unf:
            st.markdown("""<div style="min-height:60vh;display:flex;flex-direction:column;align-items:center;justify-content:center">
            <div style="font-size:64px;margin-bottom:20px">📁</div>
            <h1 style="font-size:28px;color:#1e3a5f;font-weight:800">Chargement des Fichiers</h1>
            <p style="font-size:16px;color:#718096;margin-top:8px">Veuillez charger les fichiers OT et AVIS dans le panneau de gauche.</p></div>""",unsafe_allow_html=True)
        else:
            st.markdown("""<div style="min-height:60vh;display:flex;flex-direction:column;align-items:center;justify-content:center">
            <div style="font-size:64px;margin-bottom:20px">📂</div>
            <h1 style="font-size:28px;color:#1e3a5f;font-weight:800">Fichiers Non Trouves</h1>
            <p style="font-size:16px;color:#718096;margin-top:8px">Placez les fichiers <b>ot.xlsx</b> et <b>avis.xlsx</b> dans le meme dossier que ce script.</p></div>""",unsafe_allow_html=True)

if __name__ == "__main__":
    main()
