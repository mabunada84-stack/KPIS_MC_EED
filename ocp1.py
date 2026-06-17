# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import numpy as np
import io, locale, random, time, os, hashlib, json
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
st.set_page_config(layout="wide", page_title="Dashboard KPI")
# ============================================================

QK = ["TAUX_REALISATION_CORRECTIF/PT","OT préparation <1 mois","OT préparation >3 mois",
      "OT préparation 1mois< <3mois","OT planification <1 mois","OT planification >3 mois",
      "OT planification 1mois< <3mois","OT exécution <1 mois","OT exécution >3 mois",
      "OT exécution 1mois< <3mois",
      "Performance Graissage","Performance Inspection","Performance Appels Systématiques"]
PK = ["appel avis approuvé","OT LANC ESTIME","Backlog préparation caractérisé",
      "Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL",
      "OT Fiabilité","Total Avis de Panne"]
ALL_KPI = QK + PK

CIBLE = {"TAUX_REALISATION_CORRECTIF/PT":85,"OT préparation <1 mois":80,"OT préparation >3 mois":5,
         "OT préparation 1mois< <3mois":15,"OT planification <1 mois":80,"OT planification >3 mois":5,
         "OT planification 1mois< <3mois":15,"OT exécution <1 mois":80,"OT exécution >3 mois":5,
         "OT exécution 1mois< <3mois":15,"appel avis approuvé":95,"OT LANC ESTIME":100,
         "Backlog préparation caractérisé":100,"Backlog planification caractérisé":100,
         "OT CONFIME":100,"OT_COR_EGAL":100,
         "Performance Graissage":95,"Performance Inspection":95,"Performance Appels Systématiques":95,
         "OT Fiabilité":100,"Total Avis de Panne":100}

ACT_MAP = {"TAUX_REALISATION_CORRECTIF/PT":"Ameliorer le taux de realisation des OT.",
           "OT préparation <1 mois":"Reduire l'age de preparation des OT (< 1 mois).",
           "OT préparation >3 mois":"Traiter les OT avec preparation > 3 mois.",
           "OT planification <1 mois":"Reduire l'age de planification des OT (< 1 mois).",
           "OT planification >3 mois":"Traiter les OT avec planification > 3 mois.",
           "OT exécution <1 mois":"Reduire l'age d'execution des OT (< 1 mois).",
           "OT exécution >3 mois":"Traiter les OT avec execution > 3 mois.",
           "OT LANC ESTIME":"Estimer les couts des OT lances.",
           "Backlog préparation caractérisé":"Caracteriser le backlog de preparation.",
           "Backlog planification caractérisé":"Caracteriser le backlog de planification.",
           "OT CONFIME":"Confirmer les OT termines.",
           "OT_COR_EGAL":"Rapprocher les couts reels et budgetes.",
           "appel avis approuvé":"Creer un OT pour les avis sans ordre.",
           "OT préparation 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT planification 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "OT exécution 1mois< <3mois":"Reduire les OT entre 1 et 3 mois.",
           "Performance Graissage":"Ameliorer le taux de realisation des OT de graissage (Type 350).",
           "Performance Inspection":"Ameliorer le taux de realisation des OT d'inspection (Types 290,300,310).",
           "Performance Appels Systématiques":"Ameliorer le taux de realisation des appels systematiques (Type 360).",
           "OT Fiabilité":"Maintenir la fiabilite des OT a 100%.",
           "Total Avis de Panne":"Maintenir le suivi des avis de panne a 100%."}

LOWER_BETTER = ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois",
                "OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]

MP_KW = ["CRPR ATPD","CRPR ATMR","CRPR ATER","CRPR ATRS","CRPR ATMO","ATPD","ATMR","ATER","ATRS","ATMO"]
MPLAN_KW = ["ATPL ATEI","ATPL ATAL","ATPL ATER","ATPL AGAR","ATPL ATHS","ATEI","ATAL","ATAS","AGAR","ATHS"]

CHANGELOG = [
    {"version":"2.2","date":"2025-06-20","changes":[
        "Page Anomalies : inversion lignes/colonnes (Postes en lignes, KPI en colonnes)",
        "Nouvelle page Backlog dediee avec 3 sections : OT OMS, OT Thermographie, Tous les OT",
        "Pie charts ameliores : regroupement secteurs minces, effet pull-out, donut, labels enrichis",
        "Remplissage des anomalies par nombre d'OT et d'Avis"
    ]},
    {"version":"2.1","date":"2025-06-18","changes":[
        "Deplacement KPI Graissage/Inspection/Systematiques de Qualite vers Performance",
        "Nouveau tableau OT OMS par Poste et Statut OT avec 2 Pie charts",
        "Nouveau tableau OT Thermographie par Poste et Statut OT avec 2 Pie charts",
        "Nouveau tableau Tous les OT par Poste et Statut OT avec 2 Pie charts",
        "Page Anomalies simplifiee : resume KPI x Poste avec coloriage",
        "Page Suivi & Evolution : synthese entre deux dates par poste"
    ]},
    {"version":"2.0","date":"2025-06-15","changes":[
        "KPI Taux realisation correctif/PT : ajout filtre SOPL=1, numerateur CLOT+TCLO, total=0 => 100%",
        "KPI Age backlog preparation : filtre Statut OT=CRE + Statut utilisateur contient CRPR",
        "KPI Age backlog planification : filtre Statut OT=LANC + Statut utilisateur contient ATPL",
        "Nouveau KPI Performance Graissage (Type 350) - Seuils V/J/R",
        "Nouveau KPI Performance Inspection (Types 290,300,310) - Exclusion dates futures - Seuils V/J/R",
        "Nouveau KPI Performance Appels Systematiques (Type 360) - Exclusion dates futures - Seuils V/J/R",
        "Nouveaux KPI Qualite Appels : OT Fiabilite (100%), Total Avis de Panne (100%)",
        "Mise en place mecanisme de cache pour eviter les recalculs systematiques",
        "Activation du suivi des ameliorations et evolutions (changelog)"
    ]}
]

CONSIGNES_HSE = [
    "Port obligatoire des EPI avant toute intervention.","Port obligatoire du casque de securite.",
    "Port obligatoire des lunettes de protection.","Port obligatoire des gants adaptes au travail.",
    "Utiliser les protections auditives dans les zones bruyantes.","Verifier l'absence de tension avant toute intervention electrique.",
    "Respecter la procedure de consignation et deconsignation.","Ne jamais intervenir sur un equipement en marche.",
    "Baliser et securiser la zone de travail.","Maintenir le poste de travail propre et ordonne.",
    "Verifier l'etat des outils avant utilisation.","Utiliser uniquement du materiel homologue.",
    "Respecter les permis de travail en vigueur.","Identifier les risques avant de commencer une tache.",
    "Signaler immediatement toute situation dangereuse.","Signaler tout incident ou presque accident.",
    "Ne jamais neutraliser un dispositif de securite.","Verifier les detecteurs de gaz avant utilisation.",
    "Verifier la bonne ventilation des zones de travail.","Respecter les regles des espaces confines.",
    "Controler l'atmosphere avant d'entrer dans un espace confine.","Utiliser les points d'ancrage pour les travaux en hauteur.",
    "Verifier l'etat des echafaudages avant utilisation.","Securiser les outils lors des travaux en hauteur.",
    "Ne pas travailler seul lors d'operations a risque.","Controler les elingues avant chaque levage.",
    "Respecter les limites de charge des equipements.","Verifier l'etat des appareils de levage.",
    "Maintenir les voies de circulation degagees.","Respecter la signalisation de securite.",
    "Verifier les extincteurs a proximite du chantier.","Connaitre les issues de secours les plus proches.",
    "Respecter les procedures d'arret d'urgence.","Verifier les flexibles et raccords avant mise en service.",
    "Controler les fuites avant demarrage d'un equipement.","Respecter les distances de securite.",
    "Ne jamais contourner une procedure HSE.","Porter les EPI adaptes au risque identifie.",
    "Prevenir son responsable avant toute intervention particuliere.","Analyser les risques avant chaque demarrage de chantier.",
    "Verifier la stabilite des equipements.","Utiliser les bons outils pour la bonne tache.",
    "Respecter les consignes specifiques du chantier.","Ne jamais prendre de raccourci au detriment de la securite.",
    "Arreter immediatement les travaux en cas de danger.","Proteger l'environnement lors des interventions.",
    "Collecter et trier correctement les dechets.","Eviter toute pollution accidentelle.",
    "Respecter les consignes de stockage des produits dangereux.","Lire les fiches de securite avant manipulation.",
    "Verifier les equipements avant chaque prise de poste.","S'assurer de la disponibilite des moyens de secours.",
    "Communiquer clairement avec l'equipe avant intervention.","Respecter les regles de circulation des engins.",
    "Garder une vigilance permanente sur son environnement.","Prendre le temps d'effectuer le travail en securite.",
    "La securite est l'affaire de tous.","Chaque incident peut etre evite par la prevention.",
    "Aucun travail n'est plus urgent que la securite.","Zero accident commence par un comportement sur."]

def compute_cache_key(file_date, filters_dict, now_str):
    key_data = {"file_date": file_date, "filters": filters_dict, "now": now_str}
    return hashlib.md5(json.dumps(key_data, sort_keys=True, default=str).encode()).hexdigest()

def get_date_from_file():
    if os.path.exists("date.txt"):
        try:
            with open("date.txt","r",encoding="utf-8") as f: return f.read().strip()
        except Exception: pass
    return datetime.now().strftime("%d/%m/%Y")

def save_kpis_to_excel(prows,pcols,qrows,qcols,ano_p_r,ano_p_c,ano_q_r,ano_q_c,sheet_name):
    kpis_dir="kpis"; os.makedirs(kpis_dir,exist_ok=True)
    filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
    sn=str(sheet_name).replace("/","-").replace("\\","-").replace("*","").replace("?","").replace("[","").replace("]","")[:31]
    hf=Font(bold=True,color="FFFFFF",size=10); hfl=PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
    tf=Font(bold=True,size=12,color="1E3A5F")
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    try: wb=load_workbook(filepath)
    except Exception: wb=Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    if sn in wb.sheetnames: del wb[sn]
    ws=wb.create_sheet(sn); rn=1
    def ws_sec(title,cols,rows,sr):
        ws.cell(row=sr,column=1,value=title).font=tf; sr+=1
        for j,c in enumerate(cols,1):
            cl=ws.cell(row=sr,column=j,value=c); cl.font=hf; cl.fill=hfl; cl.alignment=Alignment(horizontal='center'); cl.border=tb
        sr+=1
        for r in rows:
            for j,c in enumerate(cols,1):
                cl=ws.cell(row=sr,column=j,value=r.get(c,"")); cl.border=tb; cl.alignment=Alignment(horizontal='center')
            sr+=1
        return sr+1
    rn=ws_sec("INDICATEURS DE PERFORMANCE",pcols,prows,rn)
    if ano_p_c and ano_p_r: rn=ws_sec("ANOMALIES PERFORMANCE",ano_p_c,ano_p_r,rn)
    rn=ws_sec("INDICATEURS DE QUALITE",qcols,qrows,rn)
    if ano_q_c and ano_q_r: rn=ws_sec("ANOMALIES QUALITE",ano_q_c,ano_q_r,rn)
    try: wb.save(filepath)
    except Exception: pass

def load_historical_kpis(filepath):
    if not os.path.exists(filepath): return pd.DataFrame()
    try: wb=load_workbook(filepath,read_only=True,data_only=True)
    except Exception: return pd.DataFrame()
    records=[]; section=None; headers=None
    for sheet_name in wb.sheetnames:
        try:
            ws=wb[sheet_name]; rows_data=list(ws.iter_rows(values_only=True))
            for row in rows_data:
                cell0=str(row[0]).strip() if row[0] else ""
                if "INDICATEURS DE PERFORMANCE" in cell0.upper(): section="perf"; headers=None; continue
                elif "INDICATEURS DE QUALITE" in cell0.upper(): section="qual"; headers=None; continue
                elif "ANOMALIES" in cell0.upper(): section=None; continue
                if section and headers is None and cell0:
                    headers=[str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("CIBLE","Total general",""):
                    entry={"Date":sheet_name}
                    for j,h in enumerate(headers):
                        if j<len(row): entry[h]=row[j]
                    entry["_section"]=section; records.append(entry)
        except Exception: continue
    wb.close()
    if not records: return pd.DataFrame()
    df=pd.DataFrame(records)
    df["Date_parsed"]=pd.to_datetime(df["Date"].str.replace("-","/"),format="%d/%m/%Y",errors="coerce")
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df):
    if hist_df.empty or "Date" not in hist_df.columns: return pd.DataFrame()
    dates=sorted(hist_df["Date"].unique())
    if len(dates)<2: return pd.DataFrame()
    perf_df=hist_df[hist_df["_section"]=="perf"].copy()
    qual_df=hist_df[hist_df["_section"]=="qual"].copy()
    variations=[]
    for i in range(1,len(dates)):
        prev_date,curr_date=dates[i-1],dates[i]
        prev_perf=perf_df[perf_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf=perf_df[perf_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual=qual_df[qual_df["Date"]==prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual=qual_df[qual_df["Date"]==curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        for sec_name,prev_d,curr_d,kpi_list in [("Performance",prev_perf,curr_perf,QK+["Score Performance"]),("Qualite",prev_qual,curr_qual,PK+["Score Qualite"])]:
            for poste in set(prev_d.index)&set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns: continue
                    try: pv=float(prev_d.loc[poste,kpi])
                    except Exception: continue
                    try: cv=float(curr_d.loc[poste,kpi])
                    except Exception: continue
                    diff=cv-pv; pct=(diff/pv*100) if pv!=0 else (100 if cv!=0 else 0)
                    if abs(diff)<=0.5: trend="stabilite"
                    elif diff>0.5: trend="hausse"
                    else: trend="baisse"
                    variations.append({"Date precedente":prev_date,"Date actuelle":curr_date,"Poste":poste,
                        "Type":sec_name,"KPI":kpi,"Valeur precedente":round(pv,2),"Valeur actuelle":round(cv,2),
                        "Ecart":round(diff,2),"Ecart %":round(pct,2),"Tendance":trend})
    return pd.DataFrame(variations)

def generate_journal(var_df):
    if var_df.empty: return pd.DataFrame()
    j=var_df.copy(); j["Significatif"]=j["Ecart %"].abs()>=5
    j=j[j["Significatif"]].copy()
    j["Sens"]=j.apply(lambda r:"Amelioration" if ((r["Tendance"]=="hausse" and r["KPI"] not in LOWER_BETTER) or (r["Tendance"]=="baisse" and r["KPI"] in LOWER_BETTER)) else "Degradation",axis=1)
    return j.sort_values(["Date actuelle","Sens","Ecart %"],ascending=[True,False,False])

def calculate_rankings(var_df):
    if var_df.empty: return pd.DataFrame(),pd.DataFrame()
    scores={}
    for poste in var_df["Poste"].unique():
        pv=var_df[var_df["Poste"]==poste].copy()
        scores[poste]=sum((-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"]) for _,r in pv.iterrows())
    ranked=sorted(scores.items(),key=lambda x:x[1],reverse=True)
    return pd.DataFrame(ranked[:5],columns=["Poste","Score variation"]),pd.DataFrame(ranked[-5:][::-1],columns=["Poste","Score variation"])

def inject_custom_css():
    st.markdown("""<style>
    section[data-testid="stSidebar"]{width:250px!important}
    section[data-testid="stSidebar"][aria-expanded="false"]{width:0px!important}
    .main .block-container{max-width:100%!important;width:100%!important;padding-left:0.5rem!important;padding-right:0.5rem!important}
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    :root{--p:#1e3a5f;--pl:#2c5282;--b:#e2e8f0;--r:10px}
    *{box-sizing:border-box;margin:0;padding:0}
    .stApp{background:#edf2f7;font-family:'Inter',sans-serif}
    .main .block-container{padding-top:.8rem;padding-bottom:.8rem}
    .stTabs,.stTabs>div,.stTabs [data-baseweb="tab-list"]{width:100%!important;max-width:100%!important}
    .mh{background:linear-gradient(135deg,var(--p),var(--pl));padding:12px 20px;border-radius:var(--r);margin-bottom:6px;box-shadow:0 6px 20px rgba(0,0,0,.1);overflow:hidden}
    .mh h1{color:#fff;font-size:20px;font-weight:800;margin:0;display:inline}
    .mh .db{float:right;background:rgba(255,255,255,.15);padding:3px 12px;border-radius:14px;color:#fff;font-size:14px;font-weight:500;border:1px solid rgba(255,255,255,.2);margin-top:2px}
    .cr{display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:6px}
    .cc{background:#fff;border-radius:var(--r);padding:10px 12px;box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid var(--b);text-align:center}
    .cc .cv{font-size:26px;font-weight:900;line-height:1}
    .cc .cl{font-size:11px;color:#718096;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:2px}
    .cc.c1{border-top:3px solid #3182ce}.cc.c1 .cv{color:#2b6cb0}
    .cc.c2{border-top:3px solid #38a169}.cc.c2 .cv{color:#276749}
    .cc.c3{border-top:3px solid #805ad5}.cc.c3 .cv{color:#6b46c1}
    .cc.c4{border-top:3px solid #e53e3e}.cc.c4 .cv{color:#c53030}
    .stl{font-size:15px;font-weight:700;color:var(--p);margin:6px 0 2px 0;padding-left:10px;border-left:3px solid var(--pl)}
    .stl.q{border-left-color:#3182ce}.stl.p{border-left-color:#38a169}.stl.a{border-left-color:#e53e3e}.stl.c{border-left-color:#805ad5}.stl.s{border-left-color:#d69e2e}.stl.bl{border-left-color:#dd6b20}
    .tw{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;display:block;overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0}
    .tw thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;padding:5px 6px;border:none;white-space:nowrap;position:sticky;top:0;z-index:10}
    .tw.qt thead th{background:linear-gradient(135deg,#2b6cb0,#3182ce)}
    .tw.pt thead th{background:linear-gradient(135deg,#276749,#38a169)}
    .tw.at thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw.st thead th{background:linear-gradient(135deg,#975a16,#d69e2e)}
    .tw.omt thead th{background:linear-gradient(135deg,#6b46c1,#805ad5)}
    .tw.tht thead th{background:linear-gradient(135deg,#9b2c2c,#e53e3e)}
    .tw.tot thead th{background:linear-gradient(135deg,#1a365d,#2c5282)}
    .tw.ant thead th{background:linear-gradient(135deg,#c53030,#e53e3e)}
    .tw tbody td{padding:4px 6px;border-bottom:1px solid #edf2f7;white-space:nowrap}
    .tw tbody tr:nth-child(even) td{background:#f7fafc}
    .tw tbody tr:hover td{background:#ebf8ff!important}
    .cb td{background:#2b6cb0!important;color:#fff!important;font-weight:700!important;font-size:12px!important}
    .tr td{background:#e2e8f0!important;font-weight:800!important;font-size:12px!important}
    .stTabs [data-baseweb="tab-list"]{gap:3px;background:#e2e8f0;padding:3px;border-radius:6px;margin-bottom:4px}
    .stTabs [data-baseweb="tab"]{border-radius:5px;padding:6px 14px;font-weight:600;font-size:14px}
    .stTabs [aria-selected="true"]{background:#fff!important;color:var(--p)!important;box-shadow:0 2px 5px rgba(0,0,0,.07)}
    .ca{background:#fff;border-radius:var(--r);padding:10px;margin-top:4px;border:1px solid var(--b);box-shadow:0 1px 4px rgba(0,0,0,.02)}
    .ca .ct{font-size:14px;font-weight:700;margin-bottom:6px;padding-bottom:4px;border-bottom:1px solid var(--b)}
    .car{display:flex;align-items:center;margin-bottom:4px;font-size:12px}
    .car:last-child{margin-bottom:0}
    .car .cal{width:260px;font-weight:600;color:var(--p);text-align:right;padding-right:8px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .car .cab{flex:1;height:24px;background:#edf2f7;border-radius:4px;overflow:hidden}
    .car .caf{height:100%;border-radius:4px;transition:width .3s}
    .car .cav-out{font-size:12px;font-weight:800;color:#1a202c;min-width:55px;text-align:right;padding-left:6px}
    .gbr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .gbr:last-child{border:none}
    .gbr-l{width:160px;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:11px}
    .gbr-g{display:flex;align-items:center;gap:4px;flex:1}
    .gbr-w{flex:1;height:20px;background:#edf2f7;border-radius:3px;overflow:hidden}
    .gbr-f{height:100%;border-radius:3px}
    .gb-p{background:linear-gradient(90deg,#2b6cb0,#4299e1)}.gb-q{background:linear-gradient(90deg,#276749,#48bb78)}
    .gbr-v{font-size:11px;font-weight:800;min-width:48px;text-align:right;color:#1a202c}
    .gbr-legend{display:flex;gap:14px;margin-bottom:6px;font-size:12px;font-weight:700}
    .gbr-legend span{display:flex;align-items:center;gap:5px}
    .gbr-legend i{display:inline-block;width:14px;height:14px;border-radius:2px}
    .cg{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .cg>div{background:#fff;border-radius:var(--r);padding:8px 10px;border:1px solid var(--b)}
    .cg .ct{font-size:13px;font-weight:700;margin-bottom:4px;padding-bottom:3px;border-bottom:1px solid var(--b)}
    .cgr{display:flex;align-items:center;padding:3px 0;font-size:12px;border-bottom:1px solid #f7fafc}
    .cgr:last-child{border:none}
    .cgr .rk{width:18px;font-weight:800;text-align:center}
    .cgr .pn{flex:1;font-weight:600;color:#1a202c;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
    .cgr .ps{font-weight:800;min-width:55px;text-align:right}
    .dgrid{display:grid;grid-template-columns:1fr 1fr;gap:6px}
    .stButton>button[kind="primary"]{background:linear-gradient(135deg,var(--p),var(--pl));border:none;border-radius:6px;padding:8px 14px;font-weight:700;font-size:15px;width:100%}
    ::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:#f1f1f1}::-webkit-scrollbar-thumb{background:#cbd5e0;border-radius:3px}
    div[data-testid="stSidebar"]{background:linear-gradient(180deg,var(--p),#0f2744)}
    div[data-testid="stSidebar"]*{color:rgba(255,255,255,.9)!important}
    div[data-testid="stSidebar"] .stSelectbox label,div[data-testid="stSidebar"] .stMultiSelect label,div[data-testid="stSidebar"] .stDateInput label,div[data-testid="stSidebar"] .stCheckbox label{color:rgba(255,255,255,.8)!important;font-weight:600;font-size:13px;text-transform:uppercase;letter-spacing:.5px}
    div[data-testid="stSidebar"] div[data-testid="stWidget"]{background:rgba(255,255,255,.08);border-radius:6px;padding:3px 8px;margin-bottom:3px;border:1px solid rgba(255,255,255,.1)}
    div[data-testid="stSidebar"] .stSelectbox>div>div,div[data-testid="stSidebar"] .stMultiSelect>div>div,div[data-testid="stSidebar"] .stDateInput>div>div{background:rgba(255,255,255,.95)!important;border-radius:5px}
    .es{text-align:center;padding:14px;color:#718096;font-size:14px}
    .evol-timeline{border-left:3px solid #2c5282;margin-left:12px;padding-left:20px}
    .evol-item{position:relative;padding-bottom:18px}
    .evol-item::before{content:'';position:absolute;left:-27px;top:4px;width:12px;height:12px;border-radius:50%;background:#2c5282;border:2px solid #fff;box-shadow:0 0 0 2px #2c5282}
    .evol-ver{font-size:13px;font-weight:800;color:#2c5282;margin-bottom:2px}
    .evol-date{font-size:11px;color:#718096;margin-bottom:4px}
    .evol-change{font-size:12px;color:#4a5568;padding:2px 0;padding-left:14px;position:relative}
    .evol-change::before{content:'•';position:absolute;left:0;color:#38a169;font-weight:800}
    .synth-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px}
    .synth-tbl thead th{background:var(--p);color:#fff;font-weight:700;font-size:11px;padding:5px 8px;border:none;white-space:nowrap;position:sticky;top:0}
    .synth-tbl tbody td{padding:4px 8px;border-bottom:1px solid #edf2f7;text-align:center}
    .synth-tbl tbody tr:nth-child(even) td{background:#f7fafc}
    .synth-tbl tbody tr:hover td{background:#ebf8ff!important}
    .synth-tbl .poste-cell{text-align:left;font-weight:700;white-space:nowrap;min-width:140px}
    .ano-legend{display:flex;gap:12px;margin-bottom:8px;font-size:11px;font-weight:600}
    .ano-legend span{display:flex;align-items:center;gap:4px}
    .ano-legend i{display:inline-block;width:14px;height:14px;border-radius:3px;border:1px solid #cbd5e0}
    .backlog-section{background:#fff;border-radius:var(--r);padding:12px;margin-bottom:10px;border:1px solid var(--b);box-shadow:0 2px 8px rgba(0,0,0,.03)}
    .backlog-section .bs-title{font-size:15px;font-weight:800;margin-bottom:8px;padding:6px 12px;border-radius:6px;color:#fff}
    .backlog-section .bs-title.om{background:linear-gradient(135deg,#6b46c1,#805ad5)}
    .backlog-section .bs-title.th{background:linear-gradient(135deg,#9b2c2c,#e53e3e)}
    .backlog-section .bs-title.to{background:linear-gradient(135deg,#1a365d,#2c5282)}
    .backlog-charts{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:8px}
    @media(max-width:768px){.cr{grid-template-columns:repeat(2,1fr)}.mh h1{font-size:17px}.cg,.dgrid,.backlog-charts{grid-template-columns:1fr}.car .cal{width:120px}.gbr-l{width:100px}}
    </style>""",unsafe_allow_html=True)

# ============================================================
def main():
    try: locale.setlocale(locale.LC_ALL,'fr_FR.UTF-8')
    except Exception:
        try: locale.setlocale(locale.LC_ALL,'fr_FR')
        except Exception: pass
    inject_custom_css()
    fichier_date=get_date_from_file()

    if "hse_affiche" not in st.session_state: st.session_state.hse_affiche=False
    if not st.session_state.hse_affiche:
        c=random.choice(CONSIGNES_HSE)
        st.markdown("""<div style="min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:linear-gradient(135deg,#1a365d,#2d3748,#1a365d);padding:40px">
        <div style="font-size:64px;margin-bottom:20px">🦺</div>
        <h1 style="text-align:center;font-size:46px;color:#fff;font-weight:900;margin:0">HSE - CONSIGNE DE SECURITE</h1>
        <p style="text-align:center;color:rgba(255,255,255,.6);font-size:22px;margin-top:8px;letter-spacing:3px;text-transform:uppercase">Securite - Sante - Environnement</p>
        <div style="background:linear-gradient(135deg,#f6e05e,#ed8936);padding:36px 48px;border-radius:20px;font-size:32px;font-weight:700;text-align:center;margin:40px 0;color:#1a202c;max-width:800px;box-shadow:0 20px 60px rgba(0,0,0,.3)">⚠️ %s</div>
        <h2 style="text-align:center;color:#48bb78;font-size:36px;font-weight:900">Aucun travail n'est plus urgent que la securite</h2>
        <div style="margin-top:40px;width:200px;height:4px;background:rgba(255,255,255,.1);border-radius:2px;overflow:hidden"><div style="width:100%%;height:100%%;background:linear-gradient(90deg,#48bb78,#38a169);border-radius:2px;animation:ld 5.5s ease-in-out forwards"></div></div>
        <style>@keyframes ld{from{width:0}to{width:100%%}}</style></div>"""%c,unsafe_allow_html=True)
        time.sleep(6); st.session_state.hse_affiche=True; st.rerun(); st.stop()

    # ---- Helper functions ----
    def contient_mot(t,lm):
        t=str(t); return any(m in t for l in lm for m in l.split())
    def cat_age(a):
        if pd.isna(a): return "Inconnu"
        if a<=1: return "<1 mois"
        elif a>=3: return ">3 mois"
        return "1 mois < <3 mois"
    def ckpi(n,d,sz=100): return np.where(d==0,sz,(n/d)*100)
    def cpiv(df,f,c,p):
        return pd.pivot_table(df[f],index="Poste travail princ.",columns=c,values="Ordre",aggfunc="count",fill_value=0).reindex(p,fill_value=0)
    def excr(df):
        if "Poste travail princ." in df.columns:
            return df[~df["Poste travail princ."].astype(str).str.contains("cresseur",case=False,na=False)].copy()
        return df
    def get_text_col(df):
        for c in ["Désignation","Designation","Désignation OT","Texte ordre","Texte","Description","Libellé","Libelle"]:
            if c in df.columns: return c
        for c in df.columns:
            if df[c].dtype=='object' and any(kw in str(c).lower() for kw in ['sign','text','desc','libell']):
                return c
        return None

    def build_statut_pivot(df_sub, posts):
        if df_sub.empty:
            return pd.DataFrame(index=posts, columns=["CRÉÉ","LANC","CLOT","TCLO","Total"]).fillna(0).astype(int)
        piv=pd.pivot_table(df_sub, index="Poste travail princ.", columns="Statut OT", values="Ordre", aggfunc="count", fill_value=0)
        for s in ["CRÉÉ","LANC","CLOT","TCLO"]:
            if s not in piv.columns: piv[s]=0
        piv["Total"]=piv[["CRÉÉ","LANC","CLOT","TCLO"]].sum(axis=1)
        return piv.reindex(posts, fill_value=0).fillna(0).astype(int)

    def html_statut_pivot(piv_df, table_class):
        cols=["Poste de travail","CRÉÉ","LANC","CLOT","TCLO","Total"]
        h='<table class="tw %s"><thead><tr>'%table_class+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for poste,row in piv_df.iterrows():
            h+='<tr><td style="font-weight:600">%s</td>'%poste
            for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
                v=int(row.get(c,0))
                clr=""
                if c=="CRÉÉ" and v>0: clr="background:#fed7d7;color:#9b2c2c"
                elif c=="LANC" and v>0: clr="background:#fefcbf;color:#975a16"
                elif c=="CLOT" and v>0: clr="background:#c6f6d5;color:#276749"
                elif c=="TCLO" and v>0: clr="background:#bee3f8;color:#2b6cb0"
                h+='<td style="text-align:center;%s">%d</td>'%(clr,v)
            h+='<td style="text-align:center;font-weight:800;background:#edf2f7">%d</td>'%int(row.get("Total",0))
            h+='</tr>'
        h+='<tr class="tr"><td>Total</td>'
        for c in ["CRÉÉ","LANC","CLOT","TCLO"]:
            h+='<td style="text-align:center">%d</td>'%int(piv_df[c].sum())
        h+='<td style="text-align:center">%d</td>'%int(piv_df["Total"].sum())
        h+='</tr></tbody></table>'
        return h

    def show_improved_pie_pair(piv_df, title_prefix):
        """Pie charts ameliores : regroupement secteurs minces, donut, pull-out, labels enrichis"""
        global_counts=piv_df[["CRÉÉ","LANC","CLOT","TCLO"]].sum()
        global_counts=global_counts[global_counts>0]
        THRESHOLD=3.0

        c1,c2=st.columns(2)

        with c1:
            if not global_counts.empty:
                total=global_counts.sum()
                labels_l=global_counts.index.tolist()
                values_l=global_counts.values.tolist()
                pcts_l=[v/total*100 for v in values_l]
                main_l,main_v,main_p=[],[],[]
                other_sum=0
                for l,v,p in zip(labels_l,values_l,pcts_l):
                    if p<THRESHOLD: other_sum+=v
                    else: main_l.append(l); main_v.append(v); main_p.append(p)
                if other_sum>0:
                    main_l.append("Autres"); main_v.append(other_sum); main_p.append(other_sum/total*100)
                cmap={"CRÉÉ":"#e53e3e","LANC":"#d69e2e","CLOT":"#38a169","TCLO":"#3182ce","Autres":"#a0aec0"}
                colors=[cmap.get(l,"#a0aec0") for l in main_l]
                min_i=main_p.index(min(main_p))
                pull=[0.06 if i==min_i else 0 for i in range(len(main_l))]
                fig1=go.Figure(go.Pie(labels=main_l,values=main_v,marker_colors=colors,pull=pull,
                    textposition='inside',textinfo='label+percent+value',textfont_size=12,hole=0.38,
                    hovertemplate='<b>%%{label}</b><br>Valeur: %%{value}<br>Pourcentage: %%{percent}<extra></extra>'))
                fig1.update_layout(title=dict(text="%s — Par Statut OT"%title_prefix,font=dict(size=14,fontweight='bold')),
                    margin=dict(t=55,b=10,l=10,r=10),height=370,
                    legend=dict(font_size=11,orientation="h",yanchor="bottom",y=-0.18),showlegend=True)
                st.plotly_chart(fig1,use_container_width=True)
            else:
                st.markdown('<div class="es">Aucune donnee</div>',unsafe_allow_html=True)

        with c2:
            if global_counts.sum()>0:
                realised=global_counts.get("CLOT",0)+global_counts.get("TCLO",0)
                not_realised=total-realised
                if not_realised==0 and realised==0:
                    st.markdown('<div class="es">Aucune donnee</div>',unsafe_allow_html=True)
                else:
                    lab2=["Réalisés (CLOT+TCLO)","Non Réalisés"]
                    val2=[realised,not_realised]
                    col2=["#38a169","#e53e3e"]
                    p2=[v/sum(val2)*100 for v in val2]
                    pull2=[0.04 if p<5 else 0 for p in p2]
                    fig2=go.Figure(go.Pie(labels=lab2,values=val2,marker_colors=col2,pull=pull2,
                        textposition='inside',textinfo='label+percent+value',textfont_size=13,hole=0.38,
                        hovertemplate='<b>%%{label}</b><br>Valeur: %%{value}<br>Pourcentage: %%{percent}<extra></extra>'))
                    fig2.update_layout(title=dict(text="%s — Réalisés vs Non Réalisés"%title_prefix,font=dict(size=14,fontweight='bold')),
                        margin=dict(t=55,b=10,l=10,r=10),height=370,
                        legend=dict(font_size=11,orientation="h",yanchor="bottom",y=-0.18),showlegend=True)
                    st.plotly_chart(fig2,use_container_width=True)
            else:
                st.markdown('<div class="es">Aucune donnee</div>',unsafe_allow_html=True)

    def calc_kpis(df_i,av_i,now,posts):
        res={}; df=df_i.copy(); av=av_i.copy()
        df["Backlog preparation"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MP_KW)),"CARACTERISE","NON CARACTERISE")
        df["Backlog planification"]=np.where(df["Statut utilisateur"].apply(lambda x:contient_mot(x,MPLAN_KW)),"CARACTERISE","NON CARACTERISE")
        for dc,am,ac in [('Créé le',"amp","ap"),('Date de début planifiée',"amlp","alp"),('Date de début planifiée',"amex","aex")]:
            if dc in df.columns:
                df[dc]=pd.to_datetime(df[dc],errors='coerce')
                df[am]=((now.year-df[dc].dt.year)*12+(now.month-df[dc].dt.month)).round(2)
                df[ac]=df[am].apply(cat_age)
            else: df[am]=np.nan; df[ac]="Inconnu"
        df["OT CONFIME"]=np.where(df["Statut système"].str.contains("CLO",na=False)&df["Statut système"].str.contains("CONF",na=False),"OUI","NON")
        df["Contient SOPL"]=df["Statut utilisateur"].str.contains("SOPL",na=False).map({True:1,False:0})
        df["OT LANC ESTIME"]=np.where(df["Total coûts budgétés"].fillna(0)==0,"NON","OUI")
        df["OT_COR_EGAL"]=np.where((df["Total coûts budgétés"].fillna(0)-df["Total coûts réels"].fillna(0))==0,"OUI","NON")
        df["_tw_num"]=pd.to_numeric(df.get("Type de travail",pd.Series(dtype=float)),errors="coerce")
        res['dfp']=df
        filt_corr=(df["Nº appel pl.entret."].fillna(0)==0)&(df["Contient SOPL"]==1)
        an=cpiv(df,filt_corr,"Statut OT",posts)
        for c in ["CLOT","CRÉÉ","LANC","TCLO"]: an[c]=an.get(c,0)
        an["OT_CLOTURES"]=an["CLOT"]+an["TCLO"]
        an["TOTAL_OT"]=an[["CLOT","CRÉÉ","LANC","TCLO"]].sum(axis=1)
        an["TAUX_REALISATION_CORRECTIF/PT"]=np.where(an["TOTAL_OT"]==0,100.0,ckpi(an["OT_CLOTURES"],an["TOTAL_OT"]))
        pr=cpiv(df,(df["Statut OT"]=="CRÉÉ")&(df["Statut utilisateur"].str.contains("CRPR",na=False)),"ap",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pr[c]=pr.get(c,0)
        pr["Total"]=pr[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pr["OT préparation <1 mois"]=ckpi(pr["<1 mois"],pr["Total"]); pr["OT préparation >3 mois"]=ckpi(pr[">3 mois"],pr["Total"],0); pr["OT préparation 1mois< <3mois"]=ckpi(pr["1 mois < <3 mois"],pr["Total"],0)
        pl=cpiv(df,(df["Statut OT"]=="LANC")&(df["Statut utilisateur"].str.contains("ATPL",case=False,na=False)),"alp",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: pl[c]=pl.get(c,0)
        pl["Total"]=pl[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        pl["OT planification <1 mois"]=ckpi(pl["<1 mois"],pl["Total"]); pl["OT planification >3 mois"]=ckpi(pl[">3 mois"],pl["Total"],0); pl["OT planification 1mois< <3mois"]=ckpi(pl["1 mois < <3 mois"],pl["Total"],0)
        ex=cpiv(df,(df["Statut OT"]=="LANC")&(df["Contient SOPL"]==1),"aex",posts)
        for c in ["<1 mois",">3 mois","1 mois < <3 mois","Inconnu"]: ex[c]=ex.get(c,0)
        ex["Total"]=ex[["<1 mois","1 mois < <3 mois",">3 mois","Inconnu"]].sum(axis=1)
        ex["OT exécution <1 mois"]=ckpi(ex["<1 mois"],ex["Total"]); ex["OT exécution >3 mois"]=ckpi(ex[">3 mois"],ex["Total"],0); ex["OT exécution 1mois< <3mois"]=ckpi(ex["1 mois < <3 mois"],ex["Total"],0)
        la=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="OT LANC ESTIME",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["OUI","NON"]: la[c]=la.get(c,0)
        la["Total"]=la["OUI"]+la["NON"]; la["OT LANC ESTIME"]=ckpi(la["OUI"],la["Total"])
        pc=pd.pivot_table(df[df["Statut OT"]=="CRÉÉ"],index="Poste travail princ.",columns="Backlog preparation",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: pc[c]=pc.get(c,0)
        pc["Total"]=pc["CARACTERISE"]+pc["NON CARACTERISE"]; pc["Backlog préparation caractérisé"]=ckpi(pc["CARACTERISE"],pc["Total"])
        plc=pd.pivot_table(df[df["Statut OT"]=="LANC"],index="Poste travail princ.",columns="Backlog planification",values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["CARACTERISE","NON CARACTERISE"]: plc[c]=plc.get(c,0)
        plc["Total"]=plc["CARACTERISE"]+plc["NON CARACTERISE"]; plc["Backlog planification caractérisé"]=ckpi(plc["CARACTERISE"],plc["Total"])
        for kn,cn in [("OT CONFIME","OT CONFIME"),("OT_COR_EGAL","OT_COR_EGAL")]:
            pv=pd.pivot_table(df,index="Poste travail princ.",columns=cn,values="Ordre",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
            for c in ["OUI","NON"]: pv[c]=pv.get(c,0)
            pv["Total"]=pv["OUI"]+pv["NON"]; pv[cn]=ckpi(pv["OUI"],pv["Total"]); res[kn.lower().replace(" ","_")]=pv
        avf=av[(av["Ordre"].isna())|(av["Ordre"].astype(str).str.strip()=="")].copy(); res['avf']=avf
        tca=pd.pivot_table(avf,index="Poste travail princ.",columns="Statut utilisateur",values="Avis",aggfunc="count",fill_value=0).reindex(posts,fill_value=0)
        for c in ["APRQ","APRV","APRV AVAU","REJT"]: tca[c]=tca.get(c,0)
        tca["Total"]=tca[["APRQ","APRV","APRV AVAU","REJT"]].sum(axis=1); tca["appel avis approuvé"]=ckpi(tca["APRV"],tca["Total"])
        g_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_den=df[(df["Contient SOPL"]==1)&(df["_tw_num"]==350)].groupby("Poste travail princ.")["Ordre"].count()
        g_df=pd.DataFrame({"_n":g_num,"_d":g_den}).reindex(posts,fill_value=0)
        g_df["Performance Graissage"]=np.where(g_df["_d"]==0,100.0,(g_df["_n"]/g_df["_d"])*100)
        ins_types=[290,300,310]
        ins_base=(df["_tw_num"].isin(ins_types))&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now)
        ins_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_den=df[(df["Contient SOPL"]==1)&ins_base].groupby("Poste travail princ.")["Ordre"].count()
        ins_df=pd.DataFrame({"_n":ins_num,"_d":ins_den}).reindex(posts,fill_value=0)
        ins_df["Performance Inspection"]=np.where(ins_df["_d"]==0,100.0,(ins_df["_n"]/ins_df["_d"])*100)
        sys_base=(df["_tw_num"]==360)&(df["Date de début planifiée"].notna())&(df["Date de début planifiée"]<=now)
        sys_num=df[(df["Statut OT"].isin(["CLOT","TCLO"]))&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_den=df[(df["Contient SOPL"]==1)&sys_base].groupby("Poste travail princ.")["Ordre"].count()
        sys_df=pd.DataFrame({"_n":sys_num,"_d":sys_den}).reindex(posts,fill_value=0)
        sys_df["Performance Appels Systématiques"]=np.where(sys_df["_d"]==0,100.0,(sys_df["_n"]/sys_df["_d"])*100)
        fiab_s=pd.Series(100.0,index=posts); avpan_s=pd.Series(100.0,index=posts)
        res['ckdf']=pd.DataFrame({
            "TAUX_REALISATION_CORRECTIF/PT":an["TAUX_REALISATION_CORRECTIF/PT"],
            "OT préparation <1 mois":pr["OT préparation <1 mois"],"OT préparation >3 mois":pr["OT préparation >3 mois"],"OT préparation 1mois< <3mois":pr["OT préparation 1mois< <3mois"],
            "OT planification <1 mois":pl["OT planification <1 mois"],"OT planification >3 mois":pl["OT planification >3 mois"],"OT planification 1mois< <3mois":pl["OT planification 1mois< <3mois"],
            "OT exécution <1 mois":ex["OT exécution <1 mois"],"OT exécution >3 mois":ex["OT exécution >3 mois"],"OT exécution 1mois< <3mois":ex["OT exécution 1mois< <3mois"],
            "Performance Graissage":g_df["Performance Graissage"],"Performance Inspection":ins_df["Performance Inspection"],"Performance Appels Systématiques":sys_df["Performance Appels Systématiques"],
            "appel avis approuvé":tca["appel avis approuvé"],"OT LANC ESTIME":la["OT LANC ESTIME"],
            "Backlog préparation caractérisé":pc["Backlog préparation caractérisé"],"Backlog planification caractérisé":plc["Backlog planification caractérisé"],
            "OT CONFIME":res['ot_confime']["OT CONFIME"],"OT_COR_EGAL":res['ot_cor_egal']["OT_COR_EGAL"],
            "OT Fiabilité":fiab_s,"Total Avis de Panne":avpan_s
        })
        return res

    def ks(v,c):
        try: val=float(v)
        except Exception: return ""
        if c in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=80 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=75 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=15 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val<=5 else "background:#ffc7ce;color:#9c0006;font-weight:600"
        if c=="TAUX_REALISATION_CORRECTIF/PT":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=85 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c=="appel avis approuvé":
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=95 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>90 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        if c in ["OT Fiabilité","Total Avis de Panne"]:
            return "background:#c6efce;color:#006100;font-weight:600" if val>=100 else ("background:#ffeb9c;color:#9c6500;font-weight:600" if val>=95 else "background:#ffc7ce;color:#9c0006;font-weight:600")
        return ""
    def cs(v):
        try: val=float(str(v).replace(' %','').strip())
        except Exception: return ""
        return "background:#c6efce;color:#006100;font-weight:700" if val>=90 else ("background:#ffeb9c;color:#9c6500;font-weight:700" if val>=80 else "background:#ffc7ce;color:#9c0006;font-weight:700")
    def kas(v):
        try: val=int(v)
        except Exception: return ""
        if val==0: return "background:#c6efce;color:#006100;font-weight:600"
        if val<=3: return "background:#ffeb9c;color:#9c6500;font-weight:600"
        if val<=10: return "background:#fed7d7;color:#c53030;font-weight:600"
        return "background:#fc8181;color:#742a2a;font-weight:800"
    def gscore(k,a,t):
        if pd.isna(a) or pd.isna(t): return 0
        if k in ["OT préparation <1 mois","OT planification <1 mois","OT exécution <1 mois"]: return 1 if a>=75 else 0
        if k in ["OT préparation 1mois< <3mois","OT planification 1mois< <3mois","OT exécution 1mois< <3mois"]: return 1 if a<=15 else 0
        if k in ["OT préparation >3 mois","OT planification >3 mois","OT exécution >3 mois"]: return 1 if a<=5 else 0
        if k=="TAUX_REALISATION_CORRECTIF/PT": return 1 if a>=80 else 0
        if k=="appel avis approuvé": return 1 if a>=90 else 0
        if k in ["OT LANC ESTIME","Backlog préparation caractérisé","Backlog planification caractérisé","OT CONFIME","OT_COR_EGAL"]: return 1 if a>=95 else 0
        if k in ["Performance Graissage","Performance Inspection","Performance Appels Systématiques"]: return 1 if a>=95 else 0
        if k in ["OT Fiabilité","Total Avis de Panne"]: return 1 if a>=100 else 0
        return 0
    def is_lb(k): return k in LOWER_BETTER

    def html_table(rows,cols,tc,sc_col=None):
        h='<table class="tw %s"><thead><tr>'%tc+''.join('<th>%s</th>'%c for c in cols)+'</tr></thead><tbody>'
        for r in rows:
            rc="cb" if r.get("_t")=="cible" else ("tr" if r.get("_t")=="total" else "")
            h+='<tr class="%s">'%rc
            for c in cols:
                v=r.get(c,"")
                if r.get("_t")=="cible": h+='<td>%s</td>'%v
                else: s=cs(v) if sc_col and c in sc_col else ks(v,c); h+='<td style="%s">%s</td>'%(s or "",v)
            h+='</tr>'
        return h+'</tbody></table>'

    def html_ano_transposed(ano_counts, kpi_list, posts, table_class, title):
        """Table anomalies INVERSEE : Postes en lignes, KPI en colonnes, valeurs = nb anomalies OT/Avis"""
        h = '<div class="ca"><div class="ct">%s</div>' % title
        h += '<div class="ano-legend">'
        h += '<span><i style="background:#c6efce"></i> 0 (OK)</span>'
        h += '<span><i style="background:#ffeb9c"></i> 1-3</span>'
        h += '<span><i style="background:#fed7d7"></i> 4-10</span>'
        h += '<span><i style="background:#fc8181"></i> >10</span>'
        h += '</div>'
        h += '<div style="overflow-x:auto"><table class="tw %s"><thead><tr><th style="min-width:180px">Poste de travail</th>' % table_class
        for kpi in kpi_list:
            short = kpi.replace("OT préparation","OT prép.").replace("OT planification","OT plan.").replace("OT exécution","OT exéc.").replace("Performance ","P.").replace("Backlog ","Bk.").replace("caractérisé","caract.")
            if len(short)>22: short=short[:20]+"…"
            h += '<th style="min-width:55px;font-size:10px" title="%s">%s</th>' % (kpi, short)
        h += '<th style="min-width:50px">Total</th></tr></thead><tbody>'
        for poste in posts:
            h += '<tr><td style="font-weight:700;white-space:nowrap">%s</td>' % poste
            total = 0
            for kpi in kpi_list:
                cnt = ano_counts.get(kpi, {}).get(poste, 0)
                total += cnt
                s = kas(cnt)
                h += '<td style="%s;text-align:center">%d</td>' % (s or "", cnt)
            s = kas(total)
            h += '<td style="%s;text-align:center;font-weight:800">%d</td>' % (s or "", total)
            h += '</tr>'
        h += '<tr class="tr"><td>Total</td>'
        grand_total = 0
        for kpi in kpi_list:
            kpi_total = sum(ano_counts.get(kpi, {}).get(p, 0) for p in posts)
            grand_total += kpi_total
            h += '<td style="text-align:center">%d</td>' % kpi_total
        h += '<td style="text-align:center">%d</td>' % grand_total
        h += '</tr></tbody></table></div></div>'
        return h

    def html_actions_table(kpi_list,actuals,targets,act_map):
        h='<table class="tw at"><thead><tr><th>KPI</th><th>Valeur Actuelle</th><th>Cible</th><th>Ecart</th><th>Statut</th><th>Action Recommandee</th></tr></thead><tbody>'
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); diff=av-tv
            met=av<=tv if is_lb(k) else av>=tv
            status="ATTEINT" if met else "NON ATTEINT"
            st_s="background:#c6efce;color:#006100;font-weight:700" if met else "background:#ffc7ce;color:#9c0006;font-weight:700"
            ec_clr="#276749" if met else "#c53030"
            action="Objectif atteint" if met else act_map.get(k,"")
            h+='<tr><td style="font-weight:600">%s</td><td>%.1f%%</td><td>%.0f%%</td><td style="color:%s;font-weight:700">%+.1f%%</td><td style="%s">%s</td><td style="color:#4a5568">%s</td></tr>'%(k,av,tv,ec_clr,diff,st_s,status,action)
        return h+'</tbody></table>'

    def html_classement(scores,accent):
        sp=sorted(scores.items(),key=lambda x:x[1],reverse=True)
        met_p=[(p,s) for p,s in sp if s>=80]; not_p=[(p,s) for p,s in sp if s<80]
        t5=met_p[:5]; b5=not_p[-5:] if len(not_p)>5 else not_p
        h='<div class="cg"><div><div class="ct" style="color:#38a169">Top 5 — Objectif Atteint</div>'
        if t5:
            for i,(p,s) in enumerate(t5): h+='<div class="cgr"><span class="rk" style="color:%s">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(accent,i+1,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#718096">Aucun poste</div>'
        h+='</div><div><div class="ct" style="color:#e53e3e">Bottom 5 — Non Atteint</div>'
        if b5:
            for i,(p,s) in enumerate(reversed(b5)): h+='<div class="cgr"><span class="rk" style="color:#e53e3e">%s</span><span class="pn">%s</span><span class="ps" style="%s">%.2f%%</span></div>'%(len(b5)-i,p,cs("%.2f"%s),s)
        else: h+='<div style="padding:6px;font-size:12px;color:#38a169">Tous atteints</div>'
        h+='</div></div>'; return h

    def html_kpi_bars(kpi_list,actuals,targets,title,color_ok,color_fail):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color_ok,title)
        for k in kpi_list:
            av=actuals.get(k,0); tv=targets.get(k,100); met=av<=tv if is_lb(k) else av>=tv
            bw=min(max(av,0),100); bg=color_ok if met else color_fail
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(k,bw,bg,av)
        return h+'</div>'

    def html_bars(data,title,color):
        h='<div class="ca"><div class="ct" style="color:%s">%s</div>'%(color,title)
        for label,val in sorted(data,key=lambda x:x[1],reverse=True):
            bw=min(max(val,0),100)
            h+='<div class="car"><div class="cal">%s</div><div class="cab"><div class="caf" style="width:%s%%;background:%s"></div></div><div class="cav-out">%.1f%%</div></div>'%(label,bw,color,val)
        return h+'</div>'

    def html_grouped_bars(posts,pscores,qscores,title):
        h='<div class="ca"><div class="ct" style="color:#1e3a5f">%s</div>'%title
        h+='<div class="gbr-legend"><span><i style="background:linear-gradient(90deg,#2b6cb0,#4299e1)"></i> Performance</span><span><i style="background:linear-gradient(90deg,#276749,#48bb78)"></i> Qualite</span></div>'
        for p in sorted(posts,key=lambda x:(pscores.get(x,0)+qscores.get(x,0))/2,reverse=True):
            pv,qv=pscores.get(p,0),qscores.get(p,0)
            h+='<div class="gbr"><div class="gbr-l">%s</div><div class="gbr-g"><div class="gbr-w"><div class="gbr-f gb-p" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div><div class="gbr-w"><div class="gbr-f gb-q" style="width:%s%%"></div></div><div class="gbr-v">%.1f%%</div></div></div>'%(p,min(max(pv,0),100),pv,min(max(qv,0),100),qv)
        return h+'</div>'

    def html_synthese_table(synth_data,kpi_list,posts,sec_type):
        colors_map={"Amelioration":"#276749","Degradation":"#c53030","Stable":"#718096"}
        arrow_map={"Amelioration":"▲","Degradation":"▼","Stable":"—"}
        h='<table class="synth-tbl"><thead><tr><th rowspan="2" style="min-width:160px;text-align:left">Poste de travail</th>'
        h+='<th colspan="%d" style="background:#276749">%s</th>'%(len(kpi_list),"Valeur Precedente")
        h+='<th colspan="%d" style="background:#2c5282">%s</th>'%(len(kpi_list),"Valeur Actuelle")
        h+='<th colspan="%d" style="background:#975a16">%s</th>'%(len(kpi_list),"Ecart")
        h+='</tr><tr>'
        for _ in kpi_list: h+='<th style="background:#276749;font-size:10px">%%</th>'
        for _ in kpi_list: h+='<th style="background:#2c5282;font-size:10px">%%</th>'
        for _ in kpi_list: h+='<th style="background:#975a16;font-size:10px">%%</th>'
        h+='</tr></thead><tbody>'
        for poste in posts:
            h+='<tr><td class="poste-cell">%s</td>'%poste
            for kpi in kpi_list:
                prev_v=synth_data.get(poste,{}).get(kpi,{}).get("prev","—")
                h+='<td>%s</td>'%prev_v
            for kpi in kpi_list:
                curr_v=synth_data.get(poste,{}).get(kpi,{}).get("curr","—")
                h+='<td>%s</td>'%curr_v
            for kpi in kpi_list:
                evol=synth_data.get(poste,{}).get(kpi,{}).get("evol",{})
                sens=evol.get("sens","Stable"); ecart=evol.get("ecart","—")
                clr=colors_map.get(sens,"#718096"); arrow=arrow_map.get(sens,"—")
                h+='<td style="color:%s;font-weight:700">%s %s</td>'%(clr,arrow,ecart)
            h+='</tr>'
        h+='</tbody></table>'
        return h

    # ================================================================
    # FILE UPLOAD & SIDEBAR
    # ================================================================
    with st.sidebar:
        st.markdown('<div style="text-align:center;padding:8px 0"><span style="font-size:28px">⚙️</span><br><span style="font-size:13px;font-weight:700;color:rgba(255,255,255,.7)">PARAMETRES</span></div>',unsafe_allow_html=True)
        f_ot=st.file_uploader("Fichier OT (.xlsx)",type=["xlsx"],key="fot")
        f_av=st.file_uploader("Fichier Avis (.xlsx)",type=["xlsx"],key="fav")
        st.markdown('<div style="height:1px;background:rgba(255,255,255,.15);margin:6px 0"></div>',unsafe_allow_html=True)
        secteur_list=[]; poste_list=[]; df_ot_tmp=None
        if f_ot:
            try:
                df_ot_tmp=pd.read_excel(f_ot,dtype=str)
                if "Secteur" in df_ot_tmp.columns: secteur_list=sorted(df_ot_tmp["Secteur"].dropna().unique().tolist())
                if "Poste travail princ." in df_ot_tmp.columns: poste_list=sorted(df_ot_tmp["Poste travail princ."].dropna().unique().tolist())
            except Exception: pass
        sel_sect=st.multiselect("Secteur",secteur_list if secteur_list else [],key="ssect")
        sel_poste=st.multiselect("Poste",poste_list if poste_list else [],key="sposte")
        ref_date=st.date_input("Date de reference",datetime.now(),key="rd")
        exp_excel=st.checkbox("Exporter Excel",value=False,key="expex")

    # ================================================================
    # DATA PROCESSING
    # ================================================================
    if not f_ot or not f_av:
        st.markdown("""<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:60vh">
        <div style="font-size:80px;margin-bottom:20px">📊</div>
        <h2 style="color:#1e3a5f;font-weight:800;font-size:28px">Bienvenue sur le Dashboard KPI</h2>
        <p style="color:#718096;font-size:16px;margin-top:8px;max-width:500px;text-align:center">
        Veuillez charger les fichiers Excel des OT et des Avis dans le panneau de parametres pour commencer l'analyse.</p></div>""",unsafe_allow_html=True)
        st.stop()

    try: df=pd.read_excel(f_ot,dtype=str); df=excr(df)
    except Exception: st.error("Erreur de lecture du fichier OT."); st.stop()
    try: av=pd.read_excel(f_av,dtype=str)
    except Exception: st.error("Erreur de lecture du fichier Avis."); st.stop()

    num_cols=["Total coûts budgétés","Total coûts réels","Nº appel pl.entret.","Type de travail"]
    for nc in num_cols:
        if nc in df.columns: df[nc]=pd.to_numeric(df[nc],errors="coerce")
    if "Poste travail princ." not in df.columns or "Statut OT" not in df.columns:
        st.error("Colonnes obligatoires manquantes: 'Poste travail princ.' ou 'Statut OT'"); st.stop()

    now=pd.Timestamp(ref_date)
    if sel_sect: df=df[df["Secteur"].isin(sel_sect)]
    if sel_poste: df=df[df["Poste travail princ."].isin(sel_poste)]
    if sel_sect and "Secteur" in av.columns: av=av[av["Secteur"].isin(sel_sect)]
    if sel_poste and "Poste travail princ." in av.columns: av=av[av["Poste travail princ."].isin(sel_poste)]

    posts=sorted(df["Poste travail princ."].dropna().unique().tolist())
    if not posts: st.warning("Aucun poste trouve apres filtrage."); st.stop()

    cache_key=compute_cache_key(fichier_date,{"secteurs":sorted(sel_sect),"postes":sorted(sel_poste),"date":str(ref_date)},datetime.now().strftime("%H%M%S"))
    if "cache_key" not in st.session_state or st.session_state.cache_key!=cache_key:
        st.session_state.kpis=calc_kpis(df,av,now,posts)
        st.session_state.cache_key=cache_key
    kp=st.session_state.kpis
    ckdf=kp['ckdf']

    # ================================================================
    # COMPUTE SCORES & ANOMALIES
    # ================================================================
    pscores={}; qscores={}
    for p in posts:
        pp=sum(gscore(k,ckdf.loc[p,k],CIBLE.get(k,100)) for k in QK)
        qp=sum(gscore(k,ckdf.loc[p,k],CIBLE.get(k,100)) for k in PK)
        pscores[p]=pp/len(QK)*100 if QK else 0
        qscores[p]=qp/len(PK)*100 if PK else 0

    # Anomalies par nombre d'OT ou Avis hors cible
    ano_perf={}; ano_qual={}
    for k in QK:
        ano_perf[k]={}
        t=CIBLE.get(k,100)
        for p in posts:
            v=ckdf.loc[p,k]
            if is_lb(k):
                ano_perf[k][p]=max(0,int(v-t+0.99)) if v>t else 0
            else:
                ano_perf[k][p]=max(0,int(t-v+0.99)) if v<t else 0
    for k in PK:
        ano_qual[k]={}
        t=CIBLE.get(k,100)
        for p in posts:
            v=ckdf.loc[p,k]
            if is_lb(k):
                ano_qual[k][p]=max(0,int(v-t+0.99)) if v>t else 0
            else:
                ano_qual[k][p]=max(0,int(t-v+0.99)) if v<t else 0
    # Pour "appel avis approuvé" : compter les avis sans OT
    avf=kp.get('avf',pd.DataFrame())
    if not avf.empty and "Poste travail princ." in avf.columns:
        for p in posts:
            av_cnt=len(avf[avf["Poste travail princ."]==p])
            ano_qual["appel avis approuvé"][p]=av_cnt

    tot_ot=len(df); tot_av=len(av); tot_cree=len(df[df["Statut OT"]=="CRÉÉ"]); tot_lanc=len(df[df["Statut OT"]=="LANC"])
    tot_clot=len(df[df["Statut OT"].isin(["CLOT","TCLO"])])
    avg_p=np.mean(list(pscores.values())) if pscores else 0
    avg_q=np.mean(list(qscores.values())) if qscores else 0
    tot_ano_p=sum(sum(v.values()) for v in ano_perf.values())
    tot_ano_q=sum(sum(v.values()) for v in ano_qual.values())

    # ================================================================
    # HEADER
    # ================================================================
    st.markdown('<div class="mh"><h1>📊 Dashboard KPI Maintenance</h1><span class="db">📅 %s</span></div>'%fichier_date,unsafe_allow_html=True)
    st.markdown('<div class="cr">'
        '<div class="cc c1"><div class="cv">%d</div><div class="cl">Total OT</div></div>'
        '<div class="cc c2"><div class="cv">%d</div><div class="cl">OT Clotures</div></div>'
        '<div class="cc c3"><div class="cv">%.1f%%</div><div class="cl">Score Performance</div></div>'
        '<div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies</div></div>'
        '</div>'%(tot_ot,tot_clot,avg_p,tot_ano_p+tot_ano_q),unsafe_allow_html=True)

    # ================================================================
    # TABS
    # ================================================================
    tabs=st.tabs(["📊 Dashboard","🟢 Performance","🔵 Qualité","🔴 Anomalies","🟠 Backlog","📈 Suivi & Évolution","📝 Changelog"])

    # ---- TAB DASHBOARD ----
    with tabs[0]:
        st.markdown('<p class="stl p">Performance par Poste</p>',unsafe_allow_html=True)
        st.markdown(html_grouped_bars(posts,pscores,qscores,"Comparaison Performance / Qualite par Poste"),unsafe_allow_html=True)
        st.markdown('<p class="stl c">Classement des Postes</p>',unsafe_allow_html=True)
        st.markdown(html_classement(pscores,"#276749"),unsafe_allow_html=True)
        st.markdown('<p class="stl p">Barres de Progression — Performance</p>',unsafe_allow_html=True)
        avg_perf={k:float(ckdf[k].mean()) for k in QK}
        st.markdown(html_kpi_bars(QK,avg_perf,CIBLE,"Moyenne Globale Performance","#38a169","#e53e3e"),unsafe_allow_html=True)
        st.markdown('<p class="stl q">Barres de Progression — Qualite</p>',unsafe_allow_html=True)
        avg_qual={k:float(ckdf[k].mean()) for k in PK}
        st.markdown(html_kpi_bars(PK,avg_qual,CIBLE,"Moyenne Globale Qualite","#3182ce","#e53e3e"),unsafe_allow_html=True)

    # ---- TAB PERFORMANCE ----
    with tabs[1]:
        st.markdown('<p class="stl p">Indicateurs de Performance par Poste</p>',unsafe_allow_html=True)
        pcols=["Poste de travail"]+QK+["Score Performance"]
        prows=[]
        for p in posts:
            r={"Poste de travail":p,"_t":""}
            for k in QK: r[k]="%.1f%%"%ckdf.loc[p,k]
            r["Score Performance"]="%.2f%%"%pscores[p]
            prows.append(r)
        cible_row={"Poste de travail":"CIBLE","_t":"cible"}
        for k in QK: cible_row[k]="%.0f%%"%CIBLE.get(k,100)
        cible_row["Score Performance"]="80.00%"
        prows.append(cible_row)
        total_row={"Poste de travail":"Moyenne","_t":"total"}
        for k in QK: total_row[k]="%.1f%%"%float(ckdf[k].mean())
        total_row["Score Performance"]="%.2f%%"%avg_p
        prows.append(total_row)
        st.markdown(html_table(prows,pcols,"pt",sc_col=set(QK+["Score Performance"])),unsafe_allow_html=True)
        st.markdown('<p class="stl a">Actions Recommandees — Performance</p>',unsafe_allow_html=True)
        st.markdown(html_actions_table(QK,avg_perf,CIBLE,ACT_MAP),unsafe_allow_html=True)

    # ---- TAB QUALITE ----
    with tabs[2]:
        st.markdown('<p class="stl q">Indicateurs de Qualite par Poste</p>',unsafe_allow_html=True)
        qcols=["Poste de travail"]+PK+["Score Qualite"]
        qrows=[]
        for p in posts:
            r={"Poste de travail":p,"_t":""}
            for k in PK: r[k]="%.1f%%"%ckdf.loc[p,k]
            r["Score Qualite"]="%.2f%%"%qscores[p]
            qrows.append(r)
        cible_row_q={"Poste de travail":"CIBLE","_t":"cible"}
        for k in PK: cible_row_q[k]="%.0f%%"%CIBLE.get(k,100)
        cible_row_q["Score Qualite"]="80.00%"
        qrows.append(cible_row_q)
        total_row_q={"Poste de travail":"Moyenne","_t":"total"}
        for k in PK: total_row_q[k]="%.1f%%"%float(ckdf[k].mean())
        total_row_q["Score Qualite"]="%.2f%%"%avg_q
        qrows.append(total_row_q)
        st.markdown(html_table(qrows,qcols,"qt",sc_col=set(PK+["Score Qualite"])),unsafe_allow_html=True)
        st.markdown('<p class="stl a">Actions Recommandees — Qualite</p>',unsafe_allow_html=True)
        st.markdown(html_actions_table(PK,avg_qual,CIBLE,ACT_MAP),unsafe_allow_html=True)

    # ---- TAB ANOMALIES (INVERSE : Postes en lignes, KPI en colonnes) ----
    with tabs[3]:
        st.markdown('<p class="stl a">Anomalies de Performance (nombre d\'OT hors cible par poste)</p>',unsafe_allow_html=True)
        st.markdown(html_ano_transposed(ano_perf, QK, posts, "ant", "🔴 Anomalies Performance — Postes × KPI"),unsafe_allow_html=True)
        st.markdown('<p class="stl a">Anomalies de Qualite (nombre d\'OT/Avis hors cible par poste)</p>',unsafe_allow_html=True)
        st.markdown(html_ano_transposed(ano_qual, PK, posts, "ant", "🔴 Anomalies Qualite — Postes × KPI"),unsafe_allow_html=True)
        # Resume global
        st.markdown('<p class="stl s">Resume Global des Anomalies</p>',unsafe_allow_html=True)
        st.markdown('<div class="cr">'
            '<div class="cc c1"><div class="cv">%d</div><div class="cl">Anomalies Performance</div></div>'
            '<div class="cc c4"><div class="cv">%d</div><div class="cl">Anomalies Qualite</div></div>'
            '<div class="cc c3"><div class="cv">%d</div><div class="cl">Total Anomalies</div></div>'
            '<div class="cc c2"><div class="cv">%d</div><div class="cl">Postes sans anomalie</div></div>'
            '</div>'%(tot_ano_p,tot_ano_q,tot_ano_p+tot_ano_q,
                sum(1 for p in posts if all(ano_perf[k].get(p,0)==0 for k in QK) and all(ano_qual[k].get(p,0)==0 for k in PK))),
            unsafe_allow_html=True)

    # ---- TAB BACKLOG (NOUVEAU) ----
    with tabs[4]:
        st.markdown('<p class="stl bl">📊 Backlog — Suivi des OT par Categorie</p>',unsafe_allow_html=True)
        df_work=kp['dfp']
        text_col=get_text_col(df_work)

        # --- OT OMS (Type 360 = Appels Systematiques) ---
        st.markdown('<div class="backlog-section"><div class="bs-title om">🔧 OT OMS (Ordres de Maintenance Systematique — Type 360)</div>',unsafe_allow_html=True)
        df_oms=df_work[df_work["_tw_num"]==360] if "_tw_num" in df_work.columns else pd.DataFrame()
        piv_oms=build_statut_pivot(df_oms, posts)
        st.markdown(html_statut_pivot(piv_oms, "omt"),unsafe_allow_html=True)
        if piv_oms["Total"].sum()>0:
            show_improved_pie_pair(piv_oms, "OT OMS")
        else:
            c1,c2=st.columns(2)
            with c1: st.markdown('<div class="es">Aucun OT OMS</div>',unsafe_allow_html=True)
            with c2: st.markdown('<div class="es">Aucun OT OMS</div>',unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)

        # --- OT Thermographie ---
        st.markdown('<div class="backlog-section"><div class="bs-title th">🌡️ OT Thermographie</div>',unsafe_allow_html=True)
        df_thermo=pd.DataFrame()
        if text_col:
            mask_thermo=df_work[text_col].astype(str).str.contains("THERMO|THERMOGRAPHIE|INFRAROUGE|IR ",case=False,na=False)
            df_thermo=df_work[mask_thermo]
        piv_thermo=build_statut_pivot(df_thermo, posts)
        st.markdown(html_statut_pivot(piv_thermo, "tht"),unsafe_allow_html=True)
        if piv_thermo["Total"].sum()>0:
            show_improved_pie_pair(piv_thermo, "OT Thermographie")
        else:
            c1,c2=st.columns(2)
            with c1: st.markdown('<div class="es">Aucun OT Thermographie</div>',unsafe_allow_html=True)
            with c2: st.markdown('<div class="es">Aucun OT Thermographie</div>',unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)

        # --- Tous les OT ---
        st.markdown('<div class="backlog-section"><div class="bs-title to">📋 Tous les OT</div>',unsafe_allow_html=True)
        piv_all=build_statut_pivot(df_work, posts)
        st.markdown(html_statut_pivot(piv_all, "tot"),unsafe_allow_html=True)
        if piv_all["Total"].sum()>0:
            show_improved_pie_pair(piv_all, "Tous les OT")
        else:
            c1,c2=st.columns(2)
            with c1: st.markdown('<div class="es">Aucun OT</div>',unsafe_allow_html=True)
            with c2: st.markdown('<div class="es">Aucun OT</div>',unsafe_allow_html=True)
        st.markdown('</div>',unsafe_allow_html=True)

        # --- Pie chart comparatif global ---
        st.markdown('<p class="stl c">Vue Comparative — Repartition Globale</p>',unsafe_allow_html=True)
        all_data={
            "OT OMS (Type 360)": int(piv_oms["Total"].sum()),
            "OT Thermographie": int(piv_thermo["Total"].sum()),
            "Autres OT": int(piv_all["Total"].sum()) - int(piv_oms["Total"].sum()) - int(piv_thermo["Total"].sum())
        }
        all_data={k:v for k,v in all_data.items() if v>0}
        if all_data:
            THRESH=5.0
            total_all=sum(all_data.values())
            al_l,al_v,al_p=[],[],[]
            oth=0
            for l,v in all_data.items():
                p=v/total_all*100
                if p<THRESH: oth+=v
                else: al_l.append(l); al_v.append(v); al_p.append(p)
            if oth>0: al_l.append("Autres"); al_v.append(oth); al_p.append(oth/total_all*100)
            comp_colors=["#805ad5","#e53e3e","#3182ce","#d69e2e","#38a169","#a0aec0"]
            min_ci=al_p.index(min(al_p))
            pull_c=[0.06 if i==min_ci else 0 for i in range(len(al_l))]
            fig_comp=go.Figure(go.Pie(labels=al_l,values=al_v,
                marker_colors=comp_colors[:len(al_l)],pull=pull_c,
                textposition='inside',textinfo='label+percent+value',textfont_size=12,hole=0.4,
                hovertemplate='<b>%%{label}</b><br>Nombre: %%{value}<br>Pourcentage: %%{percent}<extra></extra>'))
            fig_comp.update_layout(title=dict(text="Repartition par Categorie d'OT",font=dict(size=15,fontweight='bold')),
                margin=dict(t=55,b=10,l=10,r=10),height=400,
                legend=dict(font_size=12,orientation="h",yanchor="bottom",y=-0.12))
            st.plotly_chart(fig_comp,use_container_width=True)

    # ---- TAB SUIVI & EVOLUTION ----
    with tabs[5]:
        st.markdown('<p class="stl s">Suivi & Evolution des KPI</p>',unsafe_allow_html=True)
        kpis_dir="kpis"; filepath=os.path.join(kpis_dir,"indicateurs_kpis.xlsx")
        if os.path.exists(filepath):
            hist_df=load_historical_kpis(filepath)
            if hist_df.empty:
                st.markdown('<div class="es">Aucune donnee historique trouvee. Exportez les KPIs pour commencer le suivi.</div>',unsafe_allow_html=True)
            else:
                dates_avail=sorted(hist_df["Date"].unique())
                c_sel=st.columns(2)
                with c_sel[0]: d1=st.selectbox("Date precedente",dates_avail,index=max(0,len(dates_avail)-2),key="d1")
                with c_sel[1]: d2=st.selectbox("Date actuelle",dates_avail,index=len(dates_avail)-1,key="d2")
                if d1==d2: st.warning("Selectionnez deux dates differentes.")
                else:
                    var_df=calculate_variations(hist_df)
                    if var_df.empty:
                        st.markdown('<div class="es">Aucune variation calculee entre ces dates.</div>',unsafe_allow_html=True)
                    else:
                        filt_var=var_df[(var_df["Date precedente"]==d1)&(var_df["Date actuelle"]==d2)]
                        if filt_var.empty:
                            st.markdown('<div class="es">Aucune donnee pour ces dates.</div>',unsafe_allow_html=True)
                        else:
                            # Synthese Performance
                            st.markdown('<p class="stl p">Synthese Performance</p>',unsafe_allow_html=True)
                            synth_p={}
                            for _,r in filt_var[filt_var["Type"]=="Performance"].iterrows():
                                p=r["Poste"]; k=r["KPI"]
                                if p not in synth_p: synth_p[p]={}
                                sens="Amelioration" if ((r["Tendance"]=="hausse" and k not in LOWER_BETTER) or (r["Tendance"]=="baisse" and k in LOWER_BETTER)) else ("Degradation" if r["Tendance"]!="stabilite" else "Stable")
                                synth_p[p][k]={"prev":r["Valeur precedente"],"curr":r["Valeur actuelle"],"evol":{"sens":sens,"ecart":"%+.1f"%r["Ecart %"]}}
                            st.markdown(html_synthese_table(synth_p,QK,posts,"perf"),unsafe_allow_html=True)
                            # Synthese Qualite
                            st.markdown('<p class="stl q">Synthese Qualite</p>',unsafe_allow_html=True)
                            synth_q={}
                            for _,r in filt_var[filt_var["Type"]=="Qualite"].iterrows():
                                p=r["Poste"]; k=r["KPI"]
                                if p not in synth_q: synth_q[p]={}
                                sens="Amelioration" if ((r["Tendance"]=="hausse" and k not in LOWER_BETTER) or (r["Tendance"]=="baisse" and k in LOWER_BETTER)) else ("Degradation" if r["Tendance"]!="stabilite" else "Stable")
                                synth_q[p][k]={"prev":r["Valeur precedente"],"curr":r["Valeur actuelle"],"evol":{"sens":sens,"ecart":"%+.1f"%r["Ecart %"]}}
                            st.markdown(html_synthese_table(synth_q,PK,posts,"qual"),unsafe_allow_html=True)
                            # Journal
                            journ=generate_journal(filt_var)
                            if not journ.empty:
                                st.markdown('<p class="stl a">Journal des Evolutions Significatives (|ecart| >= 5%%)</p>',unsafe_allow_html=True)
                                jcols=["Poste","Type","KPI","Valeur precedente","Valeur actuelle","Ecart %","Sens"]
                                jrows=[]
                                for _,r in journ.iterrows():
                                    jrows.append({"Poste":r["Poste"],"Type":r["Type"],"KPI":r["KPI"],
                                        "Valeur precedente":"%.1f%%"%r["Valeur precedente"],"Valeur actuelle":"%.1f%%"%r["Valeur actuelle"],
                                        "Ecart %":"%+.1f%%"%r["Ecart %"],"Sens":r["Sens"],"_t":""})
                                st.markdown(html_table(jrows,jcols,"st"),unsafe_allow_html=True)
        else:
            st.markdown('<div class="es">Aucune donnee historique. Exportez les KPIs pour commencer le suivi.</div>',unsafe_allow_html=True)

    # ---- TAB CHANGELOG ----
    with tabs[6]:
        st.markdown('<p class="stl c">Historique des Versions & Ameliorations</p>',unsafe_allow_html=True)
        h='<div class="evol-timeline">'
        for ver in CHANGELOG:
            h+='<div class="evol-item"><div class="evol-ver">Version %s</div><div class="evol-date">%s</div>'%(ver["version"],ver["date"])
            for ch in ver["changes"]:
                h+='<div class="evol-change">%s</div>'%ch
            h+='</div>'
        h+='</div>'
        st.markdown(h,unsafe_allow_html=True)

    # ================================================================
    # EXCEL EXPORT
    # ================================================================
    if exp_excel:
        pcols_x=["Poste de travail"]+QK+["Score Performance"]
        prows_x=[{"Poste de travail":p,**{k:"%.1f%%"%ckdf.loc[p,k] for k in QK},"Score Performance":"%.2f%%"%pscores[p]} for p in posts]
        prows_x.append({"Poste de travail":"CIBLE",**{k:"%.0f%%"%CIBLE.get(k,100) for k in QK},"Score Performance":"80.00%"})
        prows_x.append({"Poste de travail":"Moyenne",**{k:"%.1f%%"%float(ckdf[k].mean()) for k in QK},"Score Performance":"%.2f%%"%avg_p})
        qcols_x=["Poste de travail"]+PK+["Score Qualite"]
        qrows_x=[{"Poste de travail":p,**{k:"%.1f%%"%ckdf.loc[p,k] for k in PK},"Score Qualite":"%.2f%%"%qscores[p]} for p in posts]
        qrows_x.append({"Poste de travail":"CIBLE",**{k:"%.0f%%"%CIBLE.get(k,100) for k in PK},"Score Qualite":"80.00%"})
        qrows_x.append({"Poste de travail":"Moyenne",**{k:"%.1f%%"%float(ckdf[k].mean()) for k in PK},"Score Qualite":"%.2f%%"%avg_q})
        # Anomalies inversees pour Excel
        ano_p_cols=["Poste de travail"]+QK+["Total"]
        ano_p_rows=[{"Poste de travail":p,**{k:ano_perf[k].get(p,0) for k in QK},"Total":sum(ano_perf[k].get(p,0) for k in QK)} for p in posts]
        ano_p_rows.append({"Poste de travail":"Total",**{k:sum(ano_perf[k].get(pp,0) for pp in posts) for k in QK},"Total":tot_ano_p})
        ano_q_cols=["Poste de travail"]+PK+["Total"]
        ano_q_rows=[{"Poste de travail":p,**{k:ano_qual[k].get(p,0) for k in PK},"Total":sum(ano_qual[k].get(p,0) for k in PK)} for p in posts]
        ano_q_rows.append({"Poste de travail":"Total",**{k:sum(ano_qual[k].get(pp,0) for pp in posts) for k in PK},"Total":tot_ano_q})
        save_kpis_to_excel(prows_x,pcols_x,qrows_x,qcols_x,ano_p_rows,ano_p_cols,ano_q_rows,ano_q_cols,fichier_date)
        st.toast("✅ Export Excel effectue!")

# ============================================================
if __name__ == "__main__":
    main()
